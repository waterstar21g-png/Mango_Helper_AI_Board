"""
P3_필터_갱신 — 더망고 검색필터(저장조건) 화면의 저장상품수 갱신.

로직 순서 (1~7단계 — 실행 로그 MAIN 도 이 7단계만 표시):
1) 망고 수집 URL 링크로 진입 (입력 항목 정보로 제공)
2) ★엑셀 자료를 첫행부터 순차로 읽어, 그 "URL KEY"로 망고의 같은 "URL KEY" 행을 찾는다
   (엑셀이 기준 — 망고 목록을 훑는 방향이 아니다. 그 주소로 화면을 열지도 않는다)
3) 망고 필터와 엑셀 필터가 불일치하면 다음 엑셀행으로 진행 (매칭 안 되는 행은 로그에 남기지 않음)
4) 상품노출수(카드수) 추출 — ★전부 주석처리. "URL 검색" 주소로는 어떤 화면도 불러오지
   않으며, 상품수는 엑셀 값만 사용한다 (추후 완성본에서 주석 해제)
5) LABEL "수집조건수정" 버튼 클릭 → 팝업의 "저장상품수"(검색결과 상위 [ ]개) 필드에
   엑셀 상품수 기준 값 입력 → 하단 LABEL "저장하기" 버튼 클릭 (★'저장'이 아닌 '저장하기')
6) 팝업되는 "수정되었습니다" 메세지 하단 LABEL "확인" 버튼 클릭
7) 2~6단계를 다음 행에 대해 반복

처리 속도 (요건):
- 처음 5개 처리행: 5)수집조건수정 클릭 · 5)저장하기 클릭 · 6)확인 클릭 뒤 각각 3초 대기
  → 실제 망고 화면 변화를 그대로 확인
- 6번째 처리행부터: 지연 없이 가장 빠른 속도

로그 원칙:
- 위 1)~7) 단계 로그만 남김 (그 외 화면전환/준비 등 세부 로그는 억제)
- 매칭된 행만 로그에 남김 — 필터/URL 불일치 행은 로그에 남기지 않음
- 성공/실패 여부와 필터명·KEY(URL)·상품수 등 핵심 정보를 포함
- 오류 발생 시 원인 파악에 필요한 상세 정보(사유·KEY·필터명·목표값)를 포함

버튼 클릭은 좌표 추정이 아니라 LABEL 텍스트로 찾은 실제 DOM 버튼 요소를 직접 클릭한다
(href 재시도 금지).

사용법:
  python update_filters.py 엑셀.xlsx --mango-url "https://..."
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
P2_DIR = ROOT / "P2"
P1_101_DIR = ROOT / "P1_101"
if str(P2_DIR) not in sys.path:
    sys.path.insert(0, str(P2_DIR))
if str(P1_101_DIR) not in sys.path:
    sys.path.insert(0, str(P1_101_DIR))

ProgressFn = Callable[[str, str], None]

STOP_FLAG_PATH = Path(__file__).resolve().parent / ".filter_stop"
P3_RUN_LOG_DIR = Path(__file__).resolve().parent / "run-logs"
P3_SHOT_MARK = "##P3SHOT##"  # ##P3SHOT##<path>##<label>
# ★요건: 망고 행 「URL 검색」 주소로는 어떤 화면도 열지 않는다(절대 금지) — False 고정.
#   4) 상품노출수(카드수) 추출 호출부는 전부 주석처리되어 있고, 「URL 검색」 주소는
#   엑셀자료 비교(KEY)에만 쓴다. 함수(click_demango_row_url·browse_store_count_cards)는
#   추후 완성본을 위해 남겨 두지만 호출하지 않는다.
ENABLE_STORE_COUNT_CALL = False

# ★요건: 보드 「더망고 URL」초기값 (검색필터·저장조건 화면)
DEFAULT_MANGO_URL = (
    "https://tmg1898.cafe24.com/mall/admin/shop/getGoodsCategory.php"
    "?pmode=filter_delete&uids=&pg=1&date_type=modify"
    "&start_yy=2026&start_mm=8&start_dd=12"
    "&end_yy=2026&end_mm=8&end_dd=12"
    "&site_id=zara_de&sales_yn=&sch_keyword="
    "&ft_num=all&ft_show=&ft_sort=modify_asc"
)
LAST_MANGO_URL_PATH = Path(__file__).resolve().parent / ".last_mango_url"


def load_mango_url_default() -> str:
    """★고정 초기값 — 검색필터(저장조건) getGoodsCategory.php URL."""
    return DEFAULT_MANGO_URL


def save_mango_url(url: str) -> None:
    """더망고 URL 기억 파일에 저장 (보드 표시용). 비어 있으면 초기값 저장."""
    u = (url or "").strip() or DEFAULT_MANGO_URL
    if not u.lower().startswith("http"):
        u = DEFAULT_MANGO_URL
    try:
        LAST_MANGO_URL_PATH.write_text(u + "\n", encoding="utf-8")
    except Exception:
        pass

URL_HEADERS = (
    "검색필터 URL",
    "최종 카테고리 URL주소",
    "최종 카테고리 URL",
    "카테고리 URL",
    "URL주소",
    "URL",
    "url",
)
# ★요건(2026-08-20): 망고 "검색필터명"에 "최종 카테고리명"이 들어가므로,
# 망고 행 매칭도 같은 열을 우선으로 본다 ("상위 최종 카테고리명"은 옛 엑셀 대비 폴백).
FILTER_HEADERS = (
    "검색필터",
    "검색필터명",
    "최종 카테고리명",
    "상위 최종 카테고리명",
    "카테고리명",
)
COLLECTIBLE_HEADERS = ("상품수집가능개수", "총상품수")


@dataclass
class ExcelRow:
    excel_row: int  # 1-based sheet row
    url: str
    filter_name: str
    collectible: int


@dataclass
class RunResult:
    ok: bool
    total_demango: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0
    errors: list[str] = field(default_factory=list)


# 실행로그: 프로그램 로직 순서·단계(1~7단계)만 남기고 나머지는 전부 억제.
#   1) 망고 수집 URL 링크로 진입
#   2) 망고 URL 클릭 → 엑셀 KEY 매칭
#   3) 필터 불일치 → 다음 행
#   4) 상품노출수(카드수) 추출 — 현재는 건너뛰고 수행 (추후 완성본에서 추가)
#   5) LABEL '수집조건수정' 클릭 → 저장상품수 입력 → '저장하기' 클릭
#   6) '수정되었습니다' 확인 클릭
#   7) 2~6단계 반복 (다음 행)
P3_MAJOR_LOG_ONLY = True


def _is_major_log(step: str, message: str) -> bool:
    """오류/중단/완료 요약만 자동으로 MAIN 취급.

    ★1)~7) 단계 로그는 전부 호출부에서 major=True 를 명시한다. 메시지 첫머리의
    "N)" 로 자동판정하면 스크린샷 라벨("6)확인 클릭 실패 -> ....png")처럼 단계번호로
    시작하는 세부 로그까지 MAIN 에 섞여 버린다(요건: MAIN은 7단계만).
    """
    return (step or "").strip() in ("오류", "중단", "완료")


# ★P2와 동일 프로토콜(board/log_protocol.py) — MAIN(1~7단계)/SUB(세부정보)/SUBSHOT(샷)
# 발생(seq)마다 MAIN 1행 + 그 seq 에 딸린 SUB·SUBSHOT 여러 행으로 보드에 표시된다.
_SEQ_STATE = {"seq": 0, "cur_seq": 0, "cur_n": 0}

# main 그리드 표시용 — 1)~7) 은 그대로, 오류/완료/중단은 숫자 밖 코드로 구분
_META_STEP_N = {"오류": 90, "완료": 91, "중단": 92}

# 단계 MAIN 로그는 그 단계의 동작을 끝낸 뒤에 남기므로, 동작 중 쌓인 세부내용은
# 아직 그 단계의 seq 를 모른다. 다른 단계의 세부내용이 앞 단계 SUB로 섞이지 않게
# 여기에 담아 두고, 해당 단계 MAIN 이 나올 때 그 seq 로 흘려보낸다.
_PENDING_SUB: list[tuple[str, str]] = []


def _next_seq() -> int:
    _SEQ_STATE["seq"] += 1
    return _SEQ_STATE["seq"]


def _current_seq() -> int:
    if not _SEQ_STATE["cur_seq"]:
        _SEQ_STATE["cur_seq"] = _next_seq()
    return _SEQ_STATE["cur_seq"]


# 보드 실행로그에서 그 줄을 적색으로 구분 표시하게 하는 표식
#   (그리드 셀 안에서 일부 글자만 색을 바꿀 수 없어 행 단위로 구분한다)
RED_PREFIX = "##RED##"


def _red(message: str) -> str:
    return f"{RED_PREFIX}{message}"


def _emit_progress_row(ordinal: int) -> None:
    """보드 엑셀 목록에 진행 화살표(▶)를 표시할 행 번호 (1-based, 0=해제)."""
    print(f"##META##진행##{int(ordinal or 0)}", flush=True)


def _step_no_of(message: str) -> int:
    """"N) ..." 형태면 N, 그 외(단계 표기 없음)는 0."""
    m = re.match(r"^(\d+)\)", message or "")
    return int(m.group(1)) if m else 0


def _step_no_in(message: str) -> int:
    """메시지 안(앞부분)의 단계 표기 N) 을 찾는다 — 없으면 0.

    실패 로그는 "행106 · 6) '확인' 버튼 클릭 실패 …" 처럼 행번호가 앞에 붙으므로
    첫머리 매칭만으로는 몇 단계에서 실패했는지 알 수 없다.
    """
    m = re.search(r"(?:^|[·\s])(\d)\)", message or "")
    if not m:
        return 0
    n = int(m.group(1))
    return n if 1 <= n <= 7 else 0


def _strip_step_no(message: str) -> str:
    """SUB 출력용 — 앞머리 "N) " 를 뗀다 (단계번호는 MAIN 에만)."""
    return re.sub(r"^\d+\)\s*", "", message or "")


def _flush_pending_sub(seq: int) -> None:
    while _PENDING_SUB:
        ts, msg = _PENDING_SUB.pop(0)
        print(f"[{ts}] ##SUB##{seq}##{_strip_step_no(msg)}", flush=True)


def _log(
    progress: ProgressFn | None,
    step: str,
    message: str,
    *,
    major: bool | None = None,
) -> None:
    """실행로그 — P2와 동일 MAIN/SUB 프로토콜로 표준출력에 남긴다.

    - MAIN(##MAIN##seq##n##msg): 1)~7) 단계 → 보드 MAIN 그리드 (7단계만)
    - SUB(##SUB##seq##msg): 그 단계의 세부내용 → 보드 SUB/스크린샷 그리드
    - 오류/중단/완료 요약은 MAIN 행을 새로 만들지 않고(보드가 SUB로 표시) 진행 중
      단계의 문맥을 유지한다.
    """
    s = (step or "").strip()
    m = (message or "").strip()
    if not m:
        return
    is_main = True if major is True else (
        False if major is False else _is_major_log(s, m)
    )
    ts = time.strftime("%H:%M:%S")
    if is_main:
        n = _step_no_of(m)
        if not n and s == "오류":
            # 실패한 단계를 MAIN 에서 바로 찾을 수 있게 그 단계 행으로 남긴다
            n = _step_no_in(m)
        if not n:
            n = _META_STEP_N.get(s, 0)
        seq = _next_seq()
        if 1 <= n <= 7:
            _SEQ_STATE["cur_seq"] = seq
            _SEQ_STATE["cur_n"] = n
            print(f"[{ts}] ##MAIN##{seq}##{n}##{m}", flush=True)
            _flush_pending_sub(seq)
        else:
            # 단계를 특정할 수 없는 오류/완료/중단 요약 — 진행 중 단계에 남은
            # 세부내용을 먼저 흘려보낸 뒤 요약을 남긴다(보드는 SUB에 표시)
            if _SEQ_STATE["cur_seq"]:
                _flush_pending_sub(int(_SEQ_STATE["cur_seq"]))
            print(f"[{ts}] ##MAIN##{seq}##{n}##{m}", flush=True)
    else:
        cur_seq = int(_SEQ_STATE["cur_seq"] or 0)
        step_no = _step_no_of(m)
        if not cur_seq or (step_no and step_no != int(_SEQ_STATE["cur_n"] or 0)):
            _PENDING_SUB.append((ts, m))
        else:
            # ★MAIN 과 섞여 보이지 않게 SUB 는 단계번호 접두어를 떼고 출력한다
            #   (그룹 판정에만 쓰고 화면에는 세부내용만 보인다)
            print(f"[{ts}] ##SUB##{cur_seq}##{_strip_step_no(m)}", flush=True)
    if progress:
        progress(s, m)


def _emit_subshot(path, label: str) -> None:
    """스크린샷을 현재 MAIN 발생(seq)에 딸린 SUB 항목으로 보드에 알린다."""
    ts = time.strftime("%H:%M:%S")
    seq = _current_seq()
    print(f"[{ts}] ##SUBSHOT##{seq}##{path}##{label}", flush=True)


# ★P2와 동일: 실제 Chrome(CDP) 창을 OS 앞으로 가져와 동작을 보여 줌
STEP_VIEW_DWELL_SEC = 0.0  # ★컴퓨터 속도 — 화면 표시만 하고 대기 없음

# ★요건: 수집조건수정 → 저장상품수 입력 → 저장하기 → 확인 순서만 지키고,
#   중간 대기 없이 컴퓨터 속도로 진행한다.
# ★요건: 지연 없이 컴퓨터 속도로 진행 (0 = 대기 없음)
SLOW_DEMO_ROWS = 0
SLOW_DEMO_DELAY_SEC = 0.0


def step_delay_sec(processed_no: int) -> float:
    """처리 순번(1-based) → 각 동작 뒤 지연(초). 6번째 처리행부터 0(최고속)."""
    if 1 <= int(processed_no or 0) <= SLOW_DEMO_ROWS:
        return SLOW_DEMO_DELAY_SEC
    return 0.0


def describe_page_state(page) -> str:
    """화면 요약 한 줄 — URL · 제목 · 탭수."""
    url = ""
    title = ""
    tabs = 0
    try:
        url = (page.url or "").strip()
    except Exception:
        url = "(url읽기실패)"
    try:
        title = (page.title() or "").strip()
    except Exception:
        title = ""
    try:
        tabs = len(list(page.context.pages))
    except Exception:
        tabs = 0
    parts = [f"url={url[:140]}"]
    if title:
        parts.append(f"title={title[:80]}")
    if tabs:
        parts.append(f"tabs={tabs}")
    return " · ".join(parts)


def _set_chrome_window_state(page, *, window_state: str = "normal") -> bool:
    """CDP로 Chrome 창 상태 변경 (normal / maximized). 성공 시 True."""
    if page is None:
        return False
    try:
        session = page.context.new_cdp_session(page)
    except Exception:
        return False
    ok = False
    try:
        info = session.send("Browser.getWindowForTarget")
        wid = info.get("windowId") if isinstance(info, dict) else None
        if wid is None:
            return False
        # maximized 는 일부 환경에서 normal→maximized 순이 안정적
        if window_state == "maximized":
            try:
                session.send(
                    "Browser.setWindowBounds",
                    {"windowId": wid, "bounds": {"windowState": "normal"}},
                )
            except Exception:
                pass
        session.send(
            "Browser.setWindowBounds",
            {"windowId": wid, "bounds": {"windowState": window_state}},
        )
        ok = True
    except Exception:
        ok = False
    finally:
        try:
            session.detach()
        except Exception:
            pass
    return ok


def _foreground_chrome_window_windows(*, maximize: bool = False) -> None:
    """Windows: Chrome/더망고 창을 앞으로(필요 시 최대화)."""
    if os.name != "nt":
        return
    try:
        import ctypes

        user32 = ctypes.windll.user32  # type: ignore[attr-defined]
        SW_RESTORE = 9
        SW_MAXIMIZE = 3
        show_cmd = SW_MAXIMIZE if maximize else SW_RESTORE
        hwnd = user32.GetForegroundWindow()
        found = ctypes.c_void_p(0)

        @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
        def _enum(h, _l):  # noqa: ANN001
            length = user32.GetWindowTextLengthW(h)
            if length <= 0:
                return True
            buf = ctypes.create_unicode_buffer(length + 1)
            user32.GetWindowTextW(h, buf, length + 1)
            title = buf.value or ""
            if "Chrome" in title or "더망고" in title or "cafe24" in title.lower():
                found.value = h
                return False
            return True

        user32.EnumWindows(_enum, 0)
        target = found.value or hwnd
        if target:
            user32.ShowWindow(target, show_cmd)
            user32.SetForegroundWindow(target)
    except Exception:
        pass


def _activate_chrome_window(page) -> None:
    """최소화/뒤로 간 Chrome 창을 복원·앞으로 (CDP Browser.setWindowBounds)."""
    _set_chrome_window_state(page, window_state="normal")
    _foreground_chrome_window_windows(maximize=False)


def maximize_mango_chrome_window(
    page,
    progress: ProgressFn | None = None,
    *,
    dwell_s: float = 0.8,
) -> None:
    """망고 Chrome 창을 반드시 최대화한다 (목록 복귀 후 · 행 재탐색 전)."""
    if page is None:
        return
    try:
        if hasattr(page, "is_closed") and page.is_closed():
            _log(progress, "화면", "망고 창 최대화 — 창이 이미 닫힘")
            return
    except Exception:
        pass
    cdp_ok = _set_chrome_window_state(page, window_state="maximized")
    _foreground_chrome_window_windows(maximize=True)
    try:
        page.bring_to_front()
    except Exception:
        pass
    try:
        page.evaluate("() => { try { window.focus(); } catch (e) {} }")
    except Exception:
        pass
    state = ""
    try:
        state = describe_page_state(page)
    except Exception:
        state = ""
    msg = "망고 창 최대화 필수"
    if cdp_ok:
        msg += " OK"
    else:
        msg += " (CDP 실패·Windows 최대화 시도)"
    if state:
        msg += f" · {state}"
    _log(progress, "화면", msg, major=False)
    if dwell_s > 0:
        time.sleep(dwell_s)


def reveal_browser_page(
    page,
    progress: ProgressFn | None,
    *,
    step_no: str,
    action: str,
    dwell_s: float | None = None,
) -> None:
    """P2와 동일 — 망고 Chrome 창·팝업을 화면에 보이게 한다."""
    if page is None:
        return
    dwell = STEP_VIEW_DWELL_SEC if dwell_s is None else max(0.0, float(dwell_s))
    try:
        if hasattr(page, "is_closed") and page.is_closed():
            _log(progress, "화면", f"{step_no}) {action} — 창이 이미 닫힘")
            return
    except Exception:
        pass
    _activate_chrome_window(page)
    try:
        page.bring_to_front()
    except Exception:
        pass
    try:
        page.evaluate("() => { try { window.focus(); } catch (e) {} }")
    except Exception:
        pass
    state = ""
    try:
        state = describe_page_state(page)
    except Exception:
        state = ""
    msg = f"{step_no}) {action}"
    if state:
        msg += f" · {state}"
    # 화면 상세는 주요 로그에서 제외 (동작만 수행)
    _log(progress, "화면", msg + " ← 망고 Chrome 창 표시", major=False)
    if dwell > 0:
        time.sleep(dwell)


def attach_mango_browser_like_p2(p2, playwright, *, progress: ProgressFn | None = None):
    """★P2와 동일: connect_browser 로 실제 Chrome을 띄우거나 연결한 뒤 창을 앞으로.

    로그인 대기는 하지 않는다. 이후 검색필터 URL로 이동한다.
    """
    _log(
        progress,
        "준비",
        "P2와 동일 — 망고 Chrome(CDP) 연결/실행 · 화면에 창 표시",
    )
    browser, page = p2.connect_browser(playwright)
    if hasattr(p2, "refresh_if_closed"):
        page = p2.refresh_if_closed(page)
    try:
        page.set_default_timeout(120_000)
    except Exception:
        pass
    reveal_browser_page(
        page,
        progress,
        step_no="0",
        action="망고 Chrome 연동 창 표시",
        dwell_s=0.3,
    )
    try:
        cur = (page.url or "").strip()
    except Exception:
        cur = ""
    _log(progress, "준비", f"연결 직후 URL={cur[:160] or '(없음)'} → 검색필터 URL로 이동")
    return browser, page


# 하위호환 별칭
attach_current_mango_page = attach_mango_browser_like_p2


def clear_stop_flag() -> None:
    try:
        STOP_FLAG_PATH.unlink(missing_ok=True)  # type: ignore[call-arg]
    except TypeError:
        if STOP_FLAG_PATH.exists():
            try:
                STOP_FLAG_PATH.unlink()
            except OSError:
                pass
    except OSError:
        pass


def stop_requested() -> bool:
    return STOP_FLAG_PATH.is_file()


def normalize_url(url: str) -> str:
    """URL 비교용 정규화 (끝 슬래시·쿼리 일부 무시하지 않음, 스킴/호스트 소문자)."""
    s = (url or "").strip()
    if not s:
        return ""
    try:
        p = urlparse(s)
        path = unquote(p.path or "")
        if path.endswith("/") and len(path) > 1:
            path = path[:-1]
        netloc = (p.netloc or "").lower()
        scheme = (p.scheme or "https").lower()
        query = p.query or ""
        return f"{scheme}://{netloc}{path}" + (f"?{query}" if query else "")
    except Exception:
        return s.rstrip("/").lower()


def url_stem(url: str) -> str:
    """쿼리(?...)를 뗀 정규화 주소 — 행 매칭 비교용 (부분일치 금지)."""
    return normalize_url(url).split("?")[0]


def map_save_count(collectible: int) -> int:
    """상품수집가능개수 → 더망고 저장상품수."""
    n = max(0, int(collectible))
    if n <= 200:
        return n
    if n <= 500:
        return 300
    return 400


def _header_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    lowered = [(h or "").strip() for h in headers]
    for cand in candidates:
        for i, h in enumerate(lowered):
            if h == cand:
                return i
    for cand in candidates:
        c = cand.lower()
        for i, h in enumerate(lowered):
            if c in h.lower():
                return i
    return None


def _parse_int(raw) -> int:
    if raw is None:
        return 0
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return max(0, int(raw))
    s = re.sub(r"[^\d]", "", str(raw))
    return int(s) if s else 0


def read_excel_rows(path: Path) -> list[ExcelRow]:
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_vals = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("엑셀이 비어 있습니다.")
    headers = [str(h).strip() if h is not None else "" for h in header_vals]
    url_i = _header_index(headers, URL_HEADERS)
    if url_i is None:
        wb.close()
        raise ValueError(
            "URL 열 없음 — '검색필터 URL' 또는 '최종 카테고리 URL주소' 필요"
        )
    filter_i = _header_index(headers, FILTER_HEADERS)
    coll_i = _header_index(headers, COLLECTIBLE_HEADERS)
    out: list[ExcelRow] = []
    for offset, vals in enumerate(rows_iter, start=2):
        cells = list(vals) if vals else []
        url = str(cells[url_i] or "").strip() if url_i < len(cells) else ""
        if not url or not url.lower().startswith("http"):
            continue
        fname = ""
        if filter_i is not None and filter_i < len(cells):
            fname = str(cells[filter_i] or "").strip()
        coll = 0
        if coll_i is not None and coll_i < len(cells):
            coll = _parse_int(cells[coll_i])
        out.append(
            ExcelRow(
                excel_row=offset,
                url=url,
                filter_name=fname,
                collectible=coll,
            )
        )
    wb.close()
    return out


def excel_by_url(rows: list[ExcelRow]) -> dict[str, ExcelRow]:
    """엑셀 URL → 행. 조회 KEY는 더망고 URL(정규화) 기준."""
    m: dict[str, ExcelRow] = {}
    for r in rows:
        key = normalize_url(r.url)
        if key and key not in m:
            m[key] = r
    return m


def find_excel_by_demango_url(
    by_url: dict[str, ExcelRow], demango_url: str
) -> ExcelRow | None:
    """더망고 URL로 엑셀 행을 찾는다 (역방향 조회 — 호환용)."""
    key = normalize_url(demango_url)
    if not key:
        return None
    return by_url.get(key)


def _compact(text: str) -> str:
    return "".join((text or "").split())


def row_done_key(url: str, filter_name: str, fuid: str = "") -> tuple[str, str]:
    """처리 완료 행 식별키.

    행 고유 id(수집조건수정 버튼의 ps_fuid)가 있으면 그것만으로 유일하다.
    없으면 (정규화 URL, 공백제거 필터이름) 을 쓴다.
    목록이 수정일 정렬로 재배치되어도 유효하도록 DOM index 는 쓰지 않는다.
    """
    fid = _compact(fuid)
    if fid:
        return ("fuid", fid)
    return (normalize_url(url), _compact(filter_name))


def find_demango_rows_for_excel(
    page,
    excel_row: ExcelRow,
    *,
    progress: ProgressFn | None = None,
    done_keys: set[tuple[str, str]] | None = None,
) -> list[dict]:
    """★요건2: 엑셀 「URL KEY」로 망고 목록에서 그 행들을 찾는다 (여러 개면 전부).

    - 매번 현재 목록을 다시 스캔한다 (수정일 정렬로 순서가 바뀌어도 정확한 행을 잡음)
    - 같은 URL이 망고에 여러 행이면 **전체를 갱신 대상으로** 돌려준다
      (필터이름이 엑셀과 같은 행을 앞에 둔다)
    - 이미 처리한 행(done_keys)은 제외한다
    """
    try:
        rows = list_demango_rows(page)
    except Exception as e:  # noqa: BLE001
        _log(progress, "경고", f"망고 목록 스캔 실패: {str(e).split(chr(10))[0][:100]}")
        return []

    done = done_keys or set()
    key = normalize_url(excel_row.url)
    # ★부분일치(substring)는 쓰지 않는다 — 짧은/잘린 엑셀 URL이 관계없는 행까지
    #   싹 잡아 엉뚱한 행을 갱신해 버린다. 쿼리만 뗀 주소의 "완전일치"만 인정.
    stem = url_stem(excel_row.url)
    want_filter = _compact(excel_row.filter_name)

    def _pool(match) -> list[dict]:
        out: list[dict] = []
        for r in rows:
            r_url = (r.get("url") or "").strip()
            if row_done_key(
                r_url, str(r.get("filterName") or ""), str(r.get("fuid") or "")
            ) in done:
                continue
            try:
                if match(r_url):
                    out.append(r)
            except Exception:  # noqa: BLE001
                continue
        return out

    pool = _pool(lambda u: bool(key) and normalize_url(u) == key)
    if not pool:
        pool = _pool(lambda u: bool(stem) and url_stem(u) == stem)
    if not pool:
        return []

    def _rank(r: dict) -> int:
        name = str(r.get("filterName") or "")
        if want_filter and _compact(name) == want_filter:
            return 0
        if filters_equal(excel_row.filter_name, name):
            return 1
        return 2

    return sorted(pool, key=_rank)


def find_demango_row_for_excel(
    page,
    excel_row: ExcelRow,
    *,
    progress: ProgressFn | None = None,
    done_keys: set[tuple[str, str]] | None = None,
) -> dict | None:
    """엑셀 「URL KEY」로 찾은 망고 행 중 첫 행 (없으면 None)."""
    found = find_demango_rows_for_excel(
        page, excel_row, progress=progress, done_keys=done_keys
    )
    return found[0] if found else None


def filters_equal(excel_filter: str, demango_filter: str) -> bool:
    """검색필터 비교.

    1) 그대로 비교
    2) 불일치이고 엑셀 검색필터 값 *중간*에 공백이 있으면 공백→'_' 치환 후 재비교
    """
    a = (excel_filter or "").strip()
    b = (demango_filter or "").strip()
    if a == b:
        return True
    # 중간에 공백이 있는 경우만 (앞뒤 trim 이후에도 공백 존재)
    if " " in a:
        if a.replace(" ", "_") == b:
            return True
    return False


def filter_compare_note(excel_filter: str, demango_filter: str) -> str:
    """로그용 — 공백→_ 재비교로 맞은 경우 메모."""
    a = (excel_filter or "").strip()
    b = (demango_filter or "").strip()
    if a == b:
        return ""
    if " " in a and a.replace(" ", "_") == b:
        return "엑셀 중간공백→'_' 재비교 일치"
    return ""


def navigate_mango_url(
    page,
    mango_url: str,
    *,
    progress: ProgressFn | None,
    p2=None,
):
    """검색필터(저장조건) URL로 이동 — 로그인 대기 없음. P2 safe_goto 사용.

    ★1) 망고 수집 URL 링크로 진입한다 (입력 항목 정보로 제공).
    """
    url = (mango_url or "").strip() or DEFAULT_MANGO_URL
    _log(progress, "로직", "1) 망고 접속", major=True)
    _log(progress, "준비", f"접속 URL={url[:160]}")
    for attempt in range(1, 3):
        try:
            if p2 is not None and hasattr(p2, "safe_goto"):
                p2.safe_goto(page, url)
                if hasattr(p2, "refresh_if_closed"):
                    page = p2.refresh_if_closed(page)
            else:
                page.goto(url, wait_until="domcontentloaded", timeout=90_000)
        except Exception as e:  # noqa: BLE001
            _log(
                progress,
                "경고",
                f"0) URL 이동 예외({attempt}/2): {str(e).split(chr(10))[0][:120]}",
            )
        time.sleep(0.2)
        try:
            cur = (page.url or "").strip()
        except Exception:
            cur = ""
        _log(progress, "준비", f"이동 후 화면 URL={cur[:180] or '(없음)'}")
        # admin.php 등에 남으면 재시도
        if "getGoodsCategory" in cur or "filter" in cur.lower():
            break
        if attempt < 2:
            _log(progress, "경고", "검색필터 화면 미도달 — URL 재이동")
    reveal_browser_page(
        page, progress, step_no="0", action="검색필터(저장조건) 화면 표시", dwell_s=0.2
    )
    try:
        body = page.locator("body").inner_text(timeout=3000) or ""
        if re.search(r"검색\s*필터|저장\s*조건|수집\s*조건|필터이름", body):
            _log(progress, "확인", "검색필터/저장조건 화면 문구 확인")
        else:
            _log(progress, "확인", "화면 문구 미검출 — 지정 URL로 계속 진행")
    except Exception:
        pass
    return page



# 더망고 목록 행 스캔 JS (스크린샷 컬럼 구조 기준) — 테스트에서도 재사용
LIST_DEMANGO_ROWS_JS = r"""() => {
  const out = [];
  const trs = Array.from(document.querySelectorAll('table tr, form tr, tr'));
  let filterCol = -1;
  let condCol = -1;
  let headerIdx = -1;

  const isBadFilterName = (v) => {
    if (!v) return true;
    const s = String(v).trim();
    if (!s) return true;
    if (/^https?:/i.test(s)) return true;
    if (/^\d+$/.test(s)) return true;
    if (/\.com(\/|$)/i.test(s)) return true;           // Zara.com/de 등 사이트열
    if (/\d+\s*개\s*\/\s*\d+\s*개/.test(s)) return true;
    if (/^\d{4}-\d{2}-\d{2}/.test(s)) return true;
    if (/URL\s*검색|수집조건|전체저장|상품확인|검색필터관리|별도관리/.test(s)) return true;
    if (s.length > 120) return true;
    return false;
  };

  const readInputValue = (inp) => {
    if (!inp) return '';
    return (inp.value || inp.getAttribute('value') || '').trim();
  };

  // 헤더: '필터이름(수정가능)' / '검색필터(저장조건)'
  for (let i = 0; i < trs.length; i++) {
    const cells = Array.from(trs[i].querySelectorAll('th, td'));
    if (cells.length < 2) continue;
    const labels = cells.map(c => (c.innerText || '').replace(/\s+/g, ''));
    const fi = labels.findIndex(t => t.includes('필터이름'));
    const ci = labels.findIndex(t =>
      t.includes('검색필터') && (t.includes('저장조건') || t.includes('저장'))
    );
    if (fi >= 0 || ci >= 0) {
      filterCol = fi;
      condCol = ci;
      headerIdx = i;
      break;
    }
  }

  for (let i = 0; i < trs.length; i++) {
    if (i === headerIdx) continue;
    const tr = trs[i];
    const t = (tr.innerText || '').replace(/\s+/g, ' ').trim();
    if (!t) continue;
    if (!/URL\s*검색|수집\s*조건\s*수정|https?:\/\//i.test(t)) continue;
    if (/필터이름\(수정가능\)/.test(t) && !/https?:\/\//i.test(t)) continue;

    const cells = Array.from(tr.querySelectorAll(':scope > td, :scope > th'));
    const cellsFallback = cells.length ? cells : Array.from(tr.querySelectorAll('td, th'));
    let filterName = '';
    let url = '';
    let editHref = '';
    let hasEdit = false;

    // 1) 검색필터 필드값 = '필터이름(수정가능)' 열의 <input> value
    const filterCell = (filterCol >= 0 && filterCol < cellsFallback.length)
      ? cellsFallback[filterCol] : null;
    if (filterCell) {
      const inp = filterCell.querySelector(
        'input[type="text"], input[type="search"], input:not([type])'
      );
      filterName = readInputValue(inp);
      if (isBadFilterName(filterName)) filterName = '';
    }
    if (!filterName) {
      // 폴백: 행 내 텍스트 input (사이트·숫자·URL 제외)
      const inputs = Array.from(tr.querySelectorAll('input')).filter(inp => {
        const ty = (inp.getAttribute('type') || 'text').toLowerCase();
        return ty === 'text' || ty === 'search' || ty === '';
      });
      for (const inp of inputs) {
        const v = readInputValue(inp);
        if (isBadFilterName(v)) continue;
        filterName = v;
        break;
      }
    }

    // 2) URL = '검색필터(저장조건)' 열의 'URL 검색:' 뒤 링크/텍스트
    const condCell = (condCol >= 0 && condCol < cellsFallback.length)
      ? cellsFallback[condCol] : tr;
    const scope = condCell || tr;
    const aHttp = scope.querySelector('a[href^="http"], a[href*="://"]');
    if (aHttp) {
      url = (aHttp.href || aHttp.getAttribute('href') || '').trim();
    }
    if (!url) {
      const raw = (scope.innerText || scope.textContent || '');
      const m = raw.match(/URL\s*검색\s*[:：]?\s*(https?:\/\/\S+)/i);
      if (m) url = m[1].replace(/[|｜].*$/, '').trim();
    }
    if (!url) {
      const m2 = t.match(/(https?:\/\/www\.zara\.com[^\s|]+)/i)
        || t.match(/(https?:\/\/[^\s|]+)/i);
      if (m2) url = m2[1].trim();
    }
    url = url.replace(/[\)\]\>,\;]+$/, '');

    // 3) 수집조건수정/수집조건저장 버튼 → ★「수집개수 … 전체저장」바로 옆 버튼 우선
    // ※ 실제 화면 버튼명이 '수집조건수정' 또는 '수집조건저장'으로 다를 수 있어 둘 다 인식
    const editNodes = Array.from(tr.querySelectorAll(
      'a, button, input[type="button"], input[type="submit"], input[value], span'
    ));
    const nearCollect = (el) => {
      let node = el;
      for (let d = 0; d < 5 && node; d++) {
        const parent = node.parentElement;
        if (!parent) break;
        let before = '';
        for (const child of Array.from(parent.childNodes)) {
          if (child === node || (child.contains && el && child.contains(el))) break;
          before += (child.innerText || child.textContent || '');
        }
        const compact = before.replace(/\s+/g, '');
        if (compact.includes('수집개수') && (
          compact.includes('전체저장') || /수집개수[:：]?\d+개/.test(compact)
        )) {
          return true;
        }
        const full = (parent.innerText || '').replace(/\s+/g, '');
        const iCnt = full.indexOf('수집개수');
        const iAll = full.indexOf('전체저장');
        const iBtn = full.search(/수집조건(수정|저장)/);
        if (iCnt >= 0 && iBtn > iCnt && (iAll < 0 || (iAll > iCnt && iAll < iBtn))) {
          return true;
        }
        node = parent;
      }
      return false;
    };
    const isEditLabel = (el) => {
      const label = (el.value || el.textContent || '').replace(/\s+/g, '');
      return /^수집조건(수정|저장)/.test(label);
    };
    const ranked = editNodes
      .filter(isEditLabel)
      .map((el, ord) => {
        const tag = (el.tagName || '').toUpperCase();
        let score = 0;
        if (nearCollect(el)) score += 100;
        if (tag === 'INPUT' || tag === 'BUTTON') score += 20;
        if (tag === 'A') score += 10;
        if (/^수집조건(수정|저장)$/.test(((el.value || el.textContent || '').replace(/\s+/g, '')))) {
          score += 15;
        }
        return { el, score, ord };
      })
      .sort((a, b) => b.score - a.score || a.ord - b.ord);
    for (const item of ranked) {
      const el = item.el;
      hasEdit = true;
      if (el.tagName === 'A') {
        editHref = el.href || el.getAttribute('href') || '';
      } else if (el.closest && el.closest('a')) {
        const a = el.closest('a');
        editHref = a.href || a.getAttribute('href') || '';
      }
      if (!editHref) {
        const oc = el.getAttribute('onclick') || '';
        const hrefInOc = oc.match(/location\.href\s*=\s*['"]([^'"]+)['"]/i)
          || oc.match(/window\.open\s*\(\s*['"]([^'"]+)['"]/i)
          || oc.match(/['"]([^'"]*admin_group_modify\.php[^'"]*)['"]/i);
        if (hrefInOc) editHref = hrefInOc[1];
        if (!editHref) {
          const fm = oc.match(/ps_fuid\s*=\s*(\d+)/)
            || oc.match(/modify_filter[^0-9]*(\d+)/i)
            || oc.match(/fuid["'\s:=]+(\d+)/i)
            || oc.match(/(?:go|open|modify|edit)\s*\(\s*(\d+)\s*\)/i);
          if (fm) {
            editHref = 'admin_group_modify.php?ps_mode=modify_filter&ps_fuid=' + fm[1];
          }
        }
      }
      break;
    }

    if (!url && !hasEdit && !filterName) continue;
    // 행 고유 id — 같은 URL·같은 필터이름 행이 여러 개여도 이것으로 구분한다
    let fuid = '';
    const fm = String(editHref || '').match(/ps_fuid=(\d+)/);
    if (fm) fuid = fm[1];
    out.push({
      index: i,
      url,
      filterName,
      hasEdit,
      editHref,
      fuid,
      // 2단계에서 망고 행 정보를 그대로 표출하므로 넉넉히 담는다
      text: t.slice(0, 400),
    });
  }
  return out;
}"""


def list_demango_rows(page) -> list[dict]:
    """더망고 검색필터 목록 행 수집.

    스크린샷 순서·구조:
    1) 검색필터 필드값 = 열 '필터이름(수정가능)' 의 input.value (예: 여성헤어_헤어)
    2) URL = 열 '검색필터(저장조건)' 안의 'URL 검색:' 뒤 링크/텍스트
    3) 수집조건수정 = 같은 영역 버튼/링크 (href 또는 ps_fuid)
    """
    data = page.evaluate(LIST_DEMANGO_ROWS_JS)
    return list(data or [])


# ★행 지정 공통 규칙 (JS) — 같은 「URL 검색」 주소가 여러 행에 있을 수 있으므로
#   URL 만 보고 첫 행을 잡으면 2번째 행부터 첫 행을 갱신해 버린다.
#   그래서 (행 index) + (필터이름) 으로 그 행만 정확히 고른다.
ROW_PICK_JS = r"""
  const __p3Trs = () => Array.from(document.querySelectorAll('table tr, form tr, tr'));
  const __p3Compact = (s) => String(s || '').replace(/\s+/g, '');
  const __p3FilterNames = (tr) => Array.from(tr.querySelectorAll('input'))
    .filter(inp => {
      const ty = (inp.getAttribute('type') || 'text').toLowerCase();
      return ty === 'text' || ty === 'search' || ty === '';
    })
    .map(inp => __p3Compact(inp.value || inp.getAttribute('value') || ''))
    .filter(Boolean);
  const __p3UrlScore = (tr, urlHint) => {
    if (!urlHint) return 0;
    const stem = urlHint.split('?')[0];
    let best = 0;
    for (const a of Array.from(tr.querySelectorAll('a[href]'))) {
      const h = a.href || a.getAttribute('href') || '';
      if (!h) continue;
      if (h === urlHint) return 100;
      if (h.split('?')[0] === stem) { best = Math.max(best, 80); continue; }
      if (h.startsWith(stem) || urlHint.startsWith(h.split('?')[0])) {
        best = Math.max(best, 40);
      }
    }
    if (!best) {
      const text = tr.innerText || '';
      if (urlHint && text.includes(urlHint)) best = 30;
      else if (stem && text.includes(stem)) best = 20;
    }
    return best;
  };
  // 행 고유 id (수집조건수정 버튼의 ps_fuid) — 같은 URL·같은 필터이름 행도 구분한다
  const __p3RowFuid = (tr) => {
    for (const el of Array.from(tr.querySelectorAll('a, button, input'))) {
      const label = __p3Compact(el.value || el.textContent || '');
      const src = (el.getAttribute('onclick') || '')
        + ' ' + (el.getAttribute('href') || '') + ' ' + (el.href || '');
      if (!/^수집조건(수정|저장)/.test(label) && !/ps_fuid=/.test(src)) continue;
      const m = src.match(/ps_fuid=(\d+)/);
      if (m) return m[1];
    }
    return '';
  };
  const __p3PickRow = (rowIndex, urlHint, filterHint, fuidHint) => {
    const trs = __p3Trs();
    const want = __p3Compact(filterHint);
    const wantFuid = String(fuidHint || '').trim();
    // 0) 행 고유 id 가 있으면 그 행이 정답 (가장 정확)
    if (wantFuid) {
      for (let i = 0; i < trs.length; i++) {
        if (__p3RowFuid(trs[i]) === wantFuid) {
          return { tr: trs[i], index: i, by: 'fuid' };
        }
      }
    }
    const fits = (tr) => {
      if (!tr) return false;
      if (urlHint && __p3UrlScore(tr, urlHint) <= 0) return false;
      if (want && !__p3FilterNames(tr).includes(want)) return false;
      return true;
    };
    // 1) 지정된 행 index 가 URL·필터이름과 맞으면 그 행 (순차처리의 기준)
    const at = (rowIndex >= 0 && rowIndex < trs.length) ? trs[rowIndex] : null;
    if (at && fits(at)) return { tr: at, index: rowIndex, by: 'index' };
    // 2) 아니면 URL·필터이름이 맞는 행 중 지정 index 에 가장 가까운 행
    const ranked = trs
      .map((tr, i) => ({ tr: tr, i: i, s: __p3UrlScore(tr, urlHint) }))
      .filter(x => fits(x.tr))
      .sort((a, b) => (b.s - a.s)
        || (Math.abs(a.i - rowIndex) - Math.abs(b.i - rowIndex)));
    if (ranked.length) {
      return { tr: ranked[0].tr, index: ranked[0].i, by: 'match' };
    }
    return null;
  };
"""


def _find_and_mark_edit_button(
    page,
    row_index: int,
    row_url: str = "",
    filter_hint: str = "",
    fuid_hint: str = "",
) -> dict:
    """URL 바로 오른쪽·수집개수|전체저장 옆 「수집조건수정」을 data-p3-edit-target 마킹."""
    info = page.evaluate(
        """(args) => {
          __ROW_PICK__
          const rowIndex = args.rowIndex;
          const urlHint = (args.urlHint || '').trim();
          const urlStem = urlHint.split('?')[0];
          document.querySelectorAll('[data-p3-edit-target]').forEach(el => {
            el.removeAttribute('data-p3-edit-target');
          });

          const isEditControl = (el) => {
            if (!el) return false;
            const tag = (el.tagName || '').toUpperCase();
            if (!(tag === 'A' || tag === 'BUTTON' || tag === 'INPUT')) return false;
            if (tag === 'INPUT') {
              const ty = (el.getAttribute('type') || 'button').toLowerCase();
              if (!(ty === 'button' || ty === 'submit' || ty === '')) return false;
            }
            const t = (el.value || el.textContent || '').replace(/\\s+/g, '');
            return /^수집조건(수정|저장)/.test(t);
          };

          const nearCollectCount = (el) => {
            let node = el;
            for (let depth = 0; depth < 5 && node; depth++) {
              const parent = node.parentElement;
              if (!parent) break;
              let before = '';
              for (const child of Array.from(parent.childNodes)) {
                if (child === node || (child.contains && child.contains(el))) break;
                before += (child.innerText || child.textContent || '');
              }
              const compact = before.replace(/\\s+/g, '');
              if (compact.includes('수집개수') && (
                compact.includes('전체저장') || /수집개수[:：]?\\d+개/.test(compact)
              )) {
                return true;
              }
              const full = (parent.innerText || '').replace(/\\s+/g, '');
              const iCnt = full.indexOf('수집개수');
              const iAll = full.indexOf('전체저장');
              const iBtn = full.search(/수집조건(수정|저장)/);
              if (iCnt >= 0 && iBtn > iCnt && (iAll < 0 || (iAll > iCnt && iAll < iBtn))) {
                return true;
              }
              node = parent;
            }
            return false;
          };

          const findUrlAnchor = (tr) => {
            const anchors = Array.from(tr.querySelectorAll('a[href]'));
            for (const a of anchors) {
              const h = a.href || a.getAttribute('href') || '';
              if (!h || h.indexOf('http') !== 0) continue;
              if (!urlHint) return a;
              if (h === urlHint || h.startsWith(urlStem) || urlHint.startsWith(h.split('?')[0])) {
                return a;
              }
            }
            return anchors.find(a => {
              const h = a.href || '';
              return h.indexOf('http') === 0 && !/수집조건(수정|저장)/.test((a.textContent||''));
            }) || null;
          };

          const rightOfUrl = (el, urlA) => {
            if (!urlA || !el) return false;
            try {
              const ub = urlA.getBoundingClientRect();
              const eb = el.getBoundingClientRect();
              // URL 앵커의 오른쪽 (같은 행 근처)
              if (eb.left >= ub.right - 12) return true;
              if (eb.left > ub.left && eb.right > ub.right + 20) return true;
            } catch (e) {}
            return false;
          };

          const score = (el, urlA) => {
            let s = 0;
            if (rightOfUrl(el, urlA)) s += 120;
            if (nearCollectCount(el)) s += 100;
            const t = (el.value || el.textContent || '').replace(/\\s+/g, '');
            if (/^수집조건(수정|저장)$/.test(t)) s += 20;
            const tag = (el.tagName || '').toUpperCase();
            if (tag === 'INPUT' || tag === 'BUTTON') s += 10;
            if (tag === 'A') s += 5;
            return s;
          };

          const picked = __p3PickRow(rowIndex, urlHint, args.filterHint || '', args.fuidHint || '');
          if (!picked) return { ok: false, reason: 'row-not-found' };

          const tr = picked.tr;
          const urlA = findUrlAnchor(tr);
          const edits = Array.from(tr.querySelectorAll('a, button, input')).filter(isEditControl);
          if (!edits.length) return { ok: false, reason: 'button-not-found' };
          edits.sort((a, b) => score(b, urlA) - score(a, urlA));
          const right = edits.filter(e => rightOfUrl(e, urlA));
          const near = edits.filter(nearCollectCount);
          const pick = (right.length ? right : (near.length ? near : edits))[0];
          try { pick.scrollIntoView({ block: 'center', inline: 'nearest' }); } catch (e) {}
          pick.setAttribute('data-p3-edit-target', '1');
          if (urlA) urlA.setAttribute('data-p3-url-target', '1');
          const oc = pick.getAttribute('onclick') || '';
          return {
            ok: true,
            tag: pick.tagName,
            text: ((pick.value || pick.textContent || '') + '').replace(/\\s+/g, ' ').trim().slice(0, 40),
            nearCollect: nearCollectCount(pick),
            rightOfUrl: rightOfUrl(pick, urlA),
            score: score(pick, urlA),
            onclick: oc.slice(0, 160),
            // 어느 행을 잡았는지 — 호출부에서 지정 행과 같은지 검증한다
            rowIndexUsed: picked.index,
            rowPickedBy: picked.by,
            rowFuid: __p3RowFuid(tr),
            rowFilterNames: __p3FilterNames(tr).slice(0, 4),
            rowHref: ((urlA && (urlA.href || urlA.getAttribute('href'))) || '').slice(0, 200),
          };
        }""".replace("__ROW_PICK__", ROW_PICK_JS),
        {
            "rowIndex": int(row_index),
            "urlHint": (row_url or "").strip(),
            "filterHint": (filter_hint or "").strip(),
            "fuidHint": str(fuid_hint or "").strip(),
        },
    )
    return info if isinstance(info, dict) else {"ok": False, "reason": "evaluate-failed"}


def _p1_browse():
    """P1_101 팝업닫기·스크롤집계 재사용 (중단 플래그는 P3)."""
    import extract as p1_extract  # noqa: WPS433

    p1_extract.stop_requested = stop_requested
    return p1_extract


def dismiss_store_layers_only(page) -> int:
    """스토어(자라) 페이지의 레이어/쿠키만 닫기.

    ★다른 탭(더망고 목록)은 절대 page.close() 하지 않음.
    P1 dismiss_popups 는 다른 창을 닫아 더망고 핸들이 끊길 수 있음.
    """
    closed = 0
    try:
        page.on("dialog", lambda d: d.dismiss())
    except Exception:
        pass
    try:
        p1 = _p1_browse()
        selectors = getattr(p1, "POPUP_CLOSE_SELECTORS", ())
    except Exception:
        selectors = (
            'button[aria-label="Close"]',
            'button:has-text("닫기")',
            'button:has-text("Accept")',
            'button:has-text("동의")',
        )
    for sel in selectors:
        try:
            loc = page.locator(sel)
            count = loc.count()
        except Exception:
            continue
        for i in range(min(count, 3)):
            try:
                el = loc.nth(i)
                if el.is_visible(timeout=300):
                    el.click(timeout=800, force=True)
                    closed += 1
                    time.sleep(0.15)
            except Exception:
                continue
    try:
        page.keyboard.press("Escape")
        time.sleep(0.1)
    except Exception:
        pass
    return closed


def find_alive_mango_page(context, mango_url: str, prefer=None):
    """열려 있는 더망고 목록 탭을 다시 찾는다 (죽은 핸들 대체)."""
    pages: list = []
    try:
        pages = list(context.pages)
    except Exception:
        pages = []
    if prefer is not None and prefer not in pages:
        pages = [prefer] + pages

    mu = (mango_url or "").strip().lower()
    mu_stem = mu.split("?")[0]

    def _score(p) -> int:
        try:
            if p.is_closed():
                return -999
        except Exception:
            return -999
        try:
            u = (p.url or "").lower()
        except Exception:
            u = ""
        s = 0
        if prefer is not None and p is prefer:
            s += 30
        if mu and (mu in u or (mu_stem and mu_stem in u)):
            s += 120
        if any(k in u for k in ("demango", "admin_group", "filter", "mango")):
            s += 60
        if "zara.com" in u:
            s -= 100
        if u in ("", "about:blank"):
            s -= 20
        return s

    ranked = sorted(pages, key=_score, reverse=True)
    for p in ranked:
        if _score(p) > 0:
            return p
    return None


def page_is_usable(page) -> bool:
    """evaluate 가능한 살아 있는 페이지인지."""
    if page is None:
        return False
    try:
        if page.is_closed():
            return False
    except Exception:
        return False
    try:
        page.evaluate("() => 1", timeout=1500)
        return True
    except TypeError:
        # 구버전 playwright 는 evaluate timeout kw 없음
        try:
            page.evaluate("() => 1")
            return True
        except Exception:
            return False
    except Exception:
        return False


def resolve_demango_row_index_by_url(
    page,
    row_url: str,
    *,
    fallback_index: int | None = None,
    filter_hint: str = "",
    fuid_hint: str = "",
    progress: ProgressFn | None = None,
) -> int | None:
    """더망고 목록에서 URL(+필터이름)로 행 index를 다시 찾는다.

    ★목록은 수정일 정렬(ft_sort=modify_asc)이라 한 행을 갱신하면 순서가 바뀔 수 있고,
    같은 URL이 여러 행에 있을 수도 있다. 그래서 URL 뿐 아니라 필터이름까지 같은 행을
    찾아야 "2번째 행을 갱신했는데 1번째 행이 바뀌는" 사고를 막을 수 있다.
    """
    url = (row_url or "").strip()
    if not url:
        return fallback_index
    try:
        rows = list_demango_rows(page)
    except Exception as e:  # noqa: BLE001
        _log(progress, "경고", f"행 재탐색 실패(목록스캔): {str(e).split(chr(10))[0][:100]}")
        return fallback_index

    want_filter = "".join((filter_hint or "").split())
    want_fuid = "".join((fuid_hint or "").split())
    if want_fuid:
        for r in rows:
            if str(r.get("fuid") or "").strip() == want_fuid:
                idx = int(r.get("index") or 0)
                _log(
                    progress,
                    "로직",
                    f"행 재탐색 OK(행id={want_fuid}) · index={idx} · "
                    f"필터={r.get('filterName') or ''}",
                )
                return idx

    def _same_filter(r: dict) -> bool:
        if not want_filter:
            return True
        return "".join(str(r.get("filterName") or "").split()) == want_filter

    target = normalize_url(url)
    stem = url_stem(url)
    # 1) URL 완전일치 + 필터이름 일치 → 2) URL stem 완전일치 + 필터이름 → 3) URL 만
    for label, match in (
        ("정확", lambda r: normalize_url((r.get("url") or "").strip()) == target
            and _same_filter(r)),
        ("stem", lambda r: stem and url_stem(r.get("url") or "") == stem
            and _same_filter(r)),
        ("URL만", lambda r: normalize_url((r.get("url") or "").strip()) == target),
    ):
        for r in rows:
            try:
                hit = bool(match(r))
            except Exception:  # noqa: BLE001
                hit = False
            if not hit:
                continue
            idx = int(r.get("index") or 0)
            _log(
                progress,
                "로직",
                f"행 재탐색 OK({label}) · index={idx} · 필터={r.get('filterName') or ''} · "
                f"url={url[:100]}",
            )
            return idx

    _log(
        progress,
        "경고",
        f"행 재탐색 실패 → fallback index={fallback_index} · 필터={filter_hint} · "
        f"url={url[:100]}",
    )
    return fallback_index


def _find_and_mark_row_url(
    page,
    row_index: int,
    row_url: str = "",
    filter_hint: str = "",
    fuid_hint: str = "",
) -> dict:
    """행의 URL 검색 링크를 data-p3-url-target 으로 마킹."""
    info = page.evaluate(
        """(args) => {
          __ROW_PICK__
          const rowIndex = args.rowIndex;
          const urlHint = (args.urlHint || '').trim();
          const urlStem = urlHint.split('?')[0];
          document.querySelectorAll('[data-p3-url-target]').forEach(el => {
            el.removeAttribute('data-p3-url-target');
          });
          const picked = __p3PickRow(rowIndex, urlHint, args.filterHint || '', args.fuidHint || '');
          if (!picked) return { ok: false, reason: 'row-not-found' };
          const tr = picked.tr;
          const anchors = Array.from(tr.querySelectorAll('a[href]'));
          let pick = null;
          for (const a of anchors) {
            const h = a.href || a.getAttribute('href') || '';
            const label = (a.textContent || '').replace(/\\s+/g, '');
            if (/수집조건(수정|저장)/.test(label)) continue;
            if (h.indexOf('http') !== 0) continue;
            if (!urlHint || h === urlHint || h.startsWith(urlStem) || urlHint.startsWith(h.split('?')[0])) {
              pick = a; break;
            }
          }
          if (!pick) {
            pick = anchors.find(a => {
              const h = a.href || '';
              return h.indexOf('http') === 0 && !/수집조건(수정|저장)/.test((a.textContent||''));
            });
          }
          if (!pick) return { ok: false, reason: 'url-not-found' };
          try { pick.scrollIntoView({ block: 'center', inline: 'nearest' }); } catch (e) {}
          pick.setAttribute('data-p3-url-target', '1');
          return {
            ok: true,
            href: (pick.href || '').slice(0, 200),
            rowIndexUsed: picked.index,
            rowPickedBy: picked.by,
          };
        }""".replace("__ROW_PICK__", ROW_PICK_JS),
        {
            "rowIndex": int(row_index),
            "urlHint": (row_url or "").strip(),
            "filterHint": (filter_hint or "").strip(),
            "fuidHint": str(fuid_hint or "").strip(),
        },
    )
    return info if isinstance(info, dict) else {"ok": False}


def click_demango_row_url(
    page,
    row_index: int,
    row_url: str = "",
    *,
    progress: ProgressFn | None = None,
):
    """1) 필터일치 행의 URL 클릭 → 스토어 페이지(팝업/새탭) 반환."""
    info = _find_and_mark_row_url(page, row_index, row_url)
    if not info.get("ok"):
        _log(progress, "오류", f"2) URL 링크 미검출 · info={info}")
        return None
    before = []
    try:
        before = list(page.context.pages)
    except Exception:
        before = [page]
    loc = page.locator('[data-p3-url-target="1"]').first
    store = None
    try:
        with page.expect_popup(timeout=10_000) as pop_info:
            loc.click(timeout=5_000, no_wait_after=True)
        store = pop_info.value
    except Exception:
        try:
            loc.click(timeout=5_000, force=True, no_wait_after=True)
        except Exception as e:
            _log(
                progress,
                "오류",
                f"2) URL 클릭 실패 · "
                f"{str(e).split(chr(10))[0][:120]}",
            )
            return None
    time.sleep(0.6)
    if store is None:
        try:
            for p in page.context.pages:
                if p not in before:
                    store = p
                    break
        except Exception:
            pass
    if store is None:
        # 같은 탭 이동된 경우 — 목록 복귀는 호출측에서 mango 로
        store = page
    try:
        store.wait_for_load_state("domcontentloaded", timeout=30_000)
    except Exception:
        pass
    # ★P2와 동일: 실제 Chrome 팝업/탭을 앞으로
    reveal_browser_page(
        store,
        progress,
        step_no="1",
        action="URL클릭 → 스토어/팝업 창 표시",
        dwell_s=STEP_VIEW_DWELL_SEC,
    )
    return store


def browse_store_count_cards(
    store_page,
    *,
    excel_count: int,
    progress: ProgressFn | None = None,
    shot_dir: Path | None = None,
    row_no: int = 0,
) -> tuple[int, bool]:
    """2)~5) 첫팝업 닫기 → 푸터↓ → 상단↑ 카드수 → 엑셀 비교.

    ★팝업 닫기 시 다른 탭(더망고)을 닫지 않음.
    ★P2와 동일: 실제 Chrome 창·팝업을 bring_to_front 후 동작.
    """
    p1 = _p1_browse()
    # 스크롤 중 P1 이 dismiss_popups 로 타 탭을 닫지 못하게 교체
    orig_dismiss = p1.dismiss_popups
    p1.dismiss_popups = dismiss_store_layers_only  # type: ignore[assignment]
    try:
        # 2) 스토어 팝업/레이어를 Chrome 앞으로 보여 준 뒤 닫기 (P2: 실제 창 표시)
        reveal_browser_page(
            store_page,
            progress,
            step_no="2",
            action="스토어 첫 화면/팝업 노출",
            dwell_s=1.0,
        )
        screenshot_step(
            store_page,
            shot_dir,
            step_tag="02_before_dismiss",
            label="2)팝업닫기 전(브라우저표시)",
            row_no=row_no,
            progress=progress,
        )
        _log(progress, "동작", "2) 첫 팝업창 닫기 (스토어 레이어만 · 더망고탭 유지)")
        closed = dismiss_store_layers_only(store_page)
        time.sleep(0.8)
        closed += dismiss_store_layers_only(store_page)
        _log(progress, "동작", f"2) 팝업 닫기 완료 · closed={closed}")
        reveal_browser_page(
            store_page,
            progress,
            step_no="2",
            action=f"팝업닫기 후 본문 표시 · closed={closed}",
        )
        screenshot_step(
            store_page,
            shot_dir,
            step_tag="02_after_dismiss",
            label=f"2)팝업닫기 후 closed={closed}",
            row_no=row_no,
            progress=progress,
        )

        _log(progress, "동작", "3) 스크롤 푸터 영역까지 내리기")
        reveal_browser_page(
            store_page, progress, step_no="3", action="푸터 스크롤 시작", dwell_s=0.4
        )
        p1.scroll_down_to_footer(store_page, progress=progress)
        reveal_browser_page(
            store_page, progress, step_no="3", action="푸터까지 스크롤 완료"
        )
        screenshot_step(
            store_page,
            shot_dir,
            step_tag="03_footer",
            label="3)푸터까지 스크롤",
            row_no=row_no,
            progress=progress,
        )

        _log(progress, "동작", "4) 하단→상단 스크롤 · 상품수 카드 갯수 집계")
        reveal_browser_page(
            store_page, progress, step_no="4", action="상단 스크롤·카드집계 시작", dwell_s=0.4
        )
        card_n = int(p1.scroll_up_count_card_images(store_page, progress=progress) or 0)
        _log(progress, "동작", f"4) 상품수 카드 갯수={card_n}")
        reveal_browser_page(
            store_page,
            progress,
            step_no="4",
            action=f"카드갯수={card_n} 화면",
        )
        screenshot_step(
            store_page,
            shot_dir,
            step_tag="04_card_count",
            label=f"4)카드갯수={card_n}",
            row_no=row_no,
            progress=progress,
        )
    finally:
        p1.dismiss_popups = orig_dismiss  # type: ignore[assignment]

    excel_n = int(excel_count or 0)
    matched = card_n == excel_n
    _log(
        progress,
        "동작",
        f"5) 비교 · 카드상품수={card_n} · 엑셀상품수={excel_n} · 일치={'Y' if matched else 'N'}",
    )
    reveal_browser_page(
        store_page,
        progress,
        step_no="5",
        action=f"비교결과 카드={card_n} 엑셀={excel_n} {'Y' if matched else 'N'}",
    )
    screenshot_step(
        store_page,
        shot_dir,
        step_tag="05_compare",
        label=f"5)비교 카드={card_n} 엑셀={excel_n} {'Y' if matched else 'N'}",
        row_no=row_no,
        progress=progress,
    )
    return card_n, matched


def close_store_return_list(
    list_page,
    store_page,
    mango_url: str,
    *,
    progress: ProgressFn | None = None,
):
    """스토어 탭만 닫고, 살아 있는 더망고 목록 탭을 다시 찾아 반환."""
    ctx = None
    try:
        if list_page is not None:
            ctx = list_page.context
        elif store_page is not None:
            ctx = store_page.context
    except Exception:
        ctx = None

    # 스토어(자라) 탭만 닫기 — 더망고와 동일 핸들이면 닫지 않음
    try:
        if store_page is not None and store_page is not list_page:
            store_url = ""
            try:
                store_url = (store_page.url or "").lower()
            except Exception:
                store_url = ""
            # 더망고 URL 이면 닫지 않음
            mu = (mango_url or "").lower()
            is_mango = bool(mu) and (mu in store_url or store_url.startswith(mu.split("?")[0]))
            if (not is_mango) and ("zara.com" in store_url or "http" in store_url):
                if not store_page.is_closed():
                    store_page.close()
                    _log(progress, "로직", "스토어 탭 닫음 → 더망고 목록 재연결")
    except Exception as e:  # noqa: BLE001
        _log(progress, "경고", f"스토어 탭 닫기 예외: {str(e).split(chr(10))[0][:80]}")

    time.sleep(0.15)

    mango = None
    if ctx is not None:
        mango = find_alive_mango_page(ctx, mango_url, prefer=list_page)

    if mango is None or not page_is_usable(mango):
        # prefer 가 죽었어도 context 에서 재탐색
        if ctx is not None:
            mango = find_alive_mango_page(ctx, mango_url, prefer=None)
        if mango is None or not page_is_usable(mango):
            _log(progress, "오류", "더망고 목록 탭 재연결 실패 (usable page 없음)")
            return None

    try:
        mango.bring_to_front()
    except Exception:
        pass

    try:
        cur = mango.url or ""
        if (
            "modify_filter" in cur
            or "admin_group_modify" in cur
            or "zara.com" in cur.lower()
        ):
            _return_to_list(mango, mango_url)
    except Exception:
        try:
            _return_to_list(mango, mango_url)
        except Exception:
            pass

    if not page_is_usable(mango):
        _log(progress, "오류", "더망고 목록 탭 evaluate 불가")
        return None

    _log(
        progress,
        "로직",
        f"더망고 목록 탭 재연결 OK · url={(mango.url or '')[:120]}",
    )
    return mango


def _modify_ui_opened(page) -> bool:
    """수집조건수정 후 수정 팝업/페이지/iframe 이 실제로 열렸는지."""
    if page_shows_not_found(page):
        return False
    try:
        if wait_for_save_count_ready(page, timeout_ms=400):
            return True
    except Exception:
        pass
    try:
        target, kind = resolve_modify_target(page)
        if kind in ("page", "frame") and kind != "main":
            if wait_for_save_count_ready(target, timeout_ms=400):
                return True
        body = ""
        try:
            body = target.locator("body").inner_text(timeout=300) or ""
        except Exception:
            body = ""
        if "저장상품수" in body and (
            "검색필터 수정" in body or "검색결과" in body or "저장하기" in body
        ):
            return True
    except Exception:
        pass
    try:
        for p in page.context.pages:
            if page_shows_not_found(p):
                continue
            bu = p.url or ""
            if "modify_filter" in bu or "admin_group_modify" in bu:
                bt = ""
                try:
                    bt = p.locator("body").inner_text(timeout=300) or ""
                except Exception:
                    pass
                if "저장상품수" in bt or "검색필터 수정" in bt or (
                    "검색결과" in bt and "저장하기" in bt
                ):
                    return True
                # URL 은 맞는데 본문 로딩 중이면 잠깐 더
                try:
                    p.wait_for_load_state("domcontentloaded", timeout=2000)
                except Exception:
                    pass
                try:
                    bt = p.locator("body").inner_text(timeout=400) or ""
                except Exception:
                    bt = ""
                if page_shows_not_found(p):
                    continue
                if "저장상품수" in bt or "검색필터 수정" in bt:
                    return True
            else:
                try:
                    bt = p.locator("body").inner_text(timeout=250) or ""
                except Exception:
                    bt = ""
                if "저장상품수" in bt and ("검색결과" in bt or "저장하기" in bt):
                    return True
    except Exception:
        pass
    return False


# 6) 수집조건수정: 「전체저장」텍스트 확인 → 「수집조건수정/저장」LABEL의 실제 버튼요소를
#    찾아 그 요소를 직접 클릭한다.
# ※ '전체저장 우측으로 N글자 이동한 좌표를 클릭'하는 방식은 완전히 삭제됨 — 좌표계산 없음.
EDIT_CLICK_MAX_TRIES = 3  # 같은(실제) 버튼 재클릭 시도 횟수


def _log_text_find_phase(
    page,
    progress: ProgressFn | None,
    shot_dir: Path | None,
    *,
    row_no: int,
    phase: str,
    kind: str,
    label: str,
    found: bool | None = None,
    detail: str = "",
) -> None:
    """텍스트/버튼명 찾기 전·후 — 텍스트 + 스크린샷을 주요 로그에 남긴다.

    kind: '텍스트' | '버튼명'
    phase: '전' | '후'
    label: '전체저장' | '수집조건수정'
    """
    if phase == "전":
        msg = f"5) {kind} 찾기 전 · {kind}={label}"
        tag = f"05_find_{label}_before"
    else:
        status = "OK" if found else "FAIL"
        extra = f" · {detail}" if detail else ""
        msg = f"5) {kind} 찾기 후 · {kind}={label} · {status}{extra}"
        tag = f"05_find_{label}_after"
    # 버튼 찾기 과정은 5) 수집조건수정 단계의 세부내용 — MAIN엔 7단계만 남긴다
    _log(progress, "주요", msg, major=False)
    screenshot_step(
        page,
        shot_dir,
        step_tag=tag,
        label=msg,
        row_no=row_no,
        progress=progress,
    )


def _check_allsave_text_present(
    page,
    row_index: int,
    row_url: str,
    filter_hint: str = "",
    fuid_hint: str = "",
) -> dict:
    """검색필터 URL 바로 우측에 '전체저장' 텍스트가 있는지만 확인 (검증·로그용).

    ★클릭 좌표 계산에는 사용하지 않는다 — 버튼 클릭은 항상 실제 버튼요소를 찾아 수행.
    """
    result = page.evaluate(
        """(args) => {
          __ROW_PICK__
          const rowIndex = args.rowIndex;
          const urlHint = (args.urlHint || '').trim();
          const needle = '전체저장';

          const picked = __p3PickRow(rowIndex, urlHint, args.filterHint || '', args.fuidHint || '');
          if (!picked) return { found: false };

          const tr = picked.tr;
          const txt = (tr.innerText || '').replace(/\\s+/g, '');
          if (!txt.includes(needle)) return { found: false };
          const walker = document.createTreeWalker(tr, NodeFilter.SHOW_TEXT);
          let node;
          while ((node = walker.nextNode())) {
            const t = node.textContent || '';
            const i = t.indexOf(needle);
            if (i < 0) continue;
            const range = document.createRange();
            range.setStart(node, i);
            range.setEnd(node, i + needle.length);
            const rects = range.getClientRects();
            if (rects && rects.length) {
              const r = rects[rects.length - 1];
              return { found: true, left: r.left, right: r.right, top: r.top, bottom: r.bottom };
            }
          }
          return { found: true };
        }""".replace("__ROW_PICK__", ROW_PICK_JS),
        {
            "rowIndex": int(row_index),
            "urlHint": (row_url or "").strip(),
            "filterHint": (filter_hint or "").strip(),
            "fuidHint": str(fuid_hint or "").strip(),
        },
    )
    return result if isinstance(result, dict) else {"found": False}


def _find_edit_button_with_log(
    page,
    row_index: int,
    row_url: str,
    *,
    filter_hint: str = "",
    fuid_hint: str = "",
    progress: ProgressFn | None = None,
    shot_dir: Path | None = None,
    row_no: int = 0,
    log_find: bool = True,
) -> dict:
    """「전체저장」텍스트 확인 → LABEL 「수집조건수정」/「수집조건저장」의 실제 버튼요소를 찾는다.

    1) 검색필터 URL 바로 우측에서 「전체저장」텍스트 존재 확인 (전/후 로그·샷)
    2) 「전체저장」옆 실제 버튼요소(a/button/input, LABEL=수집조건수정 또는 수집조건저장)를
       DOM에서 찾아 data-p3-edit-target 로 마킹 (전/후 로그·샷)
       ★좌표 계산 없음 — 찾은 요소를 그대로 클릭 대상으로 사용한다.
    """
    _find_and_mark_row_url(page, row_index, row_url, filter_hint, fuid_hint)

    if log_find:
        _log_text_find_phase(
            page,
            progress,
            shot_dir,
            row_no=row_no,
            phase="전",
            kind="텍스트",
            label="전체저장",
        )

    allsave = _check_allsave_text_present(
        page, row_index, row_url, filter_hint, fuid_hint
    )
    allsave_found = bool(allsave.get("found"))

    if log_find:
        detail = ""
        if allsave_found and "left" in allsave:
            detail = f"x={float(allsave.get('left', 0)):.0f}~{float(allsave.get('right', 0)):.0f}"
        _log_text_find_phase(
            page,
            progress,
            shot_dir,
            row_no=row_no,
            phase="후",
            kind="텍스트",
            label="전체저장",
            found=allsave_found,
            detail=detail,
        )

    if log_find:
        _log_text_find_phase(
            page,
            progress,
            shot_dir,
            row_no=row_no,
            phase="전",
            kind="버튼명",
            label="수집조건수정/수집조건저장",
        )

    info = _find_and_mark_edit_button(
        page, row_index, row_url, filter_hint, fuid_hint
    )
    btn_found = bool(info.get("ok"))
    matched_label = str(info.get("text") or "") if btn_found else ""

    if log_find:
        detail = (
            f"매칭={matched_label} · tag={info.get('tag', '')} · "
            f"행index={info.get('rowIndexUsed')}({info.get('rowPickedBy')}) · "
            f"행id={info.get('rowFuid') or '-'} · "
            f"행필터명={info.get('rowFilterNames')}"
            if btn_found
            else str(info.get("reason") or "")
        )
        _log_text_find_phase(
            page,
            progress,
            shot_dir,
            row_no=row_no,
            phase="후",
            kind="버튼명",
            label="수집조건수정/수집조건저장",
            found=btn_found,
            detail=detail,
        )

    info["allsave_found"] = allsave_found
    info["matched_label"] = matched_label
    return info


def click_edit_on_row(
    page,
    row_index: int,
    edit_href: str = "",  # 미사용 — href 재시도 금지(호출부 호환용)
    *,
    row_url: str = "",
    filter_hint: str = "",
    fuid_hint: str = "",
    progress: ProgressFn | None = None,
    shot_dir: Path | None = None,
    row_no: int = 0,
    shot_count: int = 0,
    shot_interval_s: float = 0.0,
    max_tries: int = EDIT_CLICK_MAX_TRIES,
    try_interval_s: float = 2.0,
) -> bool:
    """6) LABEL 「수집조건수정」/「수집조건저장」의 실제 버튼요소를 찾아 직접 클릭한다.

    1) 검색필터 URL 바로 우측에서 텍스트 「전체저장」존재 확인
    2) 「전체저장」옆 실제 버튼요소(a/button/input, LABEL=수집조건수정 또는 수집조건저장)를
       DOM에서 찾아 마킹
    3) 마킹된 요소를 Playwright locator로 직접 클릭 (좌표 계산·글자이동 없음, href 재시도 금지)
    """
    _ = edit_href
    tries = max(1, int(max_tries))
    gap = max(0.2, float(try_interval_s))
    logged_find = False

    for attempt in range(1, tries + 1):
        if stop_requested():
            return False

        info = _find_edit_button_with_log(
            page,
            row_index,
            row_url,
            filter_hint=filter_hint,
            fuid_hint=fuid_hint,
            progress=progress,
            shot_dir=shot_dir,
            row_no=row_no,
            log_find=not logged_find,
        )
        logged_find = True

        if not info.get("ok"):
            _log(
                progress,
                "오류",
                f"5) '수집조건수정' 버튼 미검출 · "
                f"재시도 {attempt}/{tries} · 사유={info.get('reason', '')}",
            )
            if attempt < tries:
                time.sleep(gap)
            continue

        # ★다른 행을 절대 건드리지 않는다 — 잡은 행이 이 행인지 검증
        #   (행 고유 id 우선, 없으면 필터이름)
        got_fuid = str(info.get("rowFuid") or "")
        want_fuid = str(fuid_hint or "").strip()
        wrong_row = False
        if want_fuid and got_fuid and got_fuid != want_fuid:
            wrong_row = True
        elif not want_fuid and filter_hint:
            names = [
                "".join(str(n or "").split())
                for n in (info.get("rowFilterNames") or [])
            ]
            want = "".join(filter_hint.split())
            wrong_row = bool(want and want not in names)
        if wrong_row:
            _log(
                progress,
                "오류",
                f"5) 다른 행을 잡았습니다 — 클릭하지 않고 중단 · 요청필터={filter_hint} · "
                f"요청행id={want_fuid or '-'} · 잡은행id={got_fuid or '-'} · "
                f"잡은행필터={info.get('rowFilterNames')} · 요청행index={row_index} · "
                f"잡은행index={info.get('rowIndexUsed')}({info.get('rowPickedBy')})",
            )
            return False

        locator = page.locator('[data-p3-edit-target="1"]').first
        try:
            locator.scroll_into_view_if_needed(timeout=2_000)
        except Exception:
            pass

        btn_label = str(info.get("matched_label") or "수집조건수정")

        before_pages = []
        try:
            before_pages = list(page.context.pages)
        except Exception:
            before_pages = [page]

        popup = None
        try:
            with page.expect_popup(timeout=int(gap * 1000)) as pop_info:
                locator.click(timeout=int(gap * 1000))
            popup = pop_info.value
        except Exception:
            popup = None
            try:
                locator.click(timeout=int(gap * 1000))
            except Exception as e:
                _log(
                    progress,
                    "오류",
                    f"5) '{btn_label}' 버튼 클릭 예외 · "
                    f"시도{attempt}/{tries}: {str(e).split(chr(10))[0][:120]}",
                )

        if popup is not None:
            try:
                popup.wait_for_load_state("domcontentloaded", timeout=12_000)
            except Exception:
                pass
            reveal_browser_page(
                popup,
                progress,
                step_no="5",
                action=f"{btn_label} 팝업 표시",
                dwell_s=STEP_VIEW_DWELL_SEC,
            )

        deadline = time.time() + gap
        while time.time() < deadline:
            try:
                for p in page.context.pages:
                    if p not in before_pages:
                        try:
                            p.wait_for_load_state("domcontentloaded", timeout=2_000)
                        except Exception:
                            pass
                        try:
                            p.bring_to_front()
                        except Exception:
                            pass
            except Exception:
                pass
            if page_shows_not_found(page):
                _log(
                    progress,
                    "오류",
                    f"5) '{btn_label}' 클릭 후 not found · 중단",
                )
                return False
            if _modify_ui_opened(page):
                _log(
                    progress,
                    "로직",
                    f"5) '{btn_label}' 클릭 → 팝업 확인 OK",
                    major=False,
                )
                if shot_dir is not None and shot_count > 0:
                    screenshot_after_edit_click_series(
                        page,
                        shot_dir,
                        row_no=row_no,
                        progress=progress,
                        prefer_page=popup,
                        count=shot_count,
                        interval_s=shot_interval_s,
                    )
                return True
            time.sleep(0.25)

        if page_shows_not_found(page):
            _log(
                progress,
                "오류",
                f"5) '{btn_label}' 클릭 후 not found · 중단",
            )
            return False
        if _modify_ui_opened(page) or wait_modify_page(page, timeout_ms=800):
            _log(
                progress,
                "로직",
                f"5) '{btn_label}' 클릭 → 팝업 확인 OK",
                major=False,
            )
            if shot_dir is not None and shot_count > 0:
                screenshot_after_edit_click_series(
                    page,
                    shot_dir,
                    row_no=row_no,
                    progress=progress,
                    prefer_page=popup,
                    count=shot_count,
                    interval_s=shot_interval_s,
                )
            return True

        if attempt < tries:
            time.sleep(gap * 0.2)

    _log(
        progress,
        "오류",
        f"5) '수집조건수정' 버튼 클릭 실패 · "
        "버튼은 찾았으나 팝업 미오픈 (href 재시도 없음)",
    )
    return False


def page_shows_not_found(page) -> bool:
    """수정 팝업/페이지가 'not found' 등 잘못된 진입인지."""
    targets = [page]
    try:
        targets = list(page.context.pages) or [page]
    except Exception:
        targets = [page]
    for target in targets:
        try:
            title = ""
            try:
                title = target.title() or ""
            except Exception:
                title = ""
            body = ""
            try:
                body = target.locator("body").inner_text(timeout=400) or ""
            except Exception:
                body = ""
            blob = f"{title}\n{body}"
            # 정상 수정화면이면 not-found 로 보지 않음
            if "저장상품수" in blob and (
                "검색필터 수정" in blob or "검색결과" in blob or "저장하기" in blob
            ):
                continue
            if re.search(
                r"not\s*found|404\b|찾을\s*수\s*없|존재하지\s*않|페이지를\s*찾을|잘못된\s*접근",
                blob,
                re.I,
            ):
                return True
        except Exception:
            continue
    return False


def set_save_count(
    page,
    value: int,
    *,
    shot_dir: Path | None = None,
    progress: ProgressFn | None = None,
    row_no: int = 0,
    out: dict | None = None,
) -> bool:
    """저장상품수 입력칸에 상품수값을 **한 번만** 넣는다.

    UI: 저장상품수 | 검색결과 상위 [ 3 ] 개 상품만 저장
    ★요건: 처음 입력 후 또 한번 입력하는 방식은 쓰지 않는다 (단발 입력).
    갱신 전·후 스크린샷은 성공/실패와 무관하게 항상 남긴다.
    out: 주면 {"before": 갱신전상품수, "after": 갱신후상품수} 를 채운다.
    """
    target = str(int(value))
    if out is not None:
        out.clear()
        out.update({"before": "", "after": ""})

    work, kind = resolve_modify_target(page)
    wait_for_save_count_ready(work, timeout_ms=8_000)
    shot_page = page if kind == "frame" else work

    if shot_dir is not None:
        screenshot_step(
            shot_page,
            shot_dir,
            step_tag="05_save_count_before",
            label=f"5)저장상품수 갱신 전 →목표={target}",
            row_no=row_no,
            progress=progress,
        )

    loc = find_save_count_locator(work, prefer_value="3")
    if loc is None:
        _log(progress, "오류", f"5) 저장상품수 입력칸 미검출 · 목표={target}")
        if shot_dir is not None:
            screenshot_step(
                shot_page,
                shot_dir,
                step_tag="05_save_count_after",
                label=f"5)저장상품수 갱신 후 (실패) 칸미검출 목표={target}",
                row_no=row_no,
                progress=progress,
            )
        return False

    before_val = ""
    try:
        before_val = (loc.input_value(timeout=500) or "").strip()
    except Exception:  # noqa: BLE001
        before_val = ""
    if out is not None:
        out["before"] = before_val

    # ★단발 입력 — 값 대체 후 blur 로 확정 (Escape 로 원복되는 문제 없음)
    filled = False
    try:
        loc.fill(target, timeout=3_000)
        filled = True
    except Exception:  # noqa: BLE001
        try:
            filled = bool(
                loc.evaluate(
                    """(el, want) => {
                      el.focus();
                      el.value = String(want);
                      el.dispatchEvent(new Event('input', {bubbles:true}));
                      el.dispatchEvent(new Event('change', {bubbles:true}));
                      el.blur();
                      return (el.value || '').trim() === String(want);
                    }""",
                    target,
                )
            )
        except Exception:  # noqa: BLE001
            filled = False
    if filled:
        try:
            loc.evaluate("(el) => { el.blur(); }")
        except Exception:  # noqa: BLE001
            pass

    after_val = ""
    try:
        after_val = (loc.input_value(timeout=500) or "").strip()
    except Exception:  # noqa: BLE001
        after_val = target if filled else ""
    if out is not None:
        out["after"] = after_val or (target if filled else "")

    _log(
        progress,
        "로직",
        f"5) 저장상품수 입력 {before_val or '?'} → {after_val or target} (목표={target})",
    )
    if shot_dir is not None:
        status = "성공" if filled else "실패"
        screenshot_step(
            shot_page,
            shot_dir,
            step_tag="05_save_count_after",
            label=f"5)저장상품수 갱신 후 ({status}) {before_val or '?'}→{after_val or target}",
            row_no=row_no,
            progress=progress,
        )
    return filled


def new_shot_dir() -> Path:
    """P3 실행별 스크린샷 폴더."""
    ts = time.strftime("%Y%m%d_%H%M%S")
    d = P3_RUN_LOG_DIR / ts
    d.mkdir(parents=True, exist_ok=True)
    return d


def find_save_count_locator(page, prefer_value: str = "3"):
    """저장상품수 숫자 입력칸 — 우선 현재값이 prefer_value(기본 '3')인 input.

    스크린샷: 검색결과 상위 [ 3 ] 개 상품만 저장
    """
    prefer = (prefer_value or "3").strip()

    # 0) JS로 value===prefer (문맥: 저장상품수/상위/개) 인 요소 핸들
    try:
        handle = page.evaluate_handle(
            """(prefer) => {
              const isNumInput = (inp) => {
                if (!inp) return false;
                const ty = (inp.getAttribute('type') || 'text').toLowerCase();
                if (!(ty === 'text' || ty === 'number' || ty === '')) return false;
                if (inp.disabled || inp.readOnly) return false;
                return true;
              };
              const inCtx = (inp) => {
                const tr = inp.closest('tr');
                const scope = tr || inp.closest('td,div') || inp.parentElement;
                const t = ((scope && scope.innerText) || '').replace(/\\s+/g, '');
                return t.includes('저장상품수')
                  || (t.includes('검색결과') && t.includes('상위'))
                  || (t.includes('상위') && t.includes('개'));
              };
              const all = Array.from(document.querySelectorAll(
                'input[type="text"], input[type="number"], input:not([type])'
              )).filter(isNumInput);
              return all.find(i => (i.value || '').trim() === String(prefer) && inCtx(i))
                || all.find(i => (i.value || '').trim() === String(prefer))
                || all.find(i => inCtx(i) && /^\\d+$/.test((i.value || '').trim()))
                || all.find(i => inCtx(i))
                || null;
            }""",
            prefer,
        )
        el = handle.as_element()
        if el is not None:
            # ElementHandle → Locator 대신 직접 쓰기 위해 wrapper
            # Playwright: page.locator로 재검색이 더 안정적
            pass
    except Exception:
        handle = None
        el = None

    # 1) value=prefer 정확 매칭 (저장상품수 행)
    try:
        loc = page.locator(
            "xpath=//tr[.//text()[contains(.,'저장상품수')] or "
            ".//*[contains(normalize-space(.),'저장상품수')]]"
            f"//input[(@type='text' or @type='number' or not(@type)) and @value='{prefer}']"
        ).first
        if loc.count() > 0 and loc.is_visible(timeout=400):
            return loc
    except Exception:
        pass

    # 2) '검색결과 상위' 셀 안 value=prefer
    try:
        loc = page.locator(
            "xpath=//td[contains(.,'검색결과') and contains(.,'상위') and contains(.,'개')]"
            f"//input[(@type='text' or @type='number' or not(@type)) and @value='{prefer}']"
        ).first
        if loc.count() > 0 and loc.is_visible(timeout=400):
            return loc
    except Exception:
        pass

    # 3) 화면에 보이는 input 중 value가 prefer 인 것 (런타임 value)
    try:
        cands = page.locator(
            'input[type="text"], input[type="number"], input:not([type])'
        )
        n = min(cands.count(), 40)
        for i in range(n):
            el = cands.nth(i)
            try:
                if not el.is_visible(timeout=150):
                    continue
                v = (el.input_value(timeout=200) or "").strip()
                if v != prefer:
                    continue
                # 부모 텍스트에 상위/개 있으면 채택
                ok_ctx = el.evaluate(
                    """(el) => {
                      const tr = el.closest('tr');
                      const t = ((tr && tr.innerText) || el.parentElement?.innerText || '')
                        .replace(/\\s+/g, '');
                      return t.includes('저장상품수')
                        || (t.includes('상위') && t.includes('개'))
                        || t.includes('검색결과');
                    }"""
                )
                if ok_ctx:
                    return el
            except Exception:
                continue
    except Exception:
        pass

    # 4) 기존 폴백: 저장상품수 행의 숫자 input
    selectors = (
        "xpath=//td[contains(.,'검색결과') and contains(.,'상위') and contains(.,'개')]"
        "//input[@type='text' or @type='number' or not(@type)]",
        "xpath=//tr[.//th[contains(normalize-space(.),'저장상품수')] or "
        ".//td[normalize-space()='저장상품수'] or "
        ".//td[starts-with(normalize-space(.),'저장상품수')] or "
        ".//*[contains(normalize-space(.),'저장상품수')]]"
        "//input[@type='text' or @type='number' or not(@type)]",
        "xpath=//*[contains(normalize-space(.),'개 상품만 저장')]"
        "/preceding::input[@type='text' or @type='number' or not(@type)][1]",
    )
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() == 0:
                continue
            if not loc.is_visible(timeout=400):
                continue
            try:
                v = (loc.input_value(timeout=300) or "").strip()
                if v and not re.fullmatch(r"\d+", v):
                    if len(v) > 8 or "http" in v.lower():
                        continue
            except Exception:
                pass
            return loc
        except Exception:
            continue
    return None


def _all_pages_and_frames(page):
    """열려있는 모든 page(팝업 포함) + 그 안의 frame 을 (target, kind)로 나열.

    현재 page 를 항상 먼저 반환한다. ★팝업이 page 와 다른 BrowserContext 에
    열리는 경우까지 대비해, page.context 뿐 아니라 같은 browser 의 모든
    context 를 함께 탐색한다 (admin_etc_ok.php 같은 확인창이 별도 컨텍스트로
    뜨는 경우에도 놓치지 않도록).
    """
    ordered: list[tuple] = [("page", page)]
    seen_ids = {id(page)}

    def _add_pages(pages) -> None:
        for p in pages:
            if id(p) in seen_ids:
                continue
            seen_ids.add(id(p))
            ordered.append(("page", p))

    try:
        _add_pages(page.context.pages)
    except Exception:
        pass
    try:
        for ctx in page.context.browser.contexts:
            try:
                _add_pages(ctx.pages)
            except Exception:
                continue
    except Exception:
        pass

    out: list[tuple] = list(ordered)
    for _kind, p in ordered:
        try:
            for fr in p.frames:
                out.append(("frame", fr))
        except Exception:
            continue
    return out


def resolve_modify_target(page):
    """검색필터 수정(저장상품수) 화면이 열린 page/frame 을 찾는다.

    팝업 창·iframe 모두 탐색. (page, kind) 반환. 없으면 (page, 'main').
    """
    for kind, p in _all_pages_and_frames(page):
        try:
            url = p.url or ""
            if "modify_filter" in url or "admin_group_modify" in url:
                return p, kind
        except Exception:
            pass
        try:
            body = p.locator("body").inner_text(timeout=400) or ""
            if "저장상품수" in body and (
                "검색필터 수정" in body or "검색결과" in body or "저장하기" in body
            ):
                return p, kind
        except Exception:
            pass
    return page, "main"


def wait_for_save_count_ready(target, *, timeout_ms: int = 8000) -> bool:
    """저장상품수 입력칸이 보일 때까지 대기."""
    end = time.time() + timeout_ms / 1000.0
    while time.time() < end:
        try:
            # Page or Frame both have locator
            if find_save_count_locator(target) is not None:
                return True
        except Exception:
            pass
        try:
            body = ""
            if hasattr(target, "locator"):
                body = target.locator("body").inner_text(timeout=300) or ""
            if "저장상품수" in body or "개 상품만 저장" in body:
                # 문구는 있는데 locator 실패 — 한 번 더 여유
                time.sleep(0.3)
                if find_save_count_locator(target) is not None:
                    return True
        except Exception:
            pass
        time.sleep(0.25)
    return find_save_count_locator(target) is not None


def _capture_png(page, path: Path, *, timeout_ms: int = 3000) -> bool:
    """뷰포트 PNG 캡처 — Playwright 실패 시 CDP 폴백(샷이 본작업을 오래 막지 않음)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        page.screenshot(
            path=str(path),
            timeout=timeout_ms,
            animations="disabled",
        )
        if path.is_file() and path.stat().st_size > 0:
            return True
    except Exception:
        pass
    # CDP 폴백
    try:
        session = page.context.new_cdp_session(page)
        result = session.send("Page.captureScreenshot", {"format": "png", "fromSurface": True})
        import base64

        data = base64.b64decode(result.get("data") or "")
        if data:
            path.write_bytes(data)
            return path.is_file() and path.stat().st_size > 0
    except Exception:
        pass
    return False


def screenshot_step(
    page,
    shot_dir: Path | None,
    *,
    step_tag: str,
    label: str,
    row_no: int = 0,
    progress: ProgressFn | None = None,
    full_page: bool = False,
) -> Path | None:
    """필터 일치 행의 단계 스크린샷 → 실행 로그에 출력."""
    if shot_dir is None:
        return None
    shot_dir.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^\w\-]+", "_", step_tag)[:48]
    name = f"r{row_no:03d}_{safe}.png"
    path = shot_dir / name
    time.sleep(0.05)
    ok = False
    if full_page:
        try:
            page.screenshot(
                path=str(path),
                full_page=True,
                timeout=4_000,
                animations="disabled",
            )
            ok = path.is_file() and path.stat().st_size > 0
        except Exception:
            ok = False
    if not ok:
        ok = _capture_png(page, path, timeout_ms=3_000)
    if not ok:
        _log(progress, "샷", f"[샷 실패] {label}: 캡처 불가(타임아웃/CDP)", major=False)
        return None

    # 스크린샷은 항상 세부내용(SUB) — 보드에서 SUB 하단 스크린샷 그리드에 그려진다
    _log(progress, "샷", f"{label} -> {path}", major=False)
    print(f"{P3_SHOT_MARK}{path}##{label}", flush=True)
    return path


def _page_for_edit_shot(page, prefer_page=None):
    """수집조건수정 클릭 후 샷 대상 — 팝업/수정화면 우선."""
    if prefer_page is not None:
        try:
            if not prefer_page.is_closed():
                return prefer_page
        except Exception:
            pass
    try:
        target, kind = resolve_modify_target(page)
        if kind in ("page", "frame") and target is not None:
            return target if kind == "page" else page
    except Exception:
        pass
    try:
        pages = list(page.context.pages)
        for p in reversed(pages):
            if p is page:
                continue
            try:
                if not p.is_closed():
                    return p
            except Exception:
                continue
    except Exception:
        pass
    return page


def screenshot_after_edit_click_series(
    page,
    shot_dir: Path | None,
    *,
    row_no: int = 0,
    progress: ProgressFn | None = None,
    count: int = 3,
    interval_s: float = 3.0,
    prefer_page=None,
) -> list[Path]:
    """수집조건수정 클릭 직후 — 3초 간격 스크린샷 3장 (실행로그·뷰어 출력)."""
    out: list[Path] = []
    if shot_dir is None:
        return out
    n = max(1, int(count))
    gap = max(0.0, float(interval_s))
    _log(
        progress,
        "로직",
        f"2) 수집조건수정 클릭 직후 스크린샷 {n}장 ({gap:g}초 간격)",
    )
    for i in range(1, n + 1):
        elapsed = int(round((i - 1) * gap))
        target = _page_for_edit_shot(page, prefer_page=prefer_page)
        path = screenshot_step(
            target,
            shot_dir,
            step_tag=f"05_after_edit_{i}of{n}",
            label=f"5)수집조건수정 클릭후 샷 {i}/{n} (+{elapsed}s)",
            row_no=row_no,
            progress=progress,
        )
        if path is not None:
            out.append(path)
        if i < n and gap > 0:
            time.sleep(gap)
    return out


def screenshot_save_count_grid(
    page,
    loc,
    shot_dir: Path,
    *,
    tag: str,
    row_no: int = 0,
    note: str = "",
    progress: ProgressFn | None = None,
) -> Path | None:
    """판단한 저장상품수 입력그리드(행) 근접 스크린샷 → 로그 출력."""
    shot_dir.mkdir(parents=True, exist_ok=True)
    safe_tag = re.sub(r"[^\w\-]+", "_", tag)[:40]
    name = f"r{row_no:03d}_save_count_{safe_tag}.png"
    path = shot_dir / name
    label = f"저장상품수 입력그리드/{tag}"
    if note:
        label = f"{label} ({note})"

    ok = False
    if loc is not None:
        try:
            row = loc.locator("xpath=ancestor::tr[1]")
            if row.count() > 0:
                row.first.scroll_into_view_if_needed(timeout=1000)
                time.sleep(0.08)
                row.first.screenshot(path=str(path), timeout=3_000, animations="disabled")
                ok = path.is_file() and path.stat().st_size > 0
        except Exception:
            ok = False

        if not ok:
            try:
                loc.scroll_into_view_if_needed(timeout=1000)
                box = loc.bounding_box()
                if box:
                    pad_l, pad_r, pad_y = 160, 220, 28
                    clip = {
                        "x": max(0, box["x"] - pad_l),
                        "y": max(0, box["y"] - pad_y),
                        "width": max(80, box["width"] + pad_l + pad_r),
                        "height": max(40, box["height"] + pad_y * 2),
                    }
                    page.screenshot(
                        path=str(path),
                        clip=clip,
                        timeout=3_000,
                        animations="disabled",
                    )
                    ok = path.is_file() and path.stat().st_size > 0
            except Exception:
                ok = False

    if not ok:
        ok = _capture_png(page, path, timeout_ms=3_000)

    if not ok:
        _log(progress, "샷", f"[샷 실패] {label}: 캡처 불가", major=False)
        return None

    _log(progress, "샷", f"{label} -> {path}", major=False)
    print(f"{P3_SHOT_MARK}{path}##{label}", flush=True)
    return path


def click_save_button(page) -> bool:
    """검색필터 수정 화면(팝업·프레임 포함) 하단 '저장하기' (옆에 '닫기') 클릭.

    ★수정화면이 원래 page 가 아닌 별도 팝업/프레임에서 열릴 수 있으므로,
    set_save_count 와 동일하게 resolve_modify_target 으로 실제 위치를 찾아 클릭한다.
    """
    work, _kind = resolve_modify_target(page)
    selectors = (
        'input[type="submit"][value="저장하기"]',
        'input[type="button"][value="저장하기"]',
        'input[value="저장하기"]',
        'button:has-text("저장하기")',
        'a:has-text("저장하기")',
    )
    targets = [work]
    for _kk, p in _all_pages_and_frames(page):
        if p is not work:
            targets.append(p)
    for tgt in targets:
        for sel in selectors:
            try:
                loc = tgt.locator(sel).last
                if loc.count() > 0 and loc.is_visible(timeout=500):
                    loc.click(timeout=3000, force=True)
                    return True
            except Exception:
                continue
    for tgt in targets:
        try:
            clicked = tgt.evaluate(
                """() => {
                  const nodes = Array.from(document.querySelectorAll('a,button,input'));
                  for (const el of nodes) {
                    const t = (el.value || el.textContent || '').replace(/\\s+/g, '');
                    if (t === '저장하기') { el.click(); return true; }
                  }
                  return false;
                }"""
            )
            if clicked:
                return True
        except Exception:
            continue
    return False


def attach_native_dialog_handler(page) -> dict:
    """브라우저 alert/confirm 대비 — '수정되었습니다' 등 네이티브 다이얼로그 accept.

    저장하기 클릭 *전에* 등록해야 한다.
    ★수정화면이 원래 page 가 아닌 별도 팝업에서 열릴 수 있으므로, 현재 열려있는
    모든 page 와 이후 새로 열리는 page 에도 동일하게 등록한다 — 그래야 팝업에서
    뜨는 네이티브 alert/confirm 도 놓치지 않는다.
    """
    state: dict = {"seen": False, "message": "", "accepted": False}

    def _on_dialog(dialog) -> None:  # noqa: ANN001
        try:
            state["seen"] = True
            state["message"] = dialog.message or ""
            dialog.accept()
            state["accepted"] = True
        except Exception:
            try:
                dialog.dismiss()
            except Exception:
                pass

    try:
        for p in page.context.pages:
            try:
                p.on("dialog", _on_dialog)
            except Exception:
                continue
    except Exception:
        try:
            page.on("dialog", _on_dialog)
        except Exception:
            pass

    try:
        page.context.on("page", lambda new_page: new_page.on("dialog", _on_dialog))
    except Exception:
        pass

    return state


def is_modify_page_open(page) -> bool:
    """검색필터 수정 팝업/페이지(팝업·프레임 포함)가 열려 있는지."""
    for _kind, p in _all_pages_and_frames(page):
        try:
            url = p.url or ""
            if "modify_filter" in url or "admin_group_modify" in url:
                return True
        except Exception:
            pass
        try:
            body = p.locator("body").inner_text(timeout=300) or ""
            if "검색필터 수정" in body and "저장상품수" in body:
                return True
            if "저장상품수" in body and "검색결과" in body and "저장하기" in body:
                return True
        except Exception:
            continue
    return False


def wait_modify_page_closed(page, *, timeout_ms: int = 20000) -> bool:
    """저장하기 후 '검색필터 수정' 팝업/페이지(팝업·프레임 포함) 닫힘 확인."""
    end = time.time() + timeout_ms / 1000.0
    # 잠깐은 열려 있을 수 있음 — 닫힐 때까지 대기
    while time.time() < end:
        if not is_modify_page_open(page):
            return True
        # 수정되었습니다 팝업이 뜨면 수정화면은 사실상 닫힌 것으로 본다
        for _kind, p in _all_pages_and_frames(page):
            try:
                body = p.locator("body").inner_text(timeout=300) or ""
                if "수정되었습니다" in body or "수정 되었습니다" in body:
                    return True
            except Exception:
                continue
        time.sleep(0.2)
    return not is_modify_page_open(page)


_CONFIRM_CLICK_JS = """() => {
  const nodes = Array.from(document.querySelectorAll('a,button,input'));
  for (const el of nodes) {
    const t = (el.value || el.textContent || '').replace(/\\s+/g, '');
    if (t === '확인') { el.click(); return true; }
  }
  return false;
}"""


def click_modified_confirm(
    page,
    *,
    timeout_ms: int = 8000,
    dialog_state: dict | None = None,
) -> bool:
    """★요건: '저장하기' 다음에 팝업이 뜨면 조건 없이 「확인」 을 클릭한다.

    다른 것은 일절 따지지 않는다(메시지 내용·질문형 여부·수정화면 상태 무시).
    - 네이티브 alert/confirm: attach_native_dialog_handler 가 이미 수락 → 완료
    - HTML 팝업: LABEL 이 '확인' 인 요소를 찾으면 그대로 클릭 → 완료
    """

    def _dialog_handled() -> bool:
        if not dialog_state:
            return False
        return bool(dialog_state.get("accepted") or dialog_state.get("seen"))

    if _dialog_handled():
        return True

    end = time.time() + timeout_ms / 1000.0
    while time.time() < end:
        if _dialog_handled():
            return True
        for _kind, tgt in _all_pages_and_frames(page):
            try:
                if tgt.evaluate(_CONFIRM_CLICK_JS):
                    time.sleep(0.3)
                    return True
            except Exception:  # noqa: BLE001
                continue
        time.sleep(0.2)
    return _dialog_handled()


def wait_modify_page(page, *, timeout_ms: int = 20000) -> bool:
    """수집조건수정 후 '검색필터 수정' / 저장상품수 화면 대기.

    팝업 창·iframe 포함. not found 화면이면 실패.
    """
    end = time.time() + timeout_ms / 1000.0
    saw_not_found = False
    while time.time() < end:
        try:
            if page_shows_not_found(page):
                saw_not_found = True
                # 잠깐 더 기다려 정상 화면으로 바뀌는지 확인
                time.sleep(0.35)
                if page_shows_not_found(page):
                    return False
            target, kind = resolve_modify_target(page)
            url = ""
            try:
                url = getattr(target, "url", None) or page.url or ""
            except Exception:
                url = page.url or ""
            if "modify_filter" in url or "admin_group_modify" in url:
                if not page_shows_not_found(page):
                    # URL만 맞고 본문이 not found 인 경우 제외
                    if wait_for_save_count_ready(target, timeout_ms=400):
                        return True
                    body = ""
                    try:
                        body = target.locator("body").inner_text(timeout=300) or ""
                    except Exception:
                        body = ""
                    if "저장상품수" in body or "검색필터 수정" in body:
                        return True
            if kind in ("page", "frame") and kind != "main":
                # resolve 가 저장상품수 화면을 찾음
                if wait_for_save_count_ready(target, timeout_ms=500):
                    return True
            body = ""
            try:
                body = page.locator("body").inner_text(timeout=400) or ""
            except Exception:
                pass
            if "저장상품수" in body and (
                "검색필터 수정" in body or "검색결과" in body
            ):
                return True
            # 다른 탭/팝업
            try:
                for p in page.context.pages:
                    if p is page:
                        continue
                    if page_shows_not_found(p):
                        saw_not_found = True
                        continue
                    bu = p.url or ""
                    if "modify_filter" in bu or "admin_group_modify" in bu:
                        if wait_for_save_count_ready(p, timeout_ms=400):
                            return True
                        bt = p.locator("body").inner_text(timeout=300) or ""
                        if "저장상품수" in bt or "검색필터 수정" in bt:
                            return True
                    bt = p.locator("body").inner_text(timeout=300) or ""
                    if "저장상품수" in bt:
                        return True
            except Exception:
                pass
        except Exception:
            pass
        time.sleep(0.25)
    if saw_not_found and page_shows_not_found(page):
        return False
    target, _kind = resolve_modify_target(page)
    if page_shows_not_found(page):
        return False
    return wait_for_save_count_ready(target, timeout_ms=800)


def _return_to_list(page, list_url: str) -> None:
    """저장 후 검색필터 목록으로 복귀."""
    url = (list_url or "").strip()
    if not url:
        return
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=60_000)
        time.sleep(0.5)
    except Exception:
        try:
            page.go_back(wait_until="domcontentloaded", timeout=30_000)
            time.sleep(0.4)
        except Exception:
            pass


def run_update(
    excel_path: str | Path,
    mango_url: str,
    *,
    progress: ProgressFn | None = None,
) -> RunResult:
    path = Path(excel_path).expanduser().resolve()
    result = RunResult(ok=False)
    if not path.is_file():
        result.errors.append(f"파일 없음: {path}")
        return result
    # ★고정: 비어 있으면 지정된 검색필터 URL 초기값 사용
    mango = (mango_url or "").strip() or DEFAULT_MANGO_URL

    clear_stop_flag()
    try:
        rows = read_excel_rows(path)
    except Exception as e:  # noqa: BLE001
        result.errors.append(str(e))
        return result
    if not rows:
        result.errors.append("엑셀에 URL 행이 없습니다.")
        return result
    _log(
        progress,
        "준비",
        f"엑셀 {path.name} · URL {len(rows)}건 (엑셀 첫행부터 순차 처리) · "
        f"망고URL={mango[:120]}",
    )
    _log(
        progress,
        "준비",
        "P2와 동일 — 망고 Chrome 창을 화면에 띄운 뒤 검색필터 URL로 이동 (로그인대기 없음)",
    )

    try:
        from playwright.sync_api import sync_playwright
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"Playwright 미설치: {e}")
        return result

    try:
        import collect as p2  # type: ignore
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"P2 collect 로드 실패: {e}")
        return result

    try:
        with sync_playwright() as p:
            # ★P2 connect_browser → 창 앞으로 → 검색필터 URL 이동
            try:
                browser, page = attach_mango_browser_like_p2(p2, p, progress=progress)
            except Exception as e:  # noqa: BLE001
                result.errors.append(str(e))
                return result
            page = navigate_mango_url(page, mango, progress=progress, p2=p2) or page
            reveal_browser_page(
                page,
                progress,
                step_no="0",
                action="검색필터 목록 화면 (연동 동작 시작)",
                dwell_s=0.3,
            )
            shot_dir = new_shot_dir()
            _log(progress, "준비", f"스크린샷 폴더: {shot_dir}")
            try:
                cur_after = (page.url or "").strip()
            except Exception:
                cur_after = ""
            _log(progress, "준비", f"목록스캔 전 URL={cur_after[:180] or '(없음)'}")

            demango_rows = list_demango_rows(page)
            result.total_demango = len(demango_rows)
            _log(progress, "준비", f"더망고 목록 {result.total_demango}행 검출")

            if result.total_demango == 0:
                result.errors.append(
                    "검색필터 행 0건 — "
                    f"목표URL={mango[:100]} · 현재URL={cur_after[:100] or '(없음)'}. "
                    "로그인된 망고 Chrome에서 다시 실행하세요."
                )
                return result

            processed_no = 0  # 실제 처리(매칭)한 행 순번 — 지연 적용 판단용
            done_keys: set[tuple[str, str]] = set()  # 이미 갱신한 망고 행
            # ★요건: 1) 엑셀 첫행부터 순차로 읽고 → 2) 그 「URL KEY」로 망고 행을 찾는다
            for i, ex in enumerate(rows, start=1):
                if stop_requested():
                    _log(progress, "중단", "사용자 중단 요청")
                    break

                # 화면 엑셀 목록에 지금 작업 중인 행 표시 (▶)
                _emit_progress_row(i)

                # 2) 엑셀 URL KEY → 망고행 찾기 (같은 URL 행이 여러 개면 전체 갱신)
                matches = find_demango_rows_for_excel(
                    page, ex, progress=progress, done_keys=done_keys
                )
                # ★매칭되지 않는 정보는 로그에 남기지 않는다.
                if not matches:
                    result.skipped += 1
                    continue
                if len(matches) >= 2:
                    # ★요건: 동일 URL 행이 2개 이상이면 몇 개인지·URL 을 적색으로 구분 표시
                    _log(
                        progress,
                        "로직",
                        _red(
                            f"2) 동일 URL 망고행 {len(matches)}개 — 전체 갱신 · "
                            f"URL={ex.url}"
                        ),
                        major=True,
                    )
                    _log(
                        progress,
                        "로직",
                        _red(
                            f"2) 동일 URL {len(matches)}개 행 필터이름="
                            + ", ".join(
                                str(m.get("filterName") or "?") for m in matches
                            )
                            + f" · URL={ex.url}"
                        ),
                    )

                for drow in matches:
                    if stop_requested():
                        break
                    d_url = (drow.get("url") or "").strip()
                    d_filter = (drow.get("filterName") or "").strip()
                    row_idx = int(drow.get("index") or 0)
                    edit_href = (drow.get("editHref") or "").strip()
                    # 망고 행에 보이는 정보 원문(필터명·URL 검색·수집계수·전체저장 …)
                    d_row_text = " ".join((drow.get("text") or "").split())

                    d_fuid = str(drow.get("fuid") or "").strip()
                    same_filter = filters_equal(ex.filter_name, d_filter)
                    # 동일 URL 행이 여러 개면 필터이름이 달라도 전체 갱신한다(요건).
                    # 행이 하나뿐일 때는 기존대로 필터 불일치면 건너뛴다.
                    if (
                        len(matches) == 1
                        and ex.filter_name
                        and d_filter
                        and not same_filter
                    ):
                        result.skipped += 1
                        continue
                    done_keys.add(row_done_key(d_url, d_filter, d_fuid))

                    # 처음 5개 처리행만 동작마다 3초 대기 (요건) — 건너뛴 행은 세지 않는다
                    processed_no += 1
                    # ★요건: URL 은 2단계에서 한 번만 · 나머지 단계는 URL 없이 간단히
                    _log(
                        progress,
                        "로직",
                        f"2) 망고행 · 필터={d_filter or '?'} · 상품수={ex.collectible} · "
                        f"URL={d_url}",
                        major=True,
                    )
                    _log(progress, "로직", f"2) 망고행 원문: {d_row_text or '(없음)'}")
                    _log(
                        progress,
                        "로직",
                        f"2) 엑셀 {ex.excel_row}행 매칭 · 망고행index={row_idx}",
                    )
                    note = filter_compare_note(ex.filter_name, d_filter)
                    if note:
                        _log(progress, "로직", f"2) {note}")
                    if not same_filter and ex.filter_name and d_filter:
                        _log(
                            progress,
                            "로직",
                            _red(
                                f"2) 필터이름 다름(동일 URL이라 갱신) · 엑셀={ex.filter_name} · "
                                f"망고={d_filter} · URL={d_url}"
                            ),
                        )

                    # ★요건: 4) 망고 행 「URL 검색」 주소로 상품수를 읽어오는 부분은 전부
                    #   주석처리한다. 그 주소로는 어떤 화면도 불러오지 않는다(절대 금지).
                    #   「URL 검색」 주소는 엑셀자료 비교(KEY)에만 쓴다.
                    #   (추후 완성본에서 되살릴 때 아래 주석을 해제한다)
                    #
                    # list_page = page
                    # store = click_demango_row_url(
                    #     list_page, row_idx, d_url, progress=progress
                    # )
                    # if store is None:
                    #     _log(progress, "오류", f"엑셀{ex.excel_row}행 · 4) URL 클릭 실패 · …")
                    #     _return_to_list(list_page, mango)
                    # else:
                    #     card_n, matched = browse_store_count_cards(
                    #         store, excel_count=ex.collectible, progress=progress,
                    #         shot_dir=shot_dir, row_no=i,
                    #     )
                    #     page = close_store_return_list(
                    #         list_page, store, mango, progress=progress
                    #     )
                    #     row_idx2 = resolve_demango_row_index_by_url(
                    #         page, d_url, fallback_index=row_idx, progress=progress
                    #     )
                    #     if row_idx2 is not None:
                    #         row_idx = int(row_idx2)
                    _log(
                        progress,
                        "로직",
                        f"4) 상품수 {ex.collectible} (엑셀값 사용 · URL 화면 안 열음)",
                        major=True,
                    )

                    # 5) LABEL '수집조건수정' 버튼 클릭 → 저장상품수 입력 → '저장하기' 클릭
                    target = map_save_count(ex.collectible)
                    if not page_is_usable(page):
                        result.failed += 1
                        _log(
                            progress,
                            "오류",
                            f"엑셀{ex.excel_row}행 · 필터={d_filter} · "
                            "더망고 페이지 핸들 사용불가(닫힘/크래시)",
                        )
                        continue
                    # ★목록은 수정일 정렬이라 앞 행을 갱신하면 순서가 바뀐다. 클릭 직전에
                    #   현재 화면에서 이 행(URL+필터이름)의 index 를 다시 확정한다.
                    row_idx2 = resolve_demango_row_index_by_url(
                        page,
                        d_url,
                        fallback_index=row_idx,
                        filter_hint=d_filter,
                        fuid_hint=d_fuid,
                        progress=progress,
                    )
                    if row_idx2 is not None:
                        row_idx = int(row_idx2)
                    if not click_edit_on_row(
                        page,
                        row_idx,
                        edit_href,
                        row_url=d_url,
                        filter_hint=d_filter,
                        fuid_hint=d_fuid,
                        progress=progress,
                        shot_dir=shot_dir,
                        row_no=i,
                        max_tries=EDIT_CLICK_MAX_TRIES,
                        try_interval_s=0.3,  # ★컴퓨터 속도
                    ):
                        result.failed += 1
                        _log(
                            progress,
                            "오류",
                            f"5) 수집조건수정 오류 · 필터={d_filter} · 목표={target} · "
                            "사유=버튼 미검출/팝업 미오픈",
                        )
                        screenshot_step(
                            page,
                            shot_dir,
                            step_tag="05_edit_fail",
                            label="5)수집조건수정 실패",
                            row_no=i,
                            progress=progress,
                        )
                        _return_to_list(page, mango)
                        continue
                    if not wait_modify_page(page):
                        result.failed += 1
                        reason = (
                            "not found — 잘못된 버튼/링크 가능"
                            if page_shows_not_found(page)
                            else "검색필터 수정 화면 미진입(팝업 열렸으나 화면 미표시)"
                        )
                        _log(
                            progress,
                            "오류",
                            f"5) 수집조건수정 오류(수정화면 미진입) · 필터={d_filter} · "
                            f"사유={reason}",
                        )
                        screenshot_step(
                            page,
                            shot_dir,
                            step_tag="05_modify_missing",
                            label="5)검색필터수정 미진입/notfound",
                            row_no=i,
                            progress=progress,
                        )
                        _return_to_list(page, mango)
                        continue
                    try:
                        mod_page, _kind = resolve_modify_target(page)
                    except Exception:
                        mod_page = page
                    reveal_browser_page(
                        mod_page if mod_page is not None else page,
                        progress,
                        step_no="5",
                        action="검색필터 수정 팝업/화면 표시",
                        dwell_s=0.0,
                    )
                    screenshot_step(
                        mod_page if mod_page is not None else page,
                        shot_dir,
                        step_tag="05_modify_opened",
                        label="5)검색필터 수정 화면",
                        row_no=i,
                        progress=progress,
                    )

                    # 5) 「저장상품수」란 「검색결과 상위 [ ]개만 상품만 저장」에 엑셀 상품수 입력
                    count_io: dict = {}
                    if not set_save_count(
                        page,
                        target,
                        shot_dir=shot_dir,
                        progress=progress,
                        row_no=i,
                        out=count_io,
                    ):
                        result.failed += 1
                        _log(
                            progress,
                            "오류",
                            f"5) 수집조건수정 OK → 상품수입력 오류 · 필터={d_filter} · "
                            f"목표={target}",
                        )
                        try:
                            page.keyboard.press("Escape")
                        except Exception:
                            pass
                        _return_to_list(page, mango)
                        continue

                    # 5) LABEL '저장하기' 버튼 클릭 (★'저장'이 아닌 '저장하기' 텍스트를 찾아 클릭)
                    dialog_state = attach_native_dialog_handler(page)
                    if not click_save_button(page):
                        result.failed += 1
                        _log(
                            progress,
                            "오류",
                            f"5) 수집조건수정 OK → 저장하기 오류 · 필터={d_filter} · "
                            f"목표={target}",
                        )
                        screenshot_step(
                            page,
                            shot_dir,
                            step_tag="05_save_click_fail",
                            label="5)저장하기 클릭 실패",
                            row_no=i,
                            progress=progress,
                        )
                        _return_to_list(page, mango)
                        continue
                    screenshot_step(
                        page,
                        shot_dir,
                        step_tag="05_after_save_click",
                        label="5)저장하기 클릭 후",
                        row_no=i,
                        progress=progress,
                    )
                    # ★요건: 상품수 갱신 전·후를 5단계 '저장하기' 로그에 그대로 표출
                    before_cnt = str(count_io.get("before") or "?")
                    after_cnt = str(count_io.get("after") or target)

                    if not wait_modify_page_closed(page, timeout_ms=6_000):
                        _log(
                            progress,
                            "경고",
                            f"엑셀{ex.excel_row}행 수정팝업 닫힘 대기 시간초과 — 확인 계속 시도",
                        )
                    screenshot_step(
                        page,
                        shot_dir,
                        step_tag="05_modify_closed",
                        label="5)수정팝업 닫힘/확인대기",
                        row_no=i,
                        progress=progress,
                    )

                    # 6) '수정되었습니다' 메세지 하단 LABEL '확인' 버튼 클릭
                    if not click_modified_confirm(
                        page, timeout_ms=5_000, dialog_state=dialog_state
                    ):
                        result.failed += 1
                        _log(
                            progress,
                            "오류",
                            f"5) 수집조건수정 OK → 저장하기 OK → 확인 오류 · "
                            f"필터={d_filter} · 목표={target} · "
                            f"다이얼로그감지={dialog_state.get('seen')} · "
                            f"수정화면열림={is_modify_page_open(page)}",
                        )
                        screenshot_step(
                            page,
                            shot_dir,
                            step_tag="06_confirm_fail",
                            label="6)확인 클릭 실패",
                            row_no=i,
                            progress=progress,
                        )
                        _return_to_list(page, mango)
                        continue
                    screenshot_step(
                        page,
                        shot_dir,
                        step_tag="06_confirmed",
                        label="6)확인 클릭 완료",
                        row_no=i,
                        progress=progress,
                    )
                    # ★요건: 5단계는 한 줄 요약 — 수집조건수정 → 저장하기 → 확인 OK
                    _log(
                        progress,
                        "로직",
                        f"5) 수집조건수정 → 저장하기 → 확인 OK · 상품수 {before_cnt} → "
                        f"{after_cnt}",
                        major=True,
                    )

                    result.updated += 1
                    _log(
                        progress,
                        "완료",
                        f"엑셀{ex.excel_row}행 갱신성공 · 저장상품수={target}",
                    )
                    _return_to_list(page, mango)
                    screenshot_step(
                        page,
                        shot_dir,
                        step_tag="07_back_to_list",
                        label="7)목록복귀 → 다음 행 반복",
                        row_no=i,
                        progress=progress,
                    )
                    _log(
                        progress,
                        "로직",
                        f"7) 갱신 완료 (저장상품수 {target}) → 다음 행",
                        major=True,
                    )

                    if stop_requested():
                        break

    except Exception as e:  # noqa: BLE001
        result.errors.append(f"실행 실패: {e}")
        _log(progress, "오류", str(e))
        return result

    clear_stop_flag()
    _emit_progress_row(0)
    result.ok = result.updated > 0 and not result.errors
    _log(
        progress,
        "완료",
        f"갱신 {result.updated} · 건너뜀 {result.skipped} · 실패 {result.failed} "
        f"/ 엑셀 {len(rows)}행 (망고 목록 {result.total_demango}행)",
    )
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="P3_필터_갱신 — 저장상품수 갱신")
    parser.add_argument("excel", help="엑셀 파일 경로")
    parser.add_argument(
        "--mango-url",
        default=DEFAULT_MANGO_URL,
        help="더망고 검색필터 URL (기본=getGoodsCategory.php filter_delete)",
    )
    args = parser.parse_args(argv)
    result = run_update(args.excel, args.mango_url)
    if result.errors:
        for e in result.errors:
            print(f"[오류] {e}", flush=True)
    if result.updated > 0 or (result.ok and not result.errors):
        return 0
    if result.skipped > 0 and result.failed == 0 and not result.errors:
        return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
