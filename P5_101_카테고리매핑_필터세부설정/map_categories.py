"""
P5_101_카테고리매핑_필터세부설정 — 필터별 마켓 카테고리 자동 매핑.

초기 1회
  1) 상품수집사이트 선택 (`select[name="site_id"]`)
  2) [선택조건으로 검색하기] (`onclick="search_filter('search')"`)
  3) 마켓별 카테고리 엑셀 읽어 메모리 보관 (P5 추출 결과 양식)

1단계 루프 — 체크된 행마다
  0) 행 정보 읽기 (체크박스 · 필터이름 · ftid)
  1) 필터이름(수정가능) 읽기
  2~3) 필터세부설정 열의 [설정수정] (`onclick="market_mapping_new('<ftid>')"`)
  4) 팝업 `admin_category_set.php?tm=F&ps_ftid=<ftid>`
  5) [AI 자동 매핑 시작하기] (`onclick="search_recommend_category_all(this)"`)
  6) 2단계 루프 — 마켓마다
       필터이름 ↔ 엑셀 카테고리 비교 → 최적 카테고리
       → 검색필드 입력 (`#openmarket_category_search_text_<코드>`)
       → [검색] (`search_category('<코드>','openmarket_category_search_list_<코드>','')`)
       → 결과 목록에서 일치 항목 선택 (`#openmarket_category_search_list_<코드>`)
  7) [검색필터 설정저장 (Alt+S)] (`onclick="config_save()"`)
  8) 모달 닫기 → 9) 다음 행

사용법:
  python map_categories.py --excel-dir D:\\카테고리엑셀
  python map_categories.py --site-id abcmart --excel AUC20=D:\\옥션.xlsx --excel 11ST=D:\\11번가.xlsx
  python map_categories.py --dry-run          # 화면 조작 없이 매칭 결과만 출력
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable, Sequence

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
P2_DIR = ROOT / "P2"
P5_DIR = ROOT / "P5_카테고리_엑셀추출"
for _p in (P2_DIR, P5_DIR):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

sys.path.insert(0, str(Path(__file__).resolve().parent))
import matching  # noqa: E402
import category_db  # noqa: E402
import keyword_dictionary  # noqa: E402
import category_master_db  # noqa: E402
import extended_master_db  # noqa: E402
import market_cache  # noqa: E402

ProgressFn = Callable[[str], None]

HERE = Path(__file__).resolve().parent
STOP_FLAG_PATH = HERE / ".map_stop"

DEFAULT_LIST_URL = (
    "https://tmg1898.cafe24.com/mall/admin/admin_group.php"
    "?pmode=filter_delete&uids=&pg=1&date_type=modify"
    "&start_yy=2026&start_mm=8&start_dd=22&end_yy=2026&end_mm=8&end_dd=22"
    "&site_id=&sales_yn=&ft_group=all&sch_field=title&sch_keyword="
    "&ft_num=10&ft_sort=modify_asc"
)
CATEGORY_PAGE = "admin_category_set.php"

# 매핑 대상 마켓 (P5 추출 대상과 동일)
MARKETS: dict[str, str] = {
    "AUC20": "옥션2.0",
    "11ST": "11번가",
    "GMK20": "G마켓2.0",
    "SMART": "스마트스토어",
    "COUP": "쿠팡",
    "LTON": "롯데ON",
}

# 화면 선택자 (스크린샷 기준)
SEARCH_FILTER_JS = "search_filter('search')"
SETTING_EDIT_JS = "market_mapping_new"
AI_MAPPING_JS = "search_recommend_category_all"
CONFIG_SAVE_JS = "config_save"

T_CLICK = 3_000
T_FIELD = 5_000
T_LIST = 8_000
GAP = 0.15

MIN_SCORE = 0.34  # 이 점수 미만이면 매칭 실패로 본다

# ★요건(2026-08-22): 검증 전까지 무신사에 한해 수행
ALLOWED_SITES = ("musinsa.com",)
DEFAULT_SITE = "MUSINSA.com"

# ★요건: 마켓별 카테고리 구분 라디오 (스크린샷 3·4)
#   <input type="radio" name="openmarket_seller_type2_<코드>"
#     onclick="change_category_list(...,'<코드>', this);"><span>해외직구 카테고리</span>
VARIANT_RADIO_NAME = "openmarket_seller_type2_{market}"
MARKET_VARIANTS: dict[str, tuple[str, ...]] = {
    "11ST": ("해외카테고리", "국내카테고리"),
    "LTON": ("해외직구 카테고리", "일반카테고리"),
}
BOTH = "둘다"  # 두 구분 모두 매핑

# 상품고시정보 팝업 — show_hide('#mapping_notify_<코드>')
NOTIFY_ID = "mapping_notify_{market}"
NOTIFY_RETRIES = 3  # 상품고시정보 팝업이 계속 뜨면 다른 카테고리로 재시도하는 횟수
MAP_RETRIES = 3  # ★요건: 매핑 실패 시 다른 카테고리로 최대 3회 추가 시도
VERIFY_ROUNDS = 3  # ★요건: 저장 후 재검증 — 미매핑 마켓이 있으면 최대 3회 재시도

# ★요건: 작업 행 범위 — <작업 시작 부터> ~ <작업 종료 까지> (1부터, 양끝 포함)
DEFAULT_ROW_FROM = 1
DEFAULT_ROW_TO = 5


@dataclass
class RowInfo:
    index: int
    ftid: str
    filter_name: str
    checked: bool = True


@dataclass
class MappedItem:
    market: str
    category: str
    score: float
    ok: bool = False
    reason: str = ""


@dataclass
class RunResult:
    ok: bool
    rows: int = 0
    mapped: int = 0
    failed: int = 0
    details: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def normalize_site(site: str) -> str:
    return "".join(str(site or "").split()).lower()


def is_allowed_site(site: str) -> bool:
    """무신사만 허용 (다른 사이트는 검증 완료 후 개방)."""
    s = normalize_site(site)
    if not s:
        return False
    return any(allowed in s for allowed in ALLOWED_SITES)


def row_range(row_from: int | str = DEFAULT_ROW_FROM, row_to: int | str = DEFAULT_ROW_TO) -> tuple[int, int]:
    """작업 행 범위 정리 — 1 이상, 시작 ≤ 종료."""

    def _int(value, default: int) -> int:
        try:
            n = int(str(value).strip())
        except (TypeError, ValueError):
            return default
        return n if n > 0 else default

    start = _int(row_from, DEFAULT_ROW_FROM)
    end = _int(row_to, DEFAULT_ROW_TO)
    if end < start:
        start, end = end, start
    return start, end


def slice_rows(rows: Sequence, row_from: int | str, row_to: int | str) -> list:
    """작업 범위에 해당하는 행만 (1부터 세고 양끝 포함)."""
    start, end = row_range(row_from, row_to)
    return list(rows)[start - 1 : end]


def clear_stop_flag() -> None:
    try:
        STOP_FLAG_PATH.unlink(missing_ok=True)  # type: ignore[call-arg]
    except Exception:
        pass


def stop_requested() -> bool:
    return STOP_FLAG_PATH.is_file()


def _log(progress: ProgressFn | None, message: str, *, major: bool = False) -> None:
    line = f"##MAIN##{message}" if major else message
    print(line, flush=True)
    if progress:
        progress(line)


# ── 엑셀 카테고리 자료 (초기 1회 · 메모리 보관) ────────────────────


def market_from_filename(name: str) -> str:
    """파일명에서 마켓 코드 추정 (P5 출력: 카테고리분류표_옥션2.0_....xlsx)."""
    text = str(name)
    for code, label in MARKETS.items():
        if code.lower() in text.lower() or label in text:
            return code
    return ""


def load_categories(path: str | Path) -> list[str]:
    """엑셀에서 카테고리 전체경로 목록을 읽는다.

    P5 추출 양식(마켓·구분·1~6단계·전체경로) 우선, 없으면 단계 열을 이어 붙이거나
    첫 열을 그대로 쓴다.
    """
    from openpyxl import load_workbook  # noqa: WPS433

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c or "").strip() for c in next(rows)]
    except StopIteration:
        return []

    full_idx = header.index("전체경로") if "전체경로" in header else -1
    level_idx = [i for i, h in enumerate(header) if re.fullmatch(r"\d단계", h)]

    out: list[str] = []
    seen: set[str] = set()
    for row in rows:
        if row is None:
            continue
        if full_idx >= 0 and full_idx < len(row) and row[full_idx]:
            path_text = str(row[full_idx]).strip()
        elif level_idx:
            parts = [str(row[i]).strip() for i in level_idx if i < len(row) and row[i]]
            path_text = " > ".join(parts)
        else:
            path_text = str(row[0]).strip() if row and row[0] else ""
        if not path_text or path_text in seen:
            continue
        seen.add(path_text)
        out.append(path_text)
    return out


def load_market_excels(
    paths: dict[str, str | Path],
    *,
    progress: ProgressFn | None = None,
    update_cache: bool = True,
) -> dict[str, list[str]]:
    """마켓 코드 → 카테고리 목록.

    ★요건: 엑셀을 실제로 읽었으면(성공한 마켓이 하나라도 있으면) 그
    결과를 `market_cache` 에 저장해 둔다 — 다음 실행부터는 엑셀 파일이
    없어도(또는 다시 열지 않고도) 이 캐시로 재사용할 수 있다.
    """
    data: dict[str, list[str]] = {}
    for code, path in paths.items():
        code = code.strip().upper()
        if not path:
            continue
        try:
            cats = load_categories(path)
        except Exception as e:  # noqa: BLE001
            _log(progress, f"엑셀 읽기 실패 · {code} · {e}", major=True)
            continue
        data[code] = cats
        _log(progress, f"  {MARKETS.get(code, code)} 카테고리 {len(cats)}건 로드", major=True)
    if data and update_cache:
        try:
            cache_path = market_cache.save(data)
            _log(progress, f"  카테고리 캐시 갱신 → {cache_path}", major=True)
        except OSError as e:  # noqa: BLE001
            _log(progress, f"  경고: 카테고리 캐시 저장 실패 · {e}")
    return data


def load_market_excels_or_cache(
    paths: dict[str, str | Path] | None = None,
    *,
    progress: ProgressFn | None = None,
) -> dict[str, list[str]]:
    """엑셀 경로가 있으면 엑셀에서 읽고(캐시도 갱신), 없으면 캐시에서 읽는다.

    ★요건: "매번 엑셀을 다시 읽지 않고 캐시로 재사용" — 엑셀 파일·경로가
    준비되지 않은 상태에서도(예: 서버에 엑셀이 없는 환경) 직전에 구축해
    둔 카테고리 데이터로 계속 동작할 수 있게 한다.
    """
    if paths:
        data = load_market_excels(paths, progress=progress)
        if data:
            return data
    cached = market_cache.load()
    if cached:
        _log(
            progress,
            f"  엑셀 대신 캐시 사용 — 마켓 {len(cached)}건 · "
            f"카테고리 {sum(len(v) for v in cached.values())}건 "
            f"({market_cache.DEFAULT_CACHE_PATH.name})",
            major=True,
        )
    return cached


def build_category_db(excels: dict[str, Sequence[str]]) -> category_db.CategoryDB:
    """★요건재정의(2026-08-22 B): 6개 마켓 엑셀을 교차검색한 통합정보화DB.

    `find_category` 의 5) 단계(정보화DB 연관검색어)에서 쓴다.
    """
    return category_db.CategoryDB.build(excels)


def build_master_db(*, refresh: bool = False) -> category_master_db.MasterDB:
    """★요건(2026-08-23): "카테고리매핑DB(매핑DB)" — 사용자 제공
    CATEGORY_MASTER+KEYWORD_DICTIONARY CSV(`data/category_master.csv`·
    `data/keyword_dictionary.csv`, 6개 마켓 17,591건·25,399건)를
    그대로 DB화한 것을 불러온다. JSON 캐시(`data/category_master_db.json`)가
    있으면 그것부터 읽어(빠름) 매번 CSV를 다시 파싱하지 않는다.
    """
    return category_master_db.load(refresh=refresh)


def build_extended_master_db() -> extended_master_db.ExtendedMasterDB:
    """★요건(2026-08-23): "카테고리매핑_확장형DB(확장형DB)" — 사용자 제공
    표준 카테고리(14,207건) + 표준 연관검색어(29,999건) CSV 기반 확장형 DB.
    JSON 캐시(`data/extended_master_db.json`)가 있으면 빠르게 메모리에 올린다.
    """
    return extended_master_db.ExtendedMasterDB.load()


def best_category_via_master(
    filter_name: str,
    market_display_name: str,
    master_db: category_master_db.MasterDB,
    *,
    exclude: Sequence[str] = (),
    region_type: str = "국내",
) -> tuple[str, str]:
    """★요건: 사용자 제공 마스터DB를 **1순위 기준**으로 최적 카테고리를 찾는다.

    필터명을 하위(가장 구체적)부터 중위 순으로 뜯어 각 검색어마다
    `MasterDB.resolve_ranked()` 로 여러 순위의 후보를 받는다 — CSV 의
    Priority(1=완전일치·2=유사어·3=형태소분리) 로 확실하게 찾히면 그
    즉시 확정한다. 앞 후보가 절대규칙(성별·브랜드·배제단어·지역)에 걸리면
    같은 검색어의 다음 후보로, 그래도 없으면 다음 검색어로 넘어간다.
    확실한 매칭이 전혀 없으면(전부 "4) 최근접 강제지정"뿐이면) 절대
    규칙을 지키는 후보 중 첫 번째를 최후 수단으로 쓴다.
    """
    parsed = matching.parse_filter_name(filter_name)
    gender = matching.gender_of(parsed.raw)
    filter_mentions_child = matching.mentions_child(parsed.raw)
    skip = {matching.normalize(e) for e in (exclude or []) if str(e or "").strip()}
    terms = [t for t in [*parsed.lows, parsed.mid] if t]
    if not terms:
        terms = [parsed.raw]

    # 모자 계열 필터에서 단독 '사파리' 검색어가 의류 사파리로 오인되는 것을 방지하기 위해 '사파리모자'/'사파리햇' 보강
    if matching.class_of(filter_name) == "모자":
        expanded_terms = []
        for t in terms:
            if t == "사파리":
                expanded_terms.extend(["사파리모자", "사파리햇", "사파리"])
            else:
                expanded_terms.append(t)
        terms = expanded_terms

    confident_by_idx: dict[int, tuple[str, str]] = {}
    forced_by_idx: dict[int, tuple[str, str]] = {}
    weak_only: tuple[str, str] | None = None

    TIER_ORDER = {"1": 0, "2": 1, "3": 2, "1.5": 3}

    for idx, term in enumerate(terms):
        term_cls = matching.class_of(term)
        by_tier: dict[str, list[tuple[str, str]]] = {}
        for cat_id, step in master_db.resolve_ranked(market_display_name, term, limit=50, gender=gender):
            path = master_db.full_path(cat_id)
            if not path or matching.normalize(path) in skip:
                continue
            if matching._contains_word(path, "브랜드"):
                continue  # ★절대규칙(예외 없음): 브랜드 카테고리는 절대 확정하지 않는다
            if gender and matching.violates_gender(path, filter_name):
                continue  # ★절대규칙(예외 없음): 반대 성별 카테고리는 확정하지 않는다
            # ★요건 3: 절대 배제 단어 카테고리는 확정하지 않는다 (필터명에 없는 경우)
            if matching.is_forbidden_category(path, filter_name):
                continue
            # ★요건 4: 국내 모드 시 "해외" 카테고리 완전 배제
            if region_type == "국내" and matching.has_overseas(path):
                continue
            # ★절대규칙(실사례): 필터명이 아동을 언급하지 않았는데 후보가 아동 카테고리면 배제
            if not filter_mentions_child and matching.is_child_category(path):
                continue
            # ★품목 계열 불일치 배제 (모자 필터가 의류 사파리 자켓 등으로 새는 것 방지)
            filter_cls = matching.class_of(filter_name)
            path_cls = matching.path_class(path)
            if filter_cls and path_cls and filter_cls != path_cls and not matching.is_generic_path(path):
                continue

            label = f'{step} · 검색어="{term}"'
            is_weak = matching._is_non_product_hint(path) or (
                term_cls and path_cls and term_cls != path_cls
            )
            if is_weak:
                if weak_only is None:
                    weak_only = (path, label)
                continue
            if step.startswith("4)"):
                if idx not in forced_by_idx:
                    forced_by_idx[idx] = (path, label)
                continue
            tier_key = step.split(")")[0]
            by_tier.setdefault(tier_key, []).append((path, label))

        for tier_key in sorted(by_tier, key=lambda k: TIER_ORDER.get(k, 99)):
            candidates = by_tier[tier_key]
            best = max(
                candidates,
                key=lambda pl: matching.priority_rank(
                    pl[0], parsed, market=market_display_name, region_type=region_type
                ),
            )
            confident_by_idx[idx] = best
            break
        if idx == 0 and confident_by_idx.get(0):
            break

    best_confident_idx = min(confident_by_idx) if confident_by_idx else None
    best_forced_idx = min(forced_by_idx) if forced_by_idx else None
    if best_confident_idx is not None and (
        best_forced_idx is None or best_confident_idx <= best_forced_idx
    ):
        return confident_by_idx[best_confident_idx]
    if best_forced_idx is not None:
        return forced_by_idx[best_forced_idx]
    if best_confident_idx is not None:
        return confident_by_idx[best_confident_idx]
    if weak_only:
        return weak_only
    return "", "미검출"


def build_keyword_db(excels: dict[str, Sequence[str]]) -> keyword_dictionary.KeywordDB:
    """★요건(연관검색어DB, "최신성 우선" 원칙에 따른 공식 기준):
    6개 마켓 엑셀 전체로 CATEGORY_MASTER+KEYWORD_DICTIONARY 를 구축해
    `find_category` 의 5) 단계 검색어 확장에 함께 쓴다.

    ★주의: `category_db.py`(마켓 교차색인)와 달리 이건 실제 매핑 실행
    경로에 처음 연결하는 것이다 — 이전까지는 만들어만 두고 전혀
    쓰이지 않았다는 지적을 받아 이번에 연결했다.
    """
    all_paths = [p for paths in (excels or {}).values() for p in paths]
    return keyword_dictionary.build(all_paths)


def discover_market_excels(folder: str | Path) -> dict[str, str]:
    """폴더에서 마켓별 엑셀을 파일명으로 자동 매칭."""
    found: dict[str, str] = {}
    for p in sorted(Path(folder).glob("*.xlsx")):
        code = market_from_filename(p.name)
        if code and code not in found:
            found[code] = str(p)
    return found


# ── 필터이름 ↔ 카테고리 매칭 (순수 로직) ──────────────────────────

_SPLIT_RE = re.compile(r"[\s_\-/>·,()\[\]]+")


def tokenize(text: str) -> list[str]:
    """비교용 토큰 — 구분자·공백 제거, 소문자화."""
    raw = _SPLIT_RE.split(str(text or "").strip().lower())
    return [t for t in raw if t]


def leaf_of(path: str) -> str:
    parts = [p.strip() for p in str(path or "").split(">") if p.strip()]
    return parts[-1] if parts else ""


def similarity(filter_name: str, category_path: str) -> float:
    """필터이름과 카테고리 경로의 유사도 (0~1).

    마지막 단계(리프)에 가중치를 두고, 경로 전체 토큰 겹침도 함께 본다.
    """
    ftoks = set(tokenize(filter_name))
    if not ftoks:
        return 0.0
    path_toks = set(tokenize(category_path))
    leaf_toks = set(tokenize(leaf_of(category_path)))
    if not path_toks:
        return 0.0

    path_hit = len(ftoks & path_toks) / len(ftoks)
    leaf_hit = len(ftoks & leaf_toks) / len(ftoks) if leaf_toks else 0.0

    # 문자열 포함 보너스 (예: '남성비니' ↔ '비니')
    joined_f = "".join(sorted(ftoks))
    bonus = 0.0
    for tok in leaf_toks:
        if tok and (tok in joined_f or any(tok in f or f in tok for f in ftoks)):
            bonus = max(bonus, 0.25)
    return min(1.0, 0.5 * path_hit + 0.5 * leaf_hit + bonus)


def similarity_best(
    filter_name: str, categories: Sequence[str], *, min_score: float = MIN_SCORE
) -> tuple[str, float]:
    """토큰 유사도만으로 고르는 보조 경로 (결과 목록 선택 등)."""
    best = ""
    best_score = 0.0
    for cat in categories:
        score = similarity(filter_name, cat)
        if score > best_score or (score == best_score and cat and len(cat) < len(best)):
            best, best_score = cat, score
    if best_score < min_score:
        return "", best_score
    return best, best_score


def best_category(
    filter_name: str,
    categories: Sequence[str],
    *,
    min_score: float = MIN_SCORE,
    db: category_db.CategoryDB | None = None,
    keyword_db: keyword_dictionary.KeywordDB | None = None,
    ext_db: extended_master_db.ExtendedMasterDB | None = None,
) -> tuple[str, float]:
    """★요건(2026-08-23 3단계 매핑):
    1차: 엑셀 직접 검색 -> 2차: 확장형DB 범주확장 -> 3차: 매핑DB 범주확장 -> 최근접 폴백.
    """
    cat, _step = matching.find_category(
        filter_name, list(categories), db=db, keyword_db=keyword_db, ext_db=ext_db
    )
    if cat:
        return cat, 1.0
    return similarity_best(filter_name, categories, min_score=min_score)


def best_category_with_step(
    filter_name: str,
    categories: Sequence[str],
    *,
    exclude: Sequence[str] = (),
    market: str = "",
    region_type: str = "국내",
    db: category_db.CategoryDB | None = None,
    keyword_db: keyword_dictionary.KeywordDB | None = None,
    ext_db: extended_master_db.ExtendedMasterDB | None = None,
) -> tuple[str, str]:
    """최적 카테고리 + 어느 단계에서 찾았는지 (exclude 는 이미 시도한 것)."""
    cat, step = matching.find_category(
        filter_name,
        list(categories),
        exclude=exclude,
        market=market,
        region_type=region_type,
        db=db,
        keyword_db=keyword_db,
        ext_db=ext_db,
    )
    if cat:
        return cat, step
    skip = {matching.normalize(e) for e in exclude or []}
    rest = [c for c in categories if matching.normalize(c) not in skip]
    # 안전한 후보(배제 단어 제외, 성별/지역 조건 준수) 위주로 유사도 검색
    safe_rest = matching.strip_forbidden_categories(rest, filter_name)
    if region_type == "국내":
        safe_rest = matching.filter_by_region(safe_rest, "국내")
    gender = matching.gender_of(filter_name)
    if gender:
        safe_rest = matching.strip_opposite_gender(safe_rest, gender)
    candidate_pool = safe_rest if safe_rest else rest

    cat, score = similarity_best(filter_name, candidate_pool)
    if not cat and candidate_pool:
        cat = candidate_pool[0]
        score = 0.5
    return cat, (f"유사도 {score:.2f}" if cat else "미검출")


def search_keyword_for(category_path: str) -> str:
    """카테고리 검색필드에 넣을 검색어 — ★확정된 카테고리명 전체(단계별 공백 한 칸).

    ★요건: 엑셀은 망고 마켓별 전체 카테고리를 그대로 내려받은 것이라,
    엑셀에서 확정한 값은 망고에도 100% 존재한다.
    상위·중위·하위·세부·상세 단계를 묶어 검색할 때 단계 사이에 공백을 한 글자씩 두고 입력한다.
    예) "남성의류 > 반팔티셔츠 > 카라 반팔티" → "남성의류 반팔티셔츠 카라 반팔티"
    """
    levels = [p.strip() for p in str(category_path or "").split(">") if p.strip()]
    return " ".join(levels)


def pick_option(options: Sequence[str], category_path: str) -> str:
    """검색 결과 목록에서 고를 항목 — 1순위: 완전일치, 2순위: 리프일치, 3순위: 첫 번째 유효 항목 (절대 미매핑 방지)."""
    target = str(category_path or "").strip()
    valid_opts = [o for o in options if str(o or "").strip() and not str(o).strip().startswith("-") and not "선택해주세요" in str(o)]
    if not valid_opts:
        return ""

    if not target:
        return valid_opts[0]

    norm = lambda s: "".join(str(s or "").split()).lower()  # noqa: E731
    target_norm = norm(target)

    # 1. 완전 일치
    for opt in valid_opts:
        if norm(opt) == target_norm:
            return opt

    # 2. 리프(마지막 단계) 완전 일치
    target_leaf = norm(target.split(">")[-1])
    for opt in valid_opts:
        opt_leaf = norm(str(opt).split(">")[-1])
        if opt_leaf and opt_leaf == target_leaf:
            return opt

    # 3. 부분 일치
    for opt in valid_opts:
        opt_norm = norm(opt)
        if target_leaf and (target_leaf in opt_norm or opt_norm in target_leaf):
            return opt

    # 4. 첫 번째 유효 항목
    return valid_opts[0]


# ── 화면 조작 ────────────────────────────────────────────────────


def _first(page, selectors: Iterable[str]):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0:
                return loc
        except Exception:
            continue
    return None


SITE_OPTIONS_JS = "(el) => Array.from(el.options).map(o => (o.textContent || '').trim())"


def _norm_site(text: str) -> str:
    """사이트명 비교용 — 공백 제거 + 소문자."""
    return "".join(str(text or "").split()).lower()


def match_site_option(options: Sequence[str], wanted: str) -> str | None:
    """입력한 사이트명을 실제 옵션에 맞춘다.

    정확 일치 → 대소문자·공백 무시 → 부분 포함. `MUSINSA.COM` 처럼 대소문자가
    다르게 입력돼도 `MUSINSA.com` 옵션을 찾아낸다.
    """
    want = str(wanted or "").strip()
    if not want:
        return None
    for opt in options:
        if opt == want:
            return opt
    nw = _norm_site(want)
    for opt in options:
        if _norm_site(opt) == nw:
            return opt
    for opt in options:
        no = _norm_site(opt)
        if nw and no and (nw in no or no in nw):
            return opt
    return None


def read_site_options(page) -> list[str]:
    """상품수집사이트 드롭다운의 옵션 텍스트 목록."""
    loc = _first(page, ('select[name="site_id"]',))
    if loc is None:
        return []
    try:
        raw = loc.evaluate(SITE_OPTIONS_JS) or []
    except Exception:
        return []
    return [str(t).strip() for t in raw if str(t).strip()]


def select_site(page, site_id: str, *, progress: ProgressFn | None = None) -> bool:
    """상품수집사이트 리스트박스 선택 (초기 1회)."""
    if not site_id:
        return True
    loc = _first(page, ('select[name="site_id"]',))
    if loc is None:
        _log(progress, "오류: 상품수집사이트 드롭다운 미검출", major=True)
        return False

    options = read_site_options(page)
    target = match_site_option(options, site_id) if options else site_id
    if target is None:
        _log(
            progress,
            f"오류: 상품수집사이트 미검출 · 입력={site_id} · 화면목록={options}",
            major=True,
        )
        return False

    for how in ("label", "value"):
        try:
            if how == "label":
                loc.select_option(label=target, timeout=T_CLICK)
            else:
                loc.select_option(target, timeout=T_CLICK)
            note = "" if target == site_id else f" (입력 {site_id} → 화면 {target})"
            _log(progress, f"상품수집사이트: {target}{note}", major=True)
            return True
        except Exception:
            continue
    _log(
        progress,
        f"오류: 상품수집사이트 선택 실패 · {target} · 화면목록={options}",
        major=True,
    )
    return False


def click_search_filter(page, *, progress: ProgressFn | None = None) -> bool:
    """[선택조건으로 검색하기]."""
    loc = _first(
        page,
        (
            f'a[onclick*="{SEARCH_FILTER_JS}"]',
            'xpath=//a[.//span[contains(normalize-space(.),"선택조건으로 검색하기")]]',
            'xpath=//*[contains(normalize-space(.),"선택조건으로 검색하기")]',
        ),
    )
    if loc is None:
        _log(progress, "오류: [선택조건으로 검색하기] 미검출", major=True)
        return False
    try:
        loc.click(timeout=T_CLICK)
    except Exception as e:  # noqa: BLE001
        _log(progress, f"오류: 검색 클릭 실패 · {e}", major=True)
        return False
    _log(progress, "[선택조건으로 검색하기] 클릭", major=True)
    try:
        page.wait_for_load_state("domcontentloaded", timeout=T_FIELD)
    except Exception:
        pass
    time.sleep(GAP)
    return True


LIST_ROWS_JS = r"""
() => {
  // 표 구조에 기대지 않고 [설정수정] 링크에서 거꾸로 행을 찾는다
  const out = [];
  const anchors = Array.from(document.querySelectorAll('a[onclick*="market_mapping_new"]'));
  anchors.forEach((a, idx) => {
    const onclick = a.getAttribute('onclick') || '';
    const m = onclick.match(/market_mapping_new\(\s*'?(\d+)'?/);
    const ftid = m ? m[1] : '';
    if (!ftid) return;

    const tr = a.closest('tr');
    let name = '';
    let checked = false;
    if (tr) {
      const cb = tr.querySelector('input[type="checkbox"]');
      checked = cb ? !!cb.checked : false;
      const vals = Array.from(tr.querySelectorAll('input[type="text"], input:not([type])'))
        .map(i => (i.value || '').trim())
        .filter(v => v && !/^https?:/i.test(v) && !/^\d+$/.test(v));
      if (vals.length) name = vals[0];
      if (!name) {
        const byUid = document.querySelector('input[attr-uid="' + ftid + '"]');
        if (byUid) name = (byUid.value || '').trim();
      }
      if (!name) name = ((tr.innerText || '').trim().split('\n')[0] || '').trim();
    }
    out.push({index: idx, ftid: ftid, filterName: name, checked: checked});
  });
  return out;
}
"""


LIST_DIAG_JS = r"""
() => {
  const anchors = Array.from(document.querySelectorAll('a[onclick]'));
  const mapping = anchors.filter(a => (a.getAttribute('onclick') || '').includes('market_mapping_new'));
  return {
    url: location.href,
    table: !!document.querySelector('table#search_category'),
    rows: document.querySelectorAll('tr').length,
    checkboxes: document.querySelectorAll('input[type="checkbox"]').length,
    mappingLinks: mapping.length,
    sample: mapping.slice(0, 3).map(a => (a.getAttribute('onclick') || '').slice(0, 60)),
  };
}
"""


def contexts(page) -> list:
    """페이지 + 프레임 (목록이 프레임 안에 있을 수 있다)."""
    out = [page]
    try:
        for f in page.frames:
            if f is not page and f not in out:
                out.append(f)
    except Exception:
        pass
    return out


def diagnose_list(page, *, progress: ProgressFn | None = None) -> None:
    """행을 못 찾을 때 화면 상태를 로그로 남긴다."""
    for i, ctx in enumerate(contexts(page), start=1):
        try:
            info = ctx.evaluate(LIST_DIAG_JS)
        except Exception as e:  # noqa: BLE001
            _log(progress, f"  [진단] 프레임{i}: 읽기 실패 ({e})", major=True)
            continue
        _log(
            progress,
            f"  [진단] 프레임{i} url={str(info.get('url'))[:80]}"
            f" · tr={info.get('rows')} · 체크박스={info.get('checkboxes')}"
            f" · 설정수정링크={info.get('mappingLinks')}"
            f" · search_category={info.get('table')}"
            f" · 예시={info.get('sample')}",
            major=True,
        )


def list_rows(page) -> list[RowInfo]:
    """모든 프레임에서 행을 모으고 ftid 중복은 제거한다."""
    data: list[dict] = []
    seen: set[str] = set()
    for ctx in contexts(page):
        try:
            found = ctx.evaluate(LIST_ROWS_JS) or []
        except Exception:
            found = []
        for item in found:
            ftid = str((item or {}).get("ftid") or "").strip()
            if not ftid or ftid in seen:
                continue
            seen.add(ftid)
            data.append(item)
    rows: list[RowInfo] = []
    for d in data:
        rows.append(
            RowInfo(
                index=int(d.get("index") or 0),
                ftid=str(d.get("ftid") or "").strip(),
                filter_name=str(d.get("filterName") or "").strip(),
                checked=bool(d.get("checked")),
            )
        )
    return rows


# 목록 화면 하단 페이지 로더 — <a onclick="pageLoad('2','musinsa');">2</a>
PAGE_LOAD_JS = "pageLoad"


def has_page(page, page_no: int) -> bool:
    """하단 페이지 로더에 그 페이지 번호 링크가 있는지."""
    sel = f"a[onclick*=\"{PAGE_LOAD_JS}('{page_no}'\"]"
    try:
        return page.locator(sel).first.count() > 0
    except Exception:
        return False


def click_page_load(
    page, page_no: int, site_id: str = "", *, progress: ProgressFn | None = None
) -> bool:
    """하단 페이지 로더에서 페이지 번호를 클릭한다 (`pageLoad(N, site)`)."""
    site = str(site_id or "").strip()
    selectors = []
    if site:
        selectors.append(f"a[onclick*=\"{PAGE_LOAD_JS}('{page_no}','{site}')\"]")
    selectors.append(f"a[onclick*=\"{PAGE_LOAD_JS}('{page_no}'\"]")
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() == 0:
                continue
            loc.click(timeout=T_CLICK)
            _log(progress, f"  페이지 로더 — {page_no}페이지 이동", major=True)
            return True
        except Exception:
            continue
    return False


def collect_rows_for_range(
    page,
    end: int,
    site_id: str = "",
    *,
    progress: ProgressFn | None = None,
) -> list[RowInfo]:
    """★요건: 하단 페이지 로더를 감안해 작업행 번호를 적용한다.

    한 페이지 분량(현재 페이지에 뜬 행 수)을 넘는 작업행 범위가 요청되면
    하단 페이지 로더(`pageLoad(N, site)`)를 눌러 다음 페이지로 넘어가며
    이어서 채운다. 요청 범위가 현재 페이지 안에서 끝나면 페이지를 넘기지
    않는다(빠른 경로 — 기존 동작과 동일).
    """
    rows = list_rows(page)
    if not rows or end <= len(rows):
        return rows

    all_rows = list(rows)
    page_no = 1
    while len(all_rows) < end:
        if not has_page(page, page_no + 1):
            _log(progress, f"  페이지 로더 — {page_no}페이지가 마지막", major=True)
            break
        page_no += 1
        if not click_page_load(page, page_no, site_id, progress=progress):
            _log(progress, f"  경고: {page_no}페이지 이동 실패", major=True)
            break
        more = list_rows(page)
        if not more:
            break
        all_rows.extend(more)
    return all_rows


def build_mapping_url(ftid: str, *, list_url: str = DEFAULT_LIST_URL) -> str:
    """설정수정 팝업 URL (`admin_category_set.php?tm=F&ps_ftid=<ftid>`)."""
    from urllib.parse import urlencode, urlsplit, urlunsplit

    parts = urlsplit(list_url or DEFAULT_LIST_URL)
    base_dir = parts.path.rsplit("/", 1)[0] if "/" in parts.path else ""
    query = urlencode({"tm": "F", "ps_ftid": str(ftid)})
    return urlunsplit((parts.scheme, parts.netloc, f"{base_dir}/{CATEGORY_PAGE}", query, ""))


def reveal(page, *, progress: ProgressFn | None = None) -> None:
    """★크롬 창을 화면 맨 앞으로 — 작업 과정(팝업·입력·닫기)이 눈에 보이게 한다.

    `P3_필터_갱신` 과 같은 방식. 안 해주면 실제로는 잘 돌고 있어도 보드 창
    뒤에 숨어서 사용자에게 안 보인다.
    """
    try:
        page.bring_to_front()
    except Exception as e:  # noqa: BLE001
        _log(progress, f"  (화면 앞으로 가져오기 실패: {e})")


def open_setting_popup(page, row: RowInfo, *, list_url: str, progress: ProgressFn | None = None):
    """행의 [설정수정] → 팝업. 실패 시 팝업 URL 직접 오픈.

    ★같은 버튼을 두 번 클릭해 팝업을 두 개 띄우지 않는다 — expect_popup 이
    타이밍상 이벤트를 놓치면(실제로는 창이 열렸는데 감지만 실패) 곧바로
    URL 직접열기(새 탭)로 넘어가 진짜 팝업이 2개가 됐다. 예외가 나면 먼저
    새로 열린 탭이 있는지 확인해 그걸 쓴다.
    """
    try:
        before = set(page.context.pages)
    except Exception:
        before = set()
    sel = f"a[onclick*=\"{SETTING_EDIT_JS}('{row.ftid}')\"]"
    try:
        loc = page.locator(sel).first
        if loc.count() > 0:
            with page.expect_popup(timeout=T_FIELD) as info:
                loc.click(timeout=T_CLICK)
            popup = info.value
            reveal(popup, progress=progress)
            _log(progress, f"  설정수정 팝업 (ftid={row.ftid})")
            return popup
    except Exception:
        try:
            new_pages = [p for p in page.context.pages if p not in before]
        except Exception:
            new_pages = []
        if new_pages:
            popup = new_pages[-1]
            reveal(popup, progress=progress)
            _log(progress, f"  설정수정 팝업 감지(지연) (ftid={row.ftid})")
            return popup

    url = build_mapping_url(row.ftid, list_url=list_url)
    try:
        popup = page.context.new_page()
        popup.goto(url, wait_until="domcontentloaded", timeout=30_000)
        reveal(popup, progress=progress)
        _log(progress, f"  설정수정 직접 열기 (ftid={row.ftid})")
        return popup
    except Exception as e:  # noqa: BLE001
        _log(progress, f"오류: 설정수정 팝업 실패 · {e}", major=True)
        return None


def click_ai_mapping(popup, *, progress: ProgressFn | None = None) -> bool:
    """[AI 자동 매핑 시작하기]."""
    loc = _first(
        popup,
        (
            f'a[onclick*="{AI_MAPPING_JS}"]',
            'xpath=//a[.//span[contains(normalize-space(.),"AI 자동 매핑")]]',
        ),
    )
    if loc is None:
        _log(progress, "  경고: [AI 자동 매핑 시작하기] 미검출 — 수동 매핑만 진행")
        return False
    try:
        loc.click(timeout=T_CLICK)
        _log(progress, "  AI 자동 매핑 시작")
        time.sleep(0.5)
        return True
    except Exception:
        return False


def variants_for(market: str, choice: str = "") -> list[str]:
    """이 마켓에서 매핑할 구분 목록 (구분 없는 마켓은 [''])."""
    options = MARKET_VARIANTS.get(market)
    if not options:
        return [""]
    pick = str(choice or "").strip()
    if not pick or pick == BOTH:
        return list(options)
    for opt in options:
        if normalize_site(opt) == normalize_site(pick):
            return [opt]
    return list(options)


def variant_radio_selectors(market: str, variant: str) -> tuple[str, ...]:
    name = VARIANT_RADIO_NAME.format(market=market)
    return (
        f'xpath=//label[.//span[contains(normalize-space(.),"{variant}")]]'
        f'//input[@type="radio" and @name="{name}"]',
        f'xpath=//tr[@id="mapping_category_{market}"]'
        f'//label[.//span[contains(normalize-space(.),"{variant}")]]//input[@type="radio"]',
    )


def select_variant(popup, market: str, variant: str, *, progress: ProgressFn | None = None) -> bool:
    """구분 라디오를 **클릭**해 목록을 바꾼다 (이미 체크돼 있어도 클릭)."""
    if not variant:
        return True
    loc = _first(popup, variant_radio_selectors(market, variant))
    if loc is None:
        _log(progress, f"  경고: 구분 라디오 미검출 · {variant}")
        return False
    try:
        loc.click(timeout=T_CLICK, force=True)
    except Exception:
        try:
            loc.check(timeout=T_CLICK)
        except Exception:
            return False
    _log(progress, f"  구분 선택: {variant}")
    try:
        popup.wait_for_timeout(400)
    except Exception:
        time.sleep(0.4)
    return True


NOTIFY_VISIBLE_JS = """
(id) => {
  const el = document.getElementById(id);
  if (!el) return false;
  const st = window.getComputedStyle(el);
  if (st.display === 'none' || st.visibility === 'hidden') return false;
  return (el.innerText || '').trim().length > 0;
}
"""


def notify_open(popup, market: str) -> bool:
    """상품고시정보 팝업이 떠 있는가."""
    try:
        return bool(popup.evaluate(NOTIFY_VISIBLE_JS, NOTIFY_ID.format(market=market)))
    except Exception:
        return False


def close_notify(popup, market: str, *, progress: ProgressFn | None = None) -> bool:
    """상품고시정보 팝업 닫기 — show_hide 토글 또는 닫기 버튼."""
    notify_id = NOTIFY_ID.format(market=market)
    loc = _first(
        popup,
        (
            f"a[onclick*=\"show_hide('#{notify_id}')\"]",
            f'xpath=//div[@id="{notify_id}"]//*[contains(normalize-space(.),"닫기")]',
        ),
    )
    if loc is not None:
        try:
            loc.click(timeout=T_CLICK)
            _log(progress, "  상품고시정보 팝업 닫기")
            return True
        except Exception:
            pass
    try:
        popup.evaluate(
            "(id) => { const el = document.getElementById(id); if (el) el.style.display = 'none'; }",
            notify_id,
        )
        _log(progress, "  상품고시정보 팝업 닫기(강제)")
        return True
    except Exception:
        return False


def market_search_input(popup, market: str):
    return _first(
        popup,
        (
            f"#openmarket_category_search_text_{market}",
            f'input[id="openmarket_category_search_text_{market}"]',
        ),
    )


def click_market_search(popup, market: str) -> bool:
    loc = _first(
        popup,
        (
            f"a[onclick*=\"search_category('{market}','openmarket_category_search_list_{market}',''\"]",
            f'xpath=//tr[@id="mapping_category_{market}"]'
            '//a[.//span[normalize-space()="검색"]]',
        ),
    )
    if loc is None:
        return False
    try:
        loc.click(timeout=T_CLICK)
        return True
    except Exception:
        return False


# 결과 리스트박스는 마켓에 따라 list_ / list2_ 두 벌이고 보이는 쪽이 다르다
# (11번가·롯데ON). 보이는 select 을 우선 읽고, 선택도 그 id 에 한다.
RESULT_OPTIONS_JS = """
(ids) => {
  const texts = (el) => Array.from(el.options).map(o => (o.textContent || '').trim());
  const isVisible = (el) => {
    const st = window.getComputedStyle(el);
    if (st.display === 'none' || st.visibility === 'hidden') return false;
    return el.offsetParent !== null || st.display === 'inline-block';
  };
  const cands = ids.map(id => document.getElementById(id)).filter(Boolean);
  const pick = (list) => {
    let best = {texts: [], id: ''};
    for (const el of list) {
      const t = texts(el);
      if (t.length > best.texts.length) best = {texts: t, id: el.id || ''};
    }
    return best;
  };
  const visible = pick(cands.filter(isVisible));
  return visible.texts.length ? visible : pick(cands);
}
"""


def result_select_ids(market: str) -> list[str]:
    return [
        f"openmarket_category_search_list_{market}",
        f"openmarket_category_search_list2_{market}",
    ]


def read_result_options(
    popup, market: str, *, timeout_ms: int | None = None
) -> tuple[list[str], str]:
    """(옵션 목록, 사용한 select id) — 보이는 리스트박스 기준.

    `timeout_ms` 기본값은 호출 시점의 `T_LIST` 를 그대로 읽는다 — 함수 정의
    시점 값으로 고정하면(`= T_LIST`) 테스트에서 `monkeypatch.setattr(mc,
    "T_LIST", ...)` 로 줄여도 반영되지 않아 매번 진짜 5초를 기다리게 된다.
    """
    if timeout_ms is None:
        timeout_ms = T_LIST
    deadline = time.monotonic() + timeout_ms / 1000
    while True:
        try:
            info = popup.evaluate(RESULT_OPTIONS_JS, result_select_ids(market)) or {}
        except Exception:
            info = {}
        texts = list(info.get("texts") or []) if isinstance(info, dict) else list(info or [])
        sel_id = str(info.get("id") or "") if isinstance(info, dict) else ""
        real = [t for t in texts if t and not t.startswith("-")]
        if real or time.monotonic() >= deadline:
            return real, sel_id or result_select_ids(market)[0]
        try:
            popup.wait_for_timeout(200)
        except Exception:
            time.sleep(0.2)


CLEAR_MARKET_CATEGORY_JS = """
(code) => {
  const idEl = document.getElementById('openmarket_cm_category_' + code);
  const nameEl = document.getElementsByName('openmarket_cm_category_name_' + code)[0];
  if (idEl) idEl.value = '';
  if (nameEl) nameEl.value = '';
  return true;
}
"""


def clear_market_category(popup, market: str, *, progress: ProgressFn | None = None) -> None:
    """★검증(엑셀과 완전일치) 안 된 값이 남아있지 않게 필드를 비운다.

    검색만 하고 실제로 [choose_option] 으로 고르지 않으면, 결과 select
    박스의 화면 표시는 바뀌어도 실제 저장되는 hidden 필드
    (openmarket_cm_category_<코드>)는 그대로다 — 즉 [AI 자동 매핑] 이나
    이전 실행이 이미 채워둔, 우리가 검증하지 않은 값이 그대로 저장될 수
    있다. 우리 검색이 실패하면 그 값을 명시적으로 비워서, "확인 안 된
    값이 그대로 저장"되는 사고를 막는다.
    """
    try:
        popup.evaluate(CLEAR_MARKET_CATEGORY_JS, market)
        _log(progress, f"  {MARKETS.get(market, market)}: 미확인 값 비움", major=True)
    except Exception:
        pass


CHOOSE_OPTION_JS = """
(args) => {
  const [selId, targetText, code] = args;
  const sel = document.getElementById(selId);
  if (!sel) return false;
  const norm = (s) => (s || '').replace(/\\s+/g, '').toLowerCase();
  const targetNorm = norm(targetText);
  let bestIdx = -1;
  let bestScore = -1;

  for (let i = 0; i < sel.options.length; i++) {
    const opt = sel.options[i];
    const t = (opt.textContent || '').trim();
    const v = (opt.value || '').trim();
    if (!t || t.startsWith('-') || t.includes('선택해주세요')) continue;
    const optNorm = norm(t);
    if (optNorm === targetNorm) {
      bestIdx = i;
      bestScore = 100;
      break;
    }
    const optLeaf = norm(t.split('>').pop());
    const targetLeaf = norm(targetText.split('>').pop());
    if (optLeaf && targetLeaf && optLeaf === targetLeaf && bestScore < 80) {
      bestIdx = i;
      bestScore = 80;
    } else if (targetLeaf && (optNorm.includes(targetLeaf) || targetNorm.includes(optLeaf))) {
      if (bestScore < 50) {
        bestIdx = i;
        bestScore = 50;
      }
    } else if (optNorm && (optNorm.includes(targetNorm) || targetNorm.includes(optNorm))) {
      if (bestScore < 40) {
        bestIdx = i;
        bestScore = 40;
      }
    } else if (bestIdx < 0) {
      bestIdx = i;
      bestScore = 10;
    }
  }

  if (bestIdx >= 0) {
    sel.selectedIndex = bestIdx;
    sel.dispatchEvent(new Event('change', { bubbles: true }));
    sel.dispatchEvent(new Event('input', { bubbles: true }));
    const idEl = document.getElementById('openmarket_cm_category_' + code);
    const nameEl = document.getElementsByName('openmarket_cm_category_name_' + code)[0];
    if (idEl && sel.options[bestIdx].value) idEl.value = sel.options[bestIdx].value;
    if (nameEl) nameEl.value = (sel.options[bestIdx].textContent || '').trim();
    return true;
  }
  return false;
}
"""


def choose_first_and_only_option(popup, market: str, *, select_id: str = "") -> str:
    """★요건: 엑셀에서 찾은 확정 카테고리로 검색하면 망고 목록에는 항상 그 항목 1개뿐이다.

    검색 후 select 리스트박스에 로드된 유일한 결과 항목을 바로 선택하고,
    화면 hidden 필드(openmarket_cm_category_<코드>)에 동기화한다.
    """
    ids = [select_id] if select_id else result_select_ids(market)
    for sid in ids:
        loc = _first(popup, (f"#{sid}",))
        if loc is None:
            continue
        try:
            res = popup.evaluate(CHOOSE_OPTION_JS, [sid, "", market])
            if res:
                name_el = popup.evaluate(
                    f"() => {{ const el = document.getElementsByName('openmarket_cm_category_name_{market}')[0]; return el ? el.value : ''; }}"
                )
                return str(name_el or "").strip()
        except Exception:
            continue
    return ""


def choose_option(popup, market: str, label: str, *, select_id: str = "") -> bool:
    ids = [select_id] if select_id else result_select_ids(market)
    for sid in ids:
        loc = _first(popup, (f"#{sid}",))
        if loc is None:
            continue
        try:
            loc.select_option(label=label, timeout=T_CLICK)
            return True
        except Exception:
            try:
                res = popup.evaluate(CHOOSE_OPTION_JS, [sid, label, market])
                if res:
                    return True
            except Exception:
                continue
    return False


def click_config_save(popup, *, progress: ProgressFn | None = None) -> bool:
    """[검색필터 설정저장 (Alt+S)]."""
    loc = _first(
        popup,
        (
            f'a[onclick*="{CONFIG_SAVE_JS}"]',
            'xpath=//a[.//span[contains(normalize-space(.),"검색필터 설정저장")]]',
        ),
    )
    if loc is None:
        _log(progress, "  오류: [검색필터 설정저장] 미검출", major=True)
        return False
    try:
        loc.click(timeout=T_CLICK)
        _log(progress, "  검색필터 설정저장")
        return True
    except Exception:
        try:
            popup.keyboard.press("Alt+s")
            _log(progress, "  검색필터 설정저장 (Alt+S)")
            return True
        except Exception:
            return False


def close_popup(popup) -> None:
    try:
        if not popup.is_closed():
            popup.close()
    except Exception:
        pass


def map_one_market(
    popup,
    market: str,
    filter_name: str,
    categories: Sequence[str],
    *,
    variant: str = "",
    exclude: Sequence[str] = (),
    retries: int = MAP_RETRIES,
    region_type: str = "국내",
    custom_search_keyword: str = "",
    db: category_db.CategoryDB | None = None,
    keyword_db: keyword_dictionary.KeywordDB | None = None,
    master_db: category_master_db.MasterDB | None = None,
    ext_db: extended_master_db.ExtendedMasterDB | None = None,
    progress: ProgressFn | None = None,
) -> MappedItem:
    """엑셀 확정 → 망고 제출을 **딱 한 번**만 한다."""
    return _map_once(
        popup,
        market,
        filter_name,
        categories,
        variant=variant,
        exclude=exclude,
        region_type=region_type,
        custom_search_keyword=custom_search_keyword,
        db=db,
        keyword_db=keyword_db,
        master_db=master_db,
        ext_db=ext_db,
        progress=progress,
    )


def _map_once(
    popup,
    market: str,
    filter_name: str,
    categories: Sequence[str],
    *,
    variant: str = "",
    exclude: Sequence[str] = (),
    region_type: str = "국내",
    custom_search_keyword: str = "",
    db: category_db.CategoryDB | None = None,
    keyword_db: keyword_dictionary.KeywordDB | None = None,
    master_db: category_master_db.MasterDB | None = None,
    ext_db: extended_master_db.ExtendedMasterDB | None = None,
    progress: ProgressFn | None = None,
) -> MappedItem:
    """한 마켓(+구분) 매핑 — 최적 카테고리 → 검색어 입력 → 검색 → 목록 선택."""
    label = MARKETS.get(market, market) + (f" · {variant}" if variant else "")
    if not categories:
        return MappedItem(market, "", 0.0, False, "엑셀 자료 없음")

    if variant and not select_variant(popup, market, variant, progress=progress):
        return MappedItem(market, "", 0.0, False, f"구분({variant}) 선택 실패")

    # ★요건(2026-08-23 3단계 우선순위):
    #   1차 : 엑셀 카테고리 직접 검색
    #   2차 : 확장형DB(ext_db)에서 키워드 검색으로 범주·범위 확장 후 엑셀 검색
    #   3차 : 매핑DB(master_db / db / keyword_db)에서 범주·범위 확장 후 엑셀 검색
    category, step = "", ""
    from_master = False
    if master_db is not None:
        market_name = MARKETS.get(market, market)
        category, step = best_category_via_master(
            filter_name, market_name, master_db, exclude=exclude, region_type=region_type
        )
        if category:
            step = f"[마스터DB] {step}"
            from_master = True

    if not category:
        category, step = best_category_with_step(
            filter_name,
            categories,
            exclude=exclude,
            market=market,
            region_type=region_type,
            db=db,
            keyword_db=keyword_db,
            ext_db=ext_db,
        )

    # ★절대규칙: 반대 성별 카테고리는 고르지 않는다
    if category and matching.violates_gender(category, filter_name):
        gender = matching.gender_of(filter_name)
        safe = matching.strip_opposite_gender(categories, gender)
        _log(progress, f"  {label}: 반대 성별 카테고리 배제 → 재선정", major=True)
        category, step = best_category_with_step(
            filter_name,
            safe,
            exclude=exclude,
            market=market,
            region_type=region_type,
            db=db,
            keyword_db=keyword_db,
            ext_db=ext_db,
        )
        from_master = False

    # ★요건: 최적 카테고리는 반드시 실제 존재하는 값이어야 한다
    if category and not from_master and not matching.is_from(categories, category):
        fixed = matching.ensure_from(categories, category, filter_name)
        _log(progress, f"  {label}: 엑셀 범위 밖 → 목록 내 값으로 교정 ({fixed})")
        category, step = fixed, step + " · 엑셀범위 교정"

    # ★요건 1: 미매핑 방지 강제 할당 (어떠한 경우에도 빈 카테고리로 끝나지 않게)
    if not category and categories:
        safe_pool = matching.strip_forbidden_categories(categories, filter_name)
        if region_type == "국내":
            safe_pool = matching.filter_by_region(safe_pool, "국내")
        gender = matching.gender_of(filter_name)
        if gender:
            safe_pool = matching.strip_opposite_gender(safe_pool, gender)
        category = safe_pool[0] if safe_pool else categories[0]
        step = "안전 폴백 강제지정"

    score = 1.0 if category else 0.0
    if not category:
        _log(progress, f"  {label}: 매칭 실패 ({step})")
        return MappedItem(market, "", 0.0, False, "유사 카테고리 없음")
    _log(progress, f"  {label}: 최적 카테고리(엑셀) = {category}  [{step}]")

    box = market_search_input(popup, market)
    if box is None:
        return MappedItem(market, category, score, False, "검색필드 미검출")

    # ★요건: 어떠한 경우에도 미매칭으로 넘어가서는 안 된다.
    # 1. 확정된 카테고리명(공백 구분)으로 우선 검색
    # 2. 결과가 안 나오면 리프명, 하위 키워드, 중위 키워드, 품목어로 순차 검색하여 반드시 검색 결과 목록을 확보
    search_keywords = []
    if custom_search_keyword:
        search_keywords.append(custom_search_keyword.strip())
    else:
        k_full = search_keyword_for(category)
        if k_full:
            search_keywords.append(k_full)
        k_leaf = category.split(">")[-1].strip()
        if k_leaf and k_leaf not in search_keywords:
            search_keywords.append(k_leaf)
        for part in k_leaf.replace("/", " ").replace(",", " ").split():
            if part and part not in search_keywords:
                search_keywords.append(part)
        parsed = matching.parse_filter_name(filter_name)
        for low in parsed.lows:
            if low and low not in search_keywords:
                search_keywords.append(low)
        for mid in parsed.mids:
            if mid and mid not in search_keywords:
                search_keywords.append(mid)
        cls = matching.class_of(filter_name)
        if cls and cls not in search_keywords:
            search_keywords.append(cls)
        for fb in ("의류", "신발", "잡화", "패션", "기타"):
            if fb not in search_keywords:
                search_keywords.append(fb)

    options: list[str] = []
    select_id: str = ""
    used_keyword: str = ""

    for kw in search_keywords:
        try:
            box.fill(kw, timeout=T_CLICK)
            click_market_search(popup, market)
            opts, sid = read_result_options(popup, market, timeout_ms=3000)
            if opts:
                options = opts
                select_id = sid
                used_keyword = kw
                break
        except Exception:
            continue

    if not options:
        return MappedItem(market, category, score, False, "검색 결과 없음")

    # 결과 목록에서 최적 선택: 완전일치 -> 리프일치 -> 부분일치 -> 첫 번째 항목 (절대 비우지 않음)
    picked = pick_option(options, category)
    if not picked and options:
        valid_opts = [o for o in options if o and not o.startswith("-") and not "선택해주세요" in o]
        if valid_opts:
            target_leaf = category.split(">")[-1].strip()
            picked = next((o for o in valid_opts if target_leaf and target_leaf in o), valid_opts[0])

    if not picked:
        picked = options[0]

    choose_option(popup, market, picked, select_id=select_id)
    _log(progress, f"  {label}: 선택 완료 → {picked} (검색어='{used_keyword}')")
    return MappedItem(market, picked, score, True, step)


MAPPED_STATE_JS = """
(codes) => {
  const out = {};
  for (const code of codes) {
    const idEl = document.getElementById('openmarket_cm_category_' + code);
    const nameEl = document.getElementsByName('openmarket_cm_category_name_' + code)[0];
    const sel1 = document.getElementById('openmarket_category_search_list_' + code);
    const sel2 = document.getElementById('openmarket_category_search_list2_' + code);

    let codeVal = idEl ? (idEl.value || '').trim() : '';
    let nameVal = nameEl ? (nameEl.value || '').trim() : '';

    const sel = (sel1 && (sel1.offsetParent !== null || window.getComputedStyle(sel1).display !== 'none')) ? sel1 :
                (sel2 && (sel2.offsetParent !== null || window.getComputedStyle(sel2).display !== 'none')) ? sel2 : (sel1 || sel2);

    if (sel && sel.selectedIndex >= 0 && sel.options.length > 0) {
      const opt = sel.options[sel.selectedIndex];
      const optText = (opt.textContent || '').trim();
      const optVal = (opt.value || '').trim();

      if (!optVal || optText.startsWith('-') || optText.includes('선택해주세요')) {
        codeVal = '';
        nameVal = '';
      } else {
        if (!codeVal) codeVal = optVal;
        if (!nameVal) nameVal = optText;
      }
    }

    out[code] = { code: codeVal, name: nameVal };
  }
  return out;
}
"""


def mapped_state(popup, codes: Sequence[str]) -> dict[str, dict]:
    """마켓별 매핑 결과 (hidden 값 + select 표시값) — 저장 후 재검증용."""
    try:
        data = popup.evaluate(MAPPED_STATE_JS, list(codes)) or {}
    except Exception:
        return {}
    return {str(k): dict(v or {}) for k, v in data.items()}


def unmapped_markets(popup, codes: Sequence[str]) -> list[str]:
    """카테고리가 비어 있거나 '- 전송대상...' 상태인 마켓 목록."""
    state = mapped_state(popup, codes)
    if not state:
        return []
    out: list[str] = []
    for code in codes:
        info = state.get(code) or {}
        if not (info.get("code") or info.get("name")):
            out.append(code)
    return out


def format_input_summary(
    site_id: str,
    region_type: str,
    list_url: str,
    row_from: int,
    row_to: int,
) -> list[str]:
    """실행 시작 시 화면 입력 정보를 정확히 3줄로 출력한다."""
    return [
        f"입력정보 1/3 · 사이트={site_id} · 구분={region_type}",
        f"입력정보 2/3 · 작업행={row_from}~{row_to}",
        f"입력정보 3/3 · 목록URL={list_url}",
    ]


def format_row_summary(
    row: RowInfo,
    codes: Sequence[str],
    before: dict[str, dict],
    after: dict[str, dict],
    items: Sequence[dict],
) -> list[str]:
    """한 행 처리 결과를 요건의 4줄 형식으로 만든다(2~4행 들여쓰기)."""
    latest: dict[str, str] = {}
    for item in items:
        market = str(item.get("market") or "")
        category = str(item.get("category") or "").strip()
        if market and category:
            latest[market] = category

    def value(state: dict[str, dict], code: str) -> str:
        info = state.get(code) or {}
        return str(info.get("name") or info.get("code") or "").strip()

    finals: list[str] = []
    transitions: list[str] = []
    mapped = 0
    for code in codes:
        label = MARKETS.get(code, code)
        before_value = value(before, code) or "미매핑"
        after_value = value(after, code) or latest.get(code, "") or "미매핑"
        if after_value != "미매핑":
            mapped += 1
        finals.append(f"{label}={after_value}")
        transitions.append(f"{label}:{before_value}→{after_value}")

    return [
        f"1. 망고 데이터 · ftid={row.ftid} · 필터={row.filter_name}",
        "    2. 마켓별 확정된 최종카테고리명 · " + " | ".join(finals),
        "    3. 마켓별 입력데이터 입력전 / 입력후 · " + " | ".join(transitions),
        f"    4. 6개 마켓저장 최종매핑 결과 : {mapped}건 / {len(codes)}건",
    ]


def show_live_screen(popup) -> None:
    """망고 수행 화면을 앞으로 가져와 사용자가 실시간으로 보게 한다."""
    reveal(popup)
    try:
        popup.wait_for_timeout(100)
    except Exception:
        time.sleep(0.1)


def anomalous_gender_markets(
    popup, codes: Sequence[str], filter_name: str
) -> dict[str, str]:
    """★[검색필터 설정저장] 후 실제 저장된 값 재검증 — 반대 성별로 매핑된 마켓."""
    state = mapped_state(popup, codes)
    if not state:
        return {}
    out: dict[str, str] = {}
    for code in codes:
        name = str((state.get(code) or {}).get("name") or "").strip()
        if name and matching.violates_gender(name, filter_name):
            out[code] = name
    return out


def map_one_row(
    page,
    row: RowInfo,
    excels: dict[str, list[str]],
    *,
    list_url: str,
    markets: Sequence[str] | None = None,
    variant_choice: dict[str, str] | None = None,
    region_type: str = "국내",
    db: category_db.CategoryDB | None = None,
    keyword_db: keyword_dictionary.KeywordDB | None = None,
    master_db: category_master_db.MasterDB | None = None,
    ext_db: extended_master_db.ExtendedMasterDB | None = None,
    progress: ProgressFn | None = None,
) -> dict:
    """한 행 — 설정수정 팝업 → AI 매핑 → 마켓별 매핑 → 설정저장 → 닫기.

    ★요건 (2026-08-23):
    최초: 최초 등록 -> 저장
    2차: 확인 -> 누락 확인 -> 2차 등록 -> 저장
    3차: 확인 -> 누락 확인 -> 3차 등록 -> 저장
    """
    codes = list(markets or MARKETS.keys())
    detail: dict = {"ftid": row.ftid, "filter": row.filter_name, "items": []}

    _log(progress, f"필터 [{row.filter_name}] (ftid={row.ftid})", major=True)
    popup = open_setting_popup(page, row, list_url=list_url, progress=progress)
    if popup is None:
        detail["error"] = "팝업 실패"
        return detail

    try:
        popup.on("dialog", lambda d: d.accept())
    except Exception:
        pass

    before_state = mapped_state(popup, codes)
    show_live_screen(popup)

    try:
        # ── [1단계: 최초 등록 & 저장] ───────────────────────────────────
        _log(progress, f"  [1차: 최초] 6개 마켓 카테고리 매핑 (구분: {region_type})", major=True)
        for market in codes:
            if stop_requested():
                break
            show_live_screen(popup)

            # ★요건 2026-08-23 (11번가, 롯데ON 라디오 구분 순서 및 검색어 규칙):
            # A. 입력필드 구분이 "국내"로 선택된 경우:
            #    1. 해외 또는 해외직구 :: 라디오 선택 후 => "의류" 검색 후 임의의 리스트로 저장 (ALT+S)
            #    2. 국내 또는 일반 :: 라디오 선택 후 => "최종 카테고리명" 입력 검색 후 저장 (ALT+S)
            # B. 입력필드 구분이 "해외"로 선택된 경우:
            #    1. 해외 또는 해외직구 :: 라디오 선택 후 => "해외명품" 검색 후 임의의 리스트로 저장 (ALT+S)
            #    2. 국내 또는 일반 :: 라디오 선택 후 => "최종 카테고리명" 입력 검색 후 저장 (ALT+S)
            if market in ("11ST", "LTON"):
                if market == "11ST":
                    overseas_v, domestic_v = "해외카테고리", "국내카테고리"
                else:
                    overseas_v, domestic_v = "해외직구 카테고리", "일반카테고리"

                # 1. 해외 탭 처리
                kw_for_overseas = "의류" if region_type == "국내" else "해외명품"
                _log(progress, f"  {MARKETS.get(market, market)}: 1단계 [{overseas_v}] 라디오 선택 → '{kw_for_overseas}' 검색 후 저장")
                item_ov = map_one_market(
                    popup,
                    market,
                    row.filter_name,
                    excels.get(market, []),
                    variant=overseas_v,
                    region_type=region_type,
                    custom_search_keyword=kw_for_overseas,
                    db=db,
                    keyword_db=keyword_db,
                    master_db=master_db,
                    ext_db=ext_db,
                    progress=progress,
                )
                if item_ov.ok:
                    click_config_save(popup, progress=progress)
                record_ov = dict(item_ov.__dict__)
                record_ov["variant"] = overseas_v
                record_ov["stage"] = "1차_해외탭"
                detail["items"].append(record_ov)
                time.sleep(GAP)

                # 2. 국내/일반 탭 처리
                _log(progress, f"  {MARKETS.get(market, market)}: 2단계 [{domestic_v}] 라디오 선택 → '최종 카테고리명' 검색 후 저장")
                item_dom = map_one_market(
                    popup,
                    market,
                    row.filter_name,
                    excels.get(market, []),
                    variant=domestic_v,
                    region_type=region_type,
                    custom_search_keyword="",  # 최종 카테고리명 사용
                    db=db,
                    keyword_db=keyword_db,
                    master_db=master_db,
                    ext_db=ext_db,
                    progress=progress,
                )
                if item_dom.ok:
                    click_config_save(popup, progress=progress)
                record_dom = dict(item_dom.__dict__)
                record_dom["variant"] = domestic_v
                record_dom["stage"] = "1차_국내탭"
                detail["items"].append(record_dom)
                time.sleep(GAP)
                continue

            for variant in variants_for(market, (variant_choice or {}).get(market, "")):
                if stop_requested():
                    break
                tried: list[str] = []
                item = map_one_market(
                    popup,
                    market,
                    row.filter_name,
                    excels.get(market, []),
                    variant=variant,
                    region_type=region_type,
                    db=db,
                    keyword_db=keyword_db,
                    master_db=master_db,
                    ext_db=ext_db,
                    progress=progress,
                )
                if item.ok:
                    tried.append(item.category)
                    click_config_save(popup, progress=progress)

                    # 상품고시정보 팝업이 뜨면 닫고 다른 카테고리로 다시 매핑
                    for _ in range(NOTIFY_RETRIES):
                        if not notify_open(popup, market):
                            break
                        _log(
                            progress,
                            f"  {MARKETS.get(market, market)}: 상품고시정보 팝업 — 다른 카테고리로 재매핑",
                            major=True,
                        )
                        close_notify(popup, market, progress=progress)
                        item = map_one_market(
                            popup,
                            market,
                            row.filter_name,
                            excels.get(market, []),
                            variant=variant,
                            exclude=tried,
                            region_type=region_type,
                            db=db,
                            keyword_db=keyword_db,
                            master_db=master_db,
                            ext_db=ext_db,
                            progress=progress,
                        )
                        if not item.ok:
                            break
                        tried.append(item.category)
                        click_config_save(popup, progress=progress)

                item.reason = item.reason or ""
                record = dict(item.__dict__)
                record["variant"] = variant
                record["stage"] = "1차_최초"
                detail["items"].append(record)
                time.sleep(GAP)

        # 1차 저장
        _log(progress, "  [1차] 최초 등록 완료 → 저장 (Alt+S)", major=True)
        show_live_screen(popup)
        click_config_save(popup, progress=progress)
        try:
            popup.wait_for_timeout(500)
        except Exception:
            time.sleep(0.5)

        # ── [2단계: 확인 -> 누락 확인 -> 2차 등록 -> 저장] ───────────────
        missing_2 = unmapped_markets(popup, codes)
        if missing_2:
            names_2 = " · ".join(MARKETS.get(m, m) for m in missing_2)
            _log(progress, f"  [2차: 확인] 누락 확인 ({len(missing_2)}개 마켓: {names_2}) → 2차 등록 시작", major=True)
            for market in missing_2:
                if stop_requested():
                    break
                if market in ("11ST", "LTON"):
                    v = "국내카테고리" if market == "11ST" else "일반카테고리"
                else:
                    v = variants_for(market, (variant_choice or {}).get(market, ""))[0]
                retry2 = map_one_market(
                    popup,
                    market,
                    row.filter_name,
                    excels.get(market, []),
                    variant=v,
                    region_type=region_type,
                    custom_search_keyword="",
                    db=db,
                    keyword_db=keyword_db,
                    master_db=master_db,
                    ext_db=ext_db,
                    progress=progress,
                )
                record2 = dict(retry2.__dict__)
                record2["variant"] = v
                record2["stage"] = "2차_등록"
                detail["items"].append(record2)
                time.sleep(GAP)

            _log(progress, "  [2차] 2차 등록 완료 → 저장 (Alt+S)", major=True)
            show_live_screen(popup)
            click_config_save(popup, progress=progress)
            try:
                popup.wait_for_timeout(500)
            except Exception:
                time.sleep(0.5)
        else:
            _log(progress, "  [2차: 확인] 누락 없음 — 전 마켓 매핑 완료 확인", major=True)

        # ── [3단계: 확인 -> 누락 확인 -> 3차 등록 -> 저장] ───────────────
        missing_3 = unmapped_markets(popup, codes)
        if missing_3:
            names_3 = " · ".join(MARKETS.get(m, m) for m in missing_3)
            _log(progress, f"  [3차: 확인] 누락 확인 ({len(missing_3)}개 마켓: {names_3}) → 3차 등록 시작", major=True)
            for market in missing_3:
                if stop_requested():
                    break
                if market in ("11ST", "LTON"):
                    v = "국내카테고리" if market == "11ST" else "일반카테고리"
                else:
                    v = variants_for(market, (variant_choice or {}).get(market, ""))[0]
                retry3 = map_one_market(
                    popup,
                    market,
                    row.filter_name,
                    excels.get(market, []),
                    variant=v,
                    region_type=region_type,
                    custom_search_keyword="",
                    db=db,
                    keyword_db=keyword_db,
                    master_db=master_db,
                    ext_db=ext_db,
                    progress=progress,
                )
                record3 = dict(retry3.__dict__)
                record3["variant"] = v
                record3["stage"] = "3차_등록"
                detail["items"].append(record3)
                time.sleep(GAP)

            _log(progress, "  [3차] 3차 등록 완료 → 저장 (Alt+S)", major=True)
            show_live_screen(popup)
            click_config_save(popup, progress=progress)
            try:
                popup.wait_for_timeout(500)
            except Exception:
                time.sleep(0.5)
        else:
            _log(progress, "  [3차: 확인] 누락 없음 — 전 마켓 매핑 완료 확인", major=True)

        # 성별 재검증
        bad = anomalous_gender_markets(popup, codes, row.filter_name)
        if bad:
            names = " · ".join(f"{MARKETS.get(m, m)}={n}" for m, n in bad.items())
            _log(progress, f"  ⚠ 성별 이상 재검증 — {names}", major=True)
            for market, bad_name in bad.items():
                if stop_requested():
                    break
                variant = variants_for(market, (variant_choice or {}).get(market, ""))[0]
                retry = map_one_market(
                    popup,
                    market,
                    row.filter_name,
                    excels.get(market, []),
                    variant=variant,
                    exclude=[bad_name],
                    region_type=region_type,
                    db=db,
                    keyword_db=keyword_db,
                    master_db=master_db,
                    ext_db=ext_db,
                    progress=progress,
                )
                record = dict(retry.__dict__)
                record["variant"] = variant
                record["stage"] = "성별재선정"
                detail["items"].append(record)
            click_config_save(popup, progress=progress)

        final_missing = unmapped_markets(popup, codes)
        if final_missing:
            detail["unmapped"] = final_missing
            _log(
                progress,
                "  ⚠ 최종 미매핑 마켓: " + " · ".join(MARKETS.get(m, m) for m in final_missing),
                major=True,
            )
        else:
            _log(progress, "  ★ 전 마켓 100% 매핑 확인 완료", major=True)

        final_state = mapped_state(popup, codes)
        for summary_line in format_row_summary(
            row, codes, before_state, final_state, detail["items"]
        ):
            _log(progress, summary_line, major=summary_line.startswith("1."))
        _log(progress, "다음행으로 이동", major=True)
    finally:
        close_popup(popup)
    return detail


def format_row_list(rows: Sequence[RowInfo], *, row_from: int | str = "", row_to: int | str = "") -> list[str]:
    """행 번호 확인용 — 1행부터 순서대로 `번호행: ftid=… · 필터=…`.

    범위를 주면 그 구간에 ★ 표시를 붙여 어디부터 작업되는지 한눈에 보인다.
    """
    start, end = (None, None)
    if str(row_from).strip() or str(row_to).strip():
        start, end = row_range(row_from or DEFAULT_ROW_FROM, row_to or DEFAULT_ROW_TO)
    lines: list[str] = []
    for i, row in enumerate(rows, start=1):
        mark = "★" if (start is not None and start <= i <= end) else "  "
        lines.append(f"{mark}{i:>3}행: ftid={row.ftid or '?'} · 필터={row.filter_name or '(이름 없음)'}")
    return lines


def list_rows_only(
    *,
    site_id: str = DEFAULT_SITE,
    list_url: str = "",
    row_from: int | str = "",
    row_to: int | str = "",
    progress: ProgressFn | None = None,
) -> list[RowInfo]:
    """망고 목록을 열어 행 순서만 확인한다 (매핑 없음) — '몇 번째 행인지' 검증용."""
    url = (list_url or "").strip()
    if not url:
        _log(
            progress,
            "목록 URL 을 입력하세요 — 브라우저에서 검색필터 목록 화면을 띄운 뒤"
            " 주소창 URL 을 그대로 붙여넣으면 됩니다.",
            major=True,
        )
        return []

    try:
        import collect as p2  # noqa: WPS433
        from playwright.sync_api import sync_playwright
    except ImportError as e:
        _log(progress, f"의존성 로드 실패: {e}", major=True)
        return []

    _start, _end = row_range(row_from or DEFAULT_ROW_FROM, row_to or DEFAULT_ROW_TO)
    if not (str(row_from).strip() or str(row_to).strip()):
        _end = 0  # 범위 지정이 없으면 현재 페이지만 (페이지 이동 없음)

    rows: list[RowInfo] = []
    try:
        with sync_playwright() as pw:
            _browser, page = p2.connect_browser(pw)
            page.goto(url, wait_until="domcontentloaded", timeout=60_000)
            reveal(page, progress=progress)
            _log(progress, f"목록 URL 로 화면 표시: {url}", major=True)
            if not select_site(page, site_id, progress=progress):
                _log(progress, "  사이트 선택을 건너뜁니다 — URL 화면 결과로 계속합니다", major=True)
            click_search_filter(page, progress=progress)
            rows = (
                collect_rows_for_range(page, _end, site_id, progress=progress)
                if _end > 0
                else list_rows(page)
            )
    except Exception as e:  # noqa: BLE001
        _log(progress, f"실행 오류: {e}", major=True)
        return rows

    _log(progress, f"검색 결과 {len(rows)}행 (위에서부터 1행)", major=True)
    for line in format_row_list(rows, row_from=row_from, row_to=row_to):
        _log(progress, line)
    return rows


def run_mapping(
    *,
    site_id: str = DEFAULT_SITE,
    excels: dict[str, list[str]] | None = None,
    excel_paths: dict[str, str] | None = None,
    list_url: str = "",
    markets: Sequence[str] | None = None,
    variant_choice: dict[str, str] | None = None,
    region_type: str = "국내",
    row_from: int | str = DEFAULT_ROW_FROM,
    row_to: int | str = DEFAULT_ROW_TO,
    progress: ProgressFn | None = None,
) -> RunResult:
    result = RunResult(ok=False)
    clear_stop_flag()

    site_id = (site_id or DEFAULT_SITE).strip()
    if not is_allowed_site(site_id):
        result.errors.append(
            f"수집사이트 제한: 현재는 {ALLOWED_SITES[0]} 만 수행합니다 (요청={site_id})."
        )
        _log(progress, result.errors[0], major=True)
        return result

    start, end = row_range(row_from, row_to)

    data = dict(excels or {})
    if not data:
        # ★요건: 엑셀 경로가 없으면(또는 읽기 실패) 캐시로 재사용한다.
        data = load_market_excels_or_cache(excel_paths, progress=progress)
    if not data:
        result.errors.append("마켓별 카테고리 엑셀이 없습니다.")
        _log(progress, result.errors[0], major=True)
        return result

    url = (list_url or "").strip() or DEFAULT_LIST_URL
    for input_line in format_input_summary(site_id, region_type, url, start, end):
        _log(progress, input_line, major=True)

    # ★요건재정의(2026-08-22 B): 6개 마켓 엑셀 전체를 교차검색한 통합정보화DB
    #   — 엑셀 탐색 5) 단계(연관검색어)에서 쓴다.
    db = build_category_db(data)
    _log(
        progress,
        f"통합정보화DB 구축 — 마켓 {db.market_count}건 · 카테고리 {db.path_count}건",
        major=True,
    )
    # ★요건(연관검색어DB, "최신성 우선"): keyword_dictionary 도 실제 실행
    #   경로에 연결한다 — 이전까지 만들어만 두고 전혀 쓰이지 않았다.
    keyword_db = build_keyword_db(data)
    _log(
        progress,
        f"연관검색어DB 구축 — 카테고리 {len(keyword_db.categories)}건 · "
        f"검색어 {len(keyword_db.keywords)}건",
        major=True,
    )
    # ★요건(2026-08-23): 사용자 제공 CATEGORY_MASTER+KEYWORD_DICTIONARY
    #   CSV 기반 마스터DB(매핑DB)를 구축.
    master_db = build_master_db()
    _log(
        progress,
        f"마스터DB(매핑DB) 로드 — 카테고리 {len(master_db.categories)}건 · "
        f"검색어 {len(master_db.keywords)}건",
        major=True,
    )
    # ★요건(2026-08-23): 확장형DB(표준 카테고리 14,207건 + 키워드 29,999건) 로드
    ext_db = build_extended_master_db()
    _log(
        progress,
        f"확장형DB 로드 — 표준 카테고리 {len(ext_db.categories)}건 · "
        f"표준 키워드 {len(ext_db.keywords)}건",
        major=True,
    )

    try:
        import collect as p2  # noqa: WPS433
        from playwright.sync_api import sync_playwright
    except ImportError as e:
        result.errors.append(f"의존성 로드 실패: {e}")
        _log(progress, result.errors[0], major=True)
        return result

    try:
        with sync_playwright() as pw:
            _browser, page = p2.connect_browser(pw)
            page.goto(url, wait_until="domcontentloaded", timeout=60_000)
            reveal(page, progress=progress)
            _log(progress, f"검색필터 목록 화면 (구분: {region_type})", major=True)

            if not select_site(page, site_id, progress=progress):
                result.errors.append("상품수집사이트 선택 실패")
            click_search_filter(page, progress=progress)

            scanned = collect_rows_for_range(page, end, site_id, progress=progress)
            found = [r for r in scanned if r.ftid]
            if not found:
                result.errors.append("작업 대상 행이 없습니다 (검색 결과 확인).")
                _log(progress, result.errors[0], major=True)
                diagnose_list(page, progress=progress)
                return result

            rows = slice_rows(found, start, end)
            result.rows = len(rows)
            checked = sum(1 for r in found if r.checked)
            _log(
                progress,
                f"검색 결과 {len(found)}행 중 **{start}~{end}행** → {len(rows)}건 수행"
                f" (체크 {checked}건 · 체크 무관 진행)",
                major=True,
            )
            if not rows:
                result.errors.append(
                    f"작업 범위({start}~{end})에 해당하는 행이 없습니다 (체크 {len(found)}건)."
                )
                _log(progress, result.errors[0], major=True)
                return result

            for i, row in enumerate(rows, start=1):
                if stop_requested():
                    _log(progress, "사용자 중단", major=True)
                    break

                _log(progress, f"[{i}/{len(rows)}] (전체 {start + i - 1}행)", major=True)
                detail = map_one_row(
                    page,
                    row,
                    data,
                    list_url=url,
                    markets=markets,
                    variant_choice=variant_choice,
                    region_type=region_type,
                    db=db,
                    keyword_db=keyword_db,
                    master_db=master_db,
                    ext_db=ext_db,
                    progress=progress,
                )
                result.details.append(detail)
                ok_cnt = sum(1 for it in detail.get("items", []) if it.get("ok"))
                result.mapped += ok_cnt
                result.failed += len(detail.get("items", [])) - ok_cnt
                time.sleep(GAP)
    except Exception as e:  # noqa: BLE001
        result.errors.append(str(e))
        _log(progress, f"실행 오류: {e}", major=True)
        return result
    finally:
        clear_stop_flag()

    result.ok = result.mapped > 0
    _log(
        progress,
        f"완료 — 행 {result.rows} · 매핑 성공 {result.mapped} · 실패 {result.failed}",
        major=True,
    )
    return result


def run_dry(
    filter_names: Sequence[str],
    excels: dict[str, list[str]],
    *,
    region_type: str = "국내",
    progress: ProgressFn | None = None,
) -> list[dict]:
    """브라우저 없이 매칭 결과만 확인 (검증용)."""
    db = build_category_db(excels)
    keyword_db = build_keyword_db(excels)
    master_db = build_master_db()
    ext_db = build_extended_master_db()
    out: list[dict] = []
    for name in filter_names:
        row = {"filter": name, "items": []}
        for code, cats in excels.items():
            cat, step = "", ""
            market_name = MARKETS.get(code, code)
            if master_db is not None:
                cat, step = best_category_via_master(
                    name, market_name, master_db, region_type=region_type
                )
                if cat:
                    step = f"[마스터DB] {step}"
            if not cat:
                cat, step = best_category_with_step(
                    name, cats, region_type=region_type, db=db, keyword_db=keyword_db, ext_db=ext_db
                )
            row["items"].append({"market": code, "category": cat, "step": step})
            _log(progress, f"{name} · {market_name} → {cat or '(없음)'} [{step}]")
        out.append(row)
    return out


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="P5_101_카테고리매핑_필터세부설정")
    parser.add_argument(
        "--site-id",
        default=DEFAULT_SITE,
        help=f"상품수집사이트 (현재 {ALLOWED_SITES[0]} 만 허용)",
    )
    parser.add_argument(
        "--region-type",
        default="국내",
        choices=["국내", "해외"],
        help="국내/해외 구분 (기본: 국내)",
    )
    parser.add_argument(
        "--row-from",
        type=int,
        default=DEFAULT_ROW_FROM,
        help=f"작업 시작 행 (1부터, 기본 {DEFAULT_ROW_FROM})",
    )
    parser.add_argument(
        "--row-to",
        type=int,
        default=DEFAULT_ROW_TO,
        help=f"작업 종료 행 (포함, 기본 {DEFAULT_ROW_TO})",
    )
    parser.add_argument("--list-url", default="", help=f"목록 URL (기본={DEFAULT_LIST_URL[:60]}…)")
    parser.add_argument("--excel-dir", default="", help="마켓별 엑셀 폴더 (파일명으로 자동 매칭)")
    parser.add_argument(
        "--excel",
        action="append",
        default=[],
        metavar="코드=경로",
        help="마켓별 엑셀 지정 (예: AUC20=D:\\옥션.xlsx)",
    )
    parser.add_argument("--markets", default="", help="대상 마켓 (쉼표, 기본=전체)")
    parser.add_argument(
        "--list-rows",
        action="store_true",
        help="매핑 없이 행 번호·ftid·필터명만 확인 ('몇 번째 행인지' 검증용)",
    )
    parser.add_argument(
        "--variant",
        action="append",
        default=[],
        metavar="코드=구분",
        help="마켓 구분 선택 (예: 11ST=국내카테고리 · LTON=둘다 · 기본=둘다)",
    )
    parser.add_argument("--dry-run", default="", help="매칭만 확인할 필터이름 (쉼표)")
    args = parser.parse_args(argv)

    paths: dict[str, str] = {}
    if args.excel_dir:
        paths.update(discover_market_excels(args.excel_dir))
    for item in args.excel:
        code, _, path = str(item).partition("=")
        if code and path:
            paths[code.strip().upper()] = path.strip()

    # ★요건: 엑셀 경로가 없으면 캐시(직전에 구축해 둔 카테고리)로 재사용한다.
    excels = load_market_excels_or_cache(paths)

    if args.dry_run:
        run_dry(
            [n.strip() for n in args.dry_run.split(",") if n.strip()],
            excels,
            region_type=args.region_type,
        )
        return 0

    markets = [m.strip().upper() for m in args.markets.split(",") if m.strip()] or None
    variant_choice: dict[str, str] = {}
    for item in args.variant:
        code, _, value = str(item).partition("=")
        if code and value:
            variant_choice[code.strip().upper()] = value.strip()
    result = run_mapping(
        site_id=args.site_id,
        excels=excels,
        list_url=args.list_url,
        markets=markets,
        variant_choice=variant_choice,
        region_type=args.region_type,
        row_from=args.row_from,
        row_to=args.row_to,
    )
    for e in result.errors:
        print(f"[오류] {e}", flush=True)
    return 0 if result.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
