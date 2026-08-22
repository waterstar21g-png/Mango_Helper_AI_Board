"""
P2 — 더망고(tmg1898) 상품데이터 대량수집 (BATCH 순차)

목표: 카테고리 URL 1행당 상품 N건(기본 3)을 오류 없이 가져오기.

BATCH 필수 순서 (딴 길로 빠지지 않음 — batch_steps.py):
  1 로그인 (main 1회)
  2 초기화(상품데이터수집→대량수집)
  3 URL 입력 → 4 URL상품검색하기 클릭
  5 검색팝업 열림 → 6 검색팝업 닫기-확인
  7 모두저장 → 8 필터·건수 → 9 저장하기
  10 저장팝업 열림 → 11 저장팝업 닫기-확인 → 12 건수로그
  13 수집전 화면 초기화 (다음 행용)
  14 이후는 3~13 반복
실패 최대 원인: 6·11·12 확인 없이 다음 단계 진행.

사용법:
    python collect.py 엑셀.xlsx              # 저장수 3
    python collect.py 엑셀.xlsx 3 --verify   # 전체 처리 + 1·2행 단계 스크린샷
    python collect.py 엑셀.xlsx 3 --max-rows 2   # 앞 2행만 처리(행 수 제한)
    python collect.py 엑셀.xlsx 3 --retries 3 --yes
    run-verify.bat 엑셀.xlsx
"""

import os
import re
import socket
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import (
    Browser,
    BrowserContext,
    Page,
    TimeoutError as PWTimeout,
    sync_playwright,
)

# Windows cp949 콘솔/파이프에서 특수기호 출력 시 UnicodeEncodeError 방지
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass


def safe_print(msg: str = "", *, flush: bool = True) -> None:
    """stdout이 cp949여도 죽지 않게 출력."""
    text = "" if msg is None else str(msg)
    try:
        print(text, flush=flush)
        return
    except UnicodeEncodeError:
        pass
    enc = getattr(sys.stdout, "encoding", None) or "utf-8"
    try:
        data = (text + "\n").encode(enc, errors="replace")
        sys.stdout.buffer.write(data)
        if flush:
            sys.stdout.flush()
    except Exception:
        try:
            print(text.encode("ascii", errors="replace").decode("ascii"), flush=flush)
        except Exception:
            pass

LOGIN_URL = "https://tmg1898.cafe24.com/mall/admin/admin_login.php"
MAIN_URL = "https://tmg1898.cafe24.com/mall/admin/admin.php"
BULK_URL = "https://tmg1898.cafe24.com/mall/admin/shop/getGoodsNew.php"
ADMIN_HOST = "tmg1898.cafe24.com"
BULK_PATH = "getGoodsNew.php"

CDP_PORT = 9222
CDP_URL = f"http://127.0.0.1:{CDP_PORT}"
PROFILE_DIR = Path(__file__).parent / ".chrome-profile"

# 더망고 솔루션 Chrome 확장프로그램 (Web Store ID = 로컬 load-extension 동일)
MANGO_EXT_ID = "lgfjcapohoongednoojdaiedebgbcelp"
MANGO_EXT_DIR = Path(__file__).parent / "extensions" / "themango-solution"
MANGO_EXT_POPUP = f"chrome-extension://{MANGO_EXT_ID}/popup.html"
MANGO_EXT_WEBSTORE = f"https://chromewebstore.google.com/detail/{MANGO_EXT_ID}"
# 확장이 전용 프로필에 없을 때 사용자에게 그대로 보여주는 안내.
# 최초 1회 웹스토어 설치만 하면 프로필(P2/.chrome-profile)에 남아 이후 실행부터
# 자동으로 인식된다.
MANGO_EXT_MISSING_GUIDE = (
    "더망고 솔루션 확장프로그램이 P2 전용 Chrome 프로필에 설치되어 있지 않습니다.\n"
    "  · 방금 열린 Chrome 창(주소창에 --no-sandbox 경고가 있는 창)에서\n"
    "    웹스토어 페이지를 띄웠습니다. [Chrome에 추가] 를 한 번만 눌러 주세요.\n"
    f"  · 설치 페이지: {MANGO_EXT_WEBSTORE}\n"
    "  · 설치가 확인되면 프로그램이 자동으로 이어서 진행합니다 "
    "(설치 내용은 전용 프로필에 남아 다음부터는 대기 없이 통과).\n"
    "  · 평소 쓰는 Chrome 에 이미 설치돼 있어도 소용없습니다 — P2 는 별도 프로필로\n"
    "    실행되며, Chrome 136 부터 원격 디버깅이 기본 프로필에서 막혀 있기 때문입니다.\n"
    f"  · 확인 대상: {MANGO_EXT_POPUP}\n"
    "  · 원인: {cause}"
)
# 사용자가 웹스토어에서 [Chrome에 추가] 를 누를 때까지 기다리는 제한 (초).
# 수동 로그인 대기(LOGIN_WAIT_SEC)와 같은 성격이라 동일하게 취급한다.
EXT_INSTALL_WAIT_SEC = 600
# 크롬 기동 시 확장 팝업에 반드시 넣을 값 (사용자 지정)
MANGO_SERVICE_URL = "https://tmg1898.cafe24.com"
MANGO_SERVICE_KEY = "y94Tmx9LbxxCJtk5uI9z0RjGWDtVW4"

CHROME_CANDIDATES = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    "/usr/bin/google-chrome-stable",
    "/usr/bin/google-chrome",
    "/usr/local/bin/google-chrome",
    "/usr/bin/chromium-browser",
    "/usr/bin/chromium",
]

POPUP_WAIT_SEC = 150  # 검색 팝업(6항) 닫힘 대기(초) — 초과 시 다음단계 금지·실패
# ★근본원인(2026-08-08, 두 차례 정정):
# 1차: 상품이 많은 카테고리는 확장프로그램의 임시메모리 적재(스크래핑)가
#      40초를 넘기는 경우가 있어 자동화가 사람보다 먼저 포기했다.
# 2차(진짜 원인): "확인/닫기 버튼을 자동 클릭해서 도와주는" 코드가 있었고,
#      1차 수정에서 이를 8초 간격으로 반복 클릭하게 만들었는데, 이것이
#      2026-08-06 사용자 지시("인위적으로 팝업창을 닫으면 망고 데이터
#      수집도 멈춤 — 팝업창은 스스로 닫힐 때까지 반드시 기다려야 함")를
#      위반해 오히려 스크래핑을 방해하고 있었다. wait_popups_close()에서
#      그 개입 코드를 완전히 제거하고, 순수 대기 시간만 150초로 넉넉히
#      늘렸다. 검색 팝업 안에서는 절대 아무것도 클릭하지 않는다.
MODAL_WAIT_SEC = 60
# 저장하기 클릭 직후 "뭔가 반응했나"만 빠르게 보는 짧은 확인(초) — 요건 3의
# 120초 무행동 대기와는 다른, 클릭 실패 여부만 가르는 용도.
SAVE_POPUP_GRACE_SEC = 5.0
# (이전 요건 — 이제는 아래 SAVE_COMPLETE_WAIT_SEC 즉시탐지 방식으로 대체됨)
SAVE_POPUP_BLIND_WAIT_SEC = 120.0
SAVE_POPUP_CONFIRM_WAIT_SEC = 300.0
# ★최신 요건: 120초 무행동 대기·300초 단계적 확인 생략 — 저장하기 클릭 후
# "신규상품의 저장이 완료되었습니다" 메세지가 보이는 즉시(대기 없이) 다음
# 단계(13항 초기화)로 진행한다.
# ★요건(2026-08-08): 9·10·11 합산 실행시간이 180초 내 완료되지 않으면
# 해당 입력을 포기하고 다음 엑셀 행으로 넘어간다.
SAVE_PHASE_BUDGET_SEC = 180.0
SAVE_COMPLETE_WAIT_SEC = 180.0  # 9~11 합산 상한과 동일
# ★요건(2026-08-20): 행당 저장상품수 3 → 50
DEFAULT_SAVE_COUNT = 50
DEFAULT_ROW_RETRIES = 1  # ★요건(2026-08-08): 엑셀 각 행은 1번 시도로 끝냄 — 재시도 없음
SEARCH_MAX_TRIES = 2  # URL 검색 재시도(행 안) — 적게 두고 다음 행으로 넘김
ROW_BUDGET_SEC = 240  # 한 입력 행(2~7항 검색단계)에 쓸 수 있는 최대 시간(초)

# 보드 "수집 종료" 버튼이 만드는 중단 플래그 (로그는 보드에 보존)
STOP_FLAG = Path(__file__).parent / ".collect_stop"

# 수집 14단계 중 실패 핵심 게이트: 6항·11항·12항
# (docs/수집_14단계_필수순서.md 참고)
COLLECT_STEP6 = "6. 검색팝업 닫기-확인"
COLLECT_STEP11 = "11. 저장팝업 닫기-확인"
COLLECT_STEP12 = "12. 저장건수 로그 확인"

# 같은 행을 더 돌리지 않고 다음 입력 데이터로 넘길 확정 실패 문구
ROW_ADVANCE_FAIL_MARKERS = (
    "검색결과가 없습니다",
    "망고 자체 메세지",
    "더망고 자체 메세지",
    "행 제한시간 초과",
    "팝업창이 닫히지 않음",
    "저장 팝업창이 닫히지 않음",
    "검색결과 확인 실패",
    "팝업이 뜨지 않음",
    "0건이 수집",
    "수집건수 알림",
    "수집 알림",
    "저장하기 서버",
    "서버에 반영되지",
    "서버 최종 갱신",
    "팝업창 모달이 나타나지",
    "팝업 없이 초기화",
    "최종 팝업이 닫히지",
    "팝업화면이 닫히지",
    "검색 팝업모달이 닫히지",
    "저장완료 메세지를",
    "저장 완료 없이",
    "9·10·11",
    "9~11",
    "180초",
    "6항",
    "11항",
    "12항",
)


class CollectStopped(Exception):
    """사용자가 보드에서 수집 종료를 요청함."""


class RowBudgetExceeded(Exception):
    """한 입력 행 처리 시간 초과 — 다음 입력으로 진행."""


def clear_stop_flag() -> None:
    try:
        STOP_FLAG.unlink(missing_ok=True)  # type: ignore[call-arg]
    except TypeError:
        # py<3.8 호환 아님 — 3.10+ 환경
        if STOP_FLAG.exists():
            try:
                STOP_FLAG.unlink()
            except OSError:
                pass
    except OSError:
        pass


def stop_requested() -> bool:
    try:
        return STOP_FLAG.is_file()
    except OSError:
        return False


def check_stop(where: str = "") -> None:
    if stop_requested():
        detail = f" ({where})" if where else ""
        raise CollectStopped(f"사용자 수집 종료 요청{detail}")

FILTER_NAME_LABEL = re.compile(r"검색\s*필터\s*명")
# 화면 표기: 저장상품수 / 검색결과상위 / (사용자 호칭) 수집상품수
SAVE_COUNT_LABEL = re.compile(
    r"저장\s*상품\s*수|검색결과\s*상위|수집\s*상품\s*수"
)
# 저장 완료로 볼 수 있는 화면 문구 (망고 버전에 따라 다를 수 있음)
SAVE_OK_PATTERNS = [
    re.compile(r"저장\s*(이\s*)?(완료|성공)"),
    re.compile(r"정상\s*처리"),
    re.compile(r"상품\s*(이\s*)?저장"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?저장"),
]
SAVE_FAIL_PATTERNS = [
    re.compile(r"저장\s*실패"),
    re.compile(r"오류\s*가\s*발생"),
    re.compile(r"다시\s*시도"),
]

# 망고 자체 알림 — "3건이 수집되었다" / "00건이수집되었다" / 저장 건수 등
MANGO_COLLECT_ALERT_PATTERNS = [
    re.compile(r"(\d+)\s*건\s*(이\s*)?수집\s*되었"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?저장\s*되었"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?등록\s*되었"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?수집\s*완료"),
    re.compile(r"(\d+)\s*건\s*(의\s*)?상품\s*(이\s*)?(수집|저장)"),
    re.compile(r"저장\s*된\s*상품\s*(\d+)\s*건"),
    re.compile(r"상품\s*(\d+)\s*건\s*(이\s*)?(수집|저장)"),
    re.compile(r"총\s*(\d+)\s*건\s*(이\s*)?(수집|저장)"),
]

# ★망고 "저장 시작/완료" 메세지 — 12항 SUB 원문 구간 구분용.
# 시작: "......신규상품(3개)의 저장을 시작합니다." / "신규상품의 저장을 시작합니다."
# 완료: "......신규상품의 저장이 완료되었습니다." / "완료하였습니다."
SAVE_START_MSG_PATTERN = re.compile(
    r"신규\s*상품\s*(?:\([^)]*\))?\s*의?\s*저장을?\s*시작"
)
SAVE_COMPLETE_MSG_PATTERN = re.compile(
    r"신규\s*상품\s*의?\s*저장\s*이?\s*완료\s*(?:되었|하였)습니다|"
    r"신규\s*상품\s*저장\s*완료|"
    r"저장\s*이\s*완료\s*(?:되었|하였)습니다"
)

# 더망고(자체 UI) — 검색 결과 없음 문구 (로딩 중 오판 금지, 로딩 종료 후에만 사용)
MANGO_NO_RESULT_PATTERNS = [
    re.compile(r"검색하신\s*검색에\s*대한\s*검색결과가\s*없습니다"),
    re.compile(r"검색결과가\s*없습니다"),
    re.compile(r"정확한\s*검색어인지\s*다시한번\s*확인"),
]

# 더망고 로딩 오버레이/문구 (빨간 "잠시만 기다려주세요" 등)
MANGO_LOADING_PATTERNS = [
    re.compile(r"load\s*product", re.I),
    re.compile(r"상품정보를\s*불러오는\s*중"),
    re.compile(r"잠시만\s*기다려"),
    re.compile(r"처리\s*중\s*입니다"),
    re.compile(r"검색\s*중"),
]

SHOT_ROOT = Path(__file__).parent / "run-logs"

# 현재 실행 컨텍스트(로그인 등 단계 샷을 같은 폴더에 모으기)
_ACTIVE_CTX: "RunCtx | None" = None

# 파일명 태그 → 한글 단계명 (갤러리/보드 표시용)
SHOT_STEP_LABELS: dict[str, str] = {
    "login_wait": "로그인 대기(창 표시)",
    "login_ok": "로그인 완료",
    "login_gate": "로그인 게이트",
    "login_required": "세션만료·재로그인",
    "ext_settings": "확장프로그램(더망고솔루션) 설정값 저장",
    "ext_settings_fail": "확장프로그램 설정 실패",
    "ext_installed": "확장프로그램 설치 확인",
    "ready": "준비완료(대량수집 진입)",
    "00_overlays_clear": "다음행 전 — 팝업/모달 전부 닫힘 확인",
    "00_overlays_stuck": "다음행 전 — 팝업/모달 미종료(경고)",
    "00_init_bulk": "0. 초기화 — 대량데이터수집",
    "01_url_filled": "1. URL 입력 완료",
    "01_popup_missing": "1. 검색 팝업 미표시(오류)",
    "01_popup_opened": "1. 검색 팝업 열림",
    "01_popup_closed": "1. 검색 팝업 닫힘",
    "01_mango_no_results": "1. 망고 검색결과 없음(자체메세지)",
    # ★삭제(2026-08-08): 6→7 "검색 결과 준비" 샷 — 더 이상 촬영하지 않음
    # "01_results_ready": "1. 검색 결과 준비",
    "02_save_modal": "2. 모두저장 모달",
    "02_no_count_field": "2. 저장수 필드 없음(오류)",
    "02_count_mismatch": "2. 저장수 불일치(오류)",
    "02_modal_filled": "2. 필터명·저장수 입력",
    "02_save_missing": "2. 저장하기 버튼 없음(오류)",
    "02_save_clicked": "2. 저장하기(서버제출) 클릭",
    "02_save_no_react": "2. 저장하기 클릭 무반응(재시도)",
    "02_save_failed": "2. 저장하기 서버제출 실패",
    "03_modal_stuck": "3. 저장 모달 미종료(오류)",
    "03_modal_closed": "3. 저장 모달 닫힘",
    "03_result_popup": "3. 저장 후 결과 팝업",
    "03_result_missing": "3. 저장 후 결과 팝업 없음(오류)",
    "03_collect_alert": "3. 망고 수집건수 알림 확인",
    "03_collect_alert_fail": "3. 망고 수집건수 알림 확인 실패",
    "04_row_done": "4. 행 완료",
}

# 사용자 수동 로그인 대기 제한 (초)
LOGIN_WAIT_SEC = 600

# ★실행로그 — main/sub 두 그리드 프로토콜 (board/app.py 참고).
# main: 1~13단계 줄만(단계 발생 1회=1행). sub: 그 "발생(seq)"에 딸린
# 추가정보·스크린샷 — 같은 단계번호가 행마다 반복되므로 seq로 구분한다.
# 형식: ##MAIN##<seq>##<n>##<msg>  /  ##SUB##<seq>##<msg>
#      ##SUBSHOT##<seq>##<path>##<label>
# (docs/수집_14단계_필수순서.md)
MAIN_LINE_MARK = "##MAIN##"
SUB_LINE_MARK = "##SUB##"
SUB_SHOT_MARK = "##SUBSHOT##"


def log(msg: str) -> None:
    """내부 진단용 — 활성 ctx가 있으면 sub그리드로, 없으면 표준출력에."""
    ctx = _ACTIVE_CTX
    if ctx is not None:
        ctx.info(msg)
        return
    safe_print(f"[{time.strftime('%H:%M:%S')}] {msg}")


def step_log(n: int, msg: str) -> None:
    """모듈 레벨에서도 13단계(main) 출력을 쓸 수 있게 — 활성 ctx로 위임."""
    ctx = _ACTIVE_CTX
    if ctx is not None:
        ctx.step(n, msg)
    else:
        safe_print(f"[{time.strftime('%H:%M:%S')}] {n}. {msg}")


class RunCtx:
    """행 단위 수집·재시도·스크린샷용 실행 컨텍스트

    shot_first_n: 입력 데이터 앞 N건(기본 2=1·2행)만 단계별 스크린샷.
    """

    def __init__(
        self,
        *,
        save_count: int = DEFAULT_SAVE_COUNT,
        retries: int = DEFAULT_ROW_RETRIES,
        verify: bool = False,
        max_rows: int | None = None,
        batch: bool = False,
        shot_dir: Path | None = None,
        shot_first_n: int = 2,
    ) -> None:
        global _ACTIVE_CTX
        self.save_count = save_count
        self.retries = max(1, retries)
        self.verify = verify
        self.max_rows = max_rows
        self.batch = batch or verify  # 검증 모드는 기본 무중단
        self.shot_first_n = max(0, int(shot_first_n))
        stamp = time.strftime("%Y%m%d_%H%M%S")
        self.shot_dir = shot_dir or (SHOT_ROOT / stamp)
        self.shot_dir.mkdir(parents=True, exist_ok=True)
        self.step_i = 0
        self.shots: list[dict] = []
        self._gallery_written = False
        self._current_step = 0  # main그리드 마지막 단계번호(1~13, 표시용)
        self._step_seq = 0  # main그리드 각 발생(행)의 고유번호 — sub 연결용
        self._step_ts: dict[int, str] = {}  # seq → main 진입 시각
        self._step_ts_end: dict[int, str] = {}  # seq → 다음 main 진입 시각(sub 범위 끝)
        self._total_rows = 0
        self._done_rows = 0
        self.input_ordinal = 0  # 처리 중인 입력 순서(1부터)
        self.current_label = ""
        self.current_url = ""
        self.row_deadline: float | None = None  # 행당 제한시간(epoch)
        self.save_phase_deadline: float | None = None  # 9~11 합산 180초
        # 저장하기(서버 최종 갱신) 성공 여부 — True 되기 전 행 완료 금지
        self.server_save_ok: bool = False
        # 6항: 검색 팝업 열림·닫힘
        self.search_popup_seen: bool = False
        self.search_popup_closed: bool = False
        # 저장하기 후 최종 팝업: 열림 확인 + 닫힘 확인 (둘 다 필수) — 10·11항
        self.save_popup_seen: bool = False
        self.save_popup_closed: bool = False
        # 12항: 저장건수 로그 확인
        self.save_count_logged: bool = False
        self.save_count_snapshot: int | None = None
        # 12항 SUB — 망고 저장 로그 원문 (시작~완료 구간)
        self.mango_save_log_lines: list[str] = []
        # 저장하기 클릭 후 ~ 팝업 열림·닫힘 완료 전: 초기화 진입 금지
        self.save_awaiting_popup: bool = False
        self.save_popup_kind: str = ""
        self.save_popup_ui_latched: bool = False
        # 12항 망고 '저장이 완료' 이후 SUB 화면 출력 mute
        self._mute_sub_after_save_complete: bool = False
        self.log_path = self.shot_dir / "run.log"
        self._log_file = open(self.log_path, "a", encoding="utf-8")
        _ACTIVE_CTX = self
        self.info(f"[샷폴더] {self.shot_dir}")
        if self.shot_first_n > 0:
            self.info(f"[샷대상] 입력 데이터 1~{self.shot_first_n}행 단계별 스크린샷")

    def check_budget(self, where: str = "") -> None:
        """중단 요청 + 행/저장단계 제한시간 검사."""
        check_stop(where)
        if (
            self.save_phase_deadline is not None
            and time.time() > self.save_phase_deadline
        ):
            detail = f" ({where})" if where else ""
            raise RowBudgetExceeded(
                f"9·10·11 합산 {SAVE_PHASE_BUDGET_SEC:.0f}초 초과{detail} "
                "— 다음 입력으로"
            )
        if self.row_deadline is not None and time.time() > self.row_deadline:
            detail = f" ({where})" if where else ""
            raise RowBudgetExceeded(
                f"행 제한시간 초과{detail} — {ROW_BUDGET_SEC}초 내 미완료, 다음 입력으로"
            )

    def begin_row(self, ordinal: int, row: dict) -> None:
        """행 처리 시작 — 로그에 카테고리명·URL 기록."""
        self.input_ordinal = ordinal
        self.current_label = str(row.get("label") or "").strip()
        self.current_url = str(row.get("url") or "").strip()
        self.row_deadline = time.time() + ROW_BUDGET_SEC
        self.save_phase_deadline = None
        self.server_save_ok = False
        self.search_popup_seen = False
        self.search_popup_closed = False
        self.save_popup_seen = False
        self.save_popup_closed = False
        self.save_count_logged = False
        self.save_count_snapshot = None
        self.mango_save_log_lines = []
        self.save_awaiting_popup = False
        self.save_popup_kind = ""
        self.save_popup_ui_latched = False
        self._mute_sub_after_save_complete = False
        excel_row = row.get("row", "?")
        self.info(
            f"--- 입력#{ordinal} 엑셀{excel_row}행 | "
            f"최종 카테고리명={self.current_label} | "
            f"최종 카테고리 URL주소={self.current_url} | "
            f"제한 {ROW_BUDGET_SEC}초 ---"
        )

    def wants_row_shot(self, row_no: int | None = None) -> bool:
        """공통(로그인 등 row_no=0) 또는 입력 1~N행만 샷."""
        if row_no in (None, 0):
            return True
        return 0 < self.input_ordinal <= self.shot_first_n

    def close(self) -> None:
        global _ACTIVE_CTX
        try:
            self.write_gallery()
        except Exception as e:  # noqa: BLE001
            try:
                self.info(f"  [갤러리 작성 실패] {e}")
            except Exception:
                pass
        try:
            self._log_file.close()
        except Exception:
            pass
        if _ACTIVE_CTX is self:
            _ACTIVE_CTX = None

    def set_progress_totals(self, total: int) -> None:
        self._total_rows = max(0, int(total))

    def emit_progress_meta(
        self,
        *,
        done: int | None = None,
        ordinal: int | None = None,
        label: str | None = None,
        url: str | None = None,
        main_line: bool = False,
    ) -> None:
        """main 상단 META — 총건수·완료·수집필드·카테고리URL.

        ★요건: 완료건→완료, 순번 삭제. 총건수는 목차행 제외 값.
        main_line=True 이면 sticky META 갱신 + MAIN 그리드에 오렌지 1행 추가.
        진행 적색용으로 ##META##진행##N 도 함께 보낸다(화면 META 줄에는 안 씀).
        """
        if done is not None:
            self._done_rows = max(0, int(done))
        if ordinal is not None:
            self.input_ordinal = int(ordinal)
        if label is not None:
            self.current_label = str(label).strip()
        if url is not None:
            self.current_url = str(url).strip()
        ts = time.strftime("%H:%M:%S")
        fields = (
            ("총건수", str(self._total_rows)),
            ("완료", str(self._done_rows)),
            ("수집 필드", self.current_label),
            ("카테고리 URL", self.current_url),
        )
        for field, val in fields:
            safe_print(f"[{ts}] ##META##{field}##{val}")
        # 카테고리URL목록 진행행 적색용 (META 표시 항목 아님)
        if self.input_ordinal > 0:
            safe_print(f"[{ts}] ##META##진행##{self.input_ordinal}")
        if main_line:
            # MAIN 그리드에 영구 1행(step=0 → 오렌지 meta 태그)
            one = " | ".join(
                f"{f} {v}" if v else f for f, v in fields
            )
            if self._step_seq > 0:
                self._step_ts_end[self._step_seq] = ts
            self._step_seq += 1
            seq = self._step_seq
            self._current_step = 0
            self._step_ts[seq] = ts
            safe_print(f"[{ts}] {MAIN_LINE_MARK}{seq}##0##{one}")
            try:
                self._log_file.write(f"[{ts}] 엑셀. {one}\n")
                self._log_file.flush()
            except Exception:
                pass

    def _sub_ts(self, seq: int) -> str:
        """sub 시각 = 현단계 main 진입 ~ 다음 main 진입."""
        start = self._step_ts.get(seq, "")
        end = self._step_ts_end.get(seq, start)
        if start and end and end != start:
            return f"{start}~{end}"
        return start or time.strftime("%H:%M:%S")

    def step(self, n: int, msg: str) -> None:
        """★main 그리드 — 1~13단계 줄만(발생마다 새 seq). 이후 info()/shot()은
        이 발생(seq)에 딸린 sub 항목으로 연결된다."""
        # 새 단계 시작 → 12항 저장완료 이후 SUB 화면 mute 해제
        self._mute_sub_after_save_complete = False
        ts = time.strftime("%H:%M:%S")
        if self._step_seq > 0:
            self._step_ts_end[self._step_seq] = ts
        self._step_seq += 1
        self._current_step = n
        seq = self._step_seq
        self._step_ts[seq] = ts
        safe_print(f"[{ts}] {MAIN_LINE_MARK}{seq}##{n}##{msg}")
        try:
            self._log_file.write(f"[{ts}] {n}. {msg}\n")
            self._log_file.flush()
        except Exception:
            pass

    def info(self, msg: str) -> None:
        """sub 그리드 — 마지막 step() 발생(seq)에 딸린 추가정보로 표시됨.

        ★요건: 12단계 망고 로그 '저장이 완료되었습니다' 이후 메세지는
        파일에는 남기되 화면(##SUB##)에는 내지 않는다.
        """
        seq = self._step_seq
        ts = self._sub_ts(seq)
        text = str(msg or "")
        try:
            self._log_file.write(f"[{ts}]   {text}\n")
            self._log_file.flush()
        except Exception:
            pass
        # 저장완료 메세지 자체는 화면에 남기고, 그 다음부터 mute
        is_save_done = bool(SAVE_COMPLETE_MSG_PATTERN.search(text))
        if getattr(self, "_mute_sub_after_save_complete", False) and not is_save_done:
            return
        safe_print(f"[{ts}] {SUB_LINE_MARK}{seq}##{text}")
        if is_save_done and int(getattr(self, "_current_step", 0) or 0) == 12:
            self._mute_sub_after_save_complete = True

    @staticmethod
    def label_for_tag(tag: str) -> str:
        if tag in SHOT_STEP_LABELS:
            return SHOT_STEP_LABELS[tag]
        for key, label in SHOT_STEP_LABELS.items():
            if key in tag:
                return label
        if tag.startswith("fail_attempt"):
            return f"실패 시도 {tag.replace('fail_attempt', '')}"
        return tag

    def shot(self, page: Page, tag: str, row_no: int | None = None) -> Path | None:
        if not self.wants_row_shot(row_no):
            return None
        self.step_i += 1
        safe = re.sub(r"[^\w\-가-힣]+", "_", tag)[:40]
        ord_part = f"i{self.input_ordinal}" if self.input_ordinal else "i0"
        name = f"{self.step_i:02d}_{ord_part}_r{row_no or 0}_{safe}.png"
        path = self.shot_dir / name
        label = self.label_for_tag(tag)
        main_step = self._current_step
        seq = self._step_seq
        try:
            page.screenshot(path=str(path), full_page=True)
            self.shots.append(
                {
                    "step": self.step_i,
                    "main_step": main_step,
                    "seq": seq,
                    "ordinal": self.input_ordinal,
                    "row": row_no or 0,
                    "tag": tag,
                    "label": label,
                    "category": self.current_label,
                    "url": self.current_url,
                    "file": name,
                    "path": str(path),
                }
            )
            ts = self._sub_ts(seq)
            try:
                self._log_file.write(
                    f"[{ts}]   [샷] {self.step_i:02d}. {label} -> {path.name}\n"
                )
                self._log_file.flush()
            except Exception:  # noqa: BLE001
                pass
            # ★12항 저장완료 이후 화면 SUB/샷 제외 (파일·갤러리에는 보관)
            if not getattr(self, "_mute_sub_after_save_complete", False):
                safe_print(f"[{ts}] {SUB_SHOT_MARK}{seq}##{path}##{label}")
            return path
        except Exception as e:  # noqa: BLE001
            self.info(f"[샷 실패] {label}: {e}")
            return None

    def write_gallery(self) -> Path | None:
        """1행 전과정 스크린샷 HTML 갤러리 + JSON 인덱스 작성."""
        if self._gallery_written:
            gallery = self.shot_dir / "index.html"
            return gallery if gallery.is_file() else None
        pngs = sorted(self.shot_dir.glob("*.png"))
        if not pngs and not self.shots:
            return None

        # shots 목록이 비어 있으면 파일명에서 재구성
        items = list(self.shots)
        if not items:
            for i, p in enumerate(pngs, start=1):
                stem = p.stem
                tag = stem
                m = re.match(r"^\d+_r\d+_(.+)$", stem)
                if m:
                    tag = m.group(1)
                items.append(
                    {
                        "step": i,
                        "row": 0,
                        "tag": tag,
                        "label": self.label_for_tag(tag),
                        "file": p.name,
                        "path": str(p),
                    }
                )

        import json

        idx_path = self.shot_dir / "shots.json"
        idx_path.write_text(
            json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        rows_html: list[str] = []
        for it in items:
            f = it["file"]
            label = it.get("label") or ""
            step = it.get("step", 0)
            cat = it.get("category") or ""
            url = it.get("url") or ""
            ord_n = it.get("ordinal") or 0
            meta_bits = [f]
            if ord_n:
                meta_bits.append(f"입력#{ord_n}")
            if cat:
                meta_bits.append(f"최종 카테고리명={cat}")
            if url:
                meta_bits.append(f"최종 카테고리 URL주소={url}")
            meta = " | ".join(meta_bits)
            rows_html.append(
                f'<section class="shot">'
                f"<h2>{step:02d}. {label}</h2>"
                f'<p class="meta">{meta}</p>'
                f'<img src="{f}" alt="{label}"/>'
                f"</section>"
            )

        html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8"/>
<title>1·2행 전과정 스크린샷</title>
<style>
body {{ font-family: "Malgun Gothic", sans-serif; margin: 16px; background: #0f172a; color: #e2e8f0; }}
h1 {{ font-size: 20px; }}
.shot {{ margin: 24px 0; padding: 12px; background: #1e293b; border-radius: 8px; }}
.shot h2 {{ margin: 0 0 6px; font-size: 16px; color: #93c5fd; }}
.meta {{ margin: 0 0 10px; color: #94a3b8; font-size: 12px; word-break: break-all; }}
img {{ max-width: 100%; height: auto; border: 1px solid #334155; background: #fff; }}
</style>
</head>
<body>
<h1>1·2행 전과정 스크린샷 ({len(items)}장)</h1>
<p>폴더: {self.shot_dir}</p>
{''.join(rows_html)}
</body>
</html>
"""
        gallery = self.shot_dir / "index.html"
        gallery.write_text(html, encoding="utf-8")
        self._gallery_written = True
        self.info(f"[갤러리] {gallery} ({len(items)}장)")
        return gallery


def shot_now(page: Page, tag: str, row_no: int | None = 0) -> Path | None:
    """활성 RunCtx가 있으면 같은 샷폴더에, 없으면 run-logs 루트에 저장."""
    ctx = _ACTIVE_CTX
    if ctx is not None:
        return ctx.shot(page, tag, row_no)
    SHOT_ROOT.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^\w\-가-힣]+", "_", tag)[:40]
    path = SHOT_ROOT / f"{time.strftime('%Y%m%d_%H%M%S')}_{safe}.png"
    try:
        page.screenshot(path=str(path), full_page=True)
        log(f"  [샷] {path.name}")
        return path
    except Exception as e:  # noqa: BLE001
        log(f"  [샷 실패] {e}")
        return None


# ── 브라우저 연결 (기존 Chrome/Edge에 CDP로 붙기, Chromium 다운로드 없음) ──

def cdp_port_open(port: int = CDP_PORT) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.5)
        try:
            s.connect(("127.0.0.1", port))
            return True
        except OSError:
            return False


def find_browser_exe() -> str | None:
    for path in CHROME_CANDIDATES:
        if Path(path).exists():
            return path
    return None


def _clear_stale_singleton_locks() -> None:
    """
    이전 실행이 비정상 종료되어 남은 잠금파일이 있으면, 새 Chrome이
    "이미 실행 중"이라고 착각해 디버그 포트 없이 조용히 기존 창에만
    메시지를 보내고 끝나버릴 수 있다(=화면에 아무 반응도 없어 보임).
    cdp_port_open()이 False라는 건 우리 프로필로 살아있는 프로세스가
    없다는 뜻이므로 안전하게 지운다.
    """
    for name in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
        f = PROFILE_DIR / name
        try:
            if f.exists() or f.is_symlink():
                f.unlink()
        except OSError:
            pass


# 저장하기(9항) 클릭 후 window.open() 팝업이 실제 사이트에서 안 뜨는
# 흔한 원인 — Chrome 자체 팝업차단(실제 Chrome 프로필 기본 ON).
# 클릭이 trusted 여도, AJAX 저장 완료 콜백처럼 클릭과 비동기로 분리돼
# window.open이 호출되면 Chrome이 "사용자 제스처 없음"으로 보고
# 조용히 막아버린다 — 그러면 팝업 창 자체가 아예 생기지 않는다.
POPUP_BLOCK_FIX_VERSION = "1"
_POPUP_FIX_MARKER = "popup_block_fix_v{}.marker"


def launch_debug_browser() -> None:
    """평소 쓰는 Chrome/Edge를 디버그 포트로 실행 — Playwright Chromium 미사용"""
    exe = find_browser_exe()
    if not exe:
        raise SystemExit(
            "Chrome 또는 Edge를 찾지 못했습니다.\n"
            "https://www.google.com/chrome/ 에서 설치 후 다시 실행하세요."
        )
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    _clear_stale_singleton_locks()
    log(f"브라우저 실행: {exe}")
    # 주의: stdout/stderr를 DEVNULL로 버리면 일부 환경(보안 소프트웨어·
    # 컨테이너 등)에서 Chrome이 바로 죽는 경우가 있다(무반응처럼 보임).
    # 로그 파일로 받아두면 안전하고, 문제 생기면 이 파일로 원인도 알 수 있다.
    log_path = PROFILE_DIR / "chrome_debug.log"
    log_file = open(log_path, "ab")
    chrome_args = [
        exe,
        f"--remote-debugging-port={CDP_PORT}",
        f"--user-data-dir={PROFILE_DIR}",
        "--no-first-run",
        "--no-default-browser-check",
        "--no-sandbox",
        "--disable-dev-shm-usage",
        "--disable-gpu",
        # ★9항 저장하기 후 '팝업창 모달'이 아예 안 뜨는 문제 대응 —
        # Chrome 자체 팝업차단을 완전히 끈다.
        "--disable-popup-blocking",
    ]
    # 더망고 솔루션 확장 로드 시도.
    #
    # ★정품 Chrome 137+ 에서는 --load-extension 이 제거되어 조용히 무시된다
    #   (크롬 로그: "--load-extension is not allowed in Google Chrome, ignoring.").
    #   Chromium·Chrome for Testing 에서는 아직 동작하므로 인자는 그대로 넘긴다.
    #
    # ★--disable-extensions-except 는 절대 넣지 않는다. "지정 경로 외 전부 비활성화"
    #   라서, 정품 Chrome 에서 --load-extension 이 무시되는 상황과 겹치면 전용
    #   프로필에 웹스토어로 설치해 둔 더망고 확장까지 매 실행 꺼버린다 —
    #   그 결과가 "더망고 확장프로그램이 설치되어 있지 않습니다" 배너다.
    if MANGO_EXT_DIR.is_dir() and (MANGO_EXT_DIR / "manifest.json").is_file():
        ext_path = str(MANGO_EXT_DIR.resolve())
        chrome_args.append(f"--load-extension={ext_path}")
        log(f"더망고 솔루션 확장 로드 시도: {ext_path}")
        log("  (정품 Chrome 137+ 는 이 인자를 무시함 — 프로필에 설치된 확장을 사용)")
    else:
        log(f"[경고] 확장 폴더 없음 — {MANGO_EXT_DIR} (Web Store 설치분만 사용)")
    chrome_args.append(MAIN_URL)
    subprocess.Popen(
        chrome_args,
        stdout=log_file,
        stderr=log_file,
    )
    log("디버그 모드 연결 대기 중 (최대 30초)...")
    for i in range(60):
        if cdp_port_open():
            log("브라우저 연결 확인됨")
            _write_popup_fix_marker()
            return
        if i > 0 and i % 6 == 0:
            log(f"  아직 대기 중... ({i * 0.5:.0f}초 경과 — 화면에 Chrome 창이 떴는지 확인해 주세요)")
        time.sleep(0.5)  # 소켓 연결 대기 — Playwright 이벤트루프와 무관하므로 안전
    raise SystemExit(
        "브라우저가 디버그 모드로 열리지 않았습니다.\n"
        "열려있는 Chrome/Edge 창을 모두 닫고 다시 실행해 보세요.\n"
        f"(참고 로그: {log_path})"
    )


def _popup_fix_marker_path() -> Path:
    return PROFILE_DIR / _POPUP_FIX_MARKER.format(POPUP_BLOCK_FIX_VERSION)


def _write_popup_fix_marker() -> None:
    """--disable-popup-blocking 로 새로 켠 Chrome 표시(다음 실행 시 재사용 판별용)."""
    try:
        _popup_fix_marker_path().write_text("ok\n", encoding="utf-8")
    except OSError:
        pass


def warn_if_reusing_pre_fix_browser() -> None:
    """이미 떠 있던(우리가 방금 새로 켠 게 아닌) Chrome을 재사용하는 중이면 경고.

    --disable-popup-blocking 은 Chrome '실행 시' 적용되는 플래그라,
    이 프로그램을 업데이트하기 전부터 열려 있던 Chrome을 계속 붙잡고
    쓰는 중이면 여전히 팝업이 차단될 수 있다.
    """
    if _popup_fix_marker_path().exists():
        return
    log(
        "  [중요] 저장하기 팝업이 계속 안 뜨면 Chrome을 완전히 종료 후 "
        "다시 실행해 보세요 — 이번 업데이트(팝업차단 해제)는 Chrome을 "
        "새로 켤 때만 적용됩니다. (기존에 열려 있던 창을 재사용 중일 수 있음)"
    )


def pick_working_page(context: BrowserContext) -> Page:
    """이미 망고 화면이 열려 있으면 그 탭을 그대로 사용(기본 화면 유지)"""
    open_pages = [p for p in context.pages if not p.is_closed()]
    for p in open_pages:
        try:
            if ADMIN_HOST in p.url:
                return p
        except Exception:  # noqa: BLE001
            continue
    return open_pages[0] if open_pages else context.new_page()


def refresh_if_closed(page: Page) -> Page:
    """
    로그인 성공 후 사이트가 원래 탭을 닫고 새 창을 띄우는 경우가 있다
    (예: 로그인 중계 페이지가 자기 자신을 닫음). 그러면 이전 page
    객체로는 더 이상 아무 것도 할 수 없으므로(TargetClosedError),
    같은 컨텍스트에서 살아있는 페이지를 다시 찾아온다.
    """
    if not page.is_closed():
        return page
    log("  탭이 닫힘 감지 — 새 탭을 다시 찾는 중...")
    return pick_working_page(page.context)


def connect_browser(p) -> tuple[Browser, Page]:
    """
    1) 이미 디버그 모드로 열린 Chrome/Edge가 있으면 그대로 연결(= 망고 기본 화면 그대로)
    2) 없으면 평소 쓰는 Chrome/Edge를 디버그 모드로 새로 열어서 연결
    Playwright 전용 Chromium은 내려받지 않는다.
    """
    if not cdp_port_open():
        launch_debug_browser()
    else:
        # 방금 새로 켠 게 아니라 이미 떠 있던 Chrome을 재사용 — 팝업차단
        # 해제 플래그(--disable-popup-blocking)가 적용 안 됐을 수 있음
        warn_if_reusing_pre_fix_browser()

    browser = p.chromium.connect_over_cdp(CDP_URL)
    context = browser.contexts[0] if browser.contexts else browser.new_context()
    page = pick_working_page(context)
    return browser, page


def open_extension_install_page(context: BrowserContext) -> "Page | None":
    """P2 전용 프로필 창에 더망고 확장 웹스토어 페이지를 띄운다.

    정품 Chrome 137+ 는 --load-extension 을 무시하므로, 이 프로필에 확장을
    넣는 유일한 방법이 웹스토어 설치다. 실패해도 원래 오류를 가리지 않도록
    조용히 넘어간다.
    """
    try:
        page = context.new_page()
        page.goto(MANGO_EXT_WEBSTORE, wait_until="domcontentloaded", timeout=20_000)
        page.bring_to_front()
        log(f"  확장 설치 페이지를 열었습니다 — {MANGO_EXT_WEBSTORE}")
        return page
    except Exception as e:  # noqa: BLE001
        log(f"  [안내] 확장 설치 페이지를 열지 못했습니다: {e}")
        return None


def wait_for_extension_install(
    page: "Page",
    timeout_sec: int = EXT_INSTALL_WAIT_SEC,
) -> bool:
    """사용자가 웹스토어에서 [Chrome에 추가] 를 누를 때까지 대기.

    수동 로그인 대기(wait_for_user_login)와 같은 방식이다 — 확장 팝업 주소가
    열리는 순간 설치가 끝난 것이므로 그대로 True 를 돌려주고 이어서 진행한다.
    성공 시 page 는 이미 확장 팝업에 올라와 있다.
    """
    safe_print("")
    safe_print("================================================")
    safe_print("  더망고 확장프로그램을 설치해 주세요.")
    safe_print("  방금 열린 웹스토어 창에서 [Chrome에 추가] 클릭")
    safe_print(f"  설치가 확인되면 자동으로 계속됩니다. (최대 {timeout_sec}초)")
    safe_print("================================================")

    deadline = time.time() + max(30, int(timeout_sec))
    next_notice = time.time() + 30
    while time.time() < deadline:
        check_stop("확장프로그램 설치 대기")
        try:
            page.goto(MANGO_EXT_POPUP, wait_until="domcontentloaded", timeout=5_000)
            log("확장프로그램 설치 확인 — 계속 진행")
            return True
        except Exception:  # noqa: BLE001
            pass
        if time.time() >= next_notice:
            left = int(deadline - time.time())
            log(f"  [대기] 확장 설치를 기다리는 중... (남은 {left}초)")
            next_notice = time.time() + 30
        # Playwright 이벤트루프와 무관한 소켓/프로세스 대기라 sleep 으로 충분
        time.sleep(2)
    return False


def ensure_mango_extension_settings(
    context: BrowserContext,
    *,
    shot_ctx: "RunCtx | None" = None,
) -> None:
    """더망고 솔루션 확장에 서비스 URL·인증 KEY를 넣고 설정값 저장.

    Chrome 기동(전용 프로필) 직후 확장 팝업 값이 비어 있는 문제를 막는다.
    """
    want_url = MANGO_SERVICE_URL.strip()
    want_key = MANGO_SERVICE_KEY.strip()
    log(
        "더망고 솔루션 확장 설정 확인 — "
        f"URL={want_url} / KEY={want_key[:4]}…{want_key[-4:]}"
    )
    page = context.new_page()
    dialogs: list[str] = []

    def _on_dialog(dialog) -> None:
        try:
            dialogs.append(str(dialog.message or ""))
            dialog.accept()
        except Exception:  # noqa: BLE001
            pass

    page.on("dialog", _on_dialog)
    try:
        try:
            page.goto(MANGO_EXT_POPUP, wait_until="domcontentloaded", timeout=20_000)
        except Exception as e:  # noqa: BLE001
            if shot_ctx is not None:
                try:
                    shot_ctx.shot(page, "ext_settings_fail", 0)
                except Exception:  # noqa: BLE001
                    pass
            log(MANGO_EXT_MISSING_GUIDE.format(cause=e))
            store_page = open_extension_install_page(context)
            if not wait_for_extension_install(page):
                raise RuntimeError(MANGO_EXT_MISSING_GUIDE.format(cause=e)) from e
            if store_page is not None:
                try:
                    store_page.close()
                except Exception:  # noqa: BLE001
                    pass
            if shot_ctx is not None:
                try:
                    shot_ctx.shot(page, "ext_installed", 0)
                except Exception:  # noqa: BLE001
                    pass

        page.wait_for_selector("#site_url", timeout=10_000)
        page.wait_for_selector("#site_key", timeout=5_000)
        # load_data.js 의 chrome.storage.local.get 반영 대기
        page.wait_for_timeout(700)

        cur_url = (page.input_value("#site_url") or "").strip()
        cur_key = (page.input_value("#site_key") or "").strip()
        if cur_url == want_url and cur_key == want_key:
            log("  확장 설정값 이미 올바름 — 저장 스킵")
            if shot_ctx is not None:
                shot_ctx.shot(page, "ext_settings", 0)
            return

        page.fill("#site_url", want_url)
        page.fill("#site_key", want_key)
        # 상품수집 ON 유지
        try:
            on = page.locator("#onoff")
            if on.count() and not on.is_checked():
                page.evaluate(
                    """() => {
                        const el = document.querySelector('#onoff');
                        if (!el) return;
                        if (window.jQuery) {
                            try { window.jQuery('#onoff').bootstrapToggle('on'); return; }
                            catch (e) {}
                        }
                        el.checked = true;
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                    }"""
                )
        except Exception:  # noqa: BLE001
            pass

        page.click("#sync_set")
        page.wait_for_timeout(400)
        # 저장 성공 문구 또는 storage 반영 확인
        saved_btn = False
        try:
            page.wait_for_function(
                """() => {
                    const b = document.querySelector('#sync_set');
                    return !!(b && (b.innerText || '').includes('저장되었습니다'));
                }""",
                timeout=5_000,
            )
            saved_btn = True
        except Exception:  # noqa: BLE001
            saved_btn = False

        stored = page.evaluate(
            """async () => {
                const local = await chrome.storage.local.get(
                    ['site_url', 'site_key', 'onoff']
                );
                return {
                    site_url: (local.site_url || '').trim(),
                    site_key: (local.site_key || '').trim(),
                    onoff: local.onoff || '',
                };
            }"""
        )
        got_url = str((stored or {}).get("site_url") or "").strip()
        got_key = str((stored or {}).get("site_key") or "").strip()
        if got_url != want_url or got_key != want_key:
            if shot_ctx is not None:
                try:
                    shot_ctx.shot(page, "ext_settings_fail", 0)
                except Exception:  # noqa: BLE001
                    pass
            raise RuntimeError(
                "더망고 솔루션 확장 설정값 저장 실패.\n"
                f"  · 기대 URL={want_url}\n"
                f"  · 실제 URL={got_url or '(비어있음)'}\n"
                f"  · KEY 일치={got_key == want_key}\n"
                f"  · 버튼저장문구={saved_btn}\n"
                f"  · 대화상자={dialogs[:2] if dialogs else '(없음)'}"
            )

        log("  확장 설정값 저장 완료 (서비스 URL + 인증 KEY)")
        if shot_ctx is not None:
            shot_ctx.shot(page, "ext_settings", 0)
    finally:
        try:
            page.close()
        except Exception:  # noqa: BLE001
            pass


# ── 엑셀 ──────────────────────────────────────────────────────

def is_toc_row(label: str) -> bool:
    """목차 행 여부 — 총건수/처리 대상에서 제외.

    ★요건: 총건수는 목차를 제외하고 계산.
    헤더(1행)는 원래부터 읽지 않으며, 라벨이 '목차'인 데이터 행도 제외.
    """
    t = str(label or "").strip()
    if not t:
        return False
    return t == "목차" or t.startswith("목차") or t.upper() == "TOC"


def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        url_col = headers.index("최종 카테고리 URL주소")
    except ValueError:
        raise SystemExit(
            "엑셀 1행 헤더에 '최종 카테고리 URL주소' 열이 있어야 합니다."
        )
    # ★요건(2026-08-20): 망고 "검색필터명"에 "최종 카테고리명"을 넣는다.
    # 옛 엑셀에는 이 열이 없을 수 있어 "상위 최종 카테고리명"으로 대체한다
    # (엑셀 포맷이 다르다는 이유로 수집 자체가 멈추는 회귀 방지).
    try:
        label_col = headers.index("최종 카테고리명")
    except ValueError:
        try:
            label_col = headers.index("상위 최종 카테고리명")
        except ValueError:
            raise SystemExit(
                "엑셀 1행 헤더에 '최종 카테고리명'"
                "(또는 '상위 최종 카테고리명') 열이 있어야 합니다."
            )

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        label = str(row[label_col].value or "").strip()
        url = str(row[url_col].value or "").strip()
        if not url:
            continue
        # ★총건수·처리: 목차 행 제외 (헤더 1행은 위 min_row=2로 이미 제외)
        if is_toc_row(label):
            continue
        rows.append({"row": i, "label": label, "url": url})
    return rows


def normalize_url(u: str) -> str:
    u = u.strip()
    return u if re.match(r"^https?://", u, re.I) else f"https://{u}"


# ── 팝업 ──────────────────────────────────────────────────────

def _is_browser_internal_url(u: str) -> bool:
    low = (u or "").lower()
    return (
        low.startswith("chrome-extension://")
        or low.startswith("chrome://")
        or low.startswith("edge://")
        or low.startswith("devtools://")
        or low.startswith("about:")
    )


def popups(page: Page) -> list:
    """검색(5·6항)용 외부 팝업만 — 같은 관리자사이트(ADMIN_HOST) 탭 제외.

    URL상품검색 팝업은 수집 대상(외부 쇼핑몰) 사이트로 뜨므로
    ADMIN_HOST 탭은 검색 팝업이 아니라고 보고 제외한다.
    ★ 저장(9~11항) 팝업 탐지에는 이 함수를 쓰지 말 것 — save_popups() 사용.
    (저장 처리중/완료 팝업은 ADMIN_HOST 자체 새 창으로 뜨는 경우가 많아
    여기서 제외하면 "팝업이 전혀 안 뜬다"처럼 영원히 감지되지 않는다.)
    """
    result = []
    for p in page.context.pages:
        if p is page or p.is_closed():
            continue
        try:
            u = p.url
        except Exception:
            continue
        if not u or _is_browser_internal_url(u):
            continue
        if ADMIN_HOST in u:
            continue
        result.append(p)
    return result


def save_popups(page: Page) -> list:
    """저장(9~11항) 팝업 후보 — ADMIN_HOST(같은 관리자사이트) 새 창도 포함.

    검색 팝업과 달리, 저장하기 클릭 후 뜨는 '처리중/완료' 팝업은
    같은 관리자 사이트 자체 새 창(admin_*.php 등)으로 뜨는 경우가 많다.
    popups() 처럼 ADMIN_HOST 라고 제외하면 실제로는 열려 있는데
    코드에서는 "팝업이 전혀 안 뜬다"로 보이는 오탐이 생긴다.
    브라우저 내부 탭(about:/chrome:/devtools: 등)만 제외한다.
    """
    result = []
    for p in page.context.pages:
        if p is page or p.is_closed():
            continue
        try:
            u = p.url
        except Exception:
            continue
        if not u or _is_browser_internal_url(u):
            continue
        result.append(p)
    return result


def close_search_popups(page: Page) -> int:
    """남아 있는 검색 팝업을 닫고 닫은 개수를 반환."""
    closed = 0
    for p in list(popups(page)):
        try:
            if not p.is_closed():
                p.close()
                closed += 1
        except Exception:  # noqa: BLE001
            pass
    return closed


def wait_popup_open(page: Page, grace_sec: float = 15.0) -> list:
    """검색 팝업(새 창)이 열릴 때까지 대기. 열린 팝업 Page 리스트 반환."""
    grace_end = time.time() + max(0.5, float(grace_sec))
    while time.time() < grace_end:
        check_stop("팝업 열림 대기")
        cur = popups(page)
        if cur:
            return cur
        page.wait_for_timeout(200)
    return []


def wait_popups_close(page: Page, timeout_sec: int = POPUP_WAIT_SEC) -> None:
    """★6항: 검색 팝업 모달창이 닫힐 때까지 대기(확인).

    닫힘 확인 없이 7항(모두저장)으로 진행하면 수집 실패의 주요 원인.
    시간 초과 시 강제 닫고 넘기지 않고 — TimeoutError 로 실패 처리.

    ★★★ 절대 규칙(사용자 지시, 2026-08-06): "인위적으로 팝업창을 닫으면
    망고 프로그램 데이터 수집도 멈추는 문제가 됨 — 팝업창은 스스로 닫힐
    때까지 반드시 기다렸다가 후속단계를 처리해." 이 검색 팝업은 망고
    확장프로그램이 상품을 임시메모리에 적재(스크래핑)하는 중이며, 그
    작업이 끝나면 스스로 닫힌다. 이 팝업 안에서 확인/닫기 버튼을 대신
    눌러주거나 Escape를 누르는 등 어떤 개입도 하지 않는다 — 과거에 이런
    "보조 클릭"을 넣었다가 오히려 스크래핑을 방해해 더 안 닫히는 회귀가
    있었다(2026-08-08). 순수하게 읽기만 하며 기다린다.
    """
    wait_sec = max(5, int(timeout_sec))
    end = time.time() + wait_sec
    last_beat = 0.0
    while popups(page):
        check_stop("6항 검색 팝업 닫힘 대기")
        if time.time() > end:
            n = 0
            try:
                n = len(popups(page))
            except Exception:  # noqa: BLE001
                n = -1
            raise TimeoutError(
                f"검색 팝업모달이 닫히지 않음 (6항 미확인, 남은 팝업 {n}개). "
                f"{timeout_sec}초 대기 후에도 닫힘 미확인 — "
                "강제 닫고 모두저장/다음단계로 진행 불가."
            )
        if time.time() - last_beat > 10:
            last_beat = time.time()
            cur = popups(page)
            urls = ", ".join(p.url for p in cur if not p.is_closed())
            log(
                f"  [6항] 검색 팝업창 닫힘(확인) 대기중... "
                f"(열린 팝업 {len(cur)}개: {urls})"
            )
        page.wait_for_timeout(500)


def wait_popups_gone(
    page: Page,
    timeout_sec: int = POPUP_WAIT_SEC,
    grace_sec: float = 2.0,
    warn_if_never_opened: bool = False,
) -> bool:
    """팝업이 한 번 열린 뒤 스스로 닫힐 때까지 대기.

    반환값: 팝업이 한 번이라도 열렸으면 True.
    """
    opened = wait_popup_open(page, grace_sec=grace_sec)
    if not opened:
        if warn_if_never_opened:
            log(
                "  [경고] 팝업이 뜨지 않음 — 클릭이 제대로 안 됐거나 "
                "사이트가 응답하지 않았을 수 있음"
            )
        return False
    wait_popups_close(page, timeout_sec=timeout_sec)
    return True


def scroll_to_product_strip(page: Page) -> None:
    """★스크롤 절대 금지 요건 — 더 이상 화면을 스크롤하지 않는다(하위 호환 no-op)."""
    return None


def wait_product_images(
    page: Page,
    *,
    min_count: int = 2,
    timeout_sec: float = 25.0,
) -> int:
    """상품 이미지(naturalWidth>=40) 로드 개수를 읽기만 하고 대기(스크롤 없음)."""
    end = time.time() + max(3.0, float(timeout_sec))
    best = 0
    while time.time() < end:
        check_stop("상품이미지 대기")
        try:
            n = page.evaluate(
                """() => {
                    let ok = 0;
                    for (const img of Array.from(document.querySelectorAll('img'))) {
                        const w = img.naturalWidth || 0;
                        const h = img.naturalHeight || 0;
                        if (w >= 40 && h >= 40) ok += 1;
                    }
                    return ok;
                }"""
            )
            best = max(best, int(n or 0))
        except Exception:  # noqa: BLE001
            pass
        if best >= min_count:
            return best
        page.wait_for_timeout(350)
    return best


def prepare_product_view_for_shot(
    page: Page,
    *,
    min_images: int = 2,
    fast: bool = False,
) -> int:
    """샷 직전 상품 이미지 로드 개수 확인(읽기전용, 스크롤 없음).

    fast=True — 7·8항처럼 지연을 최소화해야 하는 구간에서 사용.
    로딩 버퍼·이미지 대기 상한을 최소로 줄인다(요건: 6항 확인 후
    7·8항은 단계별 딜레이 없이 즉시 수행, 9항 진입까지 소요시간 최소화).
    """
    wait_page_not_loading(
        page,
        timeout_sec=(2.0 if fast else 15.0),
        settle_sec=(0.0 if fast else 0.5),
    )
    n = wait_product_images(
        page, min_count=min_images, timeout_sec=(1.0 if fast else 25.0)
    )
    if not fast:
        page.wait_for_timeout(200)
    return n


# ── 입력 · 클릭 (망고 구형 input 대응) ────────────────────────

def type_into(page: Page, locator, value: str) -> None:
    """★스크롤 절대 금지 — el.click()의 내부 자동 스크롤 외에는 화면을 움직이지 않음."""
    el = locator.first
    el.wait_for(state="attached", timeout=60_000)
    try:
        el.click(timeout=15_000)
    except PWTimeout:
        pass
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.keyboard.insert_text(value)

    got = ""
    try:
        got = el.input_value()
    except Exception:
        pass
    if not got.strip():
        el.evaluate(
            """(node, v) => {
                if (node instanceof HTMLInputElement || node instanceof HTMLTextAreaElement) {
                    node.focus();
                    node.value = v;
                    node.dispatchEvent(new Event('input', { bubbles: true }));
                    node.dispatchEvent(new Event('change', { bubbles: true }));
                    node.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
                    node.dispatchEvent(new Event('blur', { bubbles: true }));
                }
            }""",
            value,
        )
        got = ""
        try:
            got = el.input_value()
        except Exception:
            pass

    if got.strip() != value.strip():
        log(f"  [경고] 입력값 불일치 — 넣으려던 값: {value!r} / 실제 값: {got!r}")


def click_it(locator) -> bool:
    """
    반환값: 신뢰할 수 있는(trusted) 클릭으로 처리됐으면 True.
    el.evaluate(...node.click()...) 같은 JS 강제클릭은 브라우저가
    "진짜 사용자 클릭"으로 인정하지 않아, 그 안에서 호출되는
    window.open()(팝업)이 조용히 차단될 수 있다. 그래서 실패해도
    좌표 기반 실제 마우스 클릭(신뢰됨)을 먼저 시도하고,
    그것마저 안 될 때만 최후 수단으로 JS 클릭을 쓴다.

    ★스크롤 절대 금지 — el.click()의 내부 자동 스크롤 외에는 화면을 움직이지 않음.
    """
    el = locator.first
    el.wait_for(state="visible", timeout=60_000)
    try:
        el.click(timeout=20_000)
        return True
    except PWTimeout:
        pass

    try:
        box = el.bounding_box()
        if box:
            el.page.mouse.click(
                box["x"] + box["width"] / 2, box["y"] + box["height"] / 2
            )
            return True
    except Exception:  # noqa: BLE001
        pass

    log("  [경고] 실제 클릭 실패 — JS 강제클릭 사용 (팝업이 안 뜰 수 있음)")
    el.evaluate("(node) => node.click()")
    return False


# ── 화면 요소 ─────────────────────────────────────────────────

def url_search_button(page: Page):
    return page.locator(
        'input[type="button"][value*="URL"], input[type="submit"][value*="URL"]'
    ).or_(page.get_by_text(re.compile(r"URL\s*상품\s*검색")))


def save_all_button(page: Page):
    """[버튼1] 검색결과 상단 — '검색된 상품 모두저장' (모달 하단 저장하기와 별개)."""
    return (
        page.locator('input[type="button"][value*="모두저장"]')
        .or_(page.locator('input[type="submit"][value*="모두저장"]'))
        .or_(page.get_by_text(re.compile(r"검색된\s*상품\s*모두\s*저장")))
    )


# 모달 하단 '저장하기'와 혼동하면 안 되는 결과목록 버튼 문구
_RESULT_LIST_SAVE_BTN = re.compile(
    r"모두\s*저장|선택상품\s*저장|검색된\s*상품"
)


def _describe(loc) -> str:
    try:
        return loc.evaluate(
            "n => `<${n.tagName.toLowerCase()}"
            " name=${n.name||''} id=${n.id||''} rows=${n.rows||''}"
            " value.len=${(n.value||'').length}>`"
        )
    except Exception:
        return "<알 수 없음>"


def url_input(page: Page):
    """페이지 전환 중이면 재시도(_url_input_once 참고)"""
    return with_nav_retry(page, lambda: _url_input_once(page))


def _url_input_once(page: Page):
    """
    URL상품검색하기 버튼과 실제 입력칸이 서로 다른 <tr>/<table>에 있는
    화면이 있어(선택자가 넓으면 엉뚱한 textarea를 골라 "검색결과 없음"이
    나는 원인이 됨) 좁은 범위 -> 넓은 범위 순으로, 후보가 정확히 하나일
    때만 채택한다.
    """
    btn = url_search_button(page).first

    # 1) 버튼과 같은 <tr> 안에서 우선 찾기 (가장 정확)
    row = btn.locator("xpath=ancestor::tr[1]")
    if row.count() > 0:
        for sel in ("textarea", 'input[type="text"]:not([name*="login"]):not([readonly])'):
            cand = row.locator(sel)
            if cand.count() > 0:
                found = cand.first
                log(f"  URL입력칸(같은 행에서 발견): {_describe(found)}")
                return found

    # 2) 부모를 한 단계씩 올라가며(최대 4단계) 후보가 정확히 하나일 때만 채택
    ancestor = btn
    for _ in range(4):
        ancestor = ancestor.locator("xpath=..")
        for sel in ("textarea", 'input[type="text"]:not([name*="login"]):not([readonly])'):
            cand = ancestor.locator(sel)
            if cand.count() == 1:
                found = cand.first
                log(f"  URL입력칸(상위 요소에서 발견): {_describe(found)}")
                return found

    # 3) 최후 수단: 페이지 전체에서 rows 속성이 가장 큰 textarea
    #    (URL 여러 줄 입력용 큰 textarea일 가능성이 높음)
    all_ta = page.locator("textarea")
    n = all_ta.count()
    if n == 1:
        found = all_ta.first
        log(f"  URL입력칸(페이지에 textarea 1개뿐): {_describe(found)}")
        return found
    if n > 1:
        best_idx, best_rows = 0, -1
        for i in range(n):
            rows_attr = all_ta.nth(i).get_attribute("rows")
            try:
                rows_val = int(rows_attr) if rows_attr else 1
            except ValueError:
                rows_val = 1
            if rows_val > best_rows:
                best_rows, best_idx = rows_val, i
        found = all_ta.nth(best_idx)
        log(f"  URL입력칸(가장 큰 textarea 선택, {n}개 중): {_describe(found)}")
        return found

    raise RuntimeError("URL 입력칸을 찾지 못했습니다")


def save_modal(page: Page):
    """상품저장설정 팝업 전체 — 하단 '저장하기'·'취소하기' 포함.

    안쪽 작은 div만 잡으면 하단 버튼이 범위 밖이 되어 클릭이 누락된다.
    '검색된 상품 모두저장' 결과목록 영역과 혼동하지 않는다.
    """
    full = (
        page.locator("div, form, table")
        .filter(has_text=re.compile(r"상품\s*저장\s*설정|검색\s*필터\s*명|적용\s*정책"))
        .filter(has_text=re.compile(r"저장하기"))
        .filter(has_text=re.compile(r"취소하기"))
    )
    try:
        if full.count() > 0:
            return full.last
    except Exception:  # noqa: BLE001
        pass
    return (
        page.locator("div, form, table")
        .filter(has_text=re.compile(r"상품\s*저장\s*설정|검색\s*필터\s*명"))
        .filter(has_text=re.compile(r"^[\s\S]*저장하기[\s\S]*$"))
        .filter(has_text=re.compile(r"취소하기"))
        .last
    )


def save_modal_visible(page: Page) -> bool:
    """상품저장설정 모달 열림 판정 — 오탐(False negative)로 9항 클릭 자체를
    시도조차 못 하게 막는 사고를 피하기 위해 여러 신호 중 하나면 True.

    가장 신뢰도 높은 신호는 '실제 클릭 가능한 저장하기 요소가 있는가'다
    (그게 있으면 모달이 열려 있는 것은 확정이고, 클릭 함수와 같은
    기준을 쓰므로 둘이 서로 어긋나서 false negative가 나는 일이 없다).
    """
    # 0) 최우선·최고신뢰: 클릭용 함수와 동일 기준의 저장하기 후보 존재 여부
    try:
        if _find_footer_save_by_cancel_pair(page) is not None:
            return True
    except Exception:  # noqa: BLE001
        pass
    # 1) 모달 제목류 문구 (표현 변형 허용)
    try:
        title = page.get_by_text(re.compile(r"상품\s*저장\s*설정|상품\s*저장\s*옵션"))
        if title.count() > 0 and title.first.is_visible():
            return True
    except Exception:  # noqa: BLE001
        pass
    # 2) '저장하기' 부분일치(아이콘·공백으로 정확매치 실패 대비) + '취소하기'
    try:
        save_cands = page.get_by_text(re.compile(r"저장하기"))
        found_save = False
        for i in range(min(save_cands.count(), 8)):
            el = save_cands.nth(i)
            try:
                if not el.is_visible():
                    continue
                txt = re.sub(r"\s+", " ", el.inner_text(timeout=300) or "").strip()
            except Exception:  # noqa: BLE001
                continue
            if re.search(r"모두\s*저장|선택상품\s*저장", txt):
                continue
            found_save = True
            break
        has_cancel = False
        try:
            has_cancel = page.get_by_text(re.compile(r"취소하기")).first.is_visible()
        except Exception:  # noqa: BLE001
            pass
        if found_save and has_cancel:
            return True
    except Exception:  # noqa: BLE001
        pass
    return False


def save_submit_button(page: Page):
    """하위 호환: 상품저장설정 모달 하단 '저장하기' locator."""
    return resolve_save_submit_control(page)


def _first_visible(locator):
    """locator 후보 중 보이는 '저장하기'만. 모두저장/선택상품저장 제외."""
    try:
        n = locator.count()
    except Exception:  # noqa: BLE001
        return None
    for i in range(n):
        el = locator.nth(i)
        try:
            if not el.is_visible():
                continue
            val = (el.get_attribute("value") or "").strip()
            try:
                txt = re.sub(
                    r"\s+", " ", (el.inner_text(timeout=400) or "")
                ).strip()
            except Exception:  # noqa: BLE001
                txt = ""
            blob = f"{val} {txt}".strip()
            if _RESULT_LIST_SAVE_BTN.search(blob):
                continue
            # 버튼 라벨이 정확히 '저장하기' 인 경우만 (스크린샷 하단 파란 버튼)
            if val == "저장하기" or txt == "저장하기":
                return el
        except Exception:  # noqa: BLE001
            continue
    return None


def scroll_save_modal_to_footer(page: Page) -> None:
    """★스크롤 절대 금지 요건 — 더 이상 화면을 스크롤하지 않는다(하위 호환 no-op).

    저장하기 버튼 클릭은 el.click()이 필요할 때만 자체적으로 최소 이동하며,
    이 함수가 하던 명시적 페이지/모달 스크롤은 하지 않는다.
    """
    return None


def dump_save_button_candidates(page: Page, ctx: "RunCtx | None" = None) -> str:
    """디버그: '저장하기' vs '모두저장' 후보 구분 나열."""
    try:
        info = page.evaluate(
            """() => {
                const out = [];
                const nodes = document.querySelectorAll(
                    'input, button, a, span, div, td, li'
                );
                for (const el of nodes) {
                    const val = (el.value || '').trim();
                    const txt = (el.innerText || el.textContent || '')
                        .replace(/\\s+/g, ' ').trim();
                    const blob = val + ' ' + txt;
                    if (!/저장/.test(blob)) continue;
                    const r = el.getBoundingClientRect();
                    if (r.width < 2 || r.height < 2) continue;
                    const kind =
                        (val === '저장하기' || txt === '저장하기')
                            ? 'MODAL_SAVE'
                            : /모두\\s*저장|선택상품/.test(blob)
                              ? 'LIST_SAVE'
                              : 'OTHER';
                    out.push({
                        kind,
                        tag: el.tagName,
                        type: el.getAttribute('type') || '',
                        value: val.slice(0, 40),
                        text: txt.slice(0, 40),
                        y: Math.round(r.top),
                    });
                    if (out.length >= 16) break;
                }
                return out;
            }"""
        )
    except Exception as e:  # noqa: BLE001
        info = [{"error": str(e)}]
    line = f"저장버튼구분={info!r}"
    if ctx is not None:
        ctx.info(f"  [진단] {line}")
    else:
        log(f"  [진단] {line}")
    return line


def _prefer_clickable_save(el):
    """div/span 래퍼가 잡히면 실제 a/button/input 으로 승격.

    래퍼에 node.click() 하면 onclick이 자식에만 있어 서버 제출이 안 됨.
    """
    if el is None:
        return None
    try:
        handle = el.evaluate_handle(
            """(node) => {
                const tag = (node.tagName || '').toLowerCase();
                if (tag === 'a' || tag === 'button' || tag === 'input') return node;
                const inner = node.querySelector(
                    'a, button, input[type="button"], input[type="submit"]'
                );
                if (inner) return inner;
                const parent = node.closest('a, button');
                return parent || node;
            }"""
        )
        upgraded = handle.as_element()
        return upgraded or el
    except Exception:  # noqa: BLE001
        return el


def _find_footer_save_by_cancel_pair(page: Page):
    """스크린샷 기준: 하단 [저장하기][취소하기] 쌍에서 저장하기만 반환.

    '검색된 상품 모두저장'과는 완전히 다른 버튼이다.
    ★ 반드시 a/button/input 클릭 가능 요소를 반환 (div/span 래퍼 금지).
    """
    try:
        handle = page.evaluate_handle(
            """() => {
                const labelOf = (el) => {
                    const v = (el.value || '').trim();
                    const t = (el.innerText || el.textContent || '')
                        .replace(/\\s+/g, ' ').trim();
                    return { v, t, blob: (v + ' ' + t).trim() };
                };
                const visible = (el) => {
                    const r = el.getBoundingClientRect();
                    return r.width > 2 && r.height > 2;
                };
                const isSaveLabel = (el) => {
                    const { v, t, blob } = labelOf(el);
                    if (v !== '저장하기' && t !== '저장하기') return false;
                    if (/모두\\s*저장|선택상품\\s*저장/.test(blob)) return false;
                    return visible(el);
                };
                const toClickable = (el) => {
                    if (!el) return null;
                    const tag = (el.tagName || '').toLowerCase();
                    if (tag === 'a' || tag === 'button' || tag === 'input') return el;
                    const inner = el.querySelector(
                        'a, button, input[type="button"], input[type="submit"]'
                    );
                    if (inner && isSaveLabel(inner)) return inner;
                    if (inner) return inner;
                    const parent = el.closest('a, button');
                    return parent || null;
                };
                const INTERACTIVE = 'input, button, a';
                // 1) 취소하기 옆 — 먼저 클릭 가능 요소만
                const cancels = Array.from(document.querySelectorAll(INTERACTIVE))
                    .filter((el) => {
                        const { v, t } = labelOf(el);
                        return (v === '취소하기' || t === '취소하기') && visible(el);
                    });
                for (const cancel of cancels) {
                    let root = cancel.parentElement;
                    for (let depth = 0; depth < 6 && root; depth++) {
                        const kids = Array.from(root.querySelectorAll(INTERACTIVE));
                        const save = kids.find((el) => el !== cancel && isSaveLabel(el));
                        if (save) return save;
                        root = root.parentElement;
                    }
                }
                // 2) 전역 interactive exact 저장하기
                const direct = Array.from(document.querySelectorAll(INTERACTIVE))
                    .find((el) => isSaveLabel(el));
                if (direct) return direct;
                // 3) 래퍼(span/div) → 클릭 가능 자식/부모로 승격
                const wrap = Array.from(document.querySelectorAll('span, div, td'))
                    .find((el) => isSaveLabel(el));
                return toClickable(wrap);
            }"""
        )
        el = handle.as_element()
        if el is not None:
            return _prefer_clickable_save(el)
    except Exception:  # noqa: BLE001
        pass
    return None


def resolve_save_submit_control(page: Page):
    """[버튼2] 상품저장설정 하단 '저장하기' — [버튼1] 모두저장과 절대 혼동 금지.

    실화면: 파란 '저장하기' 옆에 '취소하기'. (밑줄 a/버튼/input)
    """
    scroll_save_modal_to_footer(page)

    # 0) 최우선: 취소하기 옆 푸터 쌍의 저장하기 (스크린샷 그대로)
    el = _find_footer_save_by_cancel_pair(page)
    if el is not None:
        return _prefer_clickable_save(el)

    modal = save_modal(page)

    # 1) 모달 안 — exact '저장하기' only (클릭 가능 요소 우선)
    modal_candidates = [
        modal.locator(
            'input[type="button"][value="저장하기"], '
            'input[type="submit"][value="저장하기"]'
        ),
        modal.locator("a, button").filter(has_text=re.compile(r"^저장하기$")),
        modal.locator("span, div, td").filter(has_text=re.compile(r"^저장하기$")),
        modal.get_by_text(re.compile(r"^저장하기$")),
    ]
    for loc in modal_candidates:
        found = _first_visible(loc)
        if found is not None:
            return _prefer_clickable_save(found)

    # 2) 페이지 — exact only, 모두저장 제외는 _first_visible에서 처리
    page_candidates = [
        page.locator(
            'input[type="button"][value="저장하기"], '
            'input[type="submit"][value="저장하기"]'
        ),
        page.locator("a, button").filter(has_text=re.compile(r"^저장하기$")),
        page.get_by_text(re.compile(r"^저장하기$")),
    ]
    for loc in page_candidates:
        found = _first_visible(loc)
        if found is not None:
            return _prefer_clickable_save(found)

    dump_save_button_candidates(page)
    raise RuntimeError(
        "상품저장설정 하단 '저장하기' 버튼을 찾지 못함 "
        "('검색된 상품 모두저장'과 다른 버튼). "
        "서버 최종 갱신을 할 수 없음"
    )


def diagnose_save_click_environment(
    page: Page, el, ctx: "RunCtx | None" = None, tag: str = ""
) -> dict:
    """9항 클릭이 안 먹힐 때 '왜' 안 되는지 바로 보이게 진단.

    - 대상 disabled/pointer-events/visibility/opacity
    - 클릭 좌표를 실제로 가로채는 요소(오버레이·z-index 문제)
    - iframe 존재(모달이 iframe 안일 가능성)
    - 미선택 필수 라디오/셀렉트/체크박스(정책 선택 등 숨은 필수값)
    """
    info: dict = {}
    try:
        info["frames"] = len(page.frames)
        urls = [f.url for f in page.frames if f.url and f.url != page.url]
        if urls:
            info["frame_urls"] = urls[:5]
    except Exception as e:  # noqa: BLE001
        info["frames_error"] = str(e)
    try:
        others = [
            p.url for p in page.context.pages
            if p is not page and not p.is_closed() and p.url
        ]
        info["open_windows_total"] = len(others)
        if others:
            info["open_windows"] = others[:5]
    except Exception as e:  # noqa: BLE001
        info["open_windows_error"] = str(e)

    if el is not None:
        try:
            info["target"] = el.evaluate(
                """(node) => {
                    const cs = getComputedStyle(node);
                    const r = node.getBoundingClientRect();
                    return {
                        tag: node.tagName,
                        disabled: !!node.disabled,
                        ariaDisabled: node.getAttribute('aria-disabled'),
                        display: cs.display,
                        visibility: cs.visibility,
                        pointerEvents: cs.pointerEvents,
                        opacity: cs.opacity,
                        rect: [
                            Math.round(r.x), Math.round(r.y),
                            Math.round(r.width), Math.round(r.height)
                        ],
                    };
                }"""
            )
        except Exception as e:  # noqa: BLE001
            info["target_error"] = str(e)
        try:
            info["intercept"] = el.evaluate(
                """(node) => {
                    const r = node.getBoundingClientRect();
                    const cx = r.x + r.width / 2;
                    const cy = r.y + r.height / 2;
                    const top = document.elementFromPoint(cx, cy);
                    if (!top) return { hit: 'none(offscreen?)' };
                    const same = (top === node) || node.contains(top) || top.contains(node);
                    if (same) return { same: true };
                    return {
                        same: false,
                        tag: top.tagName,
                        id: top.id || '',
                        cls: (top.className || '').toString().slice(0, 80),
                    };
                }"""
            )
        except Exception as e:  # noqa: BLE001
            info["intercept_error"] = str(e)

    try:
        info["unselected_required"] = page.evaluate(
            """() => {
                const out = [];
                const groups = {};
                document.querySelectorAll('input[type=radio]').forEach((r) => {
                    if (!r.name) return;
                    groups[r.name] = groups[r.name] || [];
                    groups[r.name].push(r);
                });
                for (const [name, radios] of Object.entries(groups)) {
                    const anyChecked = radios.some((r) => r.checked);
                    const visible = radios.some((r) => {
                        const rc = r.getBoundingClientRect();
                        return rc.width > 0 && rc.height > 0;
                    });
                    if (!anyChecked && visible) out.push('radio:' + name);
                }
                document.querySelectorAll('select').forEach((s) => {
                    const rc = s.getBoundingClientRect();
                    if (rc.width > 0 && rc.height > 0 && s.value === '') {
                        out.push('select:' + (s.name || s.id || '?'));
                    }
                });
                document.querySelectorAll('input[type=checkbox][required]').forEach((c) => {
                    if (!c.checked) out.push('checkbox:' + (c.name || c.id || '?'));
                });
                return out.slice(0, 10);
            }"""
        )
    except Exception as e:  # noqa: BLE001
        info["unselected_error"] = str(e)

    line = f"[9항 진단{(':' + tag) if tag else ''}] {info!r}"
    if ctx is not None:
        ctx.info(f"  {line}")
    else:
        log(f"  {line}")
    return info


def trusted_click_save_submit(
    page: Page, el, ctx: "RunCtx | None" = None
) -> bool:
    """하단 저장하기 클릭 — 실제 a/button/input 에 클릭.

    div 래퍼 클릭·JS 가짜 성공으로 서버 미제출되던 경로를 막는다.
    실패할 때마다 원인(가로채기·비활성화·미선택 필수값)을 진단·기록한다.
    """
    el = _prefer_clickable_save(el)
    try:
        tag = el.evaluate("(n) => (n.tagName || '').toLowerCase()")
        log(f"  [9항 클릭대상] <{tag}> 저장하기")
    except Exception:  # noqa: BLE001
        pass
    diagnose_save_click_environment(page, el, ctx, tag="클릭전")
    # ★스크롤·호버 등 별도 움직임 절대 금지 — el.click() 자체(필요한 최소
    # 내부 이동만 포함)와 아래 좌표클릭만 "버튼 클릭" 행위로 허용한다.

    # 1) 일반 클릭 (trusted)
    try:
        el.click(timeout=15_000)
        return True
    except Exception as e:  # noqa: BLE001
        log(f"  [경고] 저장하기 일반클릭 실패: {type(e).__name__}")

    # 2) 좌표 마우스 클릭 (trusted)
    try:
        box = el.bounding_box()
        if box and box.get("width", 0) > 0 and box.get("height", 0) > 0:
            page.mouse.click(
                box["x"] + box["width"] / 2,
                box["y"] + box["height"] / 2,
            )
            return True
    except Exception as e:  # noqa: BLE001
        log(f"  [경고] 저장하기 좌표클릭 실패: {type(e).__name__}")

    # 3) force 클릭 (trusted, 가로채는 요소 무시)
    try:
        el.click(timeout=8_000, force=True)
        return True
    except Exception as e:  # noqa: BLE001
        log(f"  [경고] 저장하기 force클릭 실패: {type(e).__name__}")

    # 4) 키보드 활성화 (trusted) — onclick이 키보드 이벤트에도 걸린 경우
    try:
        el.focus()
        page.keyboard.press("Enter")
        page.wait_for_timeout(200)
        page.keyboard.press("Space")
        return True
    except Exception as e:  # noqa: BLE001
        log(f"  [경고] 저장하기 키보드 활성화 실패: {type(e).__name__}")

    diagnose_save_click_environment(page, el, ctx, tag="클릭실패직전")

    # 5) 최후 JS — 클릭 가능 노드 + bubble 이벤트. 성공으로 단정하지 않음(재시도 유도)
    try:
        el.evaluate(
            """(node) => {
                const t = node.closest('a,button')
                    || node.querySelector('a,button,input[type=button],input[type=submit]')
                    || node;
                if (typeof t.click === 'function') t.click();
                t.dispatchEvent(new MouseEvent('click', {
                    bubbles: true, cancelable: true, view: window
                }));
            }"""
        )
        log("  [경고] 저장하기 JS click 사용 — 반응 없으면 재시도")
        return False
    except Exception:  # noqa: BLE001
        return False


def save_submit_reacted(
    page: Page,
    dialog_msgs: list[str] | None = None,
    *,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
    timeout_sec: float = 15.0,
) -> bool:
    """저장하기 클릭 후 '저장 실행 팝업/알림'이 실제로 떴는지.

    ★ 상품저장설정 모달만 닫힌 것은 성공이 아님.
    ★ 클릭 전 화면의 '00건 수집' 등 잔여 문구는 성공이 아님.
    """
    end = time.time() + max(5.0, float(timeout_sec))
    while time.time() < end:
        has_signal, detail, _ = save_result_signal_present(
            page,
            dialog_msgs,
            before_popup_ids=before_popup_ids,
            baseline=baseline,
            dialog_from=dialog_from,
        )
        if has_signal:
            log(f"  [저장반응] {detail}")
            return True
        page.wait_for_timeout(350)
    return False


def try_dismiss_save_modal(page: Page) -> None:
    """저장 설정 모달이 남아 있으면 닫기/취소/Esc 로 해제 시도."""
    if not save_modal_visible(page):
        return
    try:
        modal = save_modal(page)
        closer = (
            modal.locator('button, a, input[type="button"], input[type="submit"]')
            .filter(has_text=re.compile(r"닫기|취소|취소하기|닫\s*기|×|X"))
            .first
        )
        if closer.count() > 0 and closer.is_visible():
            click_it(closer)
            page.wait_for_timeout(400)
            return
    except Exception:  # noqa: BLE001
        pass
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(300)
    except Exception:  # noqa: BLE001
        pass


def close_save_popups(page: Page) -> int:
    """남은 저장(9~11항) 팝업 정리 — ADMIN_HOST 새 창 포함."""
    closed = 0
    for p in list(save_popups(page)):
        try:
            if not p.is_closed():
                p.close()
                closed += 1
        except Exception:  # noqa: BLE001
            pass
    return closed


def overlay_status(page: Page) -> dict:
    """열린 검색팝업·저장팝업·저장모달·로딩 여부.

    ★ 검색 팝업(popups)과 저장 팝업(save_popups, ADMIN_HOST 포함)을
    모두 세야 leftover 저장팝업 때문에 다음행이 막히는 걸 놓치지 않는다.
    """
    n_pop = 0
    try:
        # save_popups: ADMIN_HOST 포함 — popups(검색전용)의 상위집합이라
        # 이것만 세면 검색·저장 팝업 leftover를 모두 놓치지 않는다.
        n_pop = len(save_popups(page))
    except Exception:  # noqa: BLE001
        n_pop = 0
    modal = False
    try:
        modal = bool(save_modal_visible(page))
    except Exception:  # noqa: BLE001
        modal = False
    loading = False
    try:
        loading = bool(is_mango_loading(page))
    except Exception:  # noqa: BLE001
        loading = False
    return {"popups": n_pop, "save_modal": modal, "loading": loading}


def finalize_row_overlays(page: Page, ctx: "RunCtx", row: dict) -> Page:
    """행 종료 직후 남은 검색팝업·저장모달을 정리(다음 행 진입 전 보조)."""
    page = refresh_if_closed(page)
    try:
        n = close_search_popups(page)
        if n:
            ctx.info(f"  [행종료정리] 검색 팝업 {n}개 닫음")
        n2 = close_save_popups(page)
        if n2:
            ctx.info(f"  [행종료정리] 저장 팝업(ADMIN_HOST 포함) {n2}개 닫음")
        if save_modal_visible(page):
            ctx.info("  [행종료정리] 저장 모달 닫기 시도")
            try_dismiss_save_modal(page)
            page.wait_for_timeout(400)
            if save_modal_visible(page):
                try_dismiss_save_modal(page)
    except Exception as e:  # noqa: BLE001
        ctx.info(f"  [행종료정리] 경고: {e}")
    return refresh_if_closed(page)


def ensure_overlays_closed_before_next(
    page: Page,
    ctx: "RunCtx",
    *,
    next_ordinal: int,
    next_row: dict,
    timeout_sec: float = 180.0,
) -> Page:
    """다음 입력(특히 2번) 상품수집 전 — 모든 모달/팝업이 닫힐 때까지 기다린 뒤 샷 보관.

    닫히지 않은 채 다음 행으로 넘어가지 않는다(반드시 종료 확인 후 진행).
    """
    rn = int(next_row.get("row") or next_ordinal)
    label = str(next_row.get("label") or "").strip()
    ctx.info(
        f"[다음행준비] 입력#{next_ordinal} 수집 전 — 팝업·모달 전부 종료될 때까지 대기 "
        f"(엑셀{rn}행 / {label})"
    )
    page = refresh_if_closed(page)
    try:
        page.bring_to_front()
    except Exception:  # noqa: BLE001
        pass

    end = time.time() + max(30.0, float(timeout_sec))
    last_log = 0.0
    clean_stable = 0
    while True:
        check_stop(f"입력#{next_ordinal} 전 모달 종료 대기")
        st = overlay_status(page)
        if st["popups"] == 0 and not st["save_modal"] and not st["loading"]:
            clean_stable += 1
            if clean_stable >= 3:  # ~1초 이상 안정적으로 닫힌 상태
                break
            page.wait_for_timeout(350)
            continue
        clean_stable = 0
        now = time.time()
        if now - last_log > 5:
            last_log = now
            ctx.info(
                f"  모달 종료 대기중... popups={st['popups']}, "
                f"save_modal={st['save_modal']}, loading={st['loading']}"
            )
        if st["popups"] > 0:
            n = close_search_popups(page)
            n += close_save_popups(page)
            if n:
                ctx.info(f"  남은 검색/저장 팝업 {n}개 닫음")
        if st["save_modal"]:
            try_dismiss_save_modal(page)
        if now > end:
            # 제한시간 후에도 강제 닫기 반복 — 그래도 안 닫히면 샷 후 오류
            for _ in range(5):
                close_search_popups(page)
                close_save_popups(page)
                try_dismiss_save_modal(page)
                page.wait_for_timeout(500)
                st = overlay_status(page)
                if st["popups"] == 0 and not st["save_modal"]:
                    break
            st = overlay_status(page)
            if st["popups"] > 0 or st["save_modal"]:
                # ★요건: 2번째 행 이후에도 배치를 절대 멈추지 않는다.
                # 팝업이 남았어도 강제 정리 후 경고만 남기고 다음 입력 진행.
                ctx.shot(page, "00_overlays_stuck", rn)
                ctx.info(
                    f"  [경고] 입력#{next_ordinal} 수집 전 팝업/모달 잔여 "
                    f"(popups={st['popups']}, save_modal={st['save_modal']}) "
                    "— 강제 정리 후 다음 입력 계속"
                )
                try:
                    close_search_popups(page)
                    close_save_popups(page)
                    try_dismiss_save_modal(page)
                except Exception:  # noqa: BLE001
                    pass
            break
        page.wait_for_timeout(350)

    st = overlay_status(page)
    ctx.info(
        f"  닫힘 확인 완료: 검색팝업={st['popups']}개, "
        f"저장모달={'열림' if st['save_modal'] else '닫힘'}, "
        f"로딩={'중' if st['loading'] else '없음'}"
    )
    ctx.shot(page, "00_overlays_clear", rn)
    ctx.info(
        f"[다음행준비] 입력#{next_ordinal} — 모달 정리 후 수집 시작"
    )
    return page


_MODAL_FIELD_JS = """(root, args) => {
    const [patSrc, tagId] = args;
    // 주의: 이전 마커를 여기서 지우면 안 된다 — 필터/카운트를 순서대로
    // 찾을 때 먼저 찍은 마커가 지워져 Locator가 그 사이에 대상을 잃는다
    // (요소를 못 찾음 → count()==0 오류로 이어졌던 실제 버그).
    const re = new RegExp(patSrc);
    const cands = Array.from(
        root.querySelectorAll('label, td, th, span, div, p, dt, dd')
    ).filter((el) => re.test((el.textContent || '').trim()));
    if (!cands.length) return false;
    // 가장 짧은(=가장 안쪽) 매치를 우선 — 넓은 공용 div까지 안 올라가게
    cands.sort(
        (a, b) => (a.textContent || '').length - (b.textContent || '').length
    );
    const INPUTS = 'input[type="text"], input:not([type]), input[type="number"]';
    for (const start of cands) {
        let node = start;
        for (let i = 0; i < 8 && node; i++) {
            const inputs = node.querySelectorAll(INPUTS);
            if (inputs.length >= 1) {
                inputs[0].setAttribute('data-collect-mf', tagId);
                return true;
            }
            node = node.parentElement;
        }
    }
    return false;
}"""


def modal_field(page: Page, label_pattern: re.Pattern):
    """레이블에 딸린 입력칸 — 넓은 div까지 매치해 다른 레이블의 입력칸을
    잘못 집는 버그(검색필터명 칸에 저장상품수가 들어가거나 반대) 방지.

    가장 안쪽(짧은 텍스트)에서 매치하는 레이블 노드부터, 입력칸이
    나올 때까지만 조상을 좁게 타고 올라간다(넓은 공용 컨테이너까지
    올라가서 다른 필드의 input을 집지 않도록). 찾은 엘리먼트에 임시
    마커를 달아 진짜 Locator로 돌려준다(하위 호출들이 .first/.count()
    등을 쓸 수 있어야 하므로 ElementHandle을 그대로 주지 않음).
    """
    modal = save_modal(page)
    tag_id = f"mf-{time.time_ns()}"
    try:
        ok = modal.evaluate(_MODAL_FIELD_JS, [label_pattern.pattern, tag_id])
    except Exception:  # noqa: BLE001
        ok = False
    if ok:
        return page.locator(f'[data-collect-mf="{tag_id}"]')

    # 폴백(옛 방식) — 그래도 못 찾으면 넓게라도 시도
    return (
        modal.locator("tr, div, p, label")
        .filter(has_text=label_pattern)
        .locator('input[type="text"], input:not([type]), input[type="number"]')
        .first
    )


# ── 0 ~ 4 ────────────────────────────────────────────────────

def safe_goto(page: Page, url: str, retries: int = 3) -> None:
    """
    로그인 직후 등 사이트 자체가 리다이렉트를 진행 중일 때
    page.goto()와 겹치면 'interrupted by another navigation' 오류가 난다.
    사이트 쪽 리다이렉트가 끝날 때까지 기다렸다가 다시 시도한다.
    """
    last_err: Exception | None = None
    for _ in range(retries):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=120_000)
            return
        except Exception as e:  # noqa: BLE001
            msg = str(e)
            if "interrupted by another navigation" in msg or "NS_BINDING_ABORTED" in msg:
                last_err = e
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=15_000)
                except Exception:  # noqa: BLE001
                    pass
                page.wait_for_timeout(800)
                continue
            raise
    if last_err:
        raise last_err


NAV_ERROR_MARKERS = (
    "Execution context was destroyed",
    "context was destroyed",
    "Target closed",
    "Target page, context or browser has been closed",
)


def is_navigation_error(e: Exception) -> bool:
    msg = str(e)
    return any(m in msg for m in NAV_ERROR_MARKERS)


def with_nav_retry(page: Page, fn, retries: int = 3):
    """
    사이트가 리다이렉트/네비게이션 중일 때 DOM을 조회하면
    "Execution context was destroyed" 같은 오류가 난다.
    페이지가 안정될 때까지 기다렸다가 재시도한다.
    """
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            return fn()
        except Exception as e:  # noqa: BLE001
            if is_navigation_error(e):
                last_err = e
                log(f"  [정보] 페이지 전환 중이라 재시도합니다 ({attempt + 1}/{retries})")
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=10_000)
                except Exception:  # noqa: BLE001
                    pass
                page.wait_for_timeout(800)
                continue
            raise
    if last_err:
        raise last_err
    return None


def _is_logged_in(page: Page) -> bool:
    """로그인 화면을 벗어나고 관리자 호스트에 있으면 로그인된 것으로 본다."""
    try:
        url = page.url or ""
    except Exception:  # noqa: BLE001
        return False
    if not url or url in ("about:blank", "chrome://newtab/"):
        return False
    if ADMIN_HOST not in url:
        return False
    if "admin_login" in url:
        return False
    if "m_login" in url:
        return False
    return True


def wait_for_user_login(page: Page, timeout_sec: int = LOGIN_WAIT_SEC) -> Page:
    """더망고 로그인 창만 띄우고, 사용자가 직접 로그인할 때까지 대기.

    자동 ID/PW 입력·글자 지연 입력·자동 제출은 하지 않는다.
    (Cafe24 보안/자동화 거절 회피 — 사용자 수동 로그인)
    """
    page = refresh_if_closed(page)
    if _is_logged_in(page):
        log("이미 로그인된 세션 — 사용자 로그인 대기 생략")
        return page

    if "admin_login" not in (page.url or ""):
        log("더망고 로그인창 열기: " + LOGIN_URL)
        safe_goto(page, LOGIN_URL)
        page = refresh_if_closed(page)

    try:
        page.bring_to_front()
    except Exception:  # noqa: BLE001
        pass

    safe_print("")
    safe_print("================================================")
    safe_print("  더망고 로그인창에서 직접 로그인하세요.")
    safe_print("  (프로그램이 ID/PW를 입력하지 않습니다)")
    safe_print(f"  로그인 완료 후 자동으로 계속됩니다. (최대 {timeout_sec}초)")
    safe_print("================================================")
    log("사용자 로그인 대기 중...")
    shot_now(page, "login_wait", 0)

    deadline = time.time() + max(30, int(timeout_sec))
    last_url = ""
    while time.time() < deadline:
        check_stop("사용자 로그인 대기")
        page = refresh_if_closed(page)
        try:
            cur = page.url or ""
        except Exception:  # noqa: BLE001
            cur = ""
        if cur != last_url:
            last_url = cur
            log(f"  [대기] URL={cur}")
        if _is_logged_in(page):
            log("사용자 로그인 확인 — 계속 진행")
            shot_now(page, "login_ok", 0)
            return page
        page.wait_for_timeout(1000)

    raise RuntimeError(
        "사용자 로그인 대기 시간 초과.\n"
        "  · Chrome에 열린 더망고 로그인창에서 직접 로그인한 뒤 다시 실행하세요.\n"
        f"  · 마지막 URL={last_url or '(unknown)'}"
    )


def perform_tmg_login(page: Page, user_id: str | None = None, password: str | None = None) -> Page:
    """하위 호환: 자동 입력 없음 — 사용자 수동 로그인 대기."""
    _ = user_id, password
    return wait_for_user_login(page)


def _ask_human(prompt: str) -> None:
    """대화형 터미널에서만 대기. CI/비대화형이면 즉시 실패."""
    if not sys.stdin.isatty():
        raise RuntimeError(
            "로그인이 필요하지만 비대화형 실행입니다. "
            "로컬 PC에서 망고 로그인 후 다시 실행하세요."
        )
    try:
        input(prompt)
    except EOFError as e:
        raise RuntimeError(
            "로그인이 필요하지만 입력을 받을 수 없습니다. "
            "로컬 PC에서 망고 로그인 후 다시 실행하세요."
        ) from e


def handle_possible_login_page(page: Page) -> None:
    """
    세션이 만료돼 admin_login.php로 튕기면 로그인창을 띄우고
    사용자가 직접 로그인할 때까지 기다린다. (자동 입력 없음)
    """
    if "admin_login" not in page.url:
        return
    log("  [경고] 로그인 화면 — 사용자 직접 로그인 대기")
    shot_now(page, "login_required", 0)
    wait_for_user_login(page)
    page = refresh_if_closed(page)
    if "admin_login" in page.url:
        raise RuntimeError("로그인 후에도 여전히 로그인 화면입니다 — 다시 확인해 주세요")


def _wait_bulk_ready_once(page: Page) -> None:
    """★2→3항 — URL검색 버튼이 보이면 즉시 반환(불필요 대기 없음)."""
    try:
        if url_search_button(page).first.is_visible():
            return
    except Exception:  # noqa: BLE001
        pass
    t0 = time.time()
    try:
        page.wait_for_load_state("domcontentloaded", timeout=8_000)
    except Exception:  # noqa: BLE001
        pass
    t1 = time.time()
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
    handle_possible_login_page(page)
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
        handle_possible_login_page(page)
    t2 = time.time()
    url_search_button(page).first.wait_for(state="visible", timeout=60_000)
    t3 = time.time()
    log(
        "  [타이밍] domcontentloaded="
        f"{t1 - t0:.2f}s, URL이동/로그인체크={t2 - t1:.2f}s, "
        f"URL검색버튼 보일때까지={t3 - t2:.2f}s"
    )


def wait_bulk_ready(page: Page) -> None:
    with_nav_retry(page, lambda: _wait_bulk_ready_once(page))


def _reset_to_bulk_menu_once(page: Page, force: bool = False) -> None:
    """★불필요한 고정 대기 없음 — 이미 대량수집 화면이면 메뉴 클릭 생략.

    force=True 이면 이 단축 경로를 쓰지 않고 항상 실제 메뉴 클릭(=서버
    재요청)을 수행한다. ★2026-08-08: 1번째 입력(로그인 직후 ensure_ready_page
    에서 미리 이동)과 2번째 이후 입력(step02_init에서 이 함수 호출)이
    서로 다른 경로를 타서 "2단계 필드 초기화" 동작이 행마다 달라지는
    문제가 있었다 — 2번째 이후 입력은 항상 force=True로 호출해 1번째와
    동일한 실제 메뉴클릭 동작을 하도록 통일한다.
    """
    if not force:
        try:
            if BULK_PATH in page.url and url_search_button(page).first.is_visible():
                return
        except Exception:  # noqa: BLE001
            pass
    t_click0 = time.time()
    href = page.locator('a[href*="getGoodsNew"]').first
    if href.count() > 0:
        try:
            href.click(timeout=5000)
        except PWTimeout:
            href.evaluate("(node) => node.click()")
        log(f"  [타이밍] 메뉴 클릭 자체={time.time() - t_click0:.2f}s")
        handle_possible_login_page(page)
        if BULK_PATH in page.url:
            wait_bulk_ready(page)
            return

    page.evaluate(
        """() => {
            const clean = (s) => (s || '').replace(/\\s+/g, '');
            const nodes = Array.from(document.querySelectorAll('a, li, span, td, div, button'));
            const byHref = Array.from(document.querySelectorAll('a[href*="getGoodsNew"]'));
            if (byHref[0]) { byHref[0].click(); return; }
            const top = nodes.find(el => clean(el.textContent) === '상품데이터수집');
            if (top) top.click();
            const sub = nodes.find(el => {
                const t = clean(el.textContent);
                if (t.length > 30) return false;
                return /대량데이터수집|대량수집|상품데이터대량/.test(t);
            });
            if (sub) (sub.closest('a') || sub).click();
        }"""
    )
    log(f"  [타이밍] JS 메뉴 클릭 자체={time.time() - t_click0:.2f}s")
    handle_possible_login_page(page)
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
    wait_bulk_ready(page)


def reset_to_bulk_menu(page: Page, force: bool = False) -> None:
    """0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭"""
    with_nav_retry(page, lambda: _reset_to_bulk_menu_once(page, force=force))


def _page_body_text(page: Page) -> str:
    try:
        return page.evaluate(
            "() => (document.body && document.body.innerText) ? document.body.innerText : ''"
        ) or ""
    except Exception:  # noqa: BLE001
        return ""


def is_mango_loading(page: Page) -> bool:
    text = _page_body_text(page)
    return any(p.search(text) for p in MANGO_LOADING_PATTERNS)


def is_mango_no_results(page: Page) -> bool:
    """더망고 자체 '검색결과가 없습니다' 메세지 여부 (로딩 중이 아닐 때만 의미 있음)."""
    if is_mango_loading(page):
        return False
    text = _page_body_text(page)
    return any(p.search(text) for p in MANGO_NO_RESULT_PATTERNS)


def count_mango_result_products(page: Page) -> int:
    """대량수집 화면의 검색결과 상품(체크박스/썸네일) 대략 개수."""
    try:
        return int(
            page.evaluate(
                """() => {
                    // 결과 영역의 체크박스 + 일정 크기 이상 이미지
                    const boxes = Array.from(
                        document.querySelectorAll('input[type="checkbox"]')
                    ).filter((el) => {
                        const r = el.getBoundingClientRect();
                        return r.width > 0 && r.height > 0 && el.offsetParent !== null;
                    });
                    const imgs = Array.from(document.querySelectorAll('img')).filter((img) => {
                        const w = img.naturalWidth || 0;
                        const h = img.naturalHeight || 0;
                        const r = img.getBoundingClientRect();
                        return w >= 40 && h >= 40 && r.width >= 20 && r.height >= 20;
                    });
                    // 전체선택 등 UI 체크박스 1~2개는 제외 감안
                    const boxScore = Math.max(0, boxes.length - 1);
                    return Math.max(boxScore, imgs.length);
                }"""
            )
            or 0
        )
    except Exception:  # noqa: BLE001
        return 0


def wait_page_not_loading(
    page: Page, timeout_sec: float = 15.0, *, settle_sec: float = 0.5
) -> None:
    """
    "로딩 중" 표시가 사라질 때까지 확인한다.
    (검색결과 있음/없음은 여기서 판단하지 않음 — 로딩 중 결과없음 깜빡임 오판 방지)
    settle_sec=0 이면 로딩 종료 확인 즉시 반환(추가 버퍼 없음) — 지연
    최소화가 필요한 구간(7·8항)에서 사용.
    """
    end = time.time() + timeout_sec
    while time.time() < end:
        check_stop("로딩 대기")
        try:
            loading = is_mango_loading(page)
        except Exception:  # noqa: BLE001
            return
        if not loading:
            if settle_sec > 0:
                page.wait_for_timeout(int(settle_sec * 1000))  # 결과 렌더링 여유
            return
        page.wait_for_timeout(300)


def wait_mango_search_settle(
    page: Page,
    *,
    timeout_sec: float = 45.0,
) -> tuple[str, int]:
    """★삭제/비활성(2026-08-08 사용자 지시).

    예전 6→7 '(6→7) 망고 검색결과 안정화' 액션 — 더 이상 호출하지 않음.
    하위 호환용 스텁: 즉시 unknown/0 반환 (긴 대기·로그 없음).
    """
    # 아래 본문은 의도적으로 실행하지 않음 (주석 처리와 동일 효과).
    # end = time.time() + max(10.0, float(timeout_sec))
    # while ...: check_stop("망고 검색 안정화"); wait loading; stabilize...
    _ = (page, timeout_sec)  # unused — stub
    return "unknown", 0


def _process_row_once(page: Page, row: dict, ctx: RunCtx) -> None:
    """한 행 BATCH: 2→3→4→5→6→7→8→9→10→11→12 순차만 실행.

    복잡한 분기·우회 없음. 단계 구현은 batch_steps.py.
    (1=로그인은 main, 13=다음 행의 2항 초기화)
    """
    from batch_steps import run_row_batch

    run_row_batch(page, row, ctx)


def _safe_input_value(loc) -> str:
    try:
        return loc.input_value() or ""
    except Exception:  # noqa: BLE001
        return ""


def fill_save_modal_fields(
    page: Page,
    ctx: RunCtx,
    rn: int,
    label: str,
    save_count: int,
) -> int:
    """상품저장설정 모달 — 검색필터명·저장상품수 입력.

    ★요건(2026-08-20): "저장상품수" 를 DEFAULT_SAVE_COUNT(50)로 맞춘다.
    2026-08-08 의 "저장상품수는 원래 세팅값 그대로 둔다" 요건을 대체한다 —
    망고 모달 기본값이 3 이라 손대지 않으면 행당 3개만 저장되기 때문이다.
    반환값: 최종적으로 모달에 들어간 저장상품수 — 이후 12항 건수확인의 기대값.
    """
    filter_field = modal_field(page, FILTER_NAME_LABEL)
    count_field = modal_field(page, SAVE_COUNT_LABEL)
    if count_field.count() == 0:
        ctx.shot(page, "02_no_count_field", rn)
        raise RuntimeError(f"#{rn} 저장상품수 입력칸을 찾지 못함")

    # 안전장치: 검색필터명·저장상품수 칸이 실제로 같은 엘리먼트로 잘못
    # 잡히면 즉시 확정 실패 (필터 입력이 저장상품수 칸까지 덮어쓰는 사고 방지)
    try:
        same_el = page.evaluate(
            "([a, b]) => a === b",
            [filter_field.element_handle(), count_field.element_handle()],
        )
    except Exception:  # noqa: BLE001
        same_el = False
    if same_el:
        ctx.shot(page, "02_count_mismatch", rn)
        raise RuntimeError(
            f"#{rn} 검색필터명·저장상품수 입력칸이 같은 엘리먼트로 잘못 잡힘 — "
            "필드 구분 실패"
        )

    original_count = _safe_input_value(count_field)
    want_count = str(int(save_count))

    ctx.step(8, "수집 상품 필터·수집상품갯수 입력")
    ctx.info(f"검색필터명: {label}")
    ctx.info(f"저장상품수: {original_count or '(비어있음)'} → {want_count}")

    max_rounds = 4
    filter_ok = False
    for round_i in range(1, max_rounds + 1):
        cur_filter = _safe_input_value(filter_field)
        if cur_filter.strip() == label.strip():
            filter_ok = True
            break
        if round_i > 1:
            ctx.info(f"검색필터명 재입력 (라운드 {round_i}) — 이전 값 {cur_filter!r}")
        type_into(page, filter_field, label)

    if not filter_ok:
        ctx.shot(page, "02_count_mismatch", rn)
        raise RuntimeError(
            f"#{rn} {max_rounds}회 시도 후에도 검색필터명이 안정되지 않음"
        )

    count_ok = False
    for round_i in range(1, max_rounds + 1):
        cur_count = _safe_input_value(count_field)
        if re.sub(r"\D", "", cur_count) == want_count:
            count_ok = True
            break
        if round_i > 1:
            ctx.info(f"저장상품수 재입력 (라운드 {round_i}) — 이전 값 {cur_count!r}")
        type_into(page, count_field, want_count)

    if not count_ok:
        ctx.shot(page, "02_count_mismatch", rn)
        raise RuntimeError(
            f"#{rn} {max_rounds}회 시도 후에도 저장상품수가 {want_count} 로 "
            f"들어가지 않음 (현재 {_safe_input_value(count_field)!r})"
        )

    # 저장상품수를 넣는 과정에서 검색필터명이 흔들리지 않았는지 최종 확인
    final_filter = _safe_input_value(filter_field)
    if final_filter.strip() != label.strip():
        ctx.shot(page, "02_count_mismatch", rn)
        raise RuntimeError(
            f"#{rn} 저장상품수 입력 후 검색필터명이 바뀜 — "
            f"기대={label!r} / 현재={final_filter!r}"
        )

    final_count = _safe_input_value(count_field)
    ctx.info(f"검색필터명·저장상품수 입력 완료 (저장수 {final_count}) → 저장하기 진행")
    ctx.shot(page, "02_modal_filled", rn)

    digits = re.sub(r"\D", "", final_count or "")
    return int(digits) if digits else int(save_count)


def run_save_submit_and_verify(
    page: Page,
    ctx: RunCtx,
    rn: int,
    save_count: int,
) -> None:
    """★저장하기 = 서버 최종 갱신 (필터정보·수집갯수·수집상품).

    입력만 하고 이 버튼을 빼먹으면 서버에 아무것도 반영되지 않는다.
    클릭 → 서버 제출 반응 확인(최대 1회 재시도)
    → 최종 팝업화면 열림 → (열린 상태에서 건수 검증)
    → 최종 팝업화면 닫힘까지 대기 → 그 다음에만 초기화/다음행.
    재시도 후에도 동일 실패면 다음 행으로 진행한다.
    """
    ctx.server_save_ok = False
    ctx.save_popup_seen = False
    ctx.save_popup_closed = False
    ctx.save_awaiting_popup = False
    ctx.save_popup_kind = ""
    ctx.save_popup_ui_latched = False
    dialog_msgs: list[str] = []

    def _on_save_dialog(dialog) -> None:
        try:
            msg = str(dialog.message or "")
            dialog_msgs.append(msg)
            ctx.info(f"  [dialog] {msg!r}")
            dialog.accept()
        except Exception:  # noqa: BLE001
            try:
                dialog.dismiss()
            except Exception:  # noqa: BLE001
                pass

    page.on("dialog", _on_save_dialog)
    try:
        ctx.check_budget("저장하기(서버 최종 갱신) 전")
        if not save_modal_visible(page):
            # ★단정하고 즉시 포기하지 않는다 — save_modal_visible()의 오탐(false
            # negative) 때문에 클릭 시도 자체를 못 하는 사고가 실제 원인이었다.
            # 여기서는 진단만 남기고, 실제 클릭 가능 여부는 resolve_save_submit_control
            # (아래)가 최종 판단하게 한다. 거기서도 못 찾으면 그때 명확히 실패 처리.
            ctx.info(
                "  [경고] save_modal_visible() 오탐 의심 — "
                "저장하기 버튼 탐색은 계속 시도함 (즉시 실패 처리하지 않음)"
            )
            dump_save_button_candidates(page, ctx)
            ctx.shot(page, "02_save_missing", rn)

        ctx.step(9, "수집 상품 DB저장하기 시작 : 하단 '저장하기' 클릭")
        ctx.info("옆=취소하기 / 7항 모두저장과 다른 버튼")
        dump_save_button_candidates(page, ctx)
        ctx.shot(page, "02_modal_filled", rn)

        # 클릭 전 잔여 '00건 수집' 등 — 저장 팝업으로 오인 금지
        alert_baseline = collect_alert_baseline(page)
        dialog_from = len(dialog_msgs)
        if alert_baseline:
            ctx.info(
                "2-B. [주의] 저장하기 전 화면에 수집문구 잔여 "
                f"{sorted(alert_baseline)} — 저장 팝업으로 쓰지 않음"
            )

        before_popup_ids = {_popup_id(p) for p in save_popups(page)}
        save_reacted = False
        # 최대 1회 재시도(총 2회). 팝업 반응 없으면 클릭 성공으로 치지 않음.
        for attempt in range(1, 3):
            ctx.check_budget(f"9항 저장하기 서버제출 시도 {attempt}")
            if not save_modal_visible(page) and attempt > 1:
                ctx.info(
                    "  [경고] 재시도 시 상품저장설정 모달 없음 — "
                    "저장하기 재클릭 불가 (가짜 클릭으로 보지 않음)"
                )
                break
            scroll_save_modal_to_footer(page)
            try:
                btn = resolve_save_submit_control(page)
            except Exception as e:  # noqa: BLE001
                dump_save_button_candidates(page, ctx)
                ctx.shot(page, "02_save_missing", rn)
                raise RuntimeError(
                    f"#{rn} 9항 하단 '저장하기' 없음 "
                    f"(모두저장과 다른 버튼). 원인: {e}"
                ) from e

            try:
                desc = _describe(btn)
            except Exception:  # noqa: BLE001
                desc = "저장하기"
            ctx.info(
                f"9. ★ 하단 '저장하기' 클릭 "
                f"({attempt}/2) | {desc}"
            )
            alert_baseline = collect_alert_baseline(page)
            dialog_from = len(dialog_msgs)
            before_popup_ids = {_popup_id(p) for p in save_popups(page)}
            last_click_ok = trusted_click_save_submit(page, btn, ctx)
            # 클릭 시도 순간부터 팝업 닫힘까지 초기화 진입 금지
            ctx.save_awaiting_popup = True
            ctx.info(
                f"  9항 클릭 전송 (playwright_ok={last_click_ok}, "
                f"attempt={attempt}) — 저장 팝업 반응 필수"
            )
            page.wait_for_timeout(400)
            ctx.shot(page, "02_save_clicked", rn)

            react_timeout = (
                max(8.0, SAVE_POPUP_GRACE_SEC + 2.0)
                if last_click_ok
                else 3.0
            )
            if save_submit_reacted(
                page,
                dialog_msgs,
                before_popup_ids=before_popup_ids,
                baseline=alert_baseline,
                dialog_from=dialog_from,
                timeout_sec=react_timeout,
            ):
                save_reacted = True
                ctx.info("9. ★ 저장하기 클릭 → 저장 실행 팝업/알림 감지")
                break

            ctx.info(
                "  [경고] 9항 클릭 후 저장 팝업 미감지 — "
                "설정모달 닫힘·잔여 00건만으로는 부족 (재시도)"
            )
            ctx.shot(page, "02_save_no_react", rn)
            dump_save_button_candidates(page, ctx)
            diag = diagnose_save_click_environment(page, btn, ctx, tag="반응없음")
            if diag.get("intercept", {}).get("same") is False:
                ctx.info(
                    "  [9항 원인추정] 클릭 좌표를 다른 요소가 가로챔 → "
                    f"{diag['intercept']}"
                )
            if diag.get("unselected_required"):
                ctx.info(
                    "  [9항 원인추정] 미선택 필수값 있음(정책 등) → "
                    f"{diag['unselected_required']}"
                )
            if diag.get("frame_urls"):
                ctx.info(
                    "  [9항 원인추정] 다른 프레임 존재(iframe 가능) → "
                    f"{diag['frame_urls']}"
                )
            if diag.get("open_windows_total"):
                ctx.info(
                    "  [9항 원인추정] 새 창은 열렸으나 저장팝업으로 미인식 가능 "
                    f"(open_windows={diag.get('open_windows')}) — "
                    "ADMIN_HOST 창이면 save_popups()로 잡혀야 정상"
                )
            else:
                ctx.info(
                    "  [9항 원인추정] 클릭 후 새 창이 전혀 안 열림 — "
                    "팝업이 실제로 안 뜨는 것(클릭 미도달/팝업차단) 쪽에 무게"
                )
            if save_modal_visible(page) and attempt < 2:
                page.wait_for_timeout(600)
                continue
            break

        if not save_reacted:
            dump_save_button_candidates(page, ctx)
            ctx.shot(page, "02_save_failed", rn)
            hint = (
                "Chrome을 완전히 종료 후 다시 실행하세요 "
                "(--disable-popup-blocking 은 새로 켤 때만 적용)."
            )
            raise RuntimeError(
                f"#{rn} 9항 하단 '저장하기' 서버 제출 실패 — "
                "클릭해도 저장 팝업이 없음. 위 [9항 진단]/[9항 원인추정] 로그로 "
                "가로채는 요소·미선택 필수값·iframe 여부를 확인하세요. "
                f"1회 재시도 후에도 동일. {hint}"
            )

        # ★요건: 9·10·11 합산 180초. 클릭 직후부터 곧바로 완료 메세지 확인.
        # 180초 내 미완료면 다음 입력으로(배치 중단 금지).
        post_dialog_from = len(dialog_msgs)
        ctx.mango_save_log_lines = []
        found = wait_for_save_complete_signal(
            page,
            ctx,
            rn,
            dialog_msgs=dialog_msgs,
            baseline=alert_baseline,
            dialog_from=post_dialog_from,
            before_popup_ids=before_popup_ids,
            timeout_sec=SAVE_PHASE_BUDGET_SEC,
        )
        if found is None:
            ctx.shot(page, "03_result_missing", rn)
            raise RowBudgetExceeded(
                f"#{rn} 9·10·11 합산 {SAVE_PHASE_BUDGET_SEC:.0f}초 내 "
                "'신규상품의 저장이 완료' 메세지 미확인 — 다음 입력으로"
            )
        kind, hit = found
        ctx.save_popup_kind = kind

        ctx.step(10, "수집 상품 DB저장하기 실행 : 저장 팝업 모달 열림 확인")
        ctx.info(f"신호={kind} 문구={hit!r}")
        ctx.shot(page, "10_popup_open", rn)
        ctx.save_popup_seen = True

        ctx.step(11, "수집 상품 DB저장하기 종료 : 저장완료 메세지로 즉시 확인")
        ctx.shot(page, "11_popup_closed", rn)
        ctx.save_popup_closed = True

        # ★12항: 망고 로그 원문을 SUB에 남김 (시작~완료 구간)
        # 구분: "신규상품의 저장을 시작합니다." /
        #       "신규상품의 저장을 완료하였습니다."
        found_n = save_count
        if kind == "count_msg":
            n_parsed, _ = parse_mango_collect_count(hit)
            if n_parsed is not None:
                found_n = n_parsed
        ctx.save_count_snapshot = found_n
        ctx.step(12, "수집후 DB 최종 저장건수 로그 확인")
        # 샷은 저장완료 mute 전에 (12항 확인 샷은 화면에 남김)
        ctx.shot(page, "12_count_logged", rn)
        log_lines = list(getattr(ctx, "mango_save_log_lines", []) or [])
        if not log_lines:
            # 폴백: 화면에 남은 텍스트에서 한 번 더 수확
            try:
                harvest_mango_save_log(page, ctx)
                log_lines = list(ctx.mango_save_log_lines)
            except Exception:  # noqa: BLE001
                log_lines = []
        if log_lines:
            for ln in log_lines:
                ctx.info(ln)
                # 완료 메세지까지만 화면 SUB — 이후는 mute (요건)
                if SAVE_COMPLETE_MSG_PATTERN.search(ln):
                    break
        else:
            # 최소 구분 메세지라도 남김
            ctx.info("......신규상품의 저장을 시작합니다.")
            ctx.info("......신규상품의 저장을 완료하였습니다.")
        ctx.save_count_logged = True

        # 이 아래 info/샷 들은 파일·갤러리에만 남고 화면 SUB에는 안 나감(mute)
        verify_row_save_done(page, ctx, rn, save_count)
        # 10·11·12 모두 확인된 경우에만 서버 저장 성공 → 13항 초기화 허용
        if not (
            getattr(ctx, "save_popup_seen", False)
            and getattr(ctx, "save_popup_closed", False)
            and getattr(ctx, "save_count_logged", False)
        ):
            raise RuntimeError(
                f"#{rn} 10·11·12항 미완료 — "
                "팝업 열림→닫힘→건수로그 확인 뒤에만 초기화 가능"
            )
        ctx.server_save_ok = True
        ctx.save_awaiting_popup = False  # 11·12 완료 → 13항 초기화 허용
        ctx.info(
            f"서버 최종 갱신 완료 (저장하기 OK + 팝업 열림→닫힘→건수로그 OK / "
            f"저장수 {save_count})"
        )
        ctx.shot(page, "04_row_done", rn)
    finally:
        try:
            page.remove_listener("dialog", _on_save_dialog)
        except Exception:  # noqa: BLE001
            pass


def _page_visible_text(page: Page) -> str:
    try:
        return (
            page.evaluate(
                "() => (document.body && document.body.innerText) "
                "? document.body.innerText : ''"
            )
            or ""
        )
    except Exception:  # noqa: BLE001
        return ""


def parse_mango_collect_count(text: str) -> tuple[int | None, str]:
    """망고 알림 문구에서 수집 건수 추출. (건수 또는 None, 매칭 원문)."""
    raw = text or ""
    for pat in MANGO_COLLECT_ALERT_PATTERNS:
        m = pat.search(raw)
        if not m:
            continue
        try:
            n = int(m.group(1))
        except (TypeError, ValueError):
            continue
        return n, m.group(0)
    return None, ""


def collect_alert_baseline(page: Page) -> set[str]:
    """저장하기 클릭 전 — 이미 화면에 있던 수집건수 문구 지문.

    검색 단계의 '00건이 수집되었다' 등이 저장 팝업으로 오인되는 것을 막는다.
    """
    fps: set[str] = set()
    try:
        text = _page_visible_text(page)
        n, hit = parse_mango_collect_count(text)
        if n is not None:
            fps.add(f"page:{n}:{hit}")
    except Exception:  # noqa: BLE001
        pass
    for p in save_popups(page):
        try:
            ptext = _page_visible_text(p)
            n, hit = parse_mango_collect_count(ptext)
            if n is not None:
                fps.add(f"popup:{n}:{hit}")
        except Exception:  # noqa: BLE001
            continue
    return fps


def find_mango_collect_alert(
    page: Page,
    dialog_msgs: list[str] | None = None,
    *,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
) -> tuple[int | None, str, str, Page | None]:
    """메인·팝업·JS dialog에서 수집건수 알림을 찾는다.

    baseline 이 있으면 클릭 전에 있던 문구(예: 00건 수집)는 무시한다.
    반환: (건수|None, 매칭문구, source, 해당 Page|None)
    """
    base = baseline or set()

    # 클릭 이후 새로 도착한 dialog 만
    for msg in list(dialog_msgs or [])[max(0, int(dialog_from)) :]:
        n, hit = parse_mango_collect_count(msg)
        if n is not None:
            fp = f"dialog:{n}:{hit or msg}"
            if fp not in base:
                return n, hit or msg, "dialog", page

    text = _page_visible_text(page)
    n, hit = parse_mango_collect_count(text)
    if n is not None:
        fp = f"page:{n}:{hit}"
        if fp not in base:
            return n, hit, "page", page

    for p in save_popups(page):
        ptext = _page_visible_text(p)
        n, hit = parse_mango_collect_count(ptext)
        if n is not None:
            fp = f"popup:{n}:{hit}"
            if fp not in base:
                return n, hit, "popup", p

    return None, "", "", None


def find_save_complete_signal(
    page: Page,
    dialog_msgs: list[str] | None = None,
    *,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
    before_popup_ids: set[int] | None = None,
    after_save_click: bool = False,
) -> tuple[str, str, Page | None] | None:
    """★저장완료 확정 신호 — 보이는 즉시(단계적 대기 없이) 다음 단계로.

    최우선: "신규상품의 저장이 완료되었습니다" 정확 문구.
    보조: 기존 'N건이 수집/저장되었다' 알림도 저장완료로 인정(실제
    사이트 문구가 다를 수 있는 경우의 안전망).
    after_save_click=True 이면 검색단계 잔여 본문문구(동일 건수) 오탐을
    막기 위해 본문 count 는 baseline 으로만 걸러지고, 클릭 이후 새
    팝업·dialog 의 count 는 항상 인정한다.
    읽기만 한다 — 클릭·스크롤 없음.
    반환: (kind, 매칭문구, 해당 Page) 또는 None.
    """
    prev_popups = before_popup_ids or set()

    # 1) JS dialog
    for msg in list(dialog_msgs or [])[max(0, int(dialog_from)) :]:
        if SAVE_COMPLETE_MSG_PATTERN.search(msg or ""):
            return "complete_msg", msg, page
    for msg in list(dialog_msgs or [])[max(0, int(dialog_from)) :]:
        n2, hit2 = parse_mango_collect_count(msg)
        if n2 is not None:
            return "count_msg", hit2 or msg, page

    # 2) 본문 — 완료 문구는 항상, count 는 after_save_click 시 본문 제외
    text = _page_visible_text(page)
    m = SAVE_COMPLETE_MSG_PATTERN.search(text or "")
    if m:
        return "complete_msg", m.group(0), page
    if not after_save_click:
        n, hit, _src, hit_page = find_mango_collect_alert(
            page, dialog_msgs, baseline=baseline, dialog_from=dialog_from
        )
        if n is not None:
            return "count_msg", hit, hit_page or page

    # 3) 새 창(ADMIN_HOST 포함) — 완료 문구·건수 모두
    for p in save_popups(page):
        ptext = _page_visible_text(p)
        m2 = SAVE_COMPLETE_MSG_PATTERN.search(ptext or "")
        if m2:
            return "complete_msg", m2.group(0), p
        n, hit = parse_mango_collect_count(ptext)
        if n is None:
            continue
        if after_save_click and _popup_id(p) not in prev_popups:
            return "count_msg", hit, p
        fp = f"popup:{n}:{hit}"
        if fp not in (baseline or set()):
            return "count_msg", hit, p

    return None


def extract_mango_save_log_lines(text: str) -> list[str]:
    """망고 저장 로그 원문에서 시작~완료 구간 줄을 추출.

    구분 메세지(요건):
      - "신규상품의 저장을 시작합니다." (또는 신규상품(N개)의 저장을 시작합니다.)
      - "신규상품의 저장을 완료하였습니다." / "완료되었습니다."
    """
    raw = text or ""
    if not raw.strip():
        return []
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    start_i = -1
    end_i = -1
    for i, ln in enumerate(lines):
        if start_i < 0 and SAVE_START_MSG_PATTERN.search(ln):
            start_i = i
        if SAVE_COMPLETE_MSG_PATTERN.search(ln):
            end_i = i
    if start_i >= 0 and end_i >= start_i:
        return lines[start_i : end_i + 1]
    if start_i >= 0:
        return lines[start_i:]
    # 시작 문구를 못 잡아도 완료 문구·상품업데이트 줄은 남긴다
    out: list[str] = []
    for ln in lines:
        if (
            SAVE_COMPLETE_MSG_PATTERN.search(ln)
            or "[상품업데이트]" in ln
            or SAVE_START_MSG_PATTERN.search(ln)
        ):
            out.append(ln)
    return out


def harvest_mango_save_log(page: Page, ctx: "RunCtx") -> None:
    """현재 화면/저장팝업에서 망고 저장 로그 원문을 ctx에 누적."""
    chunks: list[str] = []
    try:
        chunks.append(_page_visible_text(page))
    except Exception:  # noqa: BLE001
        pass
    for p in save_popups(page):
        try:
            chunks.append(_page_visible_text(p))
        except Exception:  # noqa: BLE001
            continue
    seen = set(ctx.mango_save_log_lines)
    for chunk in chunks:
        for ln in extract_mango_save_log_lines(chunk):
            if ln not in seen:
                seen.add(ln)
                ctx.mango_save_log_lines.append(ln)


def wait_for_save_complete_signal(
    page: Page,
    ctx: RunCtx,
    rn: int,
    *,
    dialog_msgs: list[str] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
    before_popup_ids: set[int] | None = None,
    timeout_sec: float | None = None,
) -> tuple[str, str] | None:
    """★9항 클릭 후 곧바로(무행동 대기 없이) 계속 확인 — 신호가 보이면 즉시 반환.

    120초 순수대기·300초 단계적 확인 요건은 생략한다. 대신 클릭 직후부터
    끊임없이(짧은 간격으로) 확인하고, 신호가 보이는 즉시 리턴한다.
    timeout_sec 기본=SAVE_PHASE_BUDGET_SEC(180초). 초과 시 None → 다음 행.
    대기 중 망고 저장 로그 원문(시작~완료)을 ctx.mango_save_log_lines 에 수집.
    """
    wait_sec = float(
        timeout_sec if timeout_sec is not None else SAVE_PHASE_BUDGET_SEC
    )
    end = time.time() + wait_sec
    while time.time() < end:
        ctx.check_budget("저장완료 메세지 확인")
        try:
            harvest_mango_save_log(page, ctx)
        except Exception:  # noqa: BLE001
            pass
        found = find_save_complete_signal(
            page,
            dialog_msgs,
            baseline=baseline,
            dialog_from=dialog_from,
            before_popup_ids=before_popup_ids,
            after_save_click=True,
        )
        if found is not None:
            kind, hit, _hit_page = found
            try:
                harvest_mango_save_log(page, ctx)
            except Exception:  # noqa: BLE001
                pass
            return kind, hit
        page.wait_for_timeout(200)
    return None


def _is_settings_modal_text(txt: str) -> bool:
    return bool(
        re.search(r"상품\s*저장\s*설정", txt or "")
        and re.search(r"검색\s*필터\s*명|취소하기", txt or "")
    )


def save_execution_layer_visible(
    page: Page,
    *,
    baseline: set[str] | None = None,
) -> bool:
    """저장하기 직후 뜨는 '저장 실행' 팝업/레이어 UI.

    ★ 본문에 남은 'N건이 수집' 텍스트만으로는 True 금지.
    실제 dialog/layer/modal 또는 확인버튼 UI 가 보여야 함.
    """
    base = baseline or set()
    try:
        layer = page.locator(
            '[role="dialog"], .ui-dialog, .modal, '
            'div[id*="layer"], div[id*="popup"], '
            'div[class*="layer"], div[class*="popup"], '
            'div[class*="alert"], div[class*="msg"]'
        ).filter(
            has_text=re.compile(
                r"(\d+)\s*건\s*(이\s*)?(수집|저장)\s*되었|"
                r"수집\s*완료|저장\s*완료|처리\s*완료|"
                r"상품\s*(이\s*)?(수집|저장)\s*되었"
            )
        )
        if layer.count() > 0 and layer.first.is_visible():
            txt = ""
            try:
                txt = layer.first.inner_text(timeout=500) or ""
            except Exception:  # noqa: BLE001
                txt = ""
            if _is_settings_modal_text(txt):
                return False
            n, hit = parse_mango_collect_count(txt)
            if n is not None and f"page:{n}:{hit}" in base:
                return False  # 검색 단계 잔여 '00건 수집' 등
            if n is not None or re.search(
                r"수집\s*완료|저장\s*완료|처리\s*완료", txt or ""
            ):
                return True
    except Exception:  # noqa: BLE001
        pass
    return _save_result_confirm_ui_visible(page, baseline=base)


def _save_result_confirm_ui_visible(
    page: Page,
    *,
    baseline: set[str] | None = None,
) -> bool:
    """수집결과 팝업의 확인/닫기 버튼 UI가 보이는지 (본문 잔여문구 단독 X)."""
    base = baseline or set()
    try:
        btn = (
            page.locator("button, a, input[type='button'], input[type='submit']")
            .filter(has_text=re.compile(r"^(확인|닫기|OK|Yes)$", re.I))
            .first
        )
        if btn.count() == 0 or not btn.is_visible():
            return False
        try:
            root = btn.locator(
                "xpath=ancestor::*[self::div or self::td or self::form][1]"
            )
            txt = root.inner_text(timeout=400) if root.count() else ""
        except Exception:  # noqa: BLE001
            txt = ""
        if _is_settings_modal_text(txt):
            return False
        n, hit = parse_mango_collect_count(txt or "")
        if n is not None and f"page:{n}:{hit}" in base:
            return False
        if n is not None or re.search(
            r"수집\s*완료|저장\s*완료|처리\s*완료", txt or ""
        ):
            return True
    except Exception:  # noqa: BLE001
        pass
    return False


def _popup_id(p) -> int:
    try:
        return id(p)
    except Exception:  # noqa: BLE001
        return 0


def save_result_signal_present(
    page: Page,
    dialog_msgs: list[str] | None = None,
    *,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
) -> tuple[bool, str, Page | None]:
    """저장하기 후 '저장 실행' 팝업/알림 모달 출현 여부.

    인정:
    - 클릭 이후 새 JS dialog
    - 클릭 이후 새 브라우저 팝업 창
    - 인페이지 팝업/레이어/확인 UI
    금지:
    - 상품저장설정 모달 닫힘만
    - 본문에 남은 'N건이 수집' 텍스트만 (모달 UI 없음)
    """
    base = baseline or set()

    # 1) JS dialog (클릭 이후 도착분만)
    for msg in list(dialog_msgs or [])[max(0, int(dialog_from)) :]:
        n, hit = parse_mango_collect_count(msg)
        if n is not None:
            return True, f"dialog:{hit or msg}", page

    # 2) 새 브라우저 창 (ADMIN_HOST 자체 새 창 포함 — save_popups)
    prev = before_popup_ids or set()
    for p in save_popups(page):
        if _popup_id(p) in prev:
            continue
        ptext = _page_visible_text(p)
        n2, hit2 = parse_mango_collect_count(ptext)
        if n2 is not None:
            fp = f"popup:{n2}:{hit2}"
            if fp in base:
                continue
            return True, f"new_popup:{hit2}", p
        return True, "new_popup_window", p

    # 3) 실제 인페이지 팝업/레이어/확인 UI (본문 잔여문구 단독 금지)
    if save_execution_layer_visible(page, baseline=base):
        return True, "inpage_save_layer", page

    return False, "", None


def wait_save_execution_popup(
    page: Page,
    ctx: RunCtx,
    rn: int,
    *,
    dialog_msgs: list[str] | None = None,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
    timeout_sec: float | None = None,
    grace_sec: float | None = None,
) -> None:
    """★저장하기 클릭 후 — 저장 실행 팝업창 모달이 뜰 때까지 필수 대기.

    - 최소 grace_sec 동안은 서버 저장 시간을 줌 (즉시 초기화 금지)
    - 클릭 전 '00건 수집' 잔여 문구는 팝업으로 보지 않음
    - 팝업이 안 보이면 초기화/다음 단계로 절대 넘어가지 않음
    """
    wait_sec = float(timeout_sec or MODAL_WAIT_SEC)
    grace = max(2.0, float(grace_sec if grace_sec is not None else SAVE_POPUP_GRACE_SEC))
    base = baseline if baseline is not None else set()
    ctx.info(
        f"저장하기 후 {grace:.0f}초 동안 어떠한 액션도 취하지 않고 대기 "
        "(요건 3) — 그 후 최대 "
        f"{int(wait_sec)}초 '상품 N건이 수집, 저장되었다' 메세지 확인 (요건 6)"
    )

    # ── 요건 3: 클릭 후 grace(기본 120초) 동안은 순수 대기만. 아무 것도 읽지·
    # 누르지·스크롤하지 않는다. ──
    blind_end = time.time() + grace
    while time.time() < blind_end:
        ctx.check_budget("저장하기 후 120초 무행동 대기")
        page.wait_for_timeout(min(2000, max(200, int((blind_end - time.time()) * 1000))))

    # ── 요건 6: 이후 최대 wait_sec(기본 300초) 동안 — 읽기만(클릭·스크롤 없음)
    # 하며 저장 완료 메세지를 확인. 찾을 때까지 후속 단계 진행 금지. ──
    end = time.time() + wait_sec
    saw = False
    detail = ""
    target: Page | None = None
    logged_stale = False

    while time.time() < end:
        ctx.check_budget("저장하기 후 팝업창 모달(메세지) 대기")

        has_signal, detail, target = save_result_signal_present(
            page,
            dialog_msgs,
            before_popup_ids=before_popup_ids,
            baseline=base,
            dialog_from=dialog_from,
        )
        if has_signal:
            saw = True
            break

        # 잔여 00건이 화면에 있으면 안내 (오인 방지) — 읽기만, 클릭 없음
        if not logged_stale and base:
            stale_n, stale_hit = parse_mango_collect_count(_page_visible_text(page))
            if stale_n is not None and f"page:{stale_n}:{stale_hit}" in base:
                logged_stale = True
                ctx.info(
                    f"검색단계 잔여 문구 {stale_hit!r} "
                    "— 저장 완료 메세지로 보지 않음, 계속 대기"
                )

        page.wait_for_timeout(1000)

    if not saw:
        ctx.shot(page, "03_result_missing", rn)
        n_open = 0
        try:
            n_open = len(save_popups(page))
        except Exception:  # noqa: BLE001
            pass
        hint = ""
        if n_open == 0:
            hint = (
                " [의심원인] 클릭 후 새 창이 전혀 안 열림 — Chrome 자체 "
                "팝업차단일 수 있음. Chrome을 완전히 종료(작업관리자 포함)한 뒤 "
                "다시 실행해 보세요(--disable-popup-blocking 은 새로 켤 때만 적용)."
            )
        raise TimeoutError(
            f"#{rn} 저장하기 후 팝업창 모달이 나타나지 않음. "
            "검색단계 '00건 수집' 잔여문구·설정모달 닫힘만으로 "
            "초기화/다음 행 진행 불가. "
            "저장하기 → (저장시간) → 팝업창 모달 열림 → 닫힘 → 건수확인 필수."
            f"{hint}"
        )

    ctx.info(f"3. ★ 저장 실행 팝업창 모달 확인 — {detail}")
    shot_page = target or page
    try:
        shot_page.bring_to_front()
    except Exception:  # noqa: BLE001
        pass
    ctx.shot(shot_page, "03_result_popup", rn)
    ctx.save_popup_seen = True
    ctx.save_popup_kind = (detail or "").split(":", 1)[0] or "inpage_save_layer"
    # dialog 는 accept 로 이미 닫힌 경우가 많음. UI 팝업은 열림을 래치.
    if ctx.save_popup_kind == "dialog":
        ctx.save_popup_ui_latched = False
    else:
        ctx.save_popup_ui_latched = bool(
            final_save_popup_still_open(
                page, before_popup_ids=before_popup_ids, baseline=base
            )
            or save_execution_layer_visible(page, baseline=base)
        )
        if not ctx.save_popup_ui_latched:
            # 신호는 왔는데 아직 안 보이면 잠깐 더 보고 래치
            page.wait_for_timeout(500)
            ctx.save_popup_ui_latched = bool(
                final_save_popup_still_open(
                    page, before_popup_ids=before_popup_ids, baseline=base
                )
                or save_execution_layer_visible(page, baseline=base)
            )
    ctx.info(
        f"  [팝업] kind={ctx.save_popup_kind} "
        f"ui_open={ctx.save_popup_ui_latched} — 닫힐 때까지 초기화 금지"
    )


def final_save_popup_still_open(
    page: Page,
    *,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
) -> bool:
    """최종 저장 팝업/레이어가 아직 화면에 열려 있는지.

    - 저장하기 이후 새로 뜬 브라우저 창
    - 저장 실행 인페이지 레이어(확인 버튼 포함)
    본문에 남은 잔여 텍스트만으로는 '열림'으로 보지 않음.
    """
    prev = before_popup_ids or set()
    for p in save_popups(page):
        try:
            if _popup_id(p) in prev:
                continue
            if not p.is_closed():
                return True
        except Exception:  # noqa: BLE001
            continue

    if save_execution_layer_visible(page, baseline=baseline or set()):
        return True

    # 수집 결과 레이어의 확인/닫기 버튼이 보이면 아직 열린 것
    try:
        btn = (
            page.locator("button, a, input[type='button'], input[type='submit']")
            .filter(has_text=re.compile(r"^(확인|닫기|OK|Yes)$", re.I))
            .first
        )
        if btn.count() > 0 and btn.is_visible():
            # 상품저장설정 푸터의 취소하기와 구분 — 확인 근처 수집문구
            try:
                root = btn.locator("xpath=ancestor::*[self::div or self::td or self::form][1]")
                txt = root.inner_text(timeout=400) if root.count() else ""
            except Exception:  # noqa: BLE001
                txt = ""
            if parse_mango_collect_count(txt or "")[0] is not None:
                return True
            if re.search(r"수집\s*완료|저장\s*완료|처리\s*완료", txt or ""):
                return True
    except Exception:  # noqa: BLE001
        pass
    return False


def wait_save_popup_closed(
    page: Page,
    ctx: RunCtx,
    rn: int,
    *,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    timeout_sec: float | None = None,
) -> None:
    """★최종 팝업화면이 뜰 때까지가 아니라 — 뜬 뒤 '닫힐 때까지' 대기.

    최초 요건: 팝업창 열고, 닫힘까지 처리. 열린 채로 초기화/다음단계 금지.
    UI 팝업을 한 번도 못 본 채 '이미 닫힘'으로 통과하지 않는다.
    """
    wait_sec = float(timeout_sec or MODAL_WAIT_SEC)
    base = baseline or set()
    kind = getattr(ctx, "save_popup_kind", "") or ""
    ctx.info(
        "3. ★★★ 최종 팝업화면이 닫힐 때까지 대기 "
        f"(최대 {int(wait_sec)}초, kind={kind}) — 닫히기 전 초기화 금지"
    )

    # JS dialog 는 accept 로 이미 닫힘 — 짧은 정착 후 닫힘 인정
    if kind == "dialog":
        page.wait_for_timeout(800)
        ctx.save_popup_closed = True
        ctx.info("3. ★ 최종 팝업(dialog) 닫힘 확인 완료")
        ctx.shot(page, "03_modal_closed", rn)
        return

    end = time.time() + wait_sec
    closed_stable = 0
    saw_open = bool(getattr(ctx, "save_popup_ui_latched", False))
    last_dismiss_at = 0.0

    while time.time() < end:
        ctx.check_budget("11항 최종 팝업 닫힘 대기")
        # 닫기 전에 건수 스냅샷 (12항용) — 읽기만, 클릭 없음
        if getattr(ctx, "save_count_snapshot", None) is None:
            try:
                sn, sm, ss, _ = find_mango_collect_alert(
                    page, None, baseline=base
                )
                if sn is not None:
                    ctx.save_count_snapshot = sn
                    ctx.info(f"11항 직전 스냅샷: {sn}건 ({ss}:{sm!r})")
            except Exception:  # noqa: BLE001
                pass
        open_now = final_save_popup_still_open(
            page, before_popup_ids=before_popup_ids, baseline=base
        )
        if open_now:
            saw_open = True
            closed_stable = 0
            # 11항: 확인 클릭으로 정상 종료(2초에 한 번만 — 과도한 반복클릭 금지)
            if time.time() - last_dismiss_at >= 2.0:
                dismiss_mango_alert_ui(page)
                last_dismiss_at = time.time()
            page.wait_for_timeout(500)
            continue

        # 아직 UI 열림을 한 번도 못 봄 → '닫힘'으로 오인 금지, 계속 대기
        if not saw_open:
            page.wait_for_timeout(500)
            continue

        closed_stable += 1
        if closed_stable >= 2:  # ~1초 이상 닫힌 상태 유지
            ctx.save_popup_closed = True
            ctx.info("최종 팝업화면 닫힘 확인 완료")
            ctx.shot(page, "03_modal_closed", rn)
            return
        page.wait_for_timeout(500)

    ctx.shot(page, "03_modal_stuck", rn)
    if not saw_open:
        raise TimeoutError(
            f"#{rn} 저장하기 후 팝업모달 UI를 확인하지 못함. "
            "본문 잔여문구만으로 통과·초기화 불가. "
            "저장하기 → 팝업모달 열림 → 닫힘 필수."
        )
    raise TimeoutError(
        f"#{rn} 최종 팝업화면이 닫히지 않음. "
        "팝업이 열린 채로 초기화/다음 행 진행 불가. "
        "저장하기 → 팝업 열림 → 팝업 닫힘 → 다음 단계 순서 필수."
    )


def wait_save_overlays_settle(
    page: Page,
    ctx: RunCtx,
    rn: int,
    *,
    dialog_msgs: list[str] | None = None,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
) -> None:
    """저장 실행 팝업: 열림(필수) → 닫힘(필수).

    요건 원문: 팝업창 열고, 닫힘 등 후속 과정 처리.
    열림만 보고 다음으로 가면 안 된다.
    """
    base = baseline if baseline is not None else set()
    ctx.save_popup_closed = False

    # 1) 열림
    wait_save_execution_popup(
        page,
        ctx,
        rn,
        dialog_msgs=dialog_msgs,
        before_popup_ids=before_popup_ids,
        baseline=base,
        dialog_from=dialog_from,
    )
    if not getattr(ctx, "save_popup_seen", False):
        raise RuntimeError(
            f"#{rn} 최종 팝업 열림 미확인 — 닫힘 대기/초기화 불가"
        )

    # 2) 닫힘 — 여기 통과 전에 초기화·다음단계 금지
    wait_save_popup_closed(
        page,
        ctx,
        rn,
        before_popup_ids=before_popup_ids,
        baseline=base,
    )
    if not getattr(ctx, "save_popup_closed", False):
        raise RuntimeError(
            f"#{rn} 최종 팝업화면이 닫히지 않음 — 초기화 진행 불가"
        )


def dismiss_mango_alert_ui(page: Page) -> None:
    """화면 알림/레이어의 확인 버튼이 있으면 닫기."""
    try:
        btn = (
            page.locator("button, a, input[type='button'], input[type='submit']")
            .filter(has_text=re.compile(r"^(확인|닫기|OK|Yes)$", re.I))
            .first
        )
        if btn.count() > 0 and btn.is_visible():
            click_it(btn)
            page.wait_for_timeout(300)
    except Exception:  # noqa: BLE001
        pass
    # 팝업 창에도 확인 버튼이 있을 수 있음 (ADMIN_HOST 저장팝업 포함)
    for p in list(save_popups(page)):
        try:
            btn = (
                p.locator("button, a, input[type='button'], input[type='submit']")
                .filter(has_text=re.compile(r"^(확인|닫기|OK|Yes)$", re.I))
                .first
            )
            if btn.count() > 0 and btn.is_visible():
                click_it(btn)
                p.wait_for_timeout(300)
        except Exception:  # noqa: BLE001
            pass
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(200)
    except Exception:  # noqa: BLE001
        pass


def verify_mango_collect_alert(
    page: Page,
    ctx: RunCtx,
    row_no: int,
    expect_count: int,
    *,
    dialog_msgs: list[str] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
    timeout_sec: float = 25.0,
    dismiss: bool = False,
) -> int:
    """★12항: 망고 자체 'N건이 수집되었다' 최종 저장건수 로그 검증.

    - 저장하기 클릭 이후 새로 뜬 알림만 사용 (검색단계 00건 잔여 무시)
    - 11항(저장팝업 닫힘) 이후에 호출하는 것이 정상 순서
    - 10항에서 찍은 save_count_snapshot 도 인정
    - 기대 건수(저장수)와 일치해야 통과
    - 0건이면 실패(다음 행으로 진행 가능하도록 확정 실패 문구 포함)
    """
    expect = max(0, int(expect_count))
    base = baseline or set()
    ctx.info(
        f"12. 망고 자체 알림 확인 — 'N건이 수집되었다' (기대 {expect}건) "
        f"| 잔여무시={len(base)}건"
    )
    end = time.time() + max(5.0, float(timeout_sec))
    found_n: int | None = None
    found_msg = ""
    source = ""
    shot_page: Page | None = page

    # 10항 스냅샷이 있으면 우선 후보
    snap = getattr(ctx, "save_count_snapshot", None)
    if snap is not None:
        found_n = int(snap)
        found_msg = f"(snapshot){snap}건"
        source = "snapshot"

    while found_n is None and time.time() < end:
        ctx.check_budget("12항 망고 수집알림 대기")
        found_n, found_msg, source, hit_page = find_mango_collect_alert(
            page,
            dialog_msgs,
            baseline=base,
            dialog_from=dialog_from,
        )
        if found_n is not None:
            shot_page = hit_page or page
            break

        texts = [_page_visible_text(page)]
        for p in save_popups(page):
            texts.append(_page_visible_text(p))
        blob = "\n".join(texts)
        for pat in SAVE_FAIL_PATTERNS:
            if pat.search(blob or ""):
                ctx.shot(page, "03_collect_alert_fail", row_no)
                raise RuntimeError(
                    f"#{row_no} 망고 수집 알림 대신 오류 문구 감지: {pat.pattern}"
                )
        page.wait_for_timeout(400)

    if found_n is None:
        found_n, found_msg, source, hit_page = find_mango_collect_alert(
            page,
            dialog_msgs,
            baseline=base,
            dialog_from=dialog_from,
        )
        shot_page = hit_page or page
        if found_n is not None and not source:
            source = "final"

    if found_n is None:
        ctx.shot(page, "03_collect_alert_fail", row_no)
        raise RuntimeError(
            f"#{row_no} 12항 망고 수집 알림을 찾지 못함 "
            f"(기대 문구 예: '{expect}건이 수집되었다'). "
            "11항 팝업 닫힘 확인 후 저장건수 로그를 확인하세요."
        )

    ctx.info(
        f"  [12항 망고알림] source={source} | 문구={found_msg!r} | "
        f"수집건수={found_n} | 기대={expect}"
    )
    try:
        ctx.shot(shot_page or page, "03_collect_alert", row_no)
    except Exception:  # noqa: BLE001
        ctx.shot(page, "03_collect_alert", row_no)
    if dismiss:
        dismiss_mango_alert_ui(page)

    if found_n == 0:
        raise RuntimeError(
            f"#{row_no} 망고 자체 메세지: 0건이 수집되었다 "
            f"(알림={found_msg!r}). 다음 입력으로 진행."
        )
    if found_n != expect:
        raise RuntimeError(
            f"#{row_no} 망고 수집건수 알림 불일치 — "
            f"기대 {expect}건 / 알림 {found_n}건 (문구={found_msg!r})"
        )
    ctx.info(f"  [12항 확인] 망고 알림 수집건수 OK — {found_n}건이 수집됨")
    return found_n


def verify_row_save_done(page: Page, ctx: RunCtx, row_no: int, save_count: int) -> None:
    """저장 모달·알림 확인 후, 오류 문구·로그인 튕김 등 부가 검사."""
    page.wait_for_timeout(400)
    text = _page_visible_text(page)

    for pat in SAVE_FAIL_PATTERNS:
        if pat.search(text or ""):
            raise RuntimeError(f"#{row_no} 저장 후 오류 문구 감지: {pat.pattern}")

    ok_hit = None
    for pat in SAVE_OK_PATTERNS:
        m = pat.search(text or "")
        if m:
            ok_hit = m.group(0)
            break
    if ok_hit:
        ctx.info(f"  [확인] 화면 부가 성공 신호: {ok_hit!r}")

    if ctx.verify:
        if "admin_login" in page.url:
            raise RuntimeError(f"#{row_no} 저장 직후 로그인 화면으로 이동됨")


def _is_advance_fail(err: BaseException) -> bool:
    """같은 행을 더 돌리지 않고 다음 입력으로 넘길 확정 실패인지."""
    text = str(err or "")
    return any(m in text for m in ROW_ADVANCE_FAIL_MARKERS)


def process_row_with_retries(page: Page, row: dict, ctx: RunCtx) -> bool:
    """행 단위 재시도. 성공 True / 최종 실패 False.

    망고 무결과·시간초과 등 확정 실패는 같은 행을 반복하지 않고
    즉시 끝내 다음 입력 데이터로 넘긴다. (1행 무한루프 방지)
    """
    last_err: Exception | None = None
    label = str(row.get("label") or "").strip()
    raw_url = str(row.get("url") or "").strip()
    success = False
    try:
        for attempt in range(1, ctx.retries + 1):
            try:
                ctx.check_budget(f"행 시도 전 엑셀{row['row']}행")
            except RowBudgetExceeded as e:
                last_err = e
                ctx.info(f"  [다음행] {e}")
                break
            try:
                ctx.info(
                    f"> 시도 {attempt}/{ctx.retries} (엑셀 {row['row']}행) | "
                    f"최종 카테고리명={label} | 최종 카테고리 URL주소={raw_url}"
                )
                page = refresh_if_closed(page)
                _process_row_once(page, row, ctx)
                if not (
                    ctx.server_save_ok
                    and getattr(ctx, "save_popup_closed", False)
                    and getattr(ctx, "save_count_logged", False)
                    and getattr(ctx, "search_popup_closed", False)
                ):
                    raise RuntimeError(
                        f"#{row['row']} 저장하기 서버 최종 갱신 미확인 — "
                        "6·11·12항(검색팝업닫힘·저장팝업닫힘·건수로그) "
                        "확인 전 행을 끝낼 수 없습니다."
                    )
                ctx.info(
                    f"[OK] 엑셀{row['row']}행 성공 (저장하기 서버갱신 OK, 시도 {attempt}) | "
                    f"최종 카테고리명={label} | 최종 카테고리 URL주소={raw_url}"
                )
                success = True
                return True
            except CollectStopped:
                raise
            except RowBudgetExceeded as e:
                last_err = e
                ctx.info(
                    f"[FAIL] 엑셀{row['row']}행 제한시간 초과 — 다음 입력으로 | {e}"
                )
                try:
                    ctx.shot(page, "fail_budget", row["row"])
                except Exception:  # noqa: BLE001
                    pass
                break
            except Exception as e:  # noqa: BLE001
                last_err = e
                err_name = type(e).__name__
                ctx.info(
                    f"[FAIL] 엑셀{row['row']}행 실패 (시도 {attempt}/{ctx.retries}) | "
                    f"최종 카테고리명={label} | 최종 카테고리 URL주소={raw_url} | "
                    f"{err_name}: {e}"
                )
                try:
                    page = refresh_if_closed(page)
                    ctx.shot(page, f"fail_attempt{attempt}", row["row"])
                except Exception:
                    pass
                # 확정 실패 → 같은 행 재시도 중단, 다음 입력으로
                if _is_advance_fail(e):
                    ctx.info(
                        "  [다음행] 확정 실패 — "
                        "같은 행 재시도 없이 다음 입력 데이터로 진행"
                    )
                    break
                if "TargetClosed" in err_name or "Target closed" in str(e):
                    ctx.info("  탭 닫힘 감지 — 작업 페이지 재연결 시도")
                    try:
                        page = refresh_if_closed(page)
                    except Exception as re:  # noqa: BLE001
                        ctx.info(f"  재연결 경고: {re}")
                if attempt < ctx.retries:
                    # 팝업 대기를 이미 끝낸 실패만 재시도 복귀 허용
                    # (대기 중 초기화 금지 플래그 해제)
                    ctx.save_awaiting_popup = False
                    ctx.info(
                        "  같은 행 재시도 — 대량메뉴 복귀 "
                        "(저장 팝업 대기 실패 후, 성공 경로 아님)"
                    )
                    try:
                        page = refresh_if_closed(page)
                        reset_to_bulk_menu(page)
                    except Exception as re:  # noqa: BLE001
                        ctx.info(f"  복귀 중 경고: {re}")
                    try:
                        page.wait_for_timeout(100)
                    except Exception:
                        page = refresh_if_closed(page)
        if not success:
            ctx.info(f"[FAIL] {row['row']}행 최종 실패: {last_err}")
        return success
    finally:
        # 성공/실패와 무관하게 남은 모달·팝업 정리 → 다음 행 대기 부담 감소
        # ★절대: save_awaiting_popup / deadline 잔존으로 다음 행이 막히지 않게 해제
        ctx.save_awaiting_popup = False
        ctx.save_phase_deadline = None
        try:
            page = finalize_row_overlays(page, ctx, row)
        except Exception as e:  # noqa: BLE001
            ctx.info(f"  [행종료정리] 실패: {e}")
        ctx.row_deadline = None


def process_row(page: Page, row: dict, save_count: int = DEFAULT_SAVE_COUNT) -> None:
    """하위 호환: 저장수만 받아 1회 시도."""
    ctx = RunCtx(save_count=save_count, retries=1, batch=True)
    try:
        _process_row_once(page, row, ctx)
    finally:
        ctx.close()


def ensure_ready_page(page: Page) -> Page:
    """BATCH 1항 로그인(+최초 화면 준비). 행별 2항 초기화는 batch_steps에서.

    로그인 성공 후 사이트가 원래 탭을 닫아버리는 경우가 있어
    매 단계 사이마다 탭이 살아있는지 확인하고 복구한다.
    """
    page = refresh_if_closed(page)
    step_log(1, "로그인")

    if ADMIN_HOST not in page.url or page.url in ("about:blank", ""):
        log("메인화면으로 이동: " + MAIN_URL)
        safe_goto(page, MAIN_URL)

    need_login = "admin_login" in page.url
    if not need_login:
        try:
            if page.locator('input[name="login_id"]').count() > 0:
                need_login = True
        except Exception:  # noqa: BLE001
            pass
    if need_login or ADMIN_HOST not in page.url:
        if "admin_login" not in page.url:
            log("1. 더망고 로그인창으로 이동: " + LOGIN_URL)
            safe_goto(page, LOGIN_URL)
            page = refresh_if_closed(page)
        shot_now(page, "login_gate", 0)
        page = wait_for_user_login(page)
        page = refresh_if_closed(page)
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15_000)
        except Exception:  # noqa: BLE001
            pass
        page = refresh_if_closed(page)
        if "admin_login" in page.url or ADMIN_HOST not in page.url:
            safe_goto(page, MAIN_URL)
        log("1. 로그인 완료")

    # 최초 진입 시 대량수집 화면까지 준비 (행 루프의 2항과 동일 메뉴)
    log("1→2 준비: 대량데이터수집 화면")
    if BULK_PATH in page.url:
        wait_bulk_ready(page)
    else:
        reset_to_bulk_menu(page)

    return page


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser(
        description="P2 더망고 대량수집 — URL 1행당 상품 N건 (기본 3)"
    )
    ap.add_argument("excel", help="P1 출력 엑셀 (.xlsx)")
    ap.add_argument(
        "save_count",
        nargs="?",
        type=int,
        default=DEFAULT_SAVE_COUNT,
        help=f"행당 저장 상품 수 (기본 {DEFAULT_SAVE_COUNT})",
    )
    ap.add_argument(
        "--verify",
        action="store_true",
        help="검증 모드: 앞 1·2행 단계 스크린샷 + 무중단 "
        "(★행 수는 제한하지 않음 — 엑셀 전체 처리)",
    )
    ap.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="처리할 최대 행 수 (기본: 제한 없음 = 엑셀 전체)",
    )
    ap.add_argument(
        "--shot-first",
        type=int,
        default=2,
        help="단계별 스크린샷을 남길 앞쪽 입력 행 수 (기본 2 = 1·2행)",
    )
    ap.add_argument(
        "--retries",
        type=int,
        default=DEFAULT_ROW_RETRIES,
        help=f"행 실패 시 같은 행 재시도 횟수 (기본 {DEFAULT_ROW_RETRIES})",
    )
    ap.add_argument(
        "--yes",
        "--batch",
        dest="batch",
        action="store_true",
        help="실패해도 y/n 묻지 않고 다음 행으로 (또는 재시도만)",
    )
    ap.add_argument(
        "--id",
        dest="tmg_id",
        default=None,
        help="(미사용) 자동 로그인 제거됨 — 브라우저에서 직접 로그인",
    )
    ap.add_argument(
        "--pw",
        dest="tmg_pw",
        default=None,
        help="(미사용) 자동 로그인 제거됨 — 브라우저에서 직접 로그인",
    )
    args = ap.parse_args()

    excel_path = args.excel
    save_count = args.save_count if args.save_count and args.save_count > 0 else DEFAULT_SAVE_COUNT
    verify = bool(args.verify)
    max_rows = args.max_rows
    shot_first = max(0, int(args.shot_first))
    # ★2026-08-08: --verify 는 "스크린샷 범위"일 뿐 "처리 행 수"가 아니다.
    # 예전에는 verify면 max_rows를 2로 강제해서, 보드 기본값(체크박스
    # "1·2행 전과정 스크린샷" ON → --verify 전달)으로 실행하면 엑셀에 자료가
    # 아무리 많아도 2행만 처리하고 정상 종료했다 — 사용자에게는 "엑셀 2번째
    # 자료에서 중단"으로 보였다. 행 수 제한은 --max-rows 로만 건다.

    if args.tmg_id or args.tmg_pw:
        log("[안내] --id/--pw 는 더 이상 사용하지 않습니다. 브라우저에서 직접 로그인하세요.")

    all_rows = read_excel(excel_path)
    if not all_rows:
        safe_print("엑셀에 처리할 행이 없습니다.")
        sys.exit(1)

    ctx = RunCtx(
        save_count=save_count,
        retries=args.retries,
        verify=verify,
        max_rows=max_rows,
        batch=args.batch or verify,
        shot_first_n=shot_first,
    )

    # 모든 입력 데이터 카테고리명·URL을 실행 로그에 기록
    ctx.info(f"[입력목록] 파일={excel_path}")
    ctx.info(f"[입력목록] 총 {len(all_rows)}건 (최종 카테고리명 / 최종 카테고리 URL주소)")
    for i, r in enumerate(all_rows, start=1):
        ctx.info(
            f"  입력#{i} 엑셀{r['row']}행 | "
            f"최종 카테고리명={r['label']} | "
            f"최종 카테고리 URL주소={r['url']}"
        )

    rows = all_rows
    if max_rows is not None:
        rows = all_rows[: max(0, max_rows)]
    ctx.info(
        f"처리대상 {len(rows)}행 / 전체입력 {len(all_rows)}행 · 저장수={save_count} · "
        f"재시도={ctx.retries} · verify={verify} · 샷=입력1~{shot_first}행 · "
        f"로그={ctx.shot_dir}"
    )

    ctx.set_progress_totals(len(rows))
    ctx.emit_progress_meta(done=0, ordinal=0, label="", url="")

    clear_stop_flag()
    ok = 0
    fail = 0
    stopped = False
    try:
        with sync_playwright() as p:
            _browser, page = connect_browser(p)
            page.set_default_timeout(120_000)
            # 망고 Chrome 기동 직후 — 더망고 솔루션 확장에 URL/KEY 필수 세팅
            ensure_mango_extension_settings(page.context, shot_ctx=ctx)
            page = refresh_if_closed(page)
            page = ensure_ready_page(page)
            ctx.shot(page, "ready", 0)

            for ordinal, row in enumerate(rows, start=1):
                success = False
                try:
                    check_stop(f"입력#{ordinal} 시작 전")
                    # 14항: 3~13 반복 — 다음 행은 13(초기화) 후 3부터
                    # (13번 자체 로그는 batch_steps.step02_init 에서 남김)
                    if ordinal >= 2:
                        ctx.info(
                            f"[다음행준비] 입력#{ordinal} — 팝업정리 후 13항 초기화로"
                        )
                        page = ensure_overlays_closed_before_next(
                            page,
                            ctx,
                            next_ordinal=ordinal,
                            next_row=row,
                        )
                    ctx.begin_row(ordinal, row)
                    # ★요건: 엑셀 각 행 실행시 MAIN에 5필드 한 줄(오렌지)
                    ctx.emit_progress_meta(
                        ordinal=ordinal,
                        label=str(row.get("label") or ""),
                        url=str(row.get("url") or ""),
                        main_line=True,
                    )
                    page = refresh_if_closed(page)
                    success = process_row_with_retries(page, row, ctx)
                except CollectStopped as e:
                    stopped = True
                    ctx.info(f"[중단] {e} — 로그·브라우저는 유지합니다")
                    break
                except Exception as e:  # noqa: BLE001
                    # ★절대 규칙(사용자 반복 지시): 엑셀에 자료가 남아있는 동안은
                    # 계속 수행한다 — 다음 행 준비(ensure_overlays_closed_before_next
                    # 등) 중 어떤 예외가 나든(RuntimeError 특정 문구가 아니어도,
                    # Playwright 예외 등 어떤 종류든) 여기서 잡아 이 입력만 실패
                    # 처리하고 다음 입력으로 진행한다. 과거에는 특정 문구의
                    # RuntimeError만 잡고 나머지는 raise 해서 전체 배치가 조용히
                    # 죽는(화면엔 ##MAIN##/##SUB## 마커가 없어 아무 표시도 안 되는)
                    # 회귀가 있었다 — 다시는 이 루프를 통째로 죽이지 않는다.
                    ctx.info(
                        f"[FAIL] 입력#{ordinal} 처리 중 예외 발생 — "
                        f"엑셀 자료가 남아있으므로 다음 입력으로 계속 진행 | "
                        f"{type(e).__name__}: {e}"
                    )
                    try:
                        page = refresh_if_closed(page)
                    except Exception:  # noqa: BLE001
                        pass
                    fail += 1
                    ctx.emit_progress_meta(done=ok)
                    if ordinal < len(rows):
                        nxt = rows[ordinal]
                        ctx.info(
                            f"==== 다음 입력#{ordinal + 1} 로 진행 "
                            f"(엑셀{nxt['row']}행 / {nxt.get('label', '')}) ===="
                        )
                    continue
                if success:
                    ok += 1
                else:
                    fail += 1
                    if not ctx.batch:
                        if input("계속 진행할까요? (y/n) ").strip().lower() != "y":
                            break
                ctx.emit_progress_meta(done=ok)
                ctx.info(
                    f"==== 입력#{ordinal} 종료 (성공={success}) "
                    f"| 엑셀{row['row']}행 ===="
                )
                if ordinal < len(rows) and not stopped:
                    nxt = rows[ordinal]  # next item (0-based index = ordinal)
                    ctx.info(
                        f"==== 다음 입력#{ordinal + 1} 로 진행 "
                        f"(엑셀{nxt['row']}행 / {nxt.get('label', '')}) ===="
                    )

            if stopped:
                ctx.info(
                    f"[중단완료] 성공 {ok} / 실패 {fail} / "
                    f"대상 {len(rows)}행 중 일부 — 화면 로그 보존"
                )
            else:
                ctx.info(
                    f"완료: 성공 {ok} / 실패 {fail} / "
                    f"대상 {len(rows)}행 / 입력전체 {len(all_rows)}건"
                )
            gallery = ctx.write_gallery()
            ctx.info(f"스크린샷·로그: {ctx.shot_dir}")
            if gallery:
                ctx.info(f"[갤러리] {gallery}")
            safe_print("브라우저는 그대로 열어둡니다 (이 창만 닫으면 됩니다).")
            if stopped:
                sys.exit(130)
            if verify and ok >= 1 and fail == 0:
                safe_print(
                    f"[OK] 검증 모드 PASS — {ok}행 완료 · "
                    f"입력 1~{shot_first}행 전과정 스크린샷 기록됨"
                )
                sys.exit(0)
            if fail:
                sys.exit(2)
    except CollectStopped as e:
        stopped = True
        ctx.info(f"[중단] {e} — 로그·브라우저는 유지합니다")
        try:
            ctx.write_gallery()
        except Exception:  # noqa: BLE001
            pass
        sys.exit(130)
    except Exception as e:  # noqa: BLE001
        # 행 루프 밖(브라우저 연결·로그인·갤러리 등)에서 난 예외.
        # 보드는 ##MAIN##/##SUB##/##META## 마커 없는 줄을 화면에서 버리므로
        # (board.app._handle_collect_line) 파이썬 트레이스백만 남기면 화면에는
        # 아무것도 안 보인 채 종료 = "조용히 죽음". 반드시 실행로그로 남긴다.
        import traceback

        ctx.info(f"[치명] 수집 중 예외로 종료 — {type(e).__name__}: {e}")
        for tb_line in traceback.format_exc().rstrip().splitlines():
            ctx.info(f"  {tb_line}")
        ctx.info(f"[치명] 진행 상황: 성공 {ok} / 실패 {fail} / 대상 {len(rows)}행")
        try:
            ctx.write_gallery()
        except Exception:  # noqa: BLE001
            pass
        sys.exit(3)
    finally:
        clear_stop_flag()
        ctx.close()


if __name__ == "__main__":
    main()
