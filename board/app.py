"""
망고보드 (mango board) — Mango_Helper_AI_Board
AI board 에서 망고 연동(P2·P3) 소스만 복사. AI board 원본은 삭제하지 않음.
"""

from __future__ import annotations

import importlib.util
import os
import subprocess
import sys
import threading
import time
import tkinter as tk
import webbrowser
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from types import ModuleType

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "board"))

import input_history as ih  # noqa: E402 — 사이트명·목록 URL 이력 (리스트박스용)

# 이력 저장 파일 — 프로그램별 폴더에 둔다 (.option_stop 등과 같은 방식)
P5M_SITE_HISTORY = ROOT / "P5_101_카테고리매핑_필터세부설정" / ".recent_site.json"
P5M_URL_HISTORY = ROOT / "P5_101_카테고리매핑_필터세부설정" / ".recent_url.json"
P3RST_SITE_HISTORY = ROOT / "P3_설정수정_카테고리매핑초기화" / ".recent_site.json"
P3RST_URL_HISTORY = ROOT / "P3_설정수정_카테고리매핑초기화" / ".recent_url.json"


def _load_py_module(mod_name: str, folder: str, filename: str) -> ModuleType:
    path = ROOT / folder / filename
    spec = importlib.util.spec_from_file_location(mod_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"모듈 로드 실패: {path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[mod_name] = mod
    spec.loader.exec_module(mod)
    return mod


p3_update = _load_py_module("p3_update_filters", "P3_필터_갱신", "update_filters.py")
p1_policy = _load_py_module("p1_apply_policy", "P1_필터단위_마진정책적용", "apply_policy.py")
p2_count = _load_py_module("p2_update_product_count", "P2_필터단위_상품수변경", "update_product_count.py")
p3_option = _load_py_module(
    "p3_update_collect_option", "P3_필터단위_수집조건수정", "update_collect_option.py"
)
p5_category = _load_py_module(
    "p5_extract_categories", "P5_카테고리_엑셀추출", "extract_categories.py"
)
p5_mapping = _load_py_module(
    "p5_101_map_categories", "P5_101_카테고리매핑_필터세부설정", "map_categories.py"
)
p3_fitcl = _load_py_module("p3_fitcl_detail", "P3_핏클상세페이지", "fitcl_detail.py")
p3_reset_mapping = _load_py_module(
    "p3_reset_category_mapping",
    "P3_설정수정_카테고리매핑초기화",
    "reset_category_mapping.py",
)

from library import (  # noqa: E402
    add_paths,
    default_roots,
    entries_annotated,
    is_in_library,
    load,
    read_category_url_rows,
    remove_path,
    search_xlsx,
    set_selected,
)
from log_protocol import (  # noqa: E402
    META_FIELDS,
    META_INTERNAL_FIELDS,
    format_meta_line,
    main_tag_p3,
    parse_line,
    split_red,
    step_label_p3,
    step_tag,
    step_tag_p3,
    strip_timestamp,
    sub_time_range,
)
from shot_viewer import latest_p3_shot_dir, latest_shot_dir, open_shot_viewer  # noqa: E402
from self_update import (  # noqa: E402
    latest_open_pr_url,
    launch_external_updater,
    local_version,
)
from terms import APP_NAME, APP_SHORT_EN, APP_SHORT_KO  # noqa: E402

import re  # noqa: E402


def _read_version() -> str:
    try:
        text = (ROOT / "VERSION.txt").read_text(encoding="utf-8")
        m = re.search(r"(?:버전|version)\s*([0-9]+(?:\.[0-9]+)+)", text, re.I)
        if not m:
            m = re.search(r"([0-9]+\.[0-9]+\.[0-9]+)", text)
        if m:
            return m.group(1)
    except OSError:
        pass
    return "?"


VERSION = _read_version()


class BoardApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"{APP_SHORT_KO} ({APP_SHORT_EN})  v{VERSION}")
        self.geometry("1280x900")
        self.minsize(1024, 760)
        self.configure(bg="#1a4d5c")

        self._p1_proc: subprocess.Popen | None = None
        self._p2_count_proc: subprocess.Popen | None = None
        self._p2_proc: subprocess.Popen | None = None
        self._p3_proc: subprocess.Popen | None = None
        self._p3_option_proc: subprocess.Popen | None = None
        self._p3_reset_proc: subprocess.Popen | None = None
        self._p3_option_reload_busy = False
        self._p5_proc: subprocess.Popen | None = None
        self._p5_101_proc: subprocess.Popen | None = None
        self._p3_fitcl_proc: subprocess.Popen | None = None
        self._last_shot_dir: Path | None = None
        self._merge_update_busy = False
        self._build()
        self._show("p2")
        self._refresh_p2_list()
        self._refresh_p3_list()

    def _build(self) -> None:
        head = tk.Frame(self, bg="#164a59", pady=10)
        head.pack(fill="x")
        tk.Label(
            head,
            text=APP_SHORT_KO,
            fg="white",
            bg="#164a59",
            font=("Malgun Gothic", 14, "bold"),
        ).pack()
        tk.Label(
            head,
            text=f"{APP_SHORT_EN} · P1 마진 · P2 상품수 · P2 수집 · P3 필터 · P3 핏클",
            fg="#cbd5e1",
            bg="#164a59",
            font=("Malgun Gothic", 9),
        ).pack()

        body = tk.Frame(self, bg="#1a4d5c")
        body.pack(fill="both", expand=True, padx=8, pady=8)

        side = tk.Frame(body, bg="#d9d9d9", width=180)
        side.pack(side="left", fill="y", padx=(0, 8))
        side.pack_propagate(False)

        tk.Label(
            side,
            text=f"v{VERSION}\n프로그램",
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            pady=8,
        ).pack(fill="x")

        self.btn_p1 = tk.Button(
            side,
            text="P1_필터단위\n마진정책적용",
            command=lambda: self._show("p1"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p1.pack(fill="x", padx=6, pady=6)

        self.btn_p2_count = tk.Button(
            side,
            text="P2_필터단위\n상품수변경",
            command=lambda: self._show("p2_count"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p2_count.pack(fill="x", padx=6, pady=6)

        self.btn_p2 = tk.Button(
            side,
            text="P2\n더망고 대량수집",
            command=lambda: self._show("p2"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p2.pack(fill="x", padx=6, pady=6)

        self.btn_p3 = tk.Button(
            side,
            text="P3_필터_갱신\n수집건수 갱신",
            command=lambda: self._show("p3"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p3.pack(fill="x", padx=6, pady=6)

        self.btn_p3_option = tk.Button(
            side,
            text="P3_필터단위\n수집조건수정",
            command=lambda: self._show("p3_option"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p3_option.pack(fill="x", padx=6, pady=6)

        self.btn_p3_reset = tk.Button(
            side,
            text="P3_설정수정\n카테고리매핑초기화",
            command=lambda: self._show("p3_reset"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p3_reset.pack(fill="x", padx=6, pady=6)

        self.btn_p5 = tk.Button(
            side,
            text="P5_카테고리\n엑셀추출",
            command=lambda: self._show("p5"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p5.pack(fill="x", padx=6, pady=6)

        self.btn_p5_101 = tk.Button(
            side,
            text="P5_101_카테고리\n매핑",
            command=lambda: self._show("p5_101"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p5_101.pack(fill="x", padx=6, pady=6)

        self.btn_p3_fitcl = tk.Button(
            side,
            text="P3_핏클\n상세페이지",
            command=lambda: self._show("p3_fitcl"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p3_fitcl.pack(fill="x", padx=6, pady=6)

        side_bottom = tk.Frame(side, bg="#d9d9d9")
        side_bottom.pack(side="bottom", fill="x", padx=6, pady=(4, 10))
        tk.Label(
            side_bottom,
            text="종료 후 강제 버전갱신",
            bg="#d9d9d9",
            fg="#475569",
            font=("Malgun Gothic", 8),
        ).pack(fill="x", pady=(0, 4))
        self.btn_merge_update = tk.Button(
            side_bottom,
            text="머지반영\n업데이트",
            command=self._run_merge_update,
            bg="#0f766e",
            fg="white",
            activebackground="#0d9488",
            activeforeground="white",
            font=("Malgun Gothic", 9, "bold"),
            relief="raised",
            pady=12,
            cursor="hand2",
        )
        self.btn_merge_update.pack(fill="x")
        self.lbl_update_status = tk.Label(
            side_bottom,
            text=f"현재 v{VERSION}",
            bg="#d9d9d9",
            fg="#64748b",
            font=("Malgun Gothic", 7),
            wraplength=160,
            justify="left",
        )
        self.lbl_update_status.pack(fill="x", pady=(4, 0))

        self.main = tk.Frame(body, bg="#f1f5f9")
        self.main.pack(side="left", fill="both", expand=True)

        self.frame_p1 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p2_count = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p2 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p3 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p3_option = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p3_reset = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p5 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p5_101 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p3_fitcl = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self._build_p1(self.frame_p1)
        self._build_p2_count(self.frame_p2_count)
        self._build_p2(self.frame_p2)
        self._build_p3(self.frame_p3)
        self._build_p3_option(self.frame_p3_option)
        self._build_p3_reset(self.frame_p3_reset)
        self._build_p5(self.frame_p5)
        self._build_p5_101(self.frame_p5_101)
        self._build_p3_fitcl(self.frame_p3_fitcl)

    def _show(self, which: str) -> None:
        self.frame_p1.pack_forget()
        self.frame_p2_count.pack_forget()
        self.frame_p2.pack_forget()
        self.frame_p3.pack_forget()
        self.frame_p3_option.pack_forget()
        self.frame_p3_reset.pack_forget()
        self.frame_p5.pack_forget()
        self.frame_p5_101.pack_forget()
        self.frame_p3_fitcl.pack_forget()
        self.btn_p1.configure(bg="#ececec")
        self.btn_p2_count.configure(bg="#ececec")
        self.btn_p2.configure(bg="#ececec")
        self.btn_p3.configure(bg="#ececec")
        self.btn_p3_option.configure(bg="#ececec")
        self.btn_p3_reset.configure(bg="#ececec")
        self.btn_p5.configure(bg="#ececec")
        self.btn_p5_101.configure(bg="#ececec")
        self.btn_p3_fitcl.configure(bg="#ececec")
        if which == "p1":
            self.frame_p1.pack(fill="both", expand=True)
            self.btn_p1.configure(bg="#dbeafe")
        elif which == "p2_count":
            self.frame_p2_count.pack(fill="both", expand=True)
            self.btn_p2_count.configure(bg="#dbeafe")
        elif which == "p3":
            self.frame_p3.pack(fill="both", expand=True)
            self.btn_p3.configure(bg="#dbeafe")
        elif which == "p3_option":
            self.frame_p3_option.pack(fill="both", expand=True)
            self.btn_p3_option.configure(bg="#dbeafe")
        elif which == "p3_reset":
            self.frame_p3_reset.pack(fill="both", expand=True)
            self.btn_p3_reset.configure(bg="#dbeafe")
        elif which == "p5":
            self.frame_p5.pack(fill="both", expand=True)
            self.btn_p5.configure(bg="#dbeafe")
        elif which == "p5_101":
            self.frame_p5_101.pack(fill="both", expand=True)
            self.btn_p5_101.configure(bg="#dbeafe")
        elif which == "p3_fitcl":
            self.frame_p3_fitcl.pack(fill="both", expand=True)
            self.btn_p3_fitcl.configure(bg="#dbeafe")
        else:
            self.frame_p2.pack(fill="both", expand=True)
            self.btn_p2.configure(bg="#dbeafe")

    # ── 좌측 하단: 머지반영 업데이트 ───────────────────
    def _run_merge_update(self) -> None:
        """망고보드를 종료한 뒤 외부 스크립트로 GitHub main 강제 반영·재시작.

        (실행 중 pull 하면 Windows 파일 잠금으로 버전이 안 바뀌는 문제 방지)
        """
        if getattr(self, "_merge_update_busy", False):
            messagebox.showinfo("안내", "이미 업데이트를 진행 중입니다.")
            return

        pr_url = ""
        try:
            pr_url = latest_open_pr_url()
        except Exception:
            pr_url = f"https://github.com/waterstar21g-png/Mango_Helper_AI_Board/pulls"

        cur = local_version(ROOT) or VERSION
        msg = (
            "망고보드를 종료한 뒤 GitHub main 을 강제 반영하고 재시작합니다.\n\n"
            f"현재 버전: v{cur}\n\n"
            "아직 PR 머지 전이면 아래 머지 URL에서 먼저 머지하세요.\n"
            f"{pr_url}\n\n"
            "계속할까요?"
        )
        if not messagebox.askyesno("머지반영 업데이트", msg, parent=self):
            return

        try:
            if pr_url.startswith("http"):
                webbrowser.open(pr_url)
        except Exception:
            pass

        self._merge_update_busy = True
        self.btn_merge_update.configure(state="disabled", text="종료 후 갱신…")
        self.lbl_update_status.configure(
            text="망고보드 종료 → 강제 버전갱신 → 재시작",
            fg="#0f172a",
        )

        # 실행 중 작업 중단
        try:
            if self._p1_proc and self._p1_proc.poll() is None:
                self._p1_stop_flag().write_text("stop\n", encoding="utf-8")
                self._p1_proc.terminate()
        except Exception:
            pass
        try:
            if self._p2_count_proc and self._p2_count_proc.poll() is None:
                self._p2_count_stop_flag().write_text("stop\n", encoding="utf-8")
                self._p2_count_proc.terminate()
        except Exception:
            pass
        try:
            if self._p2_proc and self._p2_proc.poll() is None:
                self._stop_flag_path().write_text("stop\n", encoding="utf-8")
                self._p2_proc.terminate()
        except Exception:
            pass
        try:
            if self._p3_proc and self._p3_proc.poll() is None:
                self._p3_stop_flag().write_text("stop\n", encoding="utf-8")
                self._p3_proc.terminate()
        except Exception:
            pass
        try:
            if self._p3_fitcl_proc and self._p3_fitcl_proc.poll() is None:
                self._p3_fitcl_stop_flag().write_text("stop\n", encoding="utf-8")
                self._p3_fitcl_proc.terminate()
        except Exception:
            pass

        ok, detail = launch_external_updater(ROOT, wait_pid=os.getpid())
        if not ok:
            self._merge_update_busy = False
            self.btn_merge_update.configure(state="normal", text="머지반영\n업데이트")
            self.lbl_update_status.configure(text="업데이터 실행 실패", fg="#b91c1c")
            messagebox.showerror(
                "업데이트 실패",
                "외부 버전갱신 실행에 실패했습니다.\n\n"
                f"{detail}\n\n"
                "바탕화면 버전갱신 아이콘을 사용하세요.\n"
                f"머지 URL:\n{pr_url}",
                parent=self,
            )
            return

        messagebox.showinfo(
            "버전갱신 시작",
            "망고보드를 종료합니다.\n\n"
            "이어서 자동으로:\n"
            "1) GitHub main 강제 반영\n"
            "2) 망고보드 재시작\n\n"
            f"(실패 시 바탕화면 버전갱신 아이콘)\n"
            f"머지 URL:\n{pr_url}",
            parent=self,
        )
        try:
            self.destroy()
        except Exception:
            pass
        sys.exit(0)

    # ── P1_필터단위_마진정책적용 ─────────────────────────────────────
    def _build_p1(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P1_필터단위_마진정책적용 — 정책명 입력 → 필터단위 마진정책 목록(체크 행) 순차 갱신",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        form = tk.LabelFrame(parent, text="입력", bg="#ffffff", padx=8, pady=6)
        form.pack(fill="x")

        r1 = tk.Frame(form, bg="#ffffff")
        r1.pack(fill="x", pady=4)
        tk.Label(r1, text="정책명", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p1_policy = tk.StringVar()
        tk.Entry(r1, textvariable=self.var_p1_policy, width=40).pack(
            side="left", fill="x", expand=True
        )

        r2 = tk.Frame(form, bg="#ffffff")
        r2.pack(fill="x", pady=4)
        tk.Label(r2, text="망고 URL", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p1_mango_url = tk.StringVar(value=p1_policy.DEFAULT_MANGO_URL)
        tk.Entry(r2, textvariable=self.var_p1_mango_url).pack(
            side="left", fill="x", expand=True
        )
        tk.Label(
            form,
            text="(비우면 Chrome에 열린 필터단위 마진정책 목록 화면 사용)",
            bg="#ffffff",
            fg="#64748b",
            font=("Malgun Gothic", 8),
            anchor="w",
        ).pack(fill="x")

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=8)
        tk.Button(
            actions,
            text="작업시작",
            command=self._run_p1,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left")
        tk.Button(
            actions,
            text="작업중단",
            command=self._stop_p1,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=6)

        log_frame = tk.LabelFrame(parent, text="실행 로그", bg="#ffffff", padx=6, pady=4)
        log_frame.pack(fill="both", expand=True)
        self.p1_log = tk.Text(
            log_frame,
            height=18,
            font=("Consolas", 9),
            wrap="word",
            bg="#0f172a",
            fg="#e2e8f0",
        )
        p1_sb = tk.Scrollbar(log_frame, command=self.p1_log.yview)
        self.p1_log.configure(yscrollcommand=p1_sb.set)
        self.p1_log.pack(side="left", fill="both", expand=True)
        p1_sb.pack(side="right", fill="y")

        self.p1_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p1_status.pack(fill="x", pady=4)

    def _p1_stop_flag(self) -> Path:
        return ROOT / "P1_필터단위_마진정책적용" / ".policy_stop"

    def _append_p1_log(self, line: str) -> None:
        text = (line or "").strip()
        if text.startswith("##MAIN##"):
            text = text[8:]
        self.p1_log.insert("end", text + "\n")
        self.p1_log.see("end")

    def _run_p1(self) -> None:
        policy = self.var_p1_policy.get().strip()
        if not policy:
            messagebox.showinfo("안내", "정책명을 입력하세요.")
            return
        if self._p1_proc and self._p1_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return

        apply_py = ROOT / "P1_필터단위_마진정책적용" / "apply_policy.py"
        if not apply_py.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{apply_py}")
            return

        try:
            self._p1_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        mango = self.var_p1_mango_url.get().strip()
        args = [sys.executable, str(apply_py), "--policy-name", policy]
        if mango:
            args.extend(["--mango-url", mango])

        self.p1_log.delete("1.0", "end")
        self.p1_status.configure(text=f"작업 시작 — 정책명: {policy}", fg="#15803d")

        creationflags = 0
        if os.name == "nt":
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        try:
            self._p1_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P1_필터단위_마진정책적용"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p1_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(target=self._watch_p1_proc, args=(self._p1_proc,), daemon=True).start()

    def _stop_p1(self) -> None:
        proc = self._p1_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p1_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p1_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p1_status.configure(text="작업중단 요청 중…", fg="#b45309")
        try:
            proc.terminate()
        except Exception:
            pass

    def _watch_p1_proc(self, proc: subprocess.Popen) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self.after(0, lambda t=text: self._append_p1_log(t))
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self.after(0, lambda t=text: self._append_p1_log(t))
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p1_status.configure(text=f"로그 수신 오류: {e}", fg="#b91c1c"),
            )
        code = proc.wait()
        if code == 0:
            self.after(0, lambda: self.p1_status.configure(text="작업 완료", fg="#15803d"))
        else:
            self.after(
                0,
                lambda: self.p1_status.configure(text=f"종료 (exit={code})", fg="#b91c1c"),
            )

    # ── P2_필터단위_상품수변경 ─────────────────────────────────────
    def _build_p2_count(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P2_필터단위_상품수변경 — 적용상품수 입력 → 필터 목록 순차 갱신",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        form = tk.LabelFrame(parent, text="입력", bg="#ffffff", padx=8, pady=6)
        form.pack(fill="x")

        r1 = tk.Frame(form, bg="#ffffff")
        r1.pack(fill="x", pady=4)
        tk.Label(r1, text="적용상품수", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p2_count_apply = tk.StringVar(value="50")
        tk.Entry(r1, textvariable=self.var_p2_count_apply, width=20).pack(
            side="left", fill="x", expand=True
        )

        r2 = tk.Frame(form, bg="#ffffff")
        r2.pack(fill="x", pady=4)
        tk.Label(r2, text="망고 URL", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p2_count_mango_url = tk.StringVar(value=p3_update.DEFAULT_MANGO_URL)
        tk.Entry(r2, textvariable=self.var_p2_count_mango_url).pack(
            side="left", fill="x", expand=True
        )
        tk.Label(
            form,
            text="(비우면 Chrome에 열린 필터 목록 화면 사용)",
            bg="#ffffff",
            fg="#64748b",
            font=("Malgun Gothic", 8),
            anchor="w",
        ).pack(fill="x")

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=8)
        tk.Button(
            actions,
            text="작업시작",
            command=self._run_p2_count,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left")
        tk.Button(
            actions,
            text="작업중단",
            command=self._stop_p2_count,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=6)

        log_frame = tk.LabelFrame(parent, text="실행 로그", bg="#ffffff", padx=6, pady=4)
        log_frame.pack(fill="both", expand=True)
        self.p2_count_log = tk.Text(
            log_frame,
            height=18,
            font=("Consolas", 9),
            wrap="word",
            bg="#0f172a",
            fg="#e2e8f0",
        )
        p2_count_sb = tk.Scrollbar(log_frame, command=self.p2_count_log.yview)
        self.p2_count_log.configure(yscrollcommand=p2_count_sb.set)
        self.p2_count_log.pack(side="left", fill="both", expand=True)
        p2_count_sb.pack(side="right", fill="y")

        self.p2_count_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p2_count_status.pack(fill="x", pady=4)

    def _p2_count_stop_flag(self) -> Path:
        return ROOT / "P2_필터단위_상품수변경" / ".count_stop"

    def _append_p2_count_log(self, line: str) -> None:
        text = (line or "").strip()
        if text.startswith("##MAIN##"):
            text = text[8:]
        self.p2_count_log.insert("end", text + "\n")
        self.p2_count_log.see("end")

    def _run_p2_count(self) -> None:
        apply_count = self.var_p2_count_apply.get().strip()
        if not apply_count:
            messagebox.showinfo("안내", "적용상품수를 입력하세요.")
            return
        if not apply_count.isdigit() or int(apply_count) < 0:
            messagebox.showinfo("안내", "적용상품수는 0 이상의 숫자여야 합니다.")
            return
        if self._p2_count_proc and self._p2_count_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return

        update_py = ROOT / "P2_필터단위_상품수변경" / "update_product_count.py"
        if not update_py.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{update_py}")
            return

        try:
            self._p2_count_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        mango = self.var_p2_count_mango_url.get().strip()
        args = [sys.executable, str(update_py), "--apply-count", apply_count]
        if mango:
            args.extend(["--mango-url", mango])

        self.p2_count_log.delete("1.0", "end")
        self.p2_count_status.configure(
            text=f"작업 시작 — 적용상품수: {apply_count}", fg="#15803d"
        )

        creationflags = 0
        if os.name == "nt":
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        try:
            self._p2_count_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P2_필터단위_상품수변경"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p2_count_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p2_count_proc, args=(self._p2_count_proc,), daemon=True
        ).start()

    def _stop_p2_count(self) -> None:
        proc = self._p2_count_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p2_count_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p2_count_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p2_count_status.configure(text="작업중단 요청 중…", fg="#b45309")
        try:
            proc.terminate()
        except Exception:
            pass

    def _watch_p2_count_proc(self, proc: subprocess.Popen) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self.after(0, lambda t=text: self._append_p2_count_log(t))
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self.after(0, lambda t=text: self._append_p2_count_log(t))
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p2_count_status.configure(
                    text=f"로그 수신 오류: {e}", fg="#b91c1c"
                ),
            )
        code = proc.wait()
        if code == 0:
            self.after(
                0, lambda: self.p2_count_status.configure(text="작업 완료", fg="#15803d")
            )
        else:
            self.after(
                0,
                lambda: self.p2_count_status.configure(
                    text=f"종료 (exit={code})", fg="#b91c1c"
                ),
            )

    # ── P3_필터단위_수집조건수정 ───────────────────────────────────
    def _build_p3_option(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P3_필터단위_수집조건수정 — 수집사이트·번역옵션 선택 → 필터 목록 순차 적용",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        form = tk.LabelFrame(parent, text="입력", bg="#ffffff", padx=8, pady=6)
        form.pack(fill="x")

        r0 = tk.Frame(form, bg="#ffffff")
        r0.pack(fill="x", pady=4)
        tk.Label(r0, text="수집사이트", width=10, anchor="nw", bg="#ffffff").pack(
            side="left", anchor="n"
        )
        site_wrap = tk.Frame(r0, bg="#ffffff")
        site_wrap.pack(side="left", fill="both", expand=True)
        self.p3_site_list = tk.Listbox(
            site_wrap,
            height=6,
            exportselection=False,
            font=("Malgun Gothic", 9),
            activestyle="dotbox",
        )
        site_sb = tk.Scrollbar(site_wrap, orient="vertical", command=self.p3_site_list.yview)
        self.p3_site_list.configure(yscrollcommand=site_sb.set)
        self.p3_site_list.pack(side="left", fill="both", expand=True)
        site_sb.pack(side="right", fill="y")
        self.p3_site_list.bind("<<ListboxSelect>>", lambda _e: self._on_p3_option_pick())

        r1 = tk.Frame(form, bg="#ffffff")
        r1.pack(fill="x", pady=4)
        tk.Label(r1, text="번역옵션", width=10, anchor="nw", bg="#ffffff").pack(
            side="left", anchor="n"
        )
        list_wrap = tk.Frame(r1, bg="#ffffff")
        list_wrap.pack(side="left", fill="both", expand=True)
        self.p3_option_list = tk.Listbox(
            list_wrap,
            height=6,
            exportselection=False,
            font=("Malgun Gothic", 9),
            activestyle="dotbox",
        )
        opt_sb = tk.Scrollbar(list_wrap, orient="vertical", command=self.p3_option_list.yview)
        self.p3_option_list.configure(yscrollcommand=opt_sb.set)
        self.p3_option_list.pack(side="left", fill="both", expand=True)
        opt_sb.pack(side="right", fill="y")
        self.p3_option_list.bind(
            "<<ListboxSelect>>", lambda _e: self._on_p3_option_pick()
        )

        side_btns = tk.Frame(r1, bg="#ffffff")
        side_btns.pack(side="left", padx=6, anchor="n")
        self.btn_p3_option_reload = tk.Button(
            side_btns,
            text="망고에서\n옵션 읽기",
            command=self._refresh_p3_option_choices,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 8, "bold"),
            padx=6,
            pady=4,
        )
        self.btn_p3_option_reload.pack(fill="x")

        self.lbl_p3_option_pick = tk.Label(
            form,
            text="선택: (없음)",
            bg="#ffffff",
            fg="#0f172a",
            font=("Malgun Gothic", 9, "bold"),
            anchor="w",
        )
        self.lbl_p3_option_pick.pack(fill="x", pady=(2, 0))

        r2 = tk.Frame(form, bg="#ffffff")
        r2.pack(fill="x", pady=4)
        tk.Label(r2, text="망고 URL", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p3_option_mango_url = tk.StringVar(value=p3_option.DEFAULT_LIST_URL)
        tk.Entry(r2, textvariable=self.var_p3_option_mango_url).pack(
            side="left", fill="x", expand=True
        )
        tk.Label(
            form,
            text="(비우면 Chrome에 열린 필터 목록 화면 사용 · 목록이 실제와 다르면 [망고에서 옵션 읽기])",
            bg="#ffffff",
            fg="#64748b",
            font=("Malgun Gothic", 8),
            anchor="w",
        ).pack(fill="x")

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=8)
        tk.Button(
            actions,
            text="작업시작",
            command=self._run_p3_option,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left")
        tk.Button(
            actions,
            text="작업중단",
            command=self._stop_p3_option,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=6)

        log_frame = tk.LabelFrame(parent, text="실행 로그", bg="#ffffff", padx=6, pady=4)
        log_frame.pack(fill="both", expand=True)
        self.p3_option_log = tk.Text(
            log_frame,
            height=16,
            font=("Consolas", 9),
            wrap="word",
            bg="#0f172a",
            fg="#e2e8f0",
        )
        p3_option_sb = tk.Scrollbar(log_frame, command=self.p3_option_log.yview)
        self.p3_option_log.configure(yscrollcommand=p3_option_sb.set)
        self.p3_option_log.pack(side="left", fill="both", expand=True)
        p3_option_sb.pack(side="right", fill="y")

        self.p3_option_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p3_option_status.pack(fill="x", pady=4)

        self._fill_listbox(self.p3_site_list, p3_option.load_cached_sites())
        self._fill_listbox(self.p3_option_list, p3_option.load_cached_options())

    def _fill_listbox(self, box: tk.Listbox, values: list[str]) -> None:
        box.delete(0, "end")
        for v in values:
            box.insert("end", v)
        if values:
            box.selection_clear(0, "end")
            box.selection_set(0)
            box.see(0)
        self._on_p3_option_pick()

    def _selected_translate_option(self) -> str:
        sel = self.p3_option_list.curselection()
        if not sel:
            return ""
        return str(self.p3_option_list.get(sel[0])).strip()

    def _selected_collect_site(self) -> str:
        sel = self.p3_site_list.curselection()
        if not sel:
            return ""
        return str(self.p3_site_list.get(sel[0])).strip()

    def _on_p3_option_pick(self) -> None:
        site = self._selected_collect_site() or p3_option.SITE_ALL_LABEL
        picked = self._selected_translate_option()
        self.lbl_p3_option_pick.configure(
            text=f"선택: 수집사이트={site} · 번역옵션={picked or '(없음)'}"
        )

    def _p3_option_stop_flag(self) -> Path:
        return ROOT / "P3_필터단위_수집조건수정" / ".option_stop"

    def _append_p3_option_log(self, line: str) -> None:
        text = (line or "").strip()
        if text.startswith("##MAIN##"):
            text = text[8:]
        self.p3_option_log.insert("end", text + "\n")
        self.p3_option_log.see("end")

    def _p3_option_script(self) -> Path:
        return ROOT / "P3_필터단위_수집조건수정" / "update_collect_option.py"

    def _p3_option_popen(self, extra_args: list[str], *, capture: bool) -> subprocess.Popen:
        args = [sys.executable, str(self._p3_option_script()), *extra_args]
        mango = self.var_p3_option_mango_url.get().strip()
        if mango:
            args.extend(["--mango-url", mango])
        creationflags = 0
        if os.name == "nt":
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        return subprocess.Popen(
            args,
            cwd=str(ROOT / "P3_필터단위_수집조건수정"),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=False,
            bufsize=0,
            env=env,
            creationflags=creationflags,
        )

    def _refresh_p3_option_choices(self) -> None:
        """망고 수집조건수정 화면에서 번역옵션 목록을 읽어 리스트박스를 채운다."""
        if getattr(self, "_p3_option_reload_busy", False):
            messagebox.showinfo("안내", "옵션을 읽는 중입니다.")
            return
        if not self._p3_option_script().is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{self._p3_option_script()}")
            return

        self._p3_option_reload_busy = True
        self.btn_p3_option_reload.configure(state="disabled")
        self.p3_option_status.configure(text="망고에서 번역옵션 읽는 중…", fg="#b45309")

        def worker() -> None:
            text = ""
            try:
                proc = self._p3_option_popen(["--list-options"], capture=True)
                raw = proc.stdout.read() if proc.stdout else b""
                proc.wait()
                text = self._decode_log_bytes(raw)
            except Exception as e:  # noqa: BLE001
                text = f"옵션 읽기 실패: {e}"

            options = p3_option.parse_option_lines(text)
            sites = p3_option.parse_site_lines(text)
            log_text = text

            def done() -> None:
                self._p3_option_reload_busy = False
                self.btn_p3_option_reload.configure(state="normal")
                skip = (p3_option.OPTION_LINE_PREFIX, p3_option.SITE_LINE_PREFIX)
                for line in log_text.splitlines():
                    if line.strip() and not line.startswith(skip):
                        self._append_p3_option_log(line)
                if sites:
                    self._fill_listbox(self.p3_site_list, sites)
                if options:
                    self._fill_listbox(self.p3_option_list, options)
                if options or sites:
                    self.p3_option_status.configure(
                        text=f"수집사이트 {len(sites)}개 · 번역옵션 {len(options)}개 읽음",
                        fg="#15803d",
                    )
                else:
                    self.p3_option_status.configure(
                        text="목록을 읽지 못했습니다 (망고 로그인·필터 목록 화면 확인)",
                        fg="#b91c1c",
                    )

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _run_p3_option(self) -> None:
        option = self._selected_translate_option()
        if not option:
            messagebox.showinfo("안내", "번역옵션을 리스트에서 선택하세요.")
            return
        if self._p3_option_proc and self._p3_option_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return
        if not self._p3_option_script().is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{self._p3_option_script()}")
            return

        try:
            self._p3_option_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        site = self._selected_collect_site()
        self.p3_option_log.delete("1.0", "end")
        self.p3_option_status.configure(
            text=f"작업 시작 — 수집사이트: {site or p3_option.SITE_ALL_LABEL} · "
            f"번역옵션: {option}",
            fg="#15803d",
        )

        args = ["--translate-option", option]
        if site and not p3_option.is_all_sites(site):
            args.extend(["--collect-site", site])

        try:
            self._p3_option_proc = self._p3_option_popen(args, capture=False)
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p3_option_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p3_option_proc, args=(self._p3_option_proc,), daemon=True
        ).start()

    def _stop_p3_option(self) -> None:
        proc = self._p3_option_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p3_option_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p3_option_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p3_option_status.configure(text="작업중단 요청 중…", fg="#b45309")
        try:
            proc.terminate()
        except Exception:
            pass

    def _watch_p3_option_proc(self, proc: subprocess.Popen) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self.after(0, lambda t=text: self._append_p3_option_log(t))
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self.after(0, lambda t=text: self._append_p3_option_log(t))
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p3_option_status.configure(
                    text=f"로그 수신 오류: {e}", fg="#b91c1c"
                ),
            )
        code = proc.wait()
        if code == 0:
            self.after(
                0, lambda: self.p3_option_status.configure(text="작업 완료", fg="#15803d")
            )
        else:
            self.after(
                0,
                lambda: self.p3_option_status.configure(
                    text=f"종료 (exit={code})", fg="#b91c1c"
                ),
            )

    # ── P5_카테고리_엑셀추출 ───────────────────────────────────────
    def _build_p5(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P5_카테고리_엑셀추출 — 마켓 선택 → 전체카테고리 → 1~6단계 분류표 엑셀",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        form = tk.LabelFrame(parent, text="입력", bg="#ffffff", padx=8, pady=6)
        form.pack(fill="x")

        r0 = tk.Frame(form, bg="#ffffff")
        r0.pack(fill="x", pady=4)
        tk.Label(r0, text="마켓", width=10, anchor="nw", bg="#ffffff").pack(
            side="left", anchor="n"
        )
        wrap = tk.Frame(r0, bg="#ffffff")
        wrap.pack(side="left", fill="both", expand=True)
        self.p5_market_list = tk.Listbox(
            wrap,
            height=7,
            exportselection=False,
            font=("Malgun Gothic", 9),
            activestyle="dotbox",
        )
        sb = tk.Scrollbar(wrap, orient="vertical", command=self.p5_market_list.yview)
        self.p5_market_list.configure(yscrollcommand=sb.set)
        self.p5_market_list.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._p5_market_codes = [p5_category.ALL_MARKETS, *p5_category.MARKETS.keys()]
        self.p5_market_list.insert("end", "전체 마켓 일괄  (ALL)")
        for code in list(p5_category.MARKETS.keys()):
            self.p5_market_list.insert("end", f"{p5_category.MARKETS[code]}  ({code})")
        self.p5_market_list.selection_set(1)  # 기본: 옥션2.0

        r1 = tk.Frame(form, bg="#ffffff")
        r1.pack(fill="x", pady=4)
        tk.Label(r1, text="접근 URL", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p5_url = tk.StringVar(value=p5_category.DEFAULT_URL)
        tk.Entry(r1, textvariable=self.var_p5_url).pack(side="left", fill="x", expand=True)

        r2 = tk.Frame(form, bg="#ffffff")
        r2.pack(fill="x", pady=4)
        tk.Label(r2, text="저장 경로", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p5_out = tk.StringVar(value="")
        tk.Entry(r2, textvariable=self.var_p5_out).pack(side="left", fill="x", expand=True)
        tk.Button(r2, text="…", width=3, command=self._pick_p5_out).pack(side="left", padx=4)
        tk.Label(
            form,
            text="(비우면 P5_카테고리_엑셀추출\\output 폴더에 자동 파일명으로 저장)",
            bg="#ffffff",
            fg="#64748b",
            font=("Malgun Gothic", 8),
            anchor="w",
        ).pack(fill="x")

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=8)
        tk.Button(
            actions,
            text="추출 시작",
            command=self._run_p5,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left")
        tk.Button(
            actions,
            text="작업중단",
            command=self._stop_p5,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=6)
        tk.Button(
            actions,
            text="엑셀 열기",
            command=self._open_p5_excel,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left")

        log_frame = tk.LabelFrame(parent, text="실행 로그", bg="#ffffff", padx=6, pady=4)
        log_frame.pack(fill="both", expand=True)
        self.p5_log = tk.Text(
            log_frame,
            height=16,
            font=("Consolas", 9),
            wrap="word",
            bg="#0f172a",
            fg="#e2e8f0",
        )
        p5_sb = tk.Scrollbar(log_frame, command=self.p5_log.yview)
        self.p5_log.configure(yscrollcommand=p5_sb.set)
        self.p5_log.pack(side="left", fill="both", expand=True)
        p5_sb.pack(side="right", fill="y")

        self.p5_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p5_status.pack(fill="x", pady=4)
        self._p5_last_excel = ""

    def _selected_p5_market(self) -> str:
        sel = self.p5_market_list.curselection()
        if not sel:
            return p5_category.DEFAULT_MARKET
        return self._p5_market_codes[sel[0]]

    def _pick_p5_out(self) -> None:
        path = filedialog.asksaveasfilename(
            title="카테고리분류표 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if path:
            self.var_p5_out.set(path)

    def _p5_stop_flag(self) -> Path:
        return ROOT / "P5_카테고리_엑셀추출" / ".p5_stop"

    def _append_p5_log(self, line: str) -> None:
        text = (line or "").strip()
        if text.startswith("##MAIN##"):
            text = text[8:]
        self.p5_log.insert("end", text + "\n")
        self.p5_log.see("end")
        if text.lower().endswith(".xlsx"):
            self._p5_last_excel = text

    def _run_p5(self) -> None:
        if self._p5_proc and self._p5_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return
        script = ROOT / "P5_카테고리_엑셀추출" / "extract_categories.py"
        if not script.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{script}")
            return

        try:
            self._p5_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        market = self._selected_p5_market()
        args = [sys.executable, str(script), "--market", market]
        url = self.var_p5_url.get().strip()
        if url:
            args.extend(["--url", url])
        out = self.var_p5_out.get().strip()
        if out:
            args.extend(["--out", out])

        self.p5_log.delete("1.0", "end")
        self.p5_status.configure(
            text="추출 시작 — "
            + ("전체 마켓" if market == p5_category.ALL_MARKETS
               else p5_category.MARKETS.get(market, market)),
            fg="#15803d",
        )

        creationflags = 0
        if os.name == "nt":
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        try:
            self._p5_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P5_카테고리_엑셀추출"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p5_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p5_proc, args=(self._p5_proc,), daemon=True
        ).start()

    def _stop_p5(self) -> None:
        proc = self._p5_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p5_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p5_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p5_status.configure(text="작업중단 요청 중…", fg="#b45309")
        try:
            proc.terminate()
        except Exception:
            pass

    def _open_p5_excel(self) -> None:
        target = self._p5_last_excel or self.var_p5_out.get().strip()
        if not target:
            out_dir = ROOT / "P5_카테고리_엑셀추출" / "output"
            files = sorted(out_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime) if out_dir.is_dir() else []
            if not files:
                messagebox.showinfo("안내", "저장된 엑셀이 없습니다.")
                return
            target = str(files[-1])
        try:
            if os.name == "nt":
                os.startfile(target)  # type: ignore[attr-defined]
            else:
                webbrowser.open(f"file://{target}")
        except Exception as e:  # noqa: BLE001
            messagebox.showerror("열기 실패", str(e))

    def _watch_p5_proc(self, proc: subprocess.Popen) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self.after(0, lambda t=text: self._append_p5_log(t))
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self.after(0, lambda t=text: self._append_p5_log(t))
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p5_status.configure(text=f"로그 수신 오류: {e}", fg="#b91c1c"),
            )
        code = proc.wait()
        if code == 0:
            self.after(0, lambda: self.p5_status.configure(text="추출 완료", fg="#15803d"))
        else:
            self.after(
                0,
                lambda: self.p5_status.configure(text=f"종료 (exit={code})", fg="#b91c1c"),
            )

    # ── P5_101_카테고리매핑_필터세부설정 ───────────────────────────
    def _build_p5_101(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P5_101 — 체크된 필터마다 마켓 카테고리 자동 매핑 (엑셀 기준)",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        form = tk.LabelFrame(parent, text="입력 (초기 1회)", bg="#ffffff", padx=8, pady=6)
        form.pack(fill="x")

        r0 = tk.Frame(form, bg="#ffffff")
        r0.pack(fill="x", pady=3)
        tk.Label(r0, text="상품수집사이트", width=13, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p5m_site = tk.StringVar(value=p5_mapping.DEFAULT_SITE)
        self.cbo_p5m_site = ttk.Combobox(
            r0,
            textvariable=self.var_p5m_site,
            width=20,
            values=ih.load(P5M_SITE_HISTORY),
        )
        self.cbo_p5m_site.pack(side="left")
        self.cbo_p5m_site.bind(
            "<<ComboboxSelected>>", lambda e: self._on_p5m_input_picked()
        )
        tk.Label(
            r0,
            text=f"※ 현재는 {p5_mapping.ALLOWED_SITES[0]} 만 수행 (검증 후 확대)",
            bg="#ffffff",
            fg="#b45309",
            font=("Malgun Gothic", 8),
        ).pack(side="left", padx=6)


        rv = tk.Frame(form, bg="#ffffff")
        rv.pack(fill="x", pady=3)
        tk.Label(rv, text="11번가 구분", width=13, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p5m_v11 = tk.StringVar(value=p5_mapping.BOTH)
        for label in (*p5_mapping.MARKET_VARIANTS["11ST"], p5_mapping.BOTH):
            tk.Radiobutton(
                rv,
                text=label,
                value=label,
                variable=self.var_p5m_v11,
                bg="#ffffff",
                font=("Malgun Gothic", 9),
            ).pack(side="left", padx=(0, 8))

        rv2 = tk.Frame(form, bg="#ffffff")
        rv2.pack(fill="x", pady=3)
        tk.Label(rv2, text="롯데ON 구분", width=13, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p5m_vlt = tk.StringVar(value=p5_mapping.BOTH)
        for label in (*p5_mapping.MARKET_VARIANTS["LTON"], p5_mapping.BOTH):
            tk.Radiobutton(
                rv2,
                text=label,
                value=label,
                variable=self.var_p5m_vlt,
                bg="#ffffff",
                font=("Malgun Gothic", 9),
            ).pack(side="left", padx=(0, 8))

        r1 = tk.Frame(form, bg="#ffffff")
        r1.pack(fill="x", pady=3)
        tk.Label(r1, text="카테고리 엑셀 폴더", width=13, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p5m_dir = tk.StringVar(value=str(ROOT / "P5_카테고리_엑셀추출" / "output"))
        tk.Entry(r1, textvariable=self.var_p5m_dir).pack(side="left", fill="x", expand=True)
        tk.Button(r1, text="…", width=3, command=self._pick_p5m_dir).pack(side="left", padx=4)
        tk.Button(r1, text="확인", command=self._check_p5m_excels).pack(side="left")

        self.lbl_p5m_excels = tk.Label(
            form,
            text="엑셀: (확인 전)",
            bg="#ffffff",
            fg="#0f172a",
            font=("Malgun Gothic", 9),
            anchor="w",
            justify="left",
        )
        self.lbl_p5m_excels.pack(fill="x", pady=(2, 0))

        r2 = tk.Frame(form, bg="#ffffff")
        r2.pack(fill="x", pady=3)
        tk.Label(r2, text="목록 URL", width=13, anchor="w", bg="#ffffff").pack(side="left")
        # 기본값을 넣지 않는다 — 목록 URL 은 계정·검색조건마다 달라서
        # 미리 채워두면 엉뚱한 화면(행 0건)에서 작업하게 된다.
        # 대신 이전에 입력한 값을 리스트박스(콤보박스)에서 다시 고를 수 있다.
        self.var_p5m_url = tk.StringVar(value="")
        self.cbo_p5m_url = ttk.Combobox(
            r2, textvariable=self.var_p5m_url, values=ih.load(P5M_URL_HISTORY)
        )
        self.cbo_p5m_url.pack(side="left", fill="x", expand=True)
        self.cbo_p5m_url.bind(
            "<<ComboboxSelected>>", lambda e: self._on_p5m_input_picked()
        )
        tk.Label(
            r2,
            text="필수",
            bg="#ffffff",
            fg="#b91c1c",
            font=("Malgun Gothic", 8, "bold"),
        ).pack(side="left", padx=6)

        r3 = tk.Frame(form, bg="#ffffff")
        r3.pack(fill="x", pady=3)
        tk.Label(r3, text="작업 행 범위", width=13, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p5m_from = tk.StringVar(value=str(p5_mapping.DEFAULT_ROW_FROM))
        tk.Entry(r3, textvariable=self.var_p5m_from, width=6).pack(side="left")
        tk.Label(r3, text="부터", bg="#ffffff").pack(side="left", padx=(4, 10))
        self.var_p5m_to = tk.StringVar(value=str(p5_mapping.DEFAULT_ROW_TO))
        tk.Entry(r3, textvariable=self.var_p5m_to, width=6).pack(side="left")
        tk.Label(r3, text="까지", bg="#ffffff").pack(side="left", padx=(4, 10))
        tk.Label(
            r3,
            text="※ 위 「목록 URL」 검색결과의 행 번호 기준 (1부터, 양끝 포함)",
            bg="#ffffff",
            fg="#64748b",
            font=("Malgun Gothic", 8),
        ).pack(side="left")

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=8)
        tk.Button(
            actions,
            text="행 목록 확인",
            command=self._check_p5m_rows,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=(0, 6))
        tk.Button(
            actions,
            text="매핑 시작",
            command=self._run_p5_101,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left")
        tk.Button(
            actions,
            text="작업중단",
            command=self._stop_p5_101,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=6)

        log_frame = tk.LabelFrame(parent, text="실행 로그", bg="#ffffff", padx=6, pady=4)
        log_frame.pack(fill="both", expand=True)
        self.p5_101_log = tk.Text(
            log_frame, height=16, font=("Consolas", 9), wrap="word", bg="#0f172a", fg="#e2e8f0"
        )
        sb101 = tk.Scrollbar(log_frame, command=self.p5_101_log.yview)
        self.p5_101_log.configure(yscrollcommand=sb101.set)
        self.p5_101_log.pack(side="left", fill="both", expand=True)
        sb101.pack(side="right", fill="y")

        self.p5_101_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p5_101_status.pack(fill="x", pady=4)

    def _pick_p5m_dir(self) -> None:
        path = filedialog.askdirectory(title="마켓별 카테고리 엑셀 폴더")
        if path:
            self.var_p5m_dir.set(path)
            self._check_p5m_excels()

    def _check_p5m_excels(self) -> None:
        folder = self.var_p5m_dir.get().strip()
        if not folder or not Path(folder).is_dir():
            self.lbl_p5m_excels.configure(text="엑셀: 폴더 없음", fg="#b91c1c")
            return
        found = p5_mapping.discover_market_excels(folder)
        if not found:
            self.lbl_p5m_excels.configure(
                text="엑셀: 인식된 파일 없음 (P5 로 먼저 추출하세요)", fg="#b91c1c"
            )
            return
        names = " · ".join(p5_mapping.MARKETS.get(c, c) for c in found)
        missing = [p5_mapping.MARKETS[c] for c in p5_mapping.MARKETS if c not in found]
        text = f"엑셀 {len(found)}개: {names}"
        if missing:
            text += f"   (없음: {' · '.join(missing)})"
        self.lbl_p5m_excels.configure(text=text, fg="#15803d" if not missing else "#b45309")

    def _on_p5m_input_picked(self) -> None:
        """사이트명·목록 URL 리스트박스에서 값을 고르면 즉시 망고 목록을 조회한다."""
        if self.var_p5m_url.get().strip():
            self._check_p5m_rows()

    def _remember_p5m_inputs(self) -> None:
        site = self.var_p5m_site.get().strip()
        url = self.var_p5m_url.get().strip()
        if site:
            self.cbo_p5m_site.configure(values=ih.remember(P5M_SITE_HISTORY, site))
        if url:
            self.cbo_p5m_url.configure(values=ih.remember(P5M_URL_HISTORY, url))

    def _check_p5m_rows(self) -> None:
        """매핑 없이 행 번호·ftid·필터명만 확인 ('몇 번째 행인지' 검증)."""
        if self._p5_101_proc and self._p5_101_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return
        script = ROOT / "P5_101_카테고리매핑_필터세부설정" / "map_categories.py"
        if not script.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{script}")
            return

        self._remember_p5m_inputs()
        site = self.var_p5m_site.get().strip() or p5_mapping.DEFAULT_SITE
        row_from = self.var_p5m_from.get().strip() or str(p5_mapping.DEFAULT_ROW_FROM)
        row_to = self.var_p5m_to.get().strip() or str(p5_mapping.DEFAULT_ROW_TO)
        args = [
            sys.executable,
            str(script),
            "--list-rows",
            "--site-id",
            site,
            "--row-from",
            row_from,
            "--row-to",
            row_to,
        ]
        url = self.var_p5m_url.get().strip()
        if url:
            args.extend(["--list-url", url])

        self.p5_101_log.delete("1.0", "end")
        self.p5_101_status.configure(text="행 목록 확인 중…", fg="#0f766e")

        creationflags = 0
        if os.name == "nt":
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        try:
            self._p5_101_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P5_101_카테고리매핑_필터세부설정"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p5_101_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p5_101_proc, args=(self._p5_101_proc,), daemon=True
        ).start()

    def _p5_101_stop_flag(self) -> Path:
        return ROOT / "P5_101_카테고리매핑_필터세부설정" / ".map_stop"

    def _append_p5_101_log(self, line: str) -> None:
        text = (line or "").strip()
        if text.startswith("##MAIN##"):
            text = text[8:]
        self.p5_101_log.insert("end", text + "\n")
        self.p5_101_log.see("end")

    def _run_p5_101(self) -> None:
        if self._p5_101_proc and self._p5_101_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return
        script = ROOT / "P5_101_카테고리매핑_필터세부설정" / "map_categories.py"
        if not script.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{script}")
            return
        folder = self.var_p5m_dir.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showinfo("안내", "카테고리 엑셀 폴더를 지정하세요.")
            return

        try:
            self._p5_101_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        self._remember_p5m_inputs()
        site = self.var_p5m_site.get().strip() or p5_mapping.DEFAULT_SITE
        if not p5_mapping.is_allowed_site(site):
            messagebox.showwarning(
                "수집사이트 제한",
                f"현재는 {p5_mapping.ALLOWED_SITES[0]} 만 수행합니다.\n"
                f"입력값: {site}\n\n검증이 끝나면 다른 사이트도 열겠습니다.",
            )
            return

        row_from = self.var_p5m_from.get().strip()
        row_to = self.var_p5m_to.get().strip()
        if not row_from.isdigit() or not row_to.isdigit():
            messagebox.showinfo("안내", "작업 행 범위는 1 이상의 숫자여야 합니다.")
            return
        start, end = p5_mapping.row_range(row_from, row_to)

        args = [
            sys.executable,
            str(script),
            "--excel-dir",
            folder,
            "--site-id",
            site,
            "--row-from",
            str(start),
            "--row-to",
            str(end),
            "--variant",
            f"11ST={self.var_p5m_v11.get()}",
            "--variant",
            f"LTON={self.var_p5m_vlt.get()}",
        ]
        url = self.var_p5m_url.get().strip()
        if url:
            args.extend(["--list-url", url])

        self.p5_101_log.delete("1.0", "end")
        self.p5_101_status.configure(
            text=f"매핑 시작 — {site} · {start}~{end}행", fg="#15803d"
        )

        creationflags = 0
        if os.name == "nt":
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        try:
            self._p5_101_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P5_101_카테고리매핑_필터세부설정"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p5_101_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p5_101_proc, args=(self._p5_101_proc,), daemon=True
        ).start()

    def _stop_p5_101(self) -> None:
        proc = self._p5_101_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p5_101_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p5_101_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p5_101_status.configure(text="작업중단 요청 중…", fg="#b45309")
        try:
            proc.terminate()
        except Exception:
            pass

    def _watch_p5_101_proc(self, proc: subprocess.Popen) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self.after(0, lambda t=text: self._append_p5_101_log(t))
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self.after(0, lambda t=text: self._append_p5_101_log(t))
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p5_101_status.configure(text=f"로그 수신 오류: {e}", fg="#b91c1c"),
            )
        code = proc.wait()
        if code == 0:
            self.after(0, lambda: self.p5_101_status.configure(text="매핑 완료", fg="#15803d"))
        else:
            self.after(
                0,
                lambda: self.p5_101_status.configure(text=f"종료 (exit={code})", fg="#b91c1c"),
            )

    def _build_p2(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P2 — 폴더의 엑셀 파일 선택 → 카테고리URL목록 확인 → 수집 실행",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        # 1. 디렉터리 파일 목록 (리스트박스 + 스크롤)
        search = tk.LabelFrame(
            parent, text="1. 디렉터리 파일 목록", bg="#ffffff", padx=8, pady=6
        )
        search.pack(fill="x")

        self.var_dir = tk.StringVar(value=(default_roots() or [str(Path.home())])[0])
        self.var_q = tk.StringVar(value="카테고리URL")

        r1 = tk.Frame(search, bg="#ffffff")
        r1.pack(fill="x", pady=2)
        tk.Label(r1, text="폴더", width=8, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(r1, textvariable=self.var_dir).pack(side="left", fill="x", expand=True)
        tk.Button(r1, text="…", width=3, command=self._pick_search_dir).pack(side="left", padx=4)

        r2 = tk.Frame(search, bg="#ffffff")
        r2.pack(fill="x", pady=2)
        tk.Label(r2, text="필터", width=8, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(r2, textvariable=self.var_q, width=20).pack(side="left")
        tk.Button(r2, text="파일 새로고침", command=self._search_xlsx, bg="#e2e8f0").pack(
            side="left", padx=6
        )
        tk.Button(r2, text="선택 파일 열기", command=self._add_found).pack(side="left")

        self._p2_file_list_height = 5
        self._p2_url_list_height = 8

        found_wrap = tk.Frame(search, bg="#ffffff")
        found_wrap.pack(fill="x", pady=4)
        self.found_list = tk.Listbox(
            found_wrap,
            height=self._p2_file_list_height,
            selectmode="browse",
            font=("Consolas", 9),
            exportselection=False,
        )
        found_sb = tk.Scrollbar(found_wrap, orient="vertical", command=self.found_list.yview)
        self.found_list.configure(yscrollcommand=found_sb.set)
        self.found_list.pack(side="left", fill="both", expand=True)
        found_sb.pack(side="right", fill="y")
        self.found_list.bind("<<ListboxSelect>>", self._on_found_select)
        self.found_list.bind("<Double-Button-1>", lambda _e: self._add_found())
        self.found_list.bind("<MouseWheel>", self._on_found_mousewheel)
        self.found_list.bind("<Button-4>", self._on_found_mousewheel)
        self.found_list.bind("<Button-5>", self._on_found_mousewheel)
        self._found_paths: list[str] = []

        # 실행 버튼 (좌) + 체크박스 MAIN / SUB / 1·2행 스크린샷 (최우측)
        actions = tk.LabelFrame(parent, text="실행", bg="#ffffff", padx=8, pady=6)
        actions.pack(fill="x", pady=(8, 0))

        self.var_show_main = tk.BooleanVar(value=True)
        self.var_show_sub = tk.BooleanVar(value=True)
        self.var_verify = tk.BooleanVar(value=True)
        btn_row = tk.Frame(actions, bg="#ffffff")
        btn_row.pack(fill="x")

        # ★최우측 순서: MAIN → SUB → 1·2행 스크린샷
        # (실행로그 안내 라벨은 삭제 — 요건: 스크린샷 2번째 LABEL 삭제)
        # pack 순서: right 먼저 → 항상 최우측 고정
        right_checks = tk.Frame(btn_row, bg="#ffffff")
        right_checks.pack(side="right")
        tk.Checkbutton(
            right_checks,
            text="MAIN",
            variable=self.var_show_main,
            command=self._toggle_log_panels,
            bg="#ffffff",
            font=("Malgun Gothic", 9),
        ).pack(side="left", padx=(0, 6))
        tk.Checkbutton(
            right_checks,
            text="SUB",
            variable=self.var_show_sub,
            command=self._toggle_log_panels,
            bg="#ffffff",
            font=("Malgun Gothic", 9),
        ).pack(side="left", padx=(0, 6))
        tk.Checkbutton(
            right_checks,
            text="1·2행 스크린샷",
            variable=self.var_verify,
            bg="#ffffff",
            font=("Malgun Gothic", 9),
        ).pack(side="left")

        left_btns = tk.Frame(btn_row, bg="#ffffff")
        left_btns.pack(side="left", fill="x", expand=True)
        tk.Button(
            left_btns,
            text="수집 시작",
            command=self._run_p2,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=10,
            pady=4,
        ).pack(side="left")
        tk.Button(
            left_btns,
            text="수집 종료",
            command=self._stop_p2,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=10,
            pady=4,
        ).pack(side="left", padx=6)
        tk.Button(left_btns, text="파일 목록에서 제거", command=self._remove_lib).pack(
            side="left", padx=6
        )
        tk.Button(left_btns, text="새로고침", command=self._refresh_p2_list).pack(side="left")
        tk.Button(left_btns, text="로그 지우기", command=self._clear_p2_log).pack(
            side="left", padx=6
        )
        tk.Button(
            left_btns,
            text="스크린샷 보기",
            command=self._show_shot_viewer,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=8,
            pady=4,
        ).pack(side="left", padx=6)

        # 2. 카테고리URL목록 — 엑셀 전체 행 + 진행중 행 적색
        lib = tk.LabelFrame(
            parent, text="카테고리URL목록", bg="#ffffff", padx=8, pady=4
        )
        lib.pack(fill="x", pady=(8, 0))

        lib_wrap = tk.Frame(lib, bg="#ffffff")
        lib_wrap.pack(fill="both", expand=True)
        self.lib_list = tk.Listbox(
            lib_wrap,
            height=self._p2_url_list_height,
            font=("Malgun Gothic", 10),
            exportselection=False,
            activestyle="none",
        )
        lib_sb = tk.Scrollbar(lib_wrap, orient="vertical", command=self.lib_list.yview)
        self.lib_list.configure(yscrollcommand=lib_sb.set)
        self.lib_list.pack(side="left", fill="both", expand=True)
        lib_sb.pack(side="right", fill="y")
        self.lib_list.bind("<MouseWheel>", self._on_lib_mousewheel)
        self.lib_list.bind("<Button-4>", self._on_lib_mousewheel)
        self.lib_list.bind("<Button-5>", self._on_lib_mousewheel)
        self._lib_paths: list[str] = []  # 하위호환(파일경로 1개 보관용)
        self._excel_rows: list[dict] = []
        self._current_excel_path: str = ""
        self._active_ordinal: int = 0  # 1-based, 0=없음

        self.p2_sel = tk.Label(lib, text="", bg="#ffffff", fg="#64748b", anchor="w")
        self.p2_sel.pack(fill="x", pady=(2, 0))

        # 3. 실행 로그 — main(13단계) / sub(단계별 추가정보·스크린샷) 두 그리드
        log_area = tk.Frame(parent, bg="#f1f5f9")
        log_area.pack(fill="both", expand=True, pady=(8, 0))

        style = ttk.Style(self)
        try:
            style.configure("P2Log.Treeview", rowheight=22, font=("Malgun Gothic", 9))
            style.configure("P2Log.Treeview.Heading", font=("Malgun Gothic", 9, "bold"))
        except tk.TclError:
            pass

        self._p2_log_area = log_area
        self.p2_main_frame = tk.LabelFrame(
            log_area,
            text="3-A. 실행 로그 MAIN (상단=엑셀정보 · 아래=1~13단계)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        self.p2_main_frame.pack(fill="both", expand=True)

        self.p2_main_log = ttk.Treeview(
            self.p2_main_frame,
            columns=("time", "step", "message"),
            show="headings",
            height=9,
            style="P2Log.Treeview",
        )
        self.p2_main_log.heading("time", text="시각")
        self.p2_main_log.heading("step", text="단계")
        self.p2_main_log.heading("message", text="내용 (1~13단계)")
        self.p2_main_log.column("time", width=130, minwidth=110, stretch=False, anchor="center")
        self.p2_main_log.column("step", width=44, minwidth=40, stretch=False, anchor="center")
        self.p2_main_log.column("message", width=620, minwidth=220, stretch=True, anchor="w")
        main_sb = tk.Scrollbar(
            self.p2_main_frame, orient="vertical", command=self.p2_main_log.yview
        )
        self.p2_main_log.configure(yscrollcommand=main_sb.set)
        self.p2_main_log.pack(side="left", fill="both", expand=True)
        main_sb.pack(side="right", fill="y")
        self.p2_main_log.bind("<<TreeviewSelect>>", self._on_main_log_select)
        self._setup_p2_log_tags()

        self.p2_sub_frame = tk.LabelFrame(
            log_area,
            text="3-B. 실행 로그 SUB (선택한 단계의 추가정보·스크린샷)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        self.p2_sub_frame.pack(fill="both", expand=True, pady=(6, 0))

        self.p2_sub_log = ttk.Treeview(
            self.p2_sub_frame,
            columns=("time", "message"),
            show="headings",
            height=7,
            style="P2Log.Treeview",
        )
        self.p2_sub_log.heading("time", text="시각")
        self.p2_sub_log.heading("message", text="추가정보 · [샷]은 더블클릭으로 열기")
        self.p2_sub_log.column("time", width=130, minwidth=110, stretch=False, anchor="center")
        self.p2_sub_log.column("message", width=700, minwidth=220, stretch=True, anchor="w")
        sub_sb = tk.Scrollbar(
            self.p2_sub_frame, orient="vertical", command=self.p2_sub_log.yview
        )
        self.p2_sub_log.configure(yscrollcommand=sub_sb.set)
        self.p2_sub_log.pack(side="left", fill="both", expand=True)
        sub_sb.pack(side="right", fill="y")
        self.p2_sub_log.tag_configure("shot", foreground="#0f766e")
        self.p2_sub_log.bind("<Double-Button-1>", self._on_sub_log_double_click)

        # seq(단계 발생 고유번호) 기반 main↔sub 연결 데이터
        self._sub_by_seq: dict[int, list[tuple[str, str, str]]] = {}
        self._shot_path_by_seq: dict[tuple[int, int], str] = {}
        self._main_item_by_seq: dict[int, str] = {}
        self._main_ts_end: dict[int, str] = {}
        self._seq_by_main_item: dict[str, int] = {}
        self._meta_item_id: str | None = None
        self._meta_values: dict[str, str] = {f: "" for f in META_FIELDS}
        self._selected_seq: int | None = None
        self._latest_seq: int = 0
        self._follow_latest: bool = True

        self._setup_meta_rows()

        self.p2_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p2_status.pack(fill="x", pady=4)

    # ── P3_핏클상세페이지 (FitCL 연동) ─────────────────────────────
    def _build_p3_fitcl(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P3_핏클상세페이지 — FitCL 연동 · 모델컷 10장 + 디테일컷 5장 추출",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        form = tk.LabelFrame(parent, text="입력", bg="#ffffff", padx=8, pady=6)
        form.pack(fill="x")

        r1 = tk.Frame(form, bg="#ffffff")
        r1.pack(fill="x", pady=4)
        tk.Label(r1, text="1.소싱상품", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p3_fitcl_product = tk.StringVar()
        tk.Entry(r1, textvariable=self.var_p3_fitcl_product).pack(
            side="left", fill="x", expand=True
        )
        tk.Button(r1, text="…", width=3, command=self._pick_p3_fitcl_product).pack(
            side="left", padx=4
        )

        r2 = tk.Frame(form, bg="#ffffff")
        r2.pack(fill="x", pady=4)
        tk.Label(r2, text="2.사진모델", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p3_fitcl_model = tk.StringVar()
        model_cb = ttk.Combobox(
            r2,
            textvariable=self.var_p3_fitcl_model,
            values=p3_fitcl.DEFAULT_MODELS,
            width=48,
        )
        model_cb.pack(side="left", fill="x", expand=True)
        if p3_fitcl.DEFAULT_MODELS:
            model_cb.current(0)

        r3 = tk.Frame(form, bg="#ffffff")
        r3.pack(fill="x", pady=4)
        tk.Label(r3, text="FitCL URL", width=10, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p3_fitcl_url = tk.StringVar(value=p3_fitcl.DEFAULT_FITCL_URL)
        tk.Entry(r3, textvariable=self.var_p3_fitcl_url).pack(
            side="left", fill="x", expand=True
        )

        pose_frame = tk.LabelFrame(
            form,
            text=f"3.모델포즈 — 정확히 {p3_fitcl.REQUIRED_POSE_COUNT}개 선택 (Ctrl+클릭)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        pose_frame.pack(fill="both", expand=True, pady=(6, 0))
        pose_wrap = tk.Frame(pose_frame, bg="#ffffff")
        pose_wrap.pack(fill="both", expand=True)
        self.p3_fitcl_pose_list = tk.Listbox(
            pose_wrap,
            height=8,
            selectmode=tk.EXTENDED,
            exportselection=False,
            font=("Malgun Gothic", 9),
        )
        pose_sb = tk.Scrollbar(pose_wrap, command=self.p3_fitcl_pose_list.yview)
        self.p3_fitcl_pose_list.configure(yscrollcommand=pose_sb.set)
        self.p3_fitcl_pose_list.pack(side="left", fill="both", expand=True)
        pose_sb.pack(side="right", fill="y")
        for pose in p3_fitcl.DEFAULT_POSES:
            self.p3_fitcl_pose_list.insert("end", pose)
        # 기본 10개 선택
        for i in range(min(p3_fitcl.REQUIRED_POSE_COUNT, self.p3_fitcl_pose_list.size())):
            self.p3_fitcl_pose_list.selection_set(i)

        self.p3_fitcl_pose_status = tk.Label(
            pose_frame,
            text=f"선택: {p3_fitcl.REQUIRED_POSE_COUNT}개",
            bg="#ffffff",
            fg="#64748b",
            anchor="w",
            font=("Malgun Gothic", 8),
        )
        self.p3_fitcl_pose_status.pack(fill="x", pady=(2, 0))
        self.p3_fitcl_pose_list.bind("<<ListboxSelect>>", self._on_p3_fitcl_pose_select)

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=8)
        tk.Button(
            actions,
            text="작업시작",
            command=self._run_p3_fitcl,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left")
        tk.Button(
            actions,
            text="작업중단",
            command=self._stop_p3_fitcl,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=6)

        log_frame = tk.LabelFrame(parent, text="실행 로그", bg="#ffffff", padx=6, pady=4)
        log_frame.pack(fill="both", expand=True)
        self.p3_fitcl_log = tk.Text(
            log_frame,
            height=12,
            font=("Consolas", 9),
            wrap="word",
            bg="#0f172a",
            fg="#e2e8f0",
        )
        log_sb = tk.Scrollbar(log_frame, command=self.p3_fitcl_log.yview)
        self.p3_fitcl_log.configure(yscrollcommand=log_sb.set)
        self.p3_fitcl_log.pack(side="left", fill="both", expand=True)
        log_sb.pack(side="right", fill="y")

        self.p3_fitcl_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p3_fitcl_status.pack(fill="x", pady=4)

    def _pick_p3_fitcl_product(self) -> None:
        path = filedialog.askopenfilename(
            title="소싱상품 이미지 선택",
            filetypes=[
                ("이미지", "*.jpg *.jpeg *.png *.webp *.bmp"),
                ("모든 파일", "*.*"),
            ],
        )
        if path:
            self.var_p3_fitcl_product.set(path)

    def _on_p3_fitcl_pose_select(self, _evt=None) -> None:
        n = len(self.p3_fitcl_pose_list.curselection())
        need = p3_fitcl.REQUIRED_POSE_COUNT
        color = "#15803d" if n == need else "#b45309"
        self.p3_fitcl_pose_status.configure(
            text=f"선택: {n}개 / 필요: {need}개", fg=color
        )

    def _p3_fitcl_stop_flag(self) -> Path:
        return ROOT / "P3_핏클상세페이지" / ".fitcl_stop"

    def _append_p3_fitcl_log(self, line: str) -> None:
        text = (line or "").strip()
        if text.startswith("##MAIN##"):
            text = text[8:]
        self.p3_fitcl_log.insert("end", text + "\n")
        self.p3_fitcl_log.see("end")

    def _run_p3_fitcl(self) -> None:
        product = self.var_p3_fitcl_product.get().strip()
        if not product or not os.path.isfile(product):
            messagebox.showinfo("안내", "소싱상품 이미지를 선택하세요.")
            return
        model = self.var_p3_fitcl_model.get().strip()
        if not model:
            messagebox.showinfo("안내", "사진모델을 선택하세요.")
            return
        sel = self.p3_fitcl_pose_list.curselection()
        poses = [self.p3_fitcl_pose_list.get(i) for i in sel]
        if len(poses) != p3_fitcl.REQUIRED_POSE_COUNT:
            messagebox.showinfo(
                "안내",
                f"모델포즈는 정확히 {p3_fitcl.REQUIRED_POSE_COUNT}개를 선택하세요 "
                f"(현재 {len(poses)}개).",
            )
            return
        if self._p3_fitcl_proc and self._p3_fitcl_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return

        script = ROOT / "P3_핏클상세페이지" / "fitcl_detail.py"
        if not script.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{script}")
            return

        try:
            self._p3_fitcl_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        fitcl_url = self.var_p3_fitcl_url.get().strip()
        args = [
            sys.executable,
            str(script),
            "--product",
            product,
            "--model",
            model,
            "--poses",
            ",".join(poses),
        ]
        if fitcl_url:
            args.extend(["--fitcl-url", fitcl_url])

        self.p3_fitcl_log.delete("1.0", "end")
        self.p3_fitcl_status.configure(
            text=f"작업 시작 — 모델컷 {p3_fitcl.REQUIRED_POSE_COUNT}장 + 디테일컷 5장",
            fg="#15803d",
        )

        creationflags = 0
        if os.name == "nt":
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        try:
            self._p3_fitcl_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P3_핏클상세페이지"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p3_fitcl_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p3_fitcl_proc, args=(self._p3_fitcl_proc,), daemon=True
        ).start()

    def _stop_p3_fitcl(self) -> None:
        proc = self._p3_fitcl_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p3_fitcl_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p3_fitcl_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p3_fitcl_status.configure(text="작업중단 요청 중…", fg="#b45309")
        try:
            proc.terminate()
        except Exception:
            pass

    def _watch_p3_fitcl_proc(self, proc: subprocess.Popen) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self.after(0, lambda t=text: self._append_p3_fitcl_log(t))
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self.after(0, lambda t=text: self._append_p3_fitcl_log(t))
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p3_fitcl_status.configure(
                    text=f"로그 수신 오류: {e}", fg="#b91c1c"
                ),
            )
        code = proc.wait()
        if code == 0:
            self.after(
                0, lambda: self.p3_fitcl_status.configure(text="작업 완료", fg="#15803d")
            )
        else:
            self.after(
                0,
                lambda: self.p3_fitcl_status.configure(
                    text=f"종료 (exit={code})", fg="#b91c1c"
                ),
            )

    # ── P3_필터_갱신 (UI 구조 = P2와 유사 + 더망고 URL 입력) ──
    def _build_p3(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P3_필터_갱신 — 엑셀 선택 → 더망고 URL 입력 → 저장상품수 갱신",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        search = tk.LabelFrame(
            parent, text="1. 디렉터리 파일 목록", bg="#ffffff", padx=8, pady=6
        )
        search.pack(fill="x")

        self.var_p3_dir = tk.StringVar(
            value=(default_roots() or [str(Path.home())])[0]
        )
        self.var_p3_q = tk.StringVar(value="카테고리URL")
        # ★요건: 더망고 URL = 검색필터 getGoodsCategory.php 고정 초기값
        self.var_p3_mango_url = tk.StringVar(value=p3_update.DEFAULT_MANGO_URL)
        try:
            p3_update.save_mango_url(p3_update.DEFAULT_MANGO_URL)
        except Exception:
            pass

        r1 = tk.Frame(search, bg="#ffffff")
        r1.pack(fill="x", pady=2)
        tk.Label(r1, text="폴더", width=10, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(r1, textvariable=self.var_p3_dir).pack(
            side="left", fill="x", expand=True
        )
        tk.Button(r1, text="…", width=3, command=self._pick_p3_dir).pack(
            side="left", padx=4
        )

        r2 = tk.Frame(search, bg="#ffffff")
        r2.pack(fill="x", pady=2)
        tk.Label(r2, text="필터", width=10, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(r2, textvariable=self.var_p3_q, width=20).pack(side="left")
        tk.Button(
            r2, text="파일 새로고침", command=self._search_p3_xlsx, bg="#e2e8f0"
        ).pack(side="left", padx=6)
        tk.Button(r2, text="선택 파일 열기", command=self._add_p3_found).pack(
            side="left"
        )

        r3 = tk.Frame(search, bg="#ffffff")
        r3.pack(fill="x", pady=(6, 2))
        tk.Label(
            r3, text="더망고 URL", width=10, anchor="w", bg="#ffffff",
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side="left")
        tk.Entry(r3, textvariable=self.var_p3_mango_url).pack(
            side="left", fill="x", expand=True
        )
        # ★실행로그를 최대로 키우기 위해 위쪽 목록은 최소 높이만
        self._p3_file_list_height = 2
        self._p3_url_list_height = 3

        found_wrap = tk.Frame(search, bg="#ffffff")
        found_wrap.pack(fill="x", pady=4)
        self.p3_found_list = tk.Listbox(
            found_wrap,
            height=self._p3_file_list_height,
            selectmode="browse",
            font=("Consolas", 9),
            exportselection=False,
        )
        found_sb = tk.Scrollbar(
            found_wrap, orient="vertical", command=self.p3_found_list.yview
        )
        self.p3_found_list.configure(yscrollcommand=found_sb.set)
        self.p3_found_list.pack(side="left", fill="both", expand=True)
        found_sb.pack(side="right", fill="y")
        self.p3_found_list.bind("<<ListboxSelect>>", self._on_p3_found_select)
        self.p3_found_list.bind("<Double-Button-1>", lambda _e: self._add_p3_found())
        self._p3_found_paths: list[str] = []

        actions = tk.LabelFrame(parent, text="실행", bg="#ffffff", padx=8, pady=6)
        actions.pack(fill="x", pady=(8, 0))
        btn_row = tk.Frame(actions, bg="#ffffff")
        btn_row.pack(fill="x")

        # ★실행로그 표출 체크박스 — 실행 버튼줄 최우측 (MAIN / SUB / 스크린샷)
        self.var_p3_show_main = tk.BooleanVar(value=True)
        self.var_p3_show_sub = tk.BooleanVar(value=True)
        self.var_p3_show_shot = tk.BooleanVar(value=True)
        right_checks = tk.Frame(btn_row, bg="#ffffff")
        right_checks.pack(side="right")
        for text, var in (
            ("MAIN", self.var_p3_show_main),
            ("SUB", self.var_p3_show_sub),
            ("스크린샷", self.var_p3_show_shot),
        ):
            tk.Checkbutton(
                right_checks,
                text=text,
                variable=var,
                command=self._toggle_p3_log_panels,
                bg="#ffffff",
                font=("Malgun Gothic", 9),
            ).pack(side="left", padx=(0, 6))

        left_btns = tk.Frame(btn_row, bg="#ffffff")
        left_btns.pack(side="left", fill="x", expand=True)
        self.btn_p3_run = tk.Button(
            left_btns,
            text="작업시작",
            command=self._run_p3,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=10,
            pady=4,
        )
        self.btn_p3_run.pack(side="left")
        self.btn_p3_stop = tk.Button(
            left_btns,
            text="작업중단",
            command=self._stop_p3,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=10,
            pady=4,
        )
        self.btn_p3_stop.pack(side="left", padx=6)
        tk.Button(
            left_btns,
            text="스크린샷 보기",
            command=self._show_p3_shot_viewer,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=8,
            pady=4,
        ).pack(side="left", padx=6)
        tk.Button(left_btns, text="새로고침", command=self._refresh_p3_list).pack(
            side="left", padx=6
        )
        tk.Button(left_btns, text="로그 지우기", command=self._clear_p3_log).pack(
            side="left", padx=6
        )

        lib = tk.LabelFrame(
            parent, text="카테고리URL목록 (엑셀)", bg="#ffffff", padx=8, pady=4
        )
        lib.pack(fill="x", pady=(8, 0))
        lib_wrap = tk.Frame(lib, bg="#ffffff")
        lib_wrap.pack(fill="both", expand=True)
        self.p3_lib_list = tk.Listbox(
            lib_wrap,
            height=self._p3_url_list_height,
            font=("Malgun Gothic", 10),
            exportselection=False,
            activestyle="none",
        )
        lib_sb = tk.Scrollbar(
            lib_wrap, orient="vertical", command=self.p3_lib_list.yview
        )
        self.p3_lib_list.configure(yscrollcommand=lib_sb.set)
        self.p3_lib_list.pack(side="left", fill="both", expand=True)
        lib_sb.pack(side="right", fill="y")
        self._p3_excel_rows: list[dict] = []
        self._p3_lib_texts: list[str] = []
        self._p3_active_ordinal: int = 0  # 진행 중 엑셀행 (1-based, 0=없음)
        self._p3_current_excel: str = ""
        self.p3_sel = tk.Label(lib, text="", bg="#ffffff", fg="#64748b", anchor="w")
        self.p3_sel.pack(fill="x", pady=(2, 0))

        # ★요건: 실행 로그 = MAIN(1~7단계만) / SUB(단계별 세부내용) / 스크린샷(오류·원인
        # 파악용 캡처) 3단 그리드로 명확히 구분 + 상단 체크박스 3개로 표출여부 선택
        log_wrap = tk.LabelFrame(
            parent, text="실행 로그", bg="#ffffff", padx=6, pady=4
        )
        log_wrap.pack(fill="both", expand=True, pady=(8, 0))

        log_area = tk.Frame(log_wrap, bg="#f1f5f9")
        log_area.pack(fill="both", expand=True)
        # ★MAIN/SUB/스크린샷 3단을 항상 함께 보이게 — 남는 높이를 균등 분배한다
        log_area.columnconfigure(0, weight=1)
        for _r in range(3):
            log_area.rowconfigure(_r, weight=1)

        style = ttk.Style(self)
        try:
            style.configure("P3Log.Treeview", rowheight=22, font=("Malgun Gothic", 9))
            style.configure("P3Log.Treeview.Heading", font=("Malgun Gothic", 9, "bold"))
        except tk.TclError:
            pass

        self._p3_log_area = log_area

        # 1) MAIN — 7단계만 표시
        self.p3_main_frame = tk.LabelFrame(
            log_area,
            text="MAIN (1~7단계만)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        self.p3_main_frame.grid(row=0, column=0, sticky="nsew")

        self.p3_main_log = ttk.Treeview(
            self.p3_main_frame,
            columns=("time", "step", "message"),
            show="headings",
            height=3,
            style="P3Log.Treeview",
        )
        self.p3_main_log.heading("time", text="시각")
        self.p3_main_log.heading("step", text="단계")
        self.p3_main_log.heading("message", text="내용 (1~7단계)")
        self.p3_main_log.column("time", width=90, minwidth=70, stretch=False, anchor="center")
        self.p3_main_log.column("step", width=44, minwidth=40, stretch=False, anchor="center")
        self.p3_main_log.column(
            "message", width=640, minwidth=220, stretch=True, anchor="w"
        )
        main_sb = tk.Scrollbar(
            self.p3_main_frame, orient="vertical", command=self.p3_main_log.yview
        )
        self.p3_main_log.configure(yscrollcommand=main_sb.set)
        self.p3_main_log.pack(side="left", fill="both", expand=True)
        main_sb.pack(side="right", fill="y")
        self.p3_main_log.bind("<<TreeviewSelect>>", self._on_p3_main_log_select)
        self._setup_p3_log_tags()

        # 2) SUB — 선택한 MAIN 단계의 세부내용(텍스트)
        self.p3_sub_frame = tk.LabelFrame(
            log_area,
            text="SUB (선택한 단계의 세부내용)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        self.p3_sub_frame.grid(row=1, column=0, sticky="nsew", pady=(6, 0))

        self.p3_sub_log = ttk.Treeview(
            self.p3_sub_frame,
            columns=("time", "message"),
            show="headings",
            height=3,
            style="P3Log.Treeview",
        )
        self.p3_sub_log.heading("time", text="시각")
        self.p3_sub_log.heading("message", text="세부내용")
        self.p3_sub_log.column("time", width=90, minwidth=70, stretch=False, anchor="center")
        self.p3_sub_log.column("message", width=700, minwidth=220, stretch=True, anchor="w")
        sub_sb = tk.Scrollbar(
            self.p3_sub_frame, orient="vertical", command=self.p3_sub_log.yview
        )
        self.p3_sub_log.configure(yscrollcommand=sub_sb.set)
        self.p3_sub_log.pack(side="left", fill="both", expand=True)
        sub_sb.pack(side="right", fill="y")
        self.p3_sub_log.tag_configure("err", foreground="#b91c1c")
        self.p3_sub_log.tag_configure("done", foreground="#166534")
        self.p3_sub_log.tag_configure("stop", foreground="#b45309")

        # 3) 스크린샷 — SUB 하단, 오류/원인 파악용 캡처 전용 그리드
        self.p3_shot_frame = tk.LabelFrame(
            log_area,
            text="스크린샷 (선택한 단계의 캡처 · 오류/원인 파악용 · 더블클릭으로 열기)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        self.p3_shot_frame.grid(row=2, column=0, sticky="nsew", pady=(6, 0))

        self.p3_shot_log = ttk.Treeview(
            self.p3_shot_frame,
            columns=("time", "label"),
            show="headings",
            height=3,
            style="P3Log.Treeview",
        )
        self.p3_shot_log.heading("time", text="시각")
        self.p3_shot_log.heading("label", text="스크린샷 라벨 · 더블클릭으로 열기")
        self.p3_shot_log.column("time", width=90, minwidth=70, stretch=False, anchor="center")
        self.p3_shot_log.column("label", width=700, minwidth=220, stretch=True, anchor="w")
        shot_sb = tk.Scrollbar(
            self.p3_shot_frame, orient="vertical", command=self.p3_shot_log.yview
        )
        self.p3_shot_log.configure(yscrollcommand=shot_sb.set)
        self.p3_shot_log.pack(side="left", fill="both", expand=True)
        shot_sb.pack(side="right", fill="y")
        self.p3_shot_log.tag_configure("shot", foreground="#0f766e")
        self.p3_shot_log.tag_configure("err", foreground="#b91c1c")
        self.p3_shot_log.bind("<Double-Button-1>", self._on_p3_shot_log_double_click)

        # seq(단계 발생 고유번호) 기반 main↔sub↔스크린샷 연결 데이터
        self._p3_sub_by_seq: dict[int, list[tuple[str, str, str]]] = {}
        self._p3_shots_by_seq: dict[int, list[tuple[str, str, str]]] = {}
        self._p3_shot_path_by_seq: dict[tuple[int, int], str] = {}
        self._p3_main_item_by_seq: dict[int, str] = {}
        self._p3_main_ts_end: dict[int, str] = {}
        self._p3_seq_by_main_item: dict[str, int] = {}
        self._p3_selected_seq: int | None = None
        self._p3_latest_seq: int = 0
        self._p3_follow_latest: bool = True
        self._p3_last_shot_dir: Path | None = None

        self.p3_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p3_status.pack(fill="x", pady=4)

    def _pick_p3_dir(self) -> None:
        d = filedialog.askdirectory(
            initialdir=self.var_p3_dir.get() or str(Path.home())
        )
        if d:
            self.var_p3_dir.set(d)
            self._search_p3_xlsx()

    def _search_p3_xlsx(self) -> None:
        self.p3_found_list.delete(0, "end")
        self._p3_found_paths = []
        try:
            files = search_xlsx(
                self.var_p3_dir.get().strip(), self.var_p3_q.get().strip()
            )
        except Exception as e:
            messagebox.showerror("검색 실패", str(e))
            return
        for f in files:
            self._p3_found_paths.append(f["path"])
            self.p3_found_list.insert("end", f["name"])
        self.p3_status.configure(
            text=f"파일 {len(files)}개" if files else "해당 폴더에서 .xlsx 없음",
            fg="#0f172a",
        )

    def _on_p3_found_select(self, _event=None) -> None:
        sel = self.p3_found_list.curselection()
        if not sel:
            return
        path = self._p3_found_paths[sel[0]]
        self._load_p3_category_list(path)

    def _add_p3_found(self) -> None:
        sel = self.p3_found_list.curselection()
        if not sel:
            messagebox.showinfo("안내", "목록에서 엑셀 파일을 선택하세요.")
            return
        path = self._p3_found_paths[sel[0]]
        try:
            add_paths([path])
            set_selected(path)
        except Exception:
            pass
        self._load_p3_category_list(path)

    def _load_p3_category_list(self, path: str) -> None:
        """P2와 동일 엑셀 컬럼으로 카테고리URL목록 표시."""
        self.p3_lib_list.delete(0, "end")
        self._p3_excel_rows = []
        self._p3_lib_texts = []
        self._p3_active_ordinal = 0
        self._p3_current_excel = path
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            self.p3_sel.configure(text=f"엑셀 읽기 실패: {e}")
            return
        if not rows:
            self.p3_sel.configure(text="(빈 엑셀)")
            return
        headers = [str(h or "").strip() for h in rows[0]]
        try:
            # ★요건(2026-08-20): 검색필터명 = "최종 카테고리명" (옛 엑셀은 상위로 폴백)
            label_i = (
                headers.index("최종 카테고리명")
                if "최종 카테고리명" in headers
                else headers.index("상위 최종 카테고리명")
            )
            url_i = headers.index("최종 카테고리 URL주소")
        except ValueError:
            # P3: 검색필터 URL 헤더도 허용
            label_i = next(
                (i for i, h in enumerate(headers) if "카테고리명" in h or "필터" in h),
                0,
            )
            url_i = next(
                (i for i, h in enumerate(headers) if "URL" in h.upper()),
                None,
            )
            if url_i is None:
                self.p3_sel.configure(text="URL 열을 찾지 못했습니다")
                return
        for vals in rows[1:]:
            cells = list(vals) if vals else []
            url = str(cells[url_i] or "").strip() if url_i < len(cells) else ""
            if not url.lower().startswith("http"):
                continue
            label = (
                str(cells[label_i] or "").strip() if label_i < len(cells) else ""
            )
            item = {"label": label, "url": url}
            self._p3_excel_rows.append(item)
            text = f"{len(self._p3_excel_rows):03d}  {label}  |  {url}"
            self._p3_lib_texts.append(text)
            self.p3_lib_list.insert("end", text)
        self.p3_sel.configure(
            text=f"선택: {Path(path).name} · {len(self._p3_excel_rows)}행"
        )

    def _refresh_p3_list(self) -> None:
        cur = self._p3_current_excel
        self._search_p3_xlsx()
        if cur and os.path.isfile(cur):
            self._load_p3_category_list(cur)

    def _clear_p3_log(self) -> None:
        for tv in (
            getattr(self, "p3_main_log", None),
            getattr(self, "p3_sub_log", None),
            getattr(self, "p3_shot_log", None),
        ):
            if tv is not None:
                for item in tv.get_children():
                    tv.delete(item)
        self._p3_sub_by_seq = {}
        self._p3_shots_by_seq = {}
        self._p3_shot_path_by_seq = {}
        self._p3_main_item_by_seq = {}
        self._p3_main_ts_end = {}
        self._p3_seq_by_main_item = {}
        self._p3_selected_seq = None
        self._p3_latest_seq = 0
        self._p3_follow_latest = True

    def _setup_p3_log_tags(self) -> None:
        """MAIN 실행로그 — 단계 성격별 색상 태그 (P2와 동일 팔레트)."""
        tv = self.p3_main_log
        tv.tag_configure("normal", foreground="#0f172a")
        tv.tag_configure("login", foreground="#7c3aed", background="#f5f3ff")
        tv.tag_configure("save", foreground="#5b21b6", background="#f3e8ff")
        tv.tag_configure("err", foreground="#b91c1c", background="#fef2f2")

    def _toggle_p3_log_panels(self) -> None:
        """MAIN / SUB / 스크린샷 체크박스 — 실행로그 3개 패널 표시/숨김.

        표시 순서(위→아래)는 항상 MAIN → SUB → 스크린샷으로 고정.
        """
        panels = (
            (getattr(self, "p3_main_frame", None), bool(self.var_p3_show_main.get())),
            (getattr(self, "p3_sub_frame", None), bool(self.var_p3_show_sub.get())),
            (getattr(self, "p3_shot_frame", None), bool(self.var_p3_show_shot.get())),
        )
        area = getattr(self, "_p3_log_area", None)
        for frame, _show in panels:
            if frame is not None:
                frame.grid_remove()
        row = 0
        for frame, show in panels:
            if frame is None or not show:
                continue
            frame.grid(
                row=row, column=0, sticky="nsew", pady=(0, 0) if row == 0 else (6, 0)
            )
            row += 1
        if area is not None:
            for r in range(3):
                area.rowconfigure(r, weight=1 if r < row else 0)

    def _handle_p3_line(self, message: str) -> None:
        """update_filters.py stdout 한 줄 처리 — MAIN/SUB/스크린샷 3단 그리드에 반영.

        ★요건:
        - MAIN엔 1~7단계만 (오류/완료/중단은 새 MAIN 행을 만들지 않고 SUB에 표시)
        - SUB엔 선택한 단계의 세부내용(텍스트)만
        - 스크린샷엔 그 단계의 캡처(오류/원인 파악용)만 — SUB 하단 별도 그리드
        - 마커 없는 잡다한 줄은 화면에 출력하지 않음
        """
        text = (message or "").rstrip()
        if not text:
            return
        if text.startswith("##P3SHOT##"):
            # 스크린샷 폴더 기억용(「스크린샷 보기」) — 그리드 행은 ##SUB## 줄로 그린다
            # (경로가 그 줄에 함께 들어 있어 도착 순서와 무관하다)
            parts = text.split("##")
            self._capture_p3_shot_dir_from_path(
                parts[2].strip() if len(parts) > 2 else ""
            )
            return
        t, text = strip_timestamp(text)
        parsed = parse_line(text)
        if parsed is None:
            return  # 마커 없는 줄은 화면에 출력하지 않음
        kind = parsed[0]
        if kind == "meta":
            _, field, value = parsed
            if field == "진행":
                try:
                    self._mark_p3_active_row(int(str(value).strip() or "0"))
                except ValueError:
                    pass
            return
        if kind == "main":
            _, seq, n, msg = parsed
            red, msg = split_red(msg)
            if 1 <= n <= 7:
                self._insert_p3_main_row(t, seq, n, msg, red=red)
            else:
                # 오류/완료/중단(90/91/92) — MAIN엔 표시하지 않고, 마지막 실행단계의
                # SUB에 요약을 남긴다 (요건: MAIN은 7단계만).
                target_seq = self._p3_latest_seq or seq
                label = step_label_p3(n)
                tag = step_tag_p3(n)
                self._append_p3_sub_entry(target_seq, t, tag, f"[{label}] {msg}")
        elif kind == "sub":
            _, seq, msg = parsed
            red, msg = split_red(msg)
            shot_ok = re.match(r"^(.*)\s->\s(.+\.png)\s*$", msg)
            shot_fail = re.match(r"^\[샷 실패\]\s*(.*)$", msg)
            if shot_ok:
                label, path = shot_ok.group(1).strip(), shot_ok.group(2).strip()
                self._capture_p3_shot_dir_from_path(path)
                self._append_p3_shot_entry(seq, t, label, path)
            elif shot_fail:
                self._append_p3_shot_entry(
                    seq, t, f"[샷 실패] {shot_fail.group(1).strip()}", ""
                )
            else:
                self._append_p3_sub_entry(seq, t, "err" if red else "info", msg)
        elif kind == "subshot":
            _, seq, path, label = parsed
            self._capture_p3_shot_dir_from_path(path)
            self._append_p3_shot_entry(seq, t, label, path)
        # ##META## 마커는 P3(update_filters.py)에서 사용하지 않음 — 무시

    def _p3_main_ts_for_seq(self, seq: int) -> str | None:
        item = self._p3_main_item_by_seq.get(seq)
        if not item:
            return None
        vals = self.p3_main_log.item(item, "values")
        return vals[0] if vals else None

    def _p3_ts_for_sub(self, seq: int, t: str) -> str:
        """sub/스크린샷 시각 = 현단계 MAIN 진입 ~ 다음 MAIN 진입 (P2와 동일)."""
        if "~" in (t or ""):
            return t
        start = self._p3_main_ts_for_seq(seq)
        if not start:
            return t
        end = self._p3_main_ts_end.get(seq, start)
        return sub_time_range(start, end)

    def _insert_p3_main_row(
        self, t: str, seq: int, n: int, msg: str, *, red: bool = False
    ) -> None:
        if seq > 1:
            self._p3_main_ts_end[seq - 1] = t
            if self._p3_selected_seq == seq - 1:
                self._render_p3_sub_grid(seq - 1)
                self._render_p3_shot_grid(seq - 1)
        tag = "err" if red else main_tag_p3(n, msg)
        item = self.p3_main_log.insert("", "end", values=(t, n, msg), tags=(tag,))
        self._p3_main_item_by_seq[seq] = item
        self._p3_seq_by_main_item[item] = seq
        self._p3_latest_seq = max(self._p3_latest_seq, seq)
        self.p3_main_log.see(item)
        if self._p3_follow_latest:
            self.p3_main_log.selection_set(item)
            self._p3_selected_seq = seq
            self._render_p3_sub_grid(seq)
            self._render_p3_shot_grid(seq)

    def _append_p3_sub_entry(self, seq: int, t: str, kind: str, msg: str) -> None:
        display_t = self._p3_ts_for_sub(seq, t)
        self._p3_sub_by_seq.setdefault(seq, []).append((display_t, kind, msg))
        if self._p3_selected_seq == seq:
            tag = (kind,) if kind in ("err", "done", "stop") else ()
            item = self.p3_sub_log.insert("", "end", values=(display_t, msg), tags=tag)
            self.p3_sub_log.see(item)

    def _render_p3_sub_grid(self, seq: int) -> None:
        for item in self.p3_sub_log.get_children():
            self.p3_sub_log.delete(item)
        for t, kind, msg in self._p3_sub_by_seq.get(seq, []):
            display_t = self._p3_ts_for_sub(seq, t)
            tag = (kind,) if kind in ("err", "done", "stop") else ()
            self.p3_sub_log.insert("", "end", values=(display_t, msg), tags=tag)

    def _append_p3_shot_entry(self, seq: int, t: str, label: str, path: str) -> None:
        display_t = self._p3_ts_for_sub(seq, t)
        entries = self._p3_shots_by_seq.setdefault(seq, [])
        entries.append((display_t, label, path))
        idx = len(entries) - 1
        self._p3_shot_path_by_seq[(seq, idx)] = path
        if self._p3_selected_seq == seq:
            tag = ("shot",) if path else ("err",)
            item = self.p3_shot_log.insert(
                "", "end", values=(display_t, label), tags=tag
            )
            self.p3_shot_log.see(item)

    def _render_p3_shot_grid(self, seq: int) -> None:
        for item in self.p3_shot_log.get_children():
            self.p3_shot_log.delete(item)
        for t, label, path in self._p3_shots_by_seq.get(seq, []):
            display_t = self._p3_ts_for_sub(seq, t)
            tag = ("shot",) if path else ("err",)
            self.p3_shot_log.insert("", "end", values=(display_t, label), tags=tag)

    def _on_p3_main_log_select(self, _evt=None) -> None:
        sel = self.p3_main_log.selection()
        if not sel:
            return
        seq = self._p3_seq_by_main_item.get(sel[0])
        if seq is None:
            return
        self._p3_selected_seq = seq
        self._p3_follow_latest = seq == self._p3_latest_seq
        self._render_p3_sub_grid(seq)
        self._render_p3_shot_grid(seq)

    def _mark_p3_active_row(self, ordinal: int) -> None:
        """엑셀 목록에서 지금 작업 중인 행에 진행 화살표(▶)·적색 표시."""
        lst = getattr(self, "p3_lib_list", None)
        if lst is None:
            return
        texts = getattr(self, "_p3_lib_texts", [])
        prev = getattr(self, "_p3_active_ordinal", 0)
        if prev and prev - 1 < len(texts):
            try:
                lst.delete(prev - 1)
                lst.insert(prev - 1, texts[prev - 1])
                lst.itemconfigure(prev - 1, foreground="#0f172a")
            except tk.TclError:
                pass
        self._p3_active_ordinal = 0
        idx = int(ordinal or 0) - 1
        if idx < 0 or idx >= len(texts):
            return
        try:
            lst.delete(idx)
            lst.insert(idx, f"▶ {texts[idx]}")
            lst.itemconfigure(idx, foreground="#b91c1c")
            lst.see(idx)
        except tk.TclError:
            return
        self._p3_active_ordinal = idx + 1

    def _capture_p3_shot_dir_from_path(self, path: str) -> None:
        try:
            p = Path(path)
            if p.parent.is_dir():
                self._p3_last_shot_dir = p.parent
        except Exception:  # noqa: BLE001
            pass

    def _on_p3_shot_log_double_click(self, _evt=None) -> None:
        """스크린샷 그리드 행 더블클릭 → PNG 열기."""
        sel = self.p3_shot_log.selection()
        if not sel or self._p3_selected_seq is None:
            return
        idx = self.p3_shot_log.index(sel[0])
        path = self._p3_shot_path_by_seq.get((self._p3_selected_seq, idx), "")
        if not path or not os.path.isfile(path):
            return
        try:
            self._p3_last_shot_dir = Path(path).parent
        except Exception:
            pass
        try:
            if os.name == "nt":
                os.startfile(path)  # type: ignore[attr-defined]
            else:
                import subprocess as _sp

                _sp.Popen(["xdg-open", path])
        except Exception as e:  # noqa: BLE001
            messagebox.showerror("열기 실패", f"{path}\n{e}")

    def _show_p3_shot_viewer(self) -> None:
        """P3 실행 스크린샷 뷰어 (저장상품수 전·후 포함)."""
        folder = self._p3_last_shot_dir
        if folder is None or not Path(folder).is_dir():
            folder = latest_p3_shot_dir(ROOT)
        open_shot_viewer(
            self,
            shot_dir=folder,
            root=ROOT,
            prefer_p3=True,
            empty_hint=(
                "P3 스크린샷 폴더가 없습니다.\n"
                "작업시작 후 필터 일치 행에서 단계 샷이 생성됩니다.\n"
                "(5)저장상품수 갱신 전·후 샷 포함)"
            ),
        )

    def _p3_stop_flag(self) -> Path:
        return ROOT / "P3_필터_갱신" / ".filter_stop"

    def _run_p3(self) -> None:
        path = self._p3_current_excel
        if not path:
            sel = self.p3_found_list.curselection()
            if sel:
                path = self._p3_found_paths[sel[0]]
                self._load_p3_category_list(path)
        if not path or not os.path.isfile(path):
            messagebox.showinfo("안내", "엑셀 파일을 선택하세요.")
            return
        mango = self.var_p3_mango_url.get().strip() or p3_update.DEFAULT_MANGO_URL
        if not mango.lower().startswith("http"):
            messagebox.showerror("오류", "더망고 URL은 http(s)로 시작해야 합니다.")
            return
        if self._p3_proc and self._p3_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return
        if not self._p3_excel_rows:
            messagebox.showerror("오류", "엑셀에 처리할 URL 행이 없습니다.")
            return

        update_py = ROOT / "P3_필터_갱신" / "update_filters.py"
        if not update_py.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{update_py}")
            return

        try:
            p3_update.save_mango_url(mango)
        except Exception:
            pass

        try:
            self._p3_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        args = [
            sys.executable,
            str(update_py),
            path,
            "--mango-url",
            mango,
        ]
        try:
            add_paths([path])
            set_selected(path)
        except Exception:
            pass

        self._clear_p3_log()
        self.p3_status.configure(
            text=(
                f"작업시작: {Path(path).name} / 총 {len(self._p3_excel_rows)}행 — "
                "망고 Chrome 창이 뜨는지 확인하세요 (P2와 동일 연동)"
            ),
            fg="#15803d",
        )
        try:
            creationflags = 0
            if os.name == "nt":
                creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            self._p3_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P3_필터_갱신"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p3_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p3_proc,
            args=(self._p3_proc, path),
            daemon=True,
        ).start()

    def _stop_p3(self) -> None:
        proc = self._p3_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p3_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p3_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p3_status.configure(text="작업중단 요청 중…", fg="#b45309")

    def _watch_p3_proc(self, proc: subprocess.Popen, path: str) -> None:
        """update_filters.py stdout 감시 — P2와 동일 ##MAIN##/##SUB## 프로토콜만 그리드에 반영."""
        assert proc.stdout is not None
        buf = b""
        try:
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip("\r").rstrip()
                    if text:
                        self.after(0, lambda t=text: self._handle_p3_line(t))
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip("\r").rstrip()
                if text:
                    self.after(0, lambda t=text: self._handle_p3_line(t))
        except Exception as e:  # noqa: BLE001
            # except 블록을 벗어나면 e 가 사라지므로 메시지를 먼저 만들어 넘긴다
            msg = f"로그 수신 오류: {e}"
            self.after(0, lambda m=msg: self.p3_status.configure(text=m, fg="#b91c1c"))
        code = proc.wait()
        self.after(0, lambda: self._on_p3_finished(path, code))

    def _on_p3_finished(self, path: str, code: int) -> None:
        self._p3_proc = None
        try:
            self._p3_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass
        if code == 0:
            self.p3_status.configure(
                text=f"완료 · {Path(path).name}",
                fg="#15803d",
            )
        else:
            self.p3_status.configure(
                text=f"종료 (exit={code}) · {Path(path).name}",
                fg="#b91c1c",
            )

    def _pick_search_dir(self) -> None:
        d = filedialog.askdirectory(initialdir=self.var_dir.get() or str(Path.home()))
        if d:
            self.var_dir.set(d)
            self._search_xlsx()

    def _search_xlsx(self) -> None:
        """디렉터리의 .xlsx 파일 목록을 리스트박스(+스크롤)에 표시."""
        self.found_list.delete(0, "end")
        self._found_paths = []
        try:
            files = search_xlsx(self.var_dir.get().strip(), self.var_q.get().strip())
        except Exception as e:
            messagebox.showerror("검색 실패", str(e))
            return
        for f in files:
            self._found_paths.append(f["path"])
            self.found_list.insert("end", f["name"])
        self.p2_status.configure(
            text=f"파일 {len(files)}개" if files else "해당 폴더에서 .xlsx 없음",
            fg="#0f172a",
        )
        # 마지막 선택 파일이 목록에 있으면 자동 선택·카테고리URL목록 로드
        data = load()
        last = str(data.get("last_selected") or "").strip()
        if last and last in self._found_paths:
            idx = self._found_paths.index(last)
            self.found_list.selection_clear(0, "end")
            self.found_list.selection_set(idx)
            self.found_list.see(idx)
            self._load_category_url_list(last)
        elif self._found_paths and not self._current_excel_path:
            self.found_list.selection_set(0)
            self.found_list.see(0)

    def _on_found_select(self, _evt=None) -> None:
        sel = self.found_list.curselection()
        if not sel:
            return
        path = self._found_paths[sel[0]]
        self._load_category_url_list(path)

    def _on_found_mousewheel(self, event) -> str:
        if getattr(event, "num", None) == 4 or getattr(event, "delta", 0) > 0:
            self.found_list.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5 or getattr(event, "delta", 0) < 0:
            self.found_list.yview_scroll(1, "units")
        return "break"

    def _add_found(self) -> None:
        """선택한 디렉터리 파일을 열고 카테고리URL목록에 엑셀 전체 행을 표시."""
        sel = list(self.found_list.curselection())
        if not sel:
            messagebox.showinfo("안내", "디렉터리 파일 목록에서 엑셀을 선택하세요.")
            return
        path = self._found_paths[sel[0]]
        add_paths([path])
        self._load_category_url_list(path)
        self.p2_status.configure(
            text=f"카테고리URL목록 로드: {Path(path).name} ({len(self._excel_rows)}행)",
            fg="#15803d",
        )

    def _format_category_row(self, row: dict, *, active: bool = False) -> str:
        mark = "▶ " if active else "   "
        return (
            f"{mark}{row['ordinal']:03d} | {row.get('label', '')} | {row.get('url', '')}"
        )

    def _load_category_url_list(self, path: str) -> None:
        """엑셀 전체 행을 카테고리URL목록 리스트박스에 표시."""
        self.lib_list.delete(0, "end")
        self._excel_rows = []
        self._active_ordinal = 0
        self._current_excel_path = ""
        self._lib_paths = []
        if not path or not os.path.isfile(path):
            self.p2_sel.configure(text="(파일 없음 — 위에서 엑셀을 선택하세요)")
            return
        try:
            rows = read_category_url_rows(path)
        except Exception as e:
            self.p2_sel.configure(text=f"(엑셀 읽기 실패: {e})")
            messagebox.showerror("엑셀 읽기 실패", str(e))
            return
        self._excel_rows = rows
        self._current_excel_path = path
        self._lib_paths = [path]
        for row in rows:
            self.lib_list.insert("end", self._format_category_row(row, active=False))
        set_selected(path)
        self.p2_sel.configure(text=f"{path}  ·  총 {len(rows)}행")
        if rows:
            self.lib_list.see(0)

    def _highlight_active_category_row(self, ordinal: int) -> None:
        """현재 작업 진행중인 행을 적색으로 표시."""
        try:
            ord_i = int(ordinal or 0)
        except (TypeError, ValueError):
            ord_i = 0
        if not self._excel_rows:
            return
        prev = self._active_ordinal
        self._active_ordinal = ord_i
        # 이전 활성 행 복원
        if 1 <= prev <= len(self._excel_rows):
            idx = prev - 1
            self.lib_list.delete(idx)
            self.lib_list.insert(
                idx, self._format_category_row(self._excel_rows[idx], active=False)
            )
            self.lib_list.itemconfig(idx, foreground="#0f172a", background="#ffffff")
        # 새 활성 행 적색
        if 1 <= ord_i <= len(self._excel_rows):
            idx = ord_i - 1
            self.lib_list.delete(idx)
            self.lib_list.insert(
                idx, self._format_category_row(self._excel_rows[idx], active=True)
            )
            self.lib_list.itemconfig(idx, foreground="#b91c1c", background="#fee2e2")
            self.lib_list.see(idx)

    def _refresh_p2_list(self) -> None:
        """파일 목록 새로고침 + 현재 엑셀 카테고리URL목록 재로드."""
        cur = self._current_excel_path
        self._search_xlsx()
        if cur and os.path.isfile(cur):
            self._load_category_url_list(cur)
            if self._active_ordinal:
                self._highlight_active_category_row(self._active_ordinal)
        elif not self._current_excel_path:
            # 보관 라이브러리 last_selected 우선
            data = load()
            last = str(data.get("last_selected") or "").strip()
            if last and os.path.isfile(last):
                self._load_category_url_list(last)

    def _remove_lib(self) -> None:
        path = self._current_excel_path
        if not path:
            sel = self.found_list.curselection()
            if sel:
                path = self._found_paths[sel[0]]
        if not path:
            messagebox.showinfo("안내", "제거할 파일이 없습니다.")
            return
        remove_path(path)
        self._current_excel_path = ""
        self._excel_rows = []
        self._active_ordinal = 0
        self.lib_list.delete(0, "end")
        self.p2_sel.configure(text="(비어 있음 — 위에서 엑셀을 선택하세요)")
        self._search_xlsx()

    def _on_lib_mousewheel(self, event) -> str:
        """카테고리URL목록 스크롤 (Windows/macOS/Linux)."""
        if getattr(event, "num", None) == 4 or getattr(event, "delta", 0) > 0:
            self.lib_list.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5 or getattr(event, "delta", 0) < 0:
            self.lib_list.yview_scroll(1, "units")
        return "break"

    def _clear_p2_log(self) -> None:
        for tv in (getattr(self, "p2_main_log", None), getattr(self, "p2_sub_log", None)):
            if tv is not None:
                for item in tv.get_children():
                    tv.delete(item)
        self._sub_by_seq = {}
        self._shot_path_by_seq = {}
        self._main_item_by_seq = {}
        self._main_ts_end = {}
        self._seq_by_main_item = {}
        self._meta_item_id = None
        self._meta_values = {f: "" for f in META_FIELDS}
        self._selected_seq = None
        self._latest_seq = 0
        self._follow_latest = True
        self._setup_meta_rows()

    def _setup_meta_rows(self) -> None:
        """main 상단 엑셀 진행 정보 — 5항목을 1줄(오렌지)로 표시."""
        tv = getattr(self, "p2_main_log", None)
        if tv is None:
            return
        self._meta_values = {f: "" for f in META_FIELDS}
        line = format_meta_line(self._meta_values)
        self._meta_item_id = tv.insert("", 0, values=("", "엑셀", line), tags=("meta",))

    def _update_meta_row(self, field: str, value: str) -> None:
        # ★요건: 순번 META 삭제 — 진행행 적색은 내부필드 '진행'으로만
        if field in META_INTERNAL_FIELDS:
            try:
                ord_i = int(str(value or "0").strip() or "0")
            except ValueError:
                ord_i = 0
            if ord_i > 0:
                self._highlight_active_category_row(ord_i)
            return
        if field not in META_FIELDS:
            return
        self._meta_values[field] = str(value or "").strip()
        if not self._meta_item_id:
            return
        line = format_meta_line(self._meta_values)
        self.p2_main_log.item(self._meta_item_id, values=("", "엑셀", line))

    def _setup_p2_log_tags(self) -> None:
        """main 실행로그 — 단계 성격별 색상 태그."""
        tv = self.p2_main_log
        tv.tag_configure("meta", foreground="#ea580c", background="#fff7ed")
        tv.tag_configure("normal", foreground="#0f172a")
        tv.tag_configure("login", foreground="#7c3aed", background="#f5f3ff")
        tv.tag_configure("init", foreground="#0f766e", background="#f0fdfa")
        tv.tag_configure("save", foreground="#5b21b6", background="#f3e8ff")
        tv.tag_configure("done", foreground="#166534", background="#dcfce7")

    def _toggle_log_panels(self) -> None:
        """MAIN / SUB 체크박스 — 실행로그 패널 표시/숨김 (MAIN 위 · SUB 아래)."""
        show_main = bool(self.var_show_main.get())
        show_sub = bool(self.var_show_sub.get())
        main_f = getattr(self, "p2_main_frame", None)
        sub_f = getattr(self, "p2_sub_frame", None)
        if main_f is not None:
            main_f.pack_forget()
        if sub_f is not None:
            sub_f.pack_forget()
        if show_main and main_f is not None:
            main_f.pack(fill="both", expand=True)
        if show_sub and sub_f is not None:
            sub_f.pack(
                fill="both",
                expand=True,
                pady=(6, 0) if show_main else (0, 0),
            )

    def _handle_collect_line(self, message: str) -> None:
        """collect.py stdout 한 줄 처리 — main/sub 프로토콜만 그리드에 반영.

        (요건: main엔 1~13단계만, 그 외 잡다한 로그는 화면에 출력하지 않음)
        """
        text = (message or "").rstrip()
        if not text:
            return
        t, text = strip_timestamp(text)
        parsed = parse_line(text)
        if parsed is None:
            return  # 마커 없는 줄은 화면에 출력하지 않음 — 요건 2

        kind = parsed[0]
        if kind == "meta":
            _, field, value = parsed
            self._update_meta_row(field, value)
        elif kind == "main":
            _, seq, n, msg = parsed
            self._insert_main_row(t, seq, n, msg)
        elif kind == "sub":
            _, seq, msg = parsed
            self._append_sub_entry(seq, t, "info", msg)
        elif kind == "subshot":
            _, seq, path, label = parsed
            self._capture_shot_dir_from_path(path)
            self._append_sub_entry(seq, t, "shot", f"[샷] {label} -> {Path(path).name}")
            self._shot_path_by_seq[(seq, len(self._sub_by_seq.get(seq, [])) - 1)] = path

    def _main_ts_for_seq(self, seq: int) -> str | None:
        """main 그리드에 기록된 시각 — sub와 동일하게 맞출 때 사용."""
        item = self._main_item_by_seq.get(seq)
        if not item:
            return None
        vals = self.p2_main_log.item(item, "values")
        return vals[0] if vals else None

    def _ts_for_sub(self, seq: int, t: str) -> str:
        """sub 시각 = 현단계 MAIN 진입 ~ 다음 MAIN 진입."""
        if "~" in (t or ""):
            return t
        start = self._main_ts_for_seq(seq)
        if not start:
            return t
        end = self._main_ts_end.get(seq, start)
        return sub_time_range(start, end)

    def _insert_main_row(self, t: str, seq: int, n: int, msg: str) -> None:
        if seq > 1:
            self._main_ts_end[seq - 1] = t
            if self._selected_seq == seq - 1:
                self._render_sub_grid(seq - 1)
        tag = step_tag(n)
        # step=0 → 엑셀 5필드 한 줄(오렌지). 표시는 sticky META와 동일하게 "엑셀"
        step_label: str | int = "엑셀" if n == 0 else n
        item = self.p2_main_log.insert(
            "", "end", values=(t, step_label, msg), tags=(tag,)
        )
        self._main_item_by_seq[seq] = item
        self._seq_by_main_item[item] = seq
        self._latest_seq = max(self._latest_seq, seq)
        self.p2_main_log.see(item)
        if self._follow_latest:
            self.p2_main_log.selection_set(item)
            self._selected_seq = seq
            self._render_sub_grid(seq)

    def _append_sub_entry(self, seq: int, t: str, kind: str, msg: str) -> None:
        display_t = self._ts_for_sub(seq, t)
        self._sub_by_seq.setdefault(seq, []).append((display_t, kind, msg))
        if self._selected_seq == seq:
            tag = ("shot",) if kind == "shot" else ()
            item = self.p2_sub_log.insert("", "end", values=(display_t, msg), tags=tag)
            self.p2_sub_log.see(item)

    def _render_sub_grid(self, seq: int) -> None:
        for item in self.p2_sub_log.get_children():
            self.p2_sub_log.delete(item)
        for t, kind, msg in self._sub_by_seq.get(seq, []):
            display_t = self._ts_for_sub(seq, t)
            tag = ("shot",) if kind == "shot" else ()
            self.p2_sub_log.insert("", "end", values=(display_t, msg), tags=tag)

    def _on_main_log_select(self, _evt=None) -> None:
        sel = self.p2_main_log.selection()
        if not sel:
            return
        seq = self._seq_by_main_item.get(sel[0])
        if seq is None:
            return
        if self._meta_item_id and sel[0] == self._meta_item_id:
            return
        self._selected_seq = seq
        self._follow_latest = seq == self._latest_seq
        self._render_sub_grid(seq)

    def _on_sub_log_double_click(self, _evt=None) -> None:
        sel = self.p2_sub_log.selection()
        if not sel or self._selected_seq is None:
            return
        idx = self.p2_sub_log.index(sel[0])
        path = self._shot_path_by_seq.get((self._selected_seq, idx))
        if not path:
            return
        p = Path(path)
        if not p.is_file():
            return
        try:
            if os.name == "nt":
                os.startfile(str(p))  # type: ignore[attr-defined]
            else:
                webbrowser.open(p.as_uri())
        except Exception:  # noqa: BLE001
            pass

    def _capture_shot_dir_from_path(self, path: str) -> None:
        try:
            p = Path(path)
            if p.parent.is_dir():
                self._last_shot_dir = p.parent
        except Exception:  # noqa: BLE001
            pass

    def _show_shot_viewer(self) -> None:
        folder = self._last_shot_dir
        if folder is None or not folder.is_dir():
            folder = latest_shot_dir(ROOT)
        open_shot_viewer(self, shot_dir=folder, root=ROOT)

    def _p2_log_ui(self, message: str) -> None:
        self.after(0, lambda: self._handle_collect_line(message))

    def _run_p2(self) -> None:
        path = self._current_excel_path
        if not path:
            sel = self.found_list.curselection()
            if sel:
                path = self._found_paths[sel[0]]
                self._load_category_url_list(path)
        if not path:
            messagebox.showinfo(
                "안내", "디렉터리 파일 목록에서 엑셀을 선택한 뒤 실행하세요."
            )
            return
        if not os.path.isfile(path):
            messagebox.showerror("오류", f"파일 없음:\n{path}")
            return
        # 라이브러리에 없으면 자동 등록 (선택 파일로 바로 실행 가능)
        if not is_in_library(path):
            add_paths([path])
        if self._p2_proc and self._p2_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 수집이 진행 중입니다.")
            return

        # 실행 직전 카테고리URL목록 최신화
        self._load_category_url_list(path)
        if not self._excel_rows:
            messagebox.showerror("오류", "엑셀에 처리할 카테고리URL 행이 없습니다.")
            return

        collect_py = ROOT / "P2" / "collect.py"
        stop_flag = ROOT / "P2" / ".collect_stop"
        try:
            stop_flag.unlink(missing_ok=True)  # type: ignore[call-arg]
        except TypeError:
            if stop_flag.exists():
                try:
                    stop_flag.unlink()
                except OSError:
                    pass
        except OSError:
            pass

        verify = bool(self.var_verify.get())
        args = [
            sys.executable,
            str(collect_py),
            path,
            "50",  # ★요건(2026-08-20): 행당 저장상품수 3 → 50
            "--retries",
            "1",  # ★요건: 엑셀 각 행은 1번 시도로 끝냄(재시도 없음)
            "--yes",
            "--shot-first",
            "2",
        ]
        if verify:
            # ★스크린샷만 1·2행 — 처리 행 수는 엑셀 전체 (max_rows 강제 금지)
            args.append("--verify")

        set_selected(path)
        self._clear_p2_log()
        self._highlight_active_category_row(0)
        mode = "엑셀전체수집·1·2행샷" if verify else "엑셀전체수집"
        self.p2_status.configure(
            text=(
                f"수집 시작 ({mode}): {Path(path).name} "
                f"/ 총 {len(self._excel_rows)}행 — 브라우저에서 직접 로그인하세요"
            ),
            fg="#15803d",
        )

        try:
            # 보드 하단 로그로 stdout 수신 (별도 콘솔 창 없음)
            # Windows 기본 콘솔 코드페이지(CP949)와 UTF-8 혼용 대비:
            # 자식 Python은 UTF-8 강제 + 수신 시 utf-8/cp949 폴백 디코딩
            creationflags = 0
            if os.name == "nt":
                creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            self._p2_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P2"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p2_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p2_proc,
            args=(self._p2_proc, path),
            daemon=True,
        ).start()

    def _stop_flag_path(self) -> Path:
        return ROOT / "P2" / ".collect_stop"

    def _stop_p2(self) -> None:
        """중도 수집 중단 — 화면 실행로그는 지우지 않고 보존."""
        proc = self._p2_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 수집이 없습니다.")
            return
        try:
            self._stop_flag_path().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p2_status.configure(text=f"중단 플래그 기록 실패: {e}", fg="#b91c1c")
        self.p2_status.configure(text="수집 종료 요청 중… (로그 보존)", fg="#b45309")
        threading.Thread(target=self._force_stop_p2, args=(proc,), daemon=True).start()

    def _force_stop_p2(self, proc: subprocess.Popen) -> None:
        """협조적 중단 후 응답 없으면 프로세스 종료."""
        for _ in range(24):  # ~12초
            if proc.poll() is not None:
                return
            time.sleep(0.5)
        if proc.poll() is not None:
            return
        self.after(
            0,
            lambda: self.p2_status.configure(
                text="협조 중단 지연 — 프로세스 강제 종료", fg="#b45309"
            ),
        )
        try:
            proc.terminate()
        except Exception:  # noqa: BLE001
            pass
        time.sleep(1.5)
        if proc.poll() is None:
            try:
                proc.kill()
            except Exception:  # noqa: BLE001
                pass

    @staticmethod
    def _decode_log_bytes(raw: bytes) -> str:
        """Windows CP949 / UTF-8 혼용 stdout을 깨지지 않게 디코딩."""
        if not raw:
            return ""
        data = raw.replace(b"\r\n", b"\n").replace(b"\r", b"\n")
        for enc in ("utf-8", "cp949", "mbcs"):
            try:
                return data.decode(enc)
            except UnicodeDecodeError:
                continue
        return data.decode("utf-8", errors="replace")

    def _watch_p2_proc(self, proc: subprocess.Popen, path: str) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self._p2_log_ui(text)
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self._p2_log_ui(text)
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p2_status.configure(text=f"로그 수신 오류: {e}", fg="#b91c1c"),
            )
        code = proc.wait()
        if code == 0:
            self.after(0, lambda: self._on_p2_finished(True, path, code))
        elif code == 130:
            self.after(0, lambda: self._on_p2_finished(False, path, code, stopped=True))
        else:
            self.after(0, lambda: self._on_p2_finished(False, path, code))

    def _on_p2_finished(
        self,
        ok: bool,
        path: str,
        code: int = 0,
        *,
        stopped: bool = False,
    ) -> None:
        # 중단/완료 모두 실행로그는 그대로 둔다 (_clear_p2_log 호출 없음)
        if stopped:
            self.p2_status.configure(
                text="수집 종료(사용자 중단) — 실행로그 보존됨",
                fg="#b45309",
            )
            return
        if ok:
            self.p2_status.configure(text=f"수집 완료: {path}", fg="#15803d")
            folder = self._last_shot_dir or latest_shot_dir(ROOT)
            if folder and folder.is_dir() and any(folder.glob("*.png")):
                # 1행 전과정 샷이 있으면 바로 보여 줌
                if bool(self.var_verify.get()):
                    open_shot_viewer(self, shot_dir=folder, root=ROOT)
        else:
            self.p2_status.configure(text=f"수집 실패 (exit={code})", fg="#b91c1c")
            folder = self._last_shot_dir or latest_shot_dir(ROOT)
            if folder and folder.is_dir() and any(folder.glob("*.png")):
                if messagebox.askyesno(
                    "스크린샷",
                    f"실패했지만 단계 스크린샷이 있습니다.\n{folder}\n\n지금 볼까요?",
                    parent=self,
                ):
                    open_shot_viewer(self, shot_dir=folder, root=ROOT)

    # ── P3_설정수정_카테고리매핑초기화 ────────────────────────────────
    def _build_p3_reset(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P3_설정수정_카테고리매핑초기화 — 지정 행 범위의 카테고리매핑 설정을 초기화 (되돌릴 수 없음)",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        form = tk.LabelFrame(parent, text="입력", bg="#ffffff", padx=8, pady=6)
        form.pack(fill="x")

        r0 = tk.Frame(form, bg="#ffffff")
        r0.pack(fill="x", pady=3)
        tk.Label(r0, text="사이트명", width=13, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p3rst_site = tk.StringVar(value="")
        self.cbo_p3rst_site = ttk.Combobox(
            r0,
            textvariable=self.var_p3rst_site,
            width=28,
            values=ih.load(P3RST_SITE_HISTORY),
        )
        self.cbo_p3rst_site.pack(side="left")
        self.cbo_p3rst_site.bind(
            "<<ComboboxSelected>>", lambda e: self._on_p3rst_input_picked()
        )
        tk.Label(
            r0,
            text="(비우면 현재 선택 유지)",
            bg="#ffffff",
            fg="#64748b",
            font=("Malgun Gothic", 8),
        ).pack(side="left", padx=6)

        r1 = tk.Frame(form, bg="#ffffff")
        r1.pack(fill="x", pady=3)
        tk.Label(r1, text="작업 URL", width=13, anchor="w", bg="#ffffff").pack(side="left")
        # 기본값을 넣지 않는다 — 목록 화면 URL 은 계정·검색조건마다 달라서
        # 미리 채워두면 엉뚱한 화면(행 0건)에서 작업하게 된다.
        # 대신 이전에 입력한 값을 리스트박스(콤보박스)에서 다시 고를 수 있다.
        self.var_p3rst_url = tk.StringVar(value="")
        self.cbo_p3rst_url = ttk.Combobox(
            r1, textvariable=self.var_p3rst_url, values=ih.load(P3RST_URL_HISTORY)
        )
        self.cbo_p3rst_url.pack(side="left", fill="x", expand=True)
        self.cbo_p3rst_url.bind(
            "<<ComboboxSelected>>", lambda e: self._on_p3rst_input_picked()
        )
        tk.Label(
            r1,
            text="필수",
            bg="#ffffff",
            fg="#b91c1c",
            font=("Malgun Gothic", 8, "bold"),
        ).pack(side="left", padx=6)

        r2 = tk.Frame(form, bg="#ffffff")
        r2.pack(fill="x", pady=3)
        tk.Label(r2, text="작업행 범위", width=13, anchor="w", bg="#ffffff").pack(side="left")
        self.var_p3rst_from = tk.StringVar(value=str(p3_reset_mapping.DEFAULT_ROW_FROM))
        tk.Entry(r2, textvariable=self.var_p3rst_from, width=6).pack(side="left")
        tk.Label(r2, text="부터", bg="#ffffff").pack(side="left", padx=(4, 10))
        self.var_p3rst_to = tk.StringVar(value=str(p3_reset_mapping.DEFAULT_ROW_TO))
        tk.Entry(r2, textvariable=self.var_p3rst_to, width=6).pack(side="left")
        tk.Label(r2, text="까지", bg="#ffffff").pack(side="left", padx=(4, 10))
        tk.Label(
            r2,
            text="※ 위 「작업 URL」 검색결과의 행 번호 기준 (1부터, 양끝 포함)",
            bg="#ffffff",
            fg="#64748b",
            font=("Malgun Gothic", 8),
        ).pack(side="left")

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=8)
        tk.Button(
            actions,
            text="행 목록 확인",
            command=self._check_p3rst_rows,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=(0, 6))
        tk.Button(
            actions,
            text="초기화 시작",
            command=self._run_p3_reset,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left")
        tk.Button(
            actions,
            text="작업중단",
            command=self._stop_p3_reset,
            bg="#6b7280",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=4,
        ).pack(side="left", padx=6)

        tk.Label(
            parent,
            text="⚠ 되돌릴 수 없는 초기화 작업입니다. [행 목록 확인] 으로 대상을 먼저 확인하세요.",
            bg="#f1f5f9",
            fg="#b91c1c",
            font=("Malgun Gothic", 8, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 4))

        log_frame = tk.LabelFrame(parent, text="실행 로그", bg="#ffffff", padx=6, pady=4)
        log_frame.pack(fill="both", expand=True)
        self.p3_reset_log = tk.Text(
            log_frame, height=16, font=("Consolas", 9), wrap="word", bg="#0f172a", fg="#e2e8f0"
        )
        sbr = tk.Scrollbar(log_frame, command=self.p3_reset_log.yview)
        self.p3_reset_log.configure(yscrollcommand=sbr.set)
        self.p3_reset_log.pack(side="left", fill="both", expand=True)
        sbr.pack(side="right", fill="y")

        self.p3_reset_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p3_reset_status.pack(fill="x", pady=4)

    def _p3_reset_stop_flag(self) -> Path:
        return ROOT / "P3_설정수정_카테고리매핑초기화" / ".reset_stop"

    def _append_p3_reset_log(self, line: str) -> None:
        text = (line or "").strip()
        if text.startswith("##MAIN##"):
            text = text[8:]
        self.p3_reset_log.insert("end", text + "\n")
        self.p3_reset_log.see("end")

    def _p3_reset_script(self) -> Path:
        return ROOT / "P3_설정수정_카테고리매핑초기화" / "reset_category_mapping.py"

    def _p3_reset_popen(self, extra_args: list[str]) -> subprocess.Popen:
        args = [sys.executable, str(self._p3_reset_script()), *extra_args]
        creationflags = 0
        if os.name == "nt":
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        return subprocess.Popen(
            args,
            cwd=str(ROOT / "P3_설정수정_카테고리매핑초기화"),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=False,
            bufsize=0,
            env=env,
            creationflags=creationflags,
        )

    def _p3_reset_common_args(self) -> list[str]:
        args: list[str] = []
        site = self.var_p3rst_site.get().strip()
        if site:
            args.extend(["--site-id", site])
        url = self.var_p3rst_url.get().strip()
        if url:
            args.extend(["--list-url", url])
        row_from = self.var_p3rst_from.get().strip() or str(p3_reset_mapping.DEFAULT_ROW_FROM)
        row_to = self.var_p3rst_to.get().strip() or str(p3_reset_mapping.DEFAULT_ROW_TO)
        args.extend(["--row-from", row_from, "--row-to", row_to])
        return args

    def _on_p3rst_input_picked(self) -> None:
        """사이트명·작업 URL 리스트박스에서 값을 고르면 즉시 망고 목록을 조회한다."""
        if self.var_p3rst_url.get().strip():
            self._check_p3rst_rows()

    def _remember_p3rst_inputs(self) -> None:
        site = self.var_p3rst_site.get().strip()
        url = self.var_p3rst_url.get().strip()
        if site:
            self.cbo_p3rst_site.configure(values=ih.remember(P3RST_SITE_HISTORY, site))
        if url:
            self.cbo_p3rst_url.configure(values=ih.remember(P3RST_URL_HISTORY, url))

    def _check_p3rst_rows(self) -> None:
        if self._p3_reset_proc and self._p3_reset_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return
        if not self._p3_reset_script().is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{self._p3_reset_script()}")
            return

        self._remember_p3rst_inputs()
        self.p3_reset_log.delete("1.0", "end")
        self.p3_reset_status.configure(text="행 목록 확인 중…", fg="#0f766e")
        try:
            self._p3_reset_proc = self._p3_reset_popen(
                ["--list-rows", *self._p3_reset_common_args()]
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            return
        threading.Thread(
            target=self._watch_p3_reset_proc, args=(self._p3_reset_proc,), daemon=True
        ).start()

    def _run_p3_reset(self) -> None:
        if self._p3_reset_proc and self._p3_reset_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return
        if not self._p3_reset_script().is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{self._p3_reset_script()}")
            return

        row_from = self.var_p3rst_from.get().strip()
        row_to = self.var_p3rst_to.get().strip()
        if not messagebox.askyesno(
            "되돌릴 수 없는 초기화",
            f"작업 행 {row_from}~{row_to} 의 카테고리매핑 설정을 초기화합니다.\n"
            "이 작업은 되돌릴 수 없습니다. 계속할까요?",
        ):
            return

        try:
            self._p3_reset_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        self._remember_p3rst_inputs()
        self.p3_reset_log.delete("1.0", "end")
        self.p3_reset_status.configure(
            text=f"초기화 시작 — {row_from}~{row_to}행", fg="#b91c1c"
        )
        try:
            self._p3_reset_proc = self._p3_reset_popen(self._p3_reset_common_args())
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p3_reset_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return
        threading.Thread(
            target=self._watch_p3_reset_proc, args=(self._p3_reset_proc,), daemon=True
        ).start()

    def _stop_p3_reset(self) -> None:
        proc = self._p3_reset_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p3_reset_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p3_reset_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p3_reset_status.configure(text="작업중단 요청 중…", fg="#b45309")
        try:
            proc.terminate()
        except Exception:
            pass

    def _watch_p3_reset_proc(self, proc: subprocess.Popen) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self.after(0, lambda t=text: self._append_p3_reset_log(t))
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self.after(0, lambda t=text: self._append_p3_reset_log(t))
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p3_reset_status.configure(text=f"로그 수신 오류: {e}", fg="#b91c1c"),
            )
        code = proc.wait()
        if code == 0:
            self.after(0, lambda: self.p3_reset_status.configure(text="완료", fg="#15803d"))
        else:
            self.after(
                0,
                lambda: self.p3_reset_status.configure(text=f"종료 (exit={code})", fg="#b91c1c"),
            )


def main() -> None:
    app = BoardApp()
    app.mainloop()


if __name__ == "__main__":
    main()
