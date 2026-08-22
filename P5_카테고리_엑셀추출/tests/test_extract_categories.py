"""P5_카테고리_엑셀추출 단위테스트 — 브라우저 없이 파싱·엑셀·선택자 검증."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import extract_categories as ec  # noqa: E402

# 스크린샷 4의 실제 옵션 텍스트
SAMPLE = [
    "- 카테고리를 선택해주세요 -",
    "e쿠폰/모바일상품권 > 교육/어학이용권 > 온라인교육/외국어",
    "e쿠폰/모바일상품권 > 교육/어학이용권 > 자기개발/기타",
    "e쿠폰/모바일상품권 > 기타 상품권 > 구글/아이툰즈/게임",
    "e쿠폰/모바일상품권 > 도넛/아이스크림/분식 > 간식/분식",
    "e쿠폰/모바일상품권 > 미용/뷰티/스파 > 헤어(브랜드샵)",
]


def test_parse_category_path():
    assert ec.parse_category_path("A > B > C") == ["A", "B", "C"]
    assert ec.parse_category_path(" e쿠폰/모바일상품권 >  교육/어학이용권 ") == [
        "e쿠폰/모바일상품권",
        "교육/어학이용권",
    ]
    assert ec.parse_category_path("") == []


def test_parse_handles_other_separators():
    assert ec.parse_category_path("A &gt; B") == ["A", "B"]
    assert ec.parse_category_path("A ＞ B") == ["A", "B"]


def test_placeholder_detection():
    assert ec.is_placeholder("- 카테고리를 선택해주세요 -") is True
    assert ec.is_placeholder("") is True
    assert ec.is_placeholder("e쿠폰/모바일상품권 > 기타") is False


def test_to_row_fills_six_levels():
    row = ec.to_row(["A", "B", "C"], "옥션2.0")
    assert row["마켓"] == "옥션2.0"
    assert row["구분"] == ""
    assert row["1단계"] == "A" and row["2단계"] == "B" and row["3단계"] == "C"
    assert row["4단계"] == "" and row["6단계"] == ""
    assert row["전체경로"] == "A > B > C"


def test_to_row_merges_deeper_than_six():
    row = ec.to_row(["1", "2", "3", "4", "5", "6", "7"], "옥션2.0")
    assert row["5단계"] == "5"
    assert row["6단계"] == "6 > 7"  # 6단계 양식 유지
    assert row["전체경로"].endswith("6 > 7")


def test_build_rows_skips_placeholder_and_duplicates():
    rows = ec.build_rows(SAMPLE + [SAMPLE[1]], "옥션2.0")
    assert len(rows) == 5  # 안내문구 1건 제외 · 중복 1건 제외
    assert rows[0]["1단계"] == "e쿠폰/모바일상품권"
    assert rows[0]["3단계"] == "온라인교육/외국어"


def test_headers_are_category_table_form():
    assert ec.HEADERS == [
        "마켓",
        "구분",
        "1단계",
        "2단계",
        "3단계",
        "4단계",
        "5단계",
        "6단계",
        "전체경로",
    ]


def test_deepest_level():
    assert ec.deepest_level(SAMPLE) == 3
    assert ec.deepest_level(["A > B > C > D > E > F > G"]) == 7


def test_write_excel_roundtrip(tmp_path):
    from openpyxl import load_workbook

    rows = ec.build_rows(SAMPLE, "옥션2.0")
    out = ec.write_excel(rows, tmp_path / "분류표.xlsx")
    assert out.is_file()

    ws = load_workbook(out).active
    assert [c.value for c in ws[1]] == ec.HEADERS
    assert ws.max_row == len(rows) + 1
    assert ws.cell(row=2, column=3).value == "e쿠폰/모바일상품권"  # 마켓·구분 다음
    assert ws.freeze_panes == "A2"


def test_run_from_text(tmp_path):
    src = tmp_path / "목록.txt"
    src.write_text("\n".join(SAMPLE), encoding="utf-8")
    out = tmp_path / "결과.xlsx"
    result = ec.run_from_text(src, market="AUC20", out_path=out)
    assert result.ok is True
    assert result.total == 5
    assert result.deepest == 3
    assert Path(result.excel_path) == out


def test_run_from_text_missing_file(tmp_path):
    result = ec.run_from_text(tmp_path / "없음.txt")
    assert result.ok is False
    assert "읽기 실패" in result.errors[0]


# ── 화면 선택자 (스크린샷 DOM) ────────────────────────────────────


def test_all_view_selectors_match_screenshot_dom():
    sels = ec.all_view_selectors("AUC20")
    joined = " ".join(sels)
    assert "search_category('AUC20'" in joined
    assert "allview" in joined
    assert "mapping_category_AUC20" in joined
    assert "전체카테고리" in joined


def test_list_select_id():
    assert ec.list_select_id("AUC20") == "openmarket_category_search_list_AUC20"


def test_default_url_is_category_set_page():
    assert ec.DEFAULT_URL == (
        "https://tmg1898.cafe24.com/mall/admin/admin_category_set.php?tm=F&ps_ftid=790"
    )


def test_markets_match_screenshots():
    """스크린샷 1~10 의 tr#mapping_category_<코드> 와 표기."""
    assert ec.MARKETS == {
        "AUC20": "옥션2.0",
        "11ST": "11번가",
        "GMK20": "G마켓2.0",
        "SMART": "스마트스토어",
        "COUP": "쿠팡",
        "LTON": "롯데ON",
    }
    assert ec.DEFAULT_MARKET == "AUC20"


def test_all_view_selector_per_market():
    for code in ec.MARKETS:
        joined = " ".join(ec.all_view_selectors(code))
        assert f"search_category('{code}'" in joined
        assert f"mapping_category_{code}" in joined
        assert "allview" in joined


def test_list_select_ids_cover_both_variants():
    """11번가·롯데ON 은 list_ / list2_ 중 보이는 쪽이 다르다."""
    ids = ec.list_select_ids("11ST")
    assert ids == [
        "openmarket_category_search_list_11ST",
        "openmarket_category_search_list2_11ST",
    ]


def test_markets_to_run_all():
    assert ec.markets_to_run("ALL") == list(ec.MARKETS.keys())
    assert ec.markets_to_run("coup") == ["COUP"]
    assert ec.markets_to_run("") == ["AUC20"]


def test_default_excel_path_for_all_markets():
    from datetime import datetime

    p = ec.default_excel_path("ALL", datetime(2026, 8, 22, 11, 50, 0))
    assert p.name == "카테고리분류표_전체마켓_20260822_115000.xlsx"


class DualSelectPage:
    """list_ 는 비어 있고 list2_ 에만 목록이 있는 화면 (11번가 형태)."""

    def __init__(self, options):
        self.options = list(options)
        self.script_ids = None

    def evaluate(self, script, *args):
        self.script_ids = args[0] if args else None
        return {"texts": self.options, "id": "openmarket_category_search_list2_LTON"}

    def wait_for_timeout(self, ms):
        return None


def test_read_options_asks_both_select_ids():
    page = DualSelectPage(SAMPLE)
    texts = ec.read_option_texts(page, "LTON")
    assert len(texts) == 5
    assert page.script_ids == ec.list_select_ids("LTON")


class MarketLoopPage:
    """마켓별로 서로 다른 목록을 주는 화면 — ALL 추출 확인용."""

    def __init__(self):
        self.clicked: list[str] = []
        self.current = ""

    def goto(self, url, **kwargs):
        return None

    def locator(self, selector):
        for code in ec.MARKETS:
            if f"search_category('{code}'" in selector:
                self.current = code
                return _ClickOnce(self, code)
        return _Missing()

    def evaluate(self, script, *args):
        if "search_category" in script:
            return None
        return {
            "texts": [f"{ec.MARKETS[self.current]}대분류 > 중분류"],
            "id": f"openmarket_category_search_list_{self.current}",
        }

    def wait_for_timeout(self, ms):
        return None


class _ClickOnce:
    def __init__(self, page, code):
        self.page = page
        self.code = code

    @property
    def first(self):
        return self

    def count(self):
        return 1

    def click(self, timeout=None):
        self.page.clicked.append(self.code)


class _Missing:
    @property
    def first(self):
        return self

    def count(self):
        return 0


def test_extract_one_clicks_and_reads(monkeypatch):
    monkeypatch.setattr(ec, "T_LIST", 300)  # 1건만 오는 목록에서 대기 단축
    page = MarketLoopPage()
    logs: list[str] = []
    options = ec.extract_one(page, "SMART", progress=logs.append)
    assert page.clicked == ["SMART"]
    assert options == ["스마트스토어대분류 > 중분류"]
    assert any("스마트스토어" in l for l in logs)


class FakePage:
    """전체카테고리 클릭 후 두 번째 조회에서 목록이 채워지는 화면."""

    def __init__(self):
        self.reads = 0
        self.waits = 0

    def evaluate(self, script, *args):
        self.reads += 1
        if self.reads == 1:
            return ["- 카테고리를 선택해주세요 -"]
        return SAMPLE

    def wait_for_timeout(self, ms):
        self.waits += 1


def test_read_option_texts_waits_for_ajax_fill():
    page = FakePage()
    texts = ec.read_option_texts(page, "AUC20")
    assert len(texts) == 5  # 안내문구 제외
    assert page.waits >= 1


def test_default_excel_path_has_market_and_stamp():
    from datetime import datetime

    p = ec.default_excel_path("AUC20", datetime(2026, 8, 22, 11, 30, 0))
    assert p.name == "카테고리분류표_옥션2.0_20260822_113000.xlsx"
    assert p.parent == ec.OUTPUT_DIR


# ── 구현 제외 마켓 (요건 고정) ────────────────────────────────────


def test_excluded_markets_are_not_targets():
    """LFMall · 머스트잇 · 쇼피 · 큐텐(일본) · 플레이오토(EMP) 는 대상 아님."""
    assert set(ec.EXCLUDED_MARKETS) == {
        "LFMALL",
        "MUSTIT",
        "SHOPEE",
        "QOO10JP",
        "PLAYAUTO",
    }
    for code in ec.EXCLUDED_MARKETS:
        assert code not in ec.MARKETS


def test_all_run_skips_excluded_markets():
    codes = ec.markets_to_run("ALL")
    assert codes == ["AUC20", "11ST", "GMK20", "SMART", "COUP", "LTON"]
    assert not set(codes) & set(ec.EXCLUDED_MARKETS)


def test_excluded_market_requested_directly_is_dropped():
    assert ec.markets_to_run("SHOPEE") == []


# ── 카테고리 구분 (11번가 해외/국내 · 롯데ON 해외직구/일반) ────────


def test_variants_of_markets():
    assert ec.variants_of("11ST") == ("해외카테고리", "국내카테고리")
    assert ec.variants_of("LTON") == ("해외직구 카테고리", "일반카테고리")
    assert ec.variants_of("AUC20") == ("",)  # 구분 없음


def test_variant_radio_selectors_match_screenshot_dom():
    sels = ec.variant_radio_selectors("11ST", "국내카테고리")
    joined = " ".join(sels)
    assert 'openmarket_seller_type2_11ST' in joined
    assert "국내카테고리" in joined
    assert "mapping_category_11ST" in joined


def test_build_rows_tags_variant():
    rows = ec.build_rows(SAMPLE, "11번가", "해외카테고리")
    assert all(r["구분"] == "해외카테고리" for r in rows)
    assert rows[0]["마켓"] == "11번가"


class VariantPage:
    """구분 라디오 + 전체카테고리 + 구분별 목록을 주는 화면."""

    def __init__(self, market="11ST"):
        self.market = market
        self.checked: list[str] = []
        self.clicked_all = 0
        self.current = ""

    def locator(self, selector):
        for variant in ec.MARKET_VARIANTS[self.market]:
            if variant in selector and "radio" in selector:
                return _Radio(self, variant)
        if "search_category" in selector or "전체카테고리" in selector:
            return _AllBtn(self)
        return _Missing()

    def evaluate(self, script, *args):
        return {
            "texts": [f"{self.current} 대분류 > 중분류", f"{self.current} 대분류 > 중분류2"],
            "id": "openmarket_category_search_list_LTON",
        }

    def wait_for_timeout(self, ms):
        return None


class _Radio:
    def __init__(self, page, variant, checked=False):
        self.page = page
        self.variant = variant
        self.checked_state = checked
        self.clicks = 0

    @property
    def first(self):
        return self

    def count(self):
        return 1

    def is_checked(self, timeout=None):
        return self.checked_state

    def click(self, timeout=None, force=False):
        self.clicks += 1
        self.page.checked.append(self.variant)
        self.page.current = self.variant

    def check(self, timeout=None):
        self.page.checked.append(self.variant)
        self.page.current = self.variant


class _AllBtn:
    def __init__(self, page):
        self.page = page

    @property
    def first(self):
        return self

    def count(self):
        return 1

    def click(self, timeout=None):
        self.page.clicked_all += 1


def test_extract_one_selects_variant_first():
    page = VariantPage("11ST")
    logs: list[str] = []
    options = ec.extract_one(page, "11ST", variant="국내카테고리", progress=logs.append)
    assert page.checked == ["국내카테고리"]
    assert page.clicked_all == 1
    assert options and all("국내카테고리" in o for o in options)
    assert any("구분 체크: 국내카테고리" in l for l in logs)


def test_extract_one_without_variant_skips_radio():
    page = VariantPage("11ST")
    ec.extract_one(page, "AUC20", progress=None)
    assert page.checked == []          # 라디오 건드리지 않음
    assert page.clicked_all == 1


def test_missing_variant_radio_is_reported():
    logs: list[str] = []
    assert ec.select_variant(_MissingPage(), "11ST", "해외카테고리", progress=logs.append) is False
    assert any("구분 라디오 미검출" in l for l in logs)


class _MissingPage:
    def locator(self, selector):
        return _Missing()

    def wait_for_timeout(self, ms):
        return None


# ── 카테고리설정 화면 전용 탭 (수집조건수정 팝업 침범 금지) ────────


class TabPage:
    def __init__(self, url=""):
        self.url = url
        self.goto_urls: list[str] = []
        self.fronted = False

    def goto(self, url, **kwargs):
        self.goto_urls.append(url)
        self.url = url

    def bring_to_front(self):
        self.fronted = True


class TabContext:
    def __init__(self, pages):
        self.pages = list(pages)
        self.created = 0

    def new_page(self):
        self.created += 1
        pg = TabPage()
        self.pages.append(pg)
        return pg


class HostPage(TabPage):
    def __init__(self, context):
        super().__init__("https://tmg1898.cafe24.com/mall/admin/admin_group_modify.php?ps_mode=modify_filter")
        self.context = context


def test_category_page_opens_new_tab_not_the_popup():
    popup = TabPage(
        "https://tmg1898.cafe24.com/mall/admin/admin_group_modify.php?ps_mode=modify_filter&ps_fuid=724"
    )
    ctx = TabContext([popup])
    host = HostPage(ctx)
    host.context = ctx

    page = ec.open_category_page(host, ec.DEFAULT_URL)

    assert ctx.created == 1              # 새 탭 생성
    assert page is not popup            # 팝업을 쓰지 않는다
    assert popup.goto_urls == []        # 팝업을 이동시키지 않는다
    assert page.goto_urls == [ec.DEFAULT_URL]
    assert page.fronted is True


def test_category_page_reuses_existing_category_tab():
    existing = TabPage(
        "https://tmg1898.cafe24.com/mall/admin/admin_category_set.php?tm=F&ps_ftid=721"
    )
    ctx = TabContext([existing])
    host = HostPage(ctx)
    host.context = ctx

    page = ec.open_category_page(host, ec.DEFAULT_URL)
    assert page is existing
    assert ctx.created == 0
    assert existing.goto_urls == [ec.DEFAULT_URL]


def test_build_category_url_with_ftid():
    assert ec.build_category_url("721") == (
        "https://tmg1898.cafe24.com/mall/admin/admin_category_set.php?tm=F&ps_ftid=721"
    )
    assert ec.build_category_url("") .endswith(f"ps_ftid={ec.DEFAULT_FTID}")


def test_category_page_constant():
    assert ec.CATEGORY_PAGE == "admin_category_set.php"
    assert ec.CATEGORY_PAGE in ec.DEFAULT_URL


# ── 롯데ON 일반카테고리 누락 방지 (보이는 select · 교체 대기) ──────


def test_options_js_prefers_visible_select():
    """숨은 select 에 이전 구분 목록이 남아 있어도 보이는 쪽을 읽는다."""
    js = ec._OPTIONS_JS
    assert "getComputedStyle" in js
    assert "display === 'none'" in js
    assert "offsetParent" in js


def test_fingerprint_detects_same_list():
    a = ["A > B", "A > C"]
    assert ec.fingerprint(a) == ec.fingerprint(list(a))
    assert ec.fingerprint(a) != ec.fingerprint(["X > Y"])
    assert ec.fingerprint(["- 카테고리를 선택해주세요 -"]) == ec.fingerprint([])


class StaleThenFreshPage:
    """구분 전환 직후엔 이전 목록이 남아 있고, 잠시 뒤 새 목록으로 바뀌는 화면."""

    def __init__(self, stale, fresh):
        self.stale = list(stale)
        self.fresh = list(fresh)
        self.reads = 0
        self.waits = 0

    def evaluate(self, script, *args):
        self.reads += 1
        texts = self.stale if self.reads < 3 else self.fresh
        return {"texts": texts, "id": "openmarket_category_search_list_LTON"}

    def wait_for_timeout(self, ms):
        self.waits += 1


def test_read_waits_until_list_changes_after_variant_switch():
    stale = ["해외직구 대분류 > 중분류"]
    fresh = ["일반 대분류 > 중분류", "일반 대분류 > 중분류2"]
    page = StaleThenFreshPage(stale, fresh)
    logs: list[str] = []

    got = ec.read_option_texts(
        page, "LTON", avoid=ec.fingerprint(stale), progress=logs.append
    )
    assert got == fresh          # 이전 구분 목록을 그대로 반환하지 않는다
    assert page.waits >= 1
    assert any("select=" in l for l in logs)


def test_read_without_avoid_returns_first_list():
    page = StaleThenFreshPage(["A > B"], ["C > D"])
    assert ec.read_option_texts(page, "LTON") == ["A > B"]


def test_read_reports_stale_when_never_changes(monkeypatch):
    monkeypatch.setattr(ec, "T_LIST", 400)
    same = ["A > B"]

    class Frozen:
        def evaluate(self, script, *args):
            return {"texts": same, "id": "openmarket_category_search_list_LTON"}

        def wait_for_timeout(self, ms):
            return None

    logs: list[str] = []
    got = ec.read_option_texts(Frozen(), "LTON", avoid=ec.fingerprint(same), progress=logs.append)
    assert got == same  # 그래도 데이터는 넘긴다
    assert any("이전 구분과 동일" in l for l in logs)


def test_already_checked_radio_is_clicked_again():
    """기본 체크된 구분도 클릭해 change_category_list 를 실행시킨다."""

    class PreCheckedPage(VariantPage):
        def __init__(self):
            super().__init__("LTON")
            self.radio = _Radio(self, "일반카테고리", checked=True)

        def locator(self, selector):
            if "일반카테고리" in selector and "radio" in selector:
                return self.radio
            return super().locator(selector)

    page = PreCheckedPage()
    logs: list[str] = []
    assert ec.select_variant(page, "LTON", "일반카테고리", progress=logs.append) is True
    assert page.radio.clicks == 1                       # 클릭으로 onclick 발생
    assert any("재클릭" in l for l in logs)
